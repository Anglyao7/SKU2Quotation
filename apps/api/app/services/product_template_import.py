from __future__ import annotations

import hashlib
import mimetypes
import os
import posixpath
import re
import tempfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from collections.abc import Callable
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation, localcontext
from itertools import product as cartesian_product
from pathlib import Path, PurePosixPath
from urllib.parse import urlsplit
from uuid import UUID, uuid4
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from sqlalchemy import and_, func, or_, select
from sqlalchemy.orm import selectinload

from ..adapters.object_storage import get_object_storage
from ..catalog_operation_models import CatalogImportBatchRow
from ..database import SessionLocal, set_request_context
from ..db_models import ImportJobRow, SupplierRow
from ..file_security_models import MediaObjectRow, WorkerJobRow
from ..identity_models import TenantRow
from ..model_mixins import utcnow
from ..product_center_models import SKU_TEMPLATE_SOURCE_OPTION_KEY, SkuRow
from ..product_supplier_models import ProductCategoryRow, ProductImageRow, ProductRow
from ..public_catalog_models import PublicCatalogOfferRow
from .category_template_import import category_name_key
from .import_progress import publish_runtime_import_progress
from .sku_codes import CatalogSkuCodeAllocator
from .sku_quotas import sku_quota_snapshot


PRODUCT_TEMPLATE_SHEET = "商品列表"
PRODUCT_TEMPLATE_HEADERS = (
    "商品名称",
    "商品分类",
    "商品型号",
    "供应商",
    "商品价格",
    "商品描述",
    "备注",
    "标签",
    "商品图片1",
    "商品图片2",
    "商品图片3",
    "商品图片4",
    "商品图片5",
    "商品图片6",
    "商品图片7",
    "商品图片8",
    "商品图片9",
    "商品图片10",
)
PRODUCT_TEMPLATE_BASE_HEADERS = PRODUCT_TEMPLATE_HEADERS[:8]
PRODUCT_VARIANT_TEMPLATE_HEADERS = (
    "商品名称",
    "分类名称",
    "商品型号",
    "商品价格",
    "商品描述",
    "备注",
    "是否是新品",
    "一箱个数",
    "规格名称",
    "规格价格",
    "商品图片",
)
PRODUCT_MASTER_TEMPLATE_SHEET = "Product"
SKU_DETAIL_TEMPLATE_SHEET = "SKU"
PRODUCT_MASTER_TEMPLATE_HEADERS_V3 = (
    "商品编码",
    "商品名称",
    "商品分类",
    "商品型号",
    "商品价格",
    "商品描述",
    "备注",
    "标签",
    "是否是新品",
    *(f"商品图片{index}" for index in range(1, 11)),
)
SKU_DETAIL_TEMPLATE_HEADERS_V3 = (
    "商品编码",
    "SKU编号",
    "SKU名称",
    "规格名称",
    "规格1名称",
    "规格1值",
    "规格2名称",
    "规格2值",
    "规格3名称",
    "规格3值",
    "供应商",
    "SKU价格",
    "一箱个数",
)
PRODUCT_MASTER_TEMPLATE_HEADERS = (
    "商品编码",
    "商品名称",
    "商品分类",
    "商品型号",
    "商品价格",
    "商品描述",
    "备注",
    "标签",
    *(f"商品图片{index}" for index in range(1, 11)),
)
PRODUCT_MASTER_TEMPLATE_HEADERS_V4 = PRODUCT_MASTER_TEMPLATE_HEADERS
SKU_DETAIL_TEMPLATE_HEADERS_V4 = (
    "商品编码",
    "SKU编号",
    "SKU名称",
    "规格名称",
    "规格1名称",
    "规格1值",
    "规格2名称",
    "规格2值",
    "规格3名称",
    "规格3值",
    "规格4名称",
    "规格4值",
    "规格5名称",
    "规格5值",
    "供应商",
    "SKU价格",
    "毛重",
    "起定数",
    "装箱数",
)
SKU_DETAIL_TEMPLATE_HEADERS = (
    "商品编码",
    "SKU编号",
    "SKU名称",
    "规格1名称",
    *(f"规格1值（{index}）" for index in range(1, 6)),
    "规格2名称",
    *(f"规格2值（{index}）" for index in range(1, 6)),
    "规格3名称",
    *(f"规格3值（{index}）" for index in range(1, 6)),
    "供应商",
    "SKU价格",
    "毛重",
    "起定数",
    "装箱数",
)
TEMPLATE_LAYOUT_SKU_ROWS = "SKU_ROWS"
TEMPLATE_LAYOUT_PRODUCT_VARIANTS = "PRODUCT_VARIANTS"
TEMPLATE_LAYOUT_PRODUCT_SKUS = "PRODUCT_SKUS"
MAX_TEMPLATE_ROWS = 50_000
IMPORT_FLUSH_BATCH_SIZE = 500
# One source row can intentionally define several concrete variants. Keep a
# separate, bounded expansion ceiling so near-limit catalogs are not rejected
# merely because valid color/size candidates expand beyond the source row count.
MAX_EXPANDED_SKUS = 50_000
MAX_ARCHIVE_ENTRIES = 5_000
MAX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
PRICE_QUANTUM = Decimal("0.01")
MAX_UNIT_PRICE = Decimal("999999999999999999.99")
QUANTITY_QUANTUM = Decimal("0.000001")
MAX_SKU_QUANTITY = Decimal("99999999999999.999999")
MAX_TAGS = 20
MAX_TAG_LENGTH = 80
MAX_CATEGORY_NAME_LENGTH = 200
MAX_SUPPLIER_NAME_LENGTH = 300
TEMPLATE_SOURCE_KEY = SKU_TEMPLATE_SOURCE_OPTION_KEY
TEMPLATE_SOURCE_VALUE = "PRODUCT_TEMPLATE"
# A product row without any explicit SKU is represented by one generated
# no-specification SKU. Keep the flag inside the existing internal template
# marker so exports and storefront option rendering continue to hide it.
TEMPLATE_BASE_PRODUCT_FLAG = "base_product"
TEMPLATE_IMAGE_BUCKET = "product-template"
UNCATEGORIZED_CATEGORY_NAME = "未分类"
PRODUCT_IMAGE_COLUMN_OFFSET = 8
PRODUCT_IMAGE_COLUMN_COUNT = 10
MAX_PRODUCT_IMAGE_COLUMN_COUNT = 50
OOXML_WORKBOOK_PART = "xl/workbook.xml"


@dataclass(frozen=True, slots=True)
class ProductTemplateIssue:
    row_number: int | None
    column: str
    code: str
    message: str
    value: str | None = None
    suggestion: str | None = None

    def as_dict(self) -> dict[str, object]:
        return {
            "row_number": self.row_number,
            "column": self.column,
            "code": self.code,
            "message": self.message,
            "value": self.value,
            "suggestion": self.suggestion,
        }


class ProductTemplateValidationError(ValueError):
    def __init__(
        self,
        message: str,
        *,
        issues: tuple[ProductTemplateIssue, ...] = (),
    ) -> None:
        super().__init__(message)
        self.issues = issues


@dataclass(frozen=True, slots=True)
class EmbeddedTemplateImage:
    row_number: int
    image_column: int
    sequence: int
    archive_path: str
    original_filename: str
    content_type: str
    byte_size: int
    sha256: str


@dataclass(frozen=True, slots=True)
class ProductTemplateLayout:
    kind: str
    headers: tuple[str, ...]
    image_column_offset: int
    image_column_count: int


@dataclass(frozen=True, slots=True)
class ProductVariantTemplateCandidate:
    row_number: int
    name: str
    category: str
    product_model: str | None
    product_key: str
    product_price: Decimal
    description: str | None
    note: str | None
    is_new: bool
    units_per_carton: str | None
    specification: str | None
    specification_price: Decimal | None
    image_urls: tuple[str, ...]
    image_url_columns: tuple[int, ...]
    embedded_images: tuple[EmbeddedTemplateImage, ...]


@dataclass(frozen=True, slots=True)
class ProductMasterTemplateCandidate:
    row_number: int
    product_code: str
    name: str
    category: str
    product_model: str | None
    product_price: Decimal
    description: str | None
    note: str | None
    tags: tuple[str, ...]
    is_new: bool
    image_urls: tuple[str, ...]
    image_url_columns: tuple[int, ...]
    embedded_images: tuple[EmbeddedTemplateImage, ...]


@dataclass(frozen=True, slots=True)
class ProductTemplateRow:
    row_number: int
    name: str
    category: str
    product_key: str
    product_model: str | None
    sku_code: str
    sku_name: str
    specification: str | None
    units_per_carton: str | None
    is_new: bool
    schema_version: int
    supplier_name: str | None
    unit_price: Decimal
    description: str | None
    note: str | None
    tags: tuple[str, ...]
    variant_options: tuple[tuple[str, str], ...]
    default_moq: Decimal | None
    gross_weight: Decimal | None
    image_urls: tuple[str, ...]
    image_url_columns: tuple[int, ...]
    embedded_images: tuple[EmbeddedTemplateImage, ...]
    product_only: bool = False


@dataclass(frozen=True, slots=True)
class StoredTemplateImage:
    image_column: int
    sequence: int
    object_key: str
    original_filename: str
    content_type: str
    byte_size: int
    sha256: str
    storage_provider: str
    archive_path: str | None = None


@dataclass(frozen=True, slots=True)
class ProductTemplateParseResult:
    rows: tuple[ProductTemplateRow, ...]
    warnings: tuple[str, ...]
    skipped_rows: int


@dataclass(frozen=True, slots=True)
class ProductTemplateImportResult:
    status: str
    imported: int = 0
    created: int = 0
    updated: int = 0
    unchanged: int = 0
    skipped: int = 0
    warnings: tuple[str, ...] = ()
    issues: tuple[ProductTemplateIssue, ...] = ()
    message: str | None = None


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format(value, ".15g")
    return str(value).strip()


def _issue_value(value: object) -> str | None:
    text = _cell_text(value)
    if not text:
        return None
    return f"{text[:157]}…" if len(text) > 160 else text


def _issue(
    *,
    row_number: int | None,
    column: str,
    code: str,
    message: str,
    value: object = None,
    suggestion: str | None = None,
) -> ProductTemplateIssue:
    return ProductTemplateIssue(
        row_number=row_number,
        column=column,
        code=code,
        message=message,
        value=_issue_value(value),
        suggestion=suggestion,
    )


def _effective_row_limit_error(
    *,
    sheet_label: str,
    record_label: str,
    excluded_rows: str = "表头和空白行",
) -> ProductTemplateValidationError:
    message = (
        f"{sheet_label}最多允许 {MAX_TEMPLATE_ROWS} 条有效{record_label}数据"
        f"（不含{excluded_rows}）。"
    )
    issue = _issue(
        row_number=None,
        column=sheet_label,
        code="ROW_LIMIT_EXCEEDED",
        message=message,
        suggestion=(
            f"请保留不超过 {MAX_TEMPLATE_ROWS} 条有效数据；"
            "更多数据可拆分为多个文件分批导入。"
        ),
    )
    return ProductTemplateValidationError(message, issues=(issue,))


def _validation_summary(issues: list[ProductTemplateIssue]) -> str:
    affected_rows = {
        issue.row_number for issue in issues if issue.row_number is not None
    }
    location = (
        f"，涉及 {len(affected_rows)} 行"
        if affected_rows
        else ""
    )
    first_message = issues[0].message if issues else "文件内容不符合导入要求"
    return (
        f"发现 {len(issues)} 个数据问题{location}，未执行本次商品导入。"
        f"首个问题：{first_message}"
    )


def _decimal(value: object, *, field: str, row_number: int) -> Decimal:
    text = _cell_text(value).replace(",", "")
    if len(text) > 64:
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的{field}超出 Numeric(20,2) 可存储范围。"
        )
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的{field}不是有效数字。"
        ) from exc
    if not number.is_finite() or number < 0:
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的{field}必须是大于或等于 0 的数字。"
        )
    # PostgreSQL Numeric(20,2) permits at most 18 integer digits. Quantize
    # before persistence so SQLite and PostgreSQL apply the exact same rule.
    if number >= MAX_UNIT_PRICE + Decimal("0.005"):
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的{field}超出 Numeric(20,2) 可存储范围。"
        )
    try:
        with localcontext() as context:
            context.prec = 32
            number = number.quantize(PRICE_QUANTUM, rounding=ROUND_HALF_UP)
    except InvalidOperation as exc:
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的{field}超出 Numeric(20,2) 可存储范围。"
        ) from exc
    if number > MAX_UNIT_PRICE:
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的{field}超出 Numeric(20,2) 可存储范围。"
        )
    return number


def _sku_quantity_decimal(
    value: object,
    *,
    field: str,
    row_number: int,
) -> Decimal:
    """Parse non-negative SKU measurements stored as Numeric(20,6)."""

    text = _cell_text(value).replace(",", "")
    if len(text) > 64:
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的{field}超出 Numeric(20,6) 可存储范围。"
        )
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的{field}不是有效数字。"
        ) from exc
    if not number.is_finite() or number < 0:
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的{field}必须是大于或等于 0 的数字。"
        )
    if number >= MAX_SKU_QUANTITY + Decimal("0.0000005"):
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的{field}超出 Numeric(20,6) 可存储范围。"
        )
    try:
        with localcontext() as context:
            context.prec = 32
            number = number.quantize(QUANTITY_QUANTUM, rounding=ROUND_HALF_UP)
    except InvalidOperation as exc:
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的{field}超出 Numeric(20,6) 可存储范围。"
        ) from exc
    if number > MAX_SKU_QUANTITY:
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的{field}超出 Numeric(20,6) 可存储范围。"
        )
    return number


def _decimal_option_text(value: Decimal) -> str:
    normalized = format(value, "f").rstrip("0").rstrip(".")
    return normalized or "0"


def _normalize_sku_code(value: object) -> str:
    """Return the canonical SKU identity used by XLSX and tenant lookup."""

    return _cell_text(value).strip().upper()


def _source_sku_identity(sku: SkuRow) -> str:
    return _normalize_sku_code(sku.source_sku_code or sku.sku_code)


def _generated_sku_code(
    *,
    name: str,
    category: str,
    supplier_name: str | None,
    occurrence: int,
) -> str:
    """Create a deterministic temporary SKU when the source has no model."""

    identity = "\x1f".join(
        (
            " ".join(name.split()).casefold(),
            category.casefold(),
            (supplier_name or "").casefold(),
        )
    )
    digest = hashlib.sha256(identity.encode("utf-8")).hexdigest()[:16].upper()
    base = f"AUTO-{digest}"
    return base if occurrence == 1 else f"{base}-{occurrence}"


def _normalize_supplier_name(value: object, *, row_number: int) -> str | None:
    name = " ".join(_cell_text(value).split())
    if not name:
        return None
    if len(name) > MAX_SUPPLIER_NAME_LENGTH:
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的供应商名称超过 {MAX_SUPPLIER_NAME_LENGTH} 个字符，"
            "未执行本次增量导入。"
        )
    return name


def _normalize_tags(value: object, *, row_number: int) -> tuple[str, ...]:
    text = _cell_text(value)
    if not text:
        return ()
    tags: list[str] = []
    seen: set[str] = set()
    for raw_tag in re.split(r"[,，;；、|/\r\n]+", text):
        tag = raw_tag.strip()
        if not tag:
            continue
        if len(tag) > MAX_TAG_LENGTH:
            raise ProductTemplateValidationError(
                f"第 {row_number} 行的标签\"{tag[:12]}…\"超过 {MAX_TAG_LENGTH} 个字符，"
                "未执行本次增量导入。"
            )
        normalized = tag.casefold()
        if normalized in seen:
            continue
        seen.add(normalized)
        tags.append(tag)
        if len(tags) > MAX_TAGS:
            raise ProductTemplateValidationError(
                f"第 {row_number} 行最多填写 {MAX_TAGS} 个标签，"
                "未执行本次增量导入。"
            )
    return tuple(tags)


def _normalize_category_path(value: object, *, row_number: int) -> tuple[str, ...]:
    """Normalize the template's human-readable category path.

    A single segment creates a level-one category. ``A/B`` creates level one
    ``A`` and level two ``B``. Deeper or empty paths are rejected so imported
    data can never silently create an unsupported hierarchy.
    """

    text = _cell_text(value).replace("／", "/")
    parts = tuple(part.strip() for part in text.split("/"))
    if not text or any(not part for part in parts):
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的商品分类格式无效；请填写“一级分类”或“一级分类/二级分类”。"
        )
    if len(parts) > 2:
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的商品分类超过两级；最多填写“一级分类/二级分类”。"
        )
    if any(len(part) > MAX_CATEGORY_NAME_LENGTH for part in parts):
        raise ProductTemplateValidationError(
            f"第 {row_number} 行的分类名称超过 {MAX_CATEGORY_NAME_LENGTH} 个字符。"
        )
    return parts


def _valid_image_url(value: object) -> str | None:
    url = _cell_text(value)
    if not url:
        return None
    parsed = urlsplit(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc or len(url) > 1024:
        return None
    return url


def _inspect_archive(path: Path) -> None:
    try:
        with ZipFile(path) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ARCHIVE_ENTRIES:
                raise ProductTemplateValidationError("Excel 文件包含过多内部条目，无法安全解析。")
            uncompressed = sum(max(0, entry.file_size) for entry in entries)
            if uncompressed > MAX_UNCOMPRESSED_BYTES:
                raise ProductTemplateValidationError("Excel 解压后体积过大，无法安全解析。")
            if not any(entry.filename.startswith("xl/") for entry in entries):
                raise ProductTemplateValidationError("文件不是有效的 XLSX 工作簿。")
    except BadZipFile as exc:
        raise ProductTemplateValidationError("文件不是有效的 XLSX 工作簿。") from exc


def _xml_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _xml_attribute(element: ET.Element, name: str) -> str | None:
    for key, value in element.attrib.items():
        if _xml_local_name(key) == name:
            return value
    return None


def _first_xml_child(element: ET.Element, name: str) -> ET.Element | None:
    return next(
        (child for child in element if _xml_local_name(child.tag) == name),
        None,
    )


def _relationship_part_path(source_part: str) -> str:
    return posixpath.join(
        posixpath.dirname(source_part),
        "_rels",
        f"{posixpath.basename(source_part)}.rels",
    )


def _resolve_package_target(source_part: str, target: str) -> str | None:
    normalized_target = target.replace("\\", "/").strip()
    if not normalized_target:
        return None
    if normalized_target.startswith("/"):
        resolved = posixpath.normpath(normalized_target.lstrip("/"))
    else:
        resolved = posixpath.normpath(
            posixpath.join(posixpath.dirname(source_part), normalized_target)
        )
    if resolved in {"", ".", ".."} or resolved.startswith("../"):
        return None
    return PurePosixPath(resolved).as_posix()


def _relationship_targets(
    archive: ZipFile,
    *,
    source_part: str,
) -> dict[str, str]:
    rels_path = _relationship_part_path(source_part)
    if rels_path not in archive.namelist():
        return {}
    try:
        root = ET.fromstring(archive.read(rels_path))
    except (ET.ParseError, KeyError):
        return {}
    targets: dict[str, str] = {}
    for relationship in root:
        if _xml_local_name(relationship.tag) != "Relationship":
            continue
        if str(relationship.attrib.get("TargetMode", "")).casefold() == "external":
            continue
        relationship_id = relationship.attrib.get("Id")
        target = relationship.attrib.get("Target")
        resolved = (
            _resolve_package_target(source_part, target)
            if target is not None
            else None
        )
        if relationship_id and resolved:
            targets[relationship_id] = resolved
    return targets


def _zip_member_sha256(archive: ZipFile, member: str) -> str:
    digest = hashlib.sha256()
    with archive.open(member) as source:
        while chunk := source.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _embedded_image_content_type(member: str) -> str | None:
    content_type, _encoding = mimetypes.guess_type(member)
    if content_type and content_type.startswith("image/"):
        return content_type
    return None


def _extract_embedded_template_images(
    path: Path,
    *,
    sheet_name: str,
    image_column_offset: int,
    image_column_count: int,
) -> tuple[dict[int, tuple[EmbeddedTemplateImage, ...]], tuple[str, ...]]:
    """Read XLSX drawing relationships without materializing every image.

    Excel drawing anchors are zero-based. The caller supplies the detected
    product worksheet, the first image column and its sequential column count.
    This supports both the original I:R image area and the product/variant
    template whose single image column is K. Keeping only archive references
    in the parse result prevents a large workbook from duplicating all image
    bytes in worker memory.
    """

    warnings: list[str] = []
    grouped: dict[int, list[EmbeddedTemplateImage]] = defaultdict(list)
    try:
        with ZipFile(path) as archive:
            archive_names = set(archive.namelist())
            if OOXML_WORKBOOK_PART not in archive_names:
                return {}, ()
            try:
                workbook_root = ET.fromstring(archive.read(OOXML_WORKBOOK_PART))
            except ET.ParseError:
                return {}, ("工作簿图片关系无法读取，已忽略内嵌图片。",)

            sheet_relationships = _relationship_targets(
                archive,
                source_part=OOXML_WORKBOOK_PART,
            )
            sheet_part: str | None = None
            for element in workbook_root.iter():
                if (
                    _xml_local_name(element.tag) == "sheet"
                    and element.attrib.get("name") == sheet_name
                ):
                    relationship_id = _xml_attribute(element, "id")
                    sheet_part = (
                        sheet_relationships.get(relationship_id)
                        if relationship_id
                        else None
                    )
                    break
            if sheet_part is None or sheet_part not in archive_names:
                return {}, ()

            try:
                sheet_root = ET.fromstring(archive.read(sheet_part))
            except ET.ParseError:
                return {}, ("商品工作表图片关系无法读取，已忽略内嵌图片。",)
            drawing_relationships = _relationship_targets(
                archive,
                source_part=sheet_part,
            )
            drawing_parts: list[str] = []
            for element in sheet_root.iter():
                if _xml_local_name(element.tag) != "drawing":
                    continue
                relationship_id = _xml_attribute(element, "id")
                drawing_part = (
                    drawing_relationships.get(relationship_id)
                    if relationship_id
                    else None
                )
                if drawing_part and drawing_part in archive_names:
                    drawing_parts.append(drawing_part)

            ignored_anchors = 0
            unreadable_images = 0
            drawing_sequence = 0
            for drawing_part in drawing_parts:
                try:
                    drawing_root = ET.fromstring(archive.read(drawing_part))
                except ET.ParseError:
                    unreadable_images += 1
                    continue
                image_relationships = _relationship_targets(
                    archive,
                    source_part=drawing_part,
                )
                for anchor in drawing_root:
                    if _xml_local_name(anchor.tag) not in {
                        "oneCellAnchor",
                        "twoCellAnchor",
                    }:
                        continue
                    from_anchor = _first_xml_child(anchor, "from")
                    if from_anchor is None:
                        unreadable_images += 1
                        continue
                    row_element = _first_xml_child(from_anchor, "row")
                    column_element = _first_xml_child(from_anchor, "col")
                    if (
                        row_element is None
                        or column_element is None
                        or row_element.text is None
                        or column_element.text is None
                    ):
                        unreadable_images += 1
                        continue
                    try:
                        row_number = int(row_element.text) + 1
                        zero_based_column = int(column_element.text)
                    except ValueError:
                        unreadable_images += 1
                        continue

                    blip = next(
                        (
                            element
                            for element in anchor.iter()
                            if _xml_local_name(element.tag) == "blip"
                        ),
                        None,
                    )
                    relationship_id = (
                        _xml_attribute(blip, "embed") if blip is not None else None
                    )
                    media_part = (
                        image_relationships.get(relationship_id)
                        if relationship_id
                        else None
                    )
                    if media_part is None:
                        continue
                    if not (
                        image_column_offset
                        <= zero_based_column
                        < image_column_offset + image_column_count
                    ) or row_number < 2:
                        ignored_anchors += 1
                        continue
                    if media_part not in archive_names:
                        unreadable_images += 1
                        continue
                    content_type = _embedded_image_content_type(media_part)
                    if content_type is None:
                        unreadable_images += 1
                        continue
                    info = archive.getinfo(media_part)
                    if info.file_size <= 0:
                        unreadable_images += 1
                        continue
                    drawing_sequence += 1
                    grouped[row_number].append(
                        EmbeddedTemplateImage(
                            row_number=row_number,
                            image_column=(
                                zero_based_column - image_column_offset + 1
                            ),
                            sequence=drawing_sequence,
                            archive_path=media_part,
                            original_filename=PurePosixPath(media_part).name[:500],
                            content_type=content_type,
                            byte_size=info.file_size,
                            sha256=_zip_member_sha256(archive, media_part),
                        )
                    )

            if ignored_anchors:
                image_area = (
                    "商品图片列"
                    if image_column_count == 1
                    else f"商品图片1-{image_column_count}列"
                )
                warnings.append(
                    f"有 {ignored_anchors} 张内嵌图片未放在{image_area}，已忽略。"
                )
            if unreadable_images:
                warnings.append(
                    f"有 {unreadable_images} 张内嵌图片格式或关系异常，已忽略。"
                )
    except BadZipFile:
        return {}, ()

    return (
        {
            row_number: tuple(
                sorted(
                    images,
                    key=lambda image: (image.image_column, image.sequence),
                )
            )
            for row_number, images in grouped.items()
        },
        tuple(warnings),
    )


def _product_image_column_count(sheet: object) -> int | None:
    """Return the compatible sequential image-column count for a worksheet."""

    max_column = getattr(sheet, "max_column", None) or len(PRODUCT_TEMPLATE_HEADERS)
    header_values = [
        _cell_text(sheet.cell(row=1, column=index).value)
        for index in range(1, max_column + 1)
    ]
    last_header_index = next(
        (
            index
            for index in range(len(header_values), 0, -1)
            if header_values[index - 1]
        ),
        0,
    )
    if last_header_index < len(PRODUCT_TEMPLATE_HEADERS):
        return None
    effective_headers = tuple(header_values[:last_header_index])
    if effective_headers[: len(PRODUCT_TEMPLATE_BASE_HEADERS)] != (
        PRODUCT_TEMPLATE_BASE_HEADERS
    ):
        return None
    image_column_count = (
        len(effective_headers) - len(PRODUCT_TEMPLATE_BASE_HEADERS)
    )
    if not (
        PRODUCT_IMAGE_COLUMN_COUNT
        <= image_column_count
        <= MAX_PRODUCT_IMAGE_COLUMN_COUNT
    ):
        return None
    expected_headers = PRODUCT_TEMPLATE_BASE_HEADERS + tuple(
        f"商品图片{index}" for index in range(1, image_column_count + 1)
    )
    if effective_headers != expected_headers:
        return None
    return image_column_count


def _product_template_layout(sheet: object | None) -> ProductTemplateLayout | None:
    if sheet is None:
        return None
    max_column = getattr(sheet, "max_column", None) or len(PRODUCT_TEMPLATE_HEADERS)
    header_values = tuple(
        _cell_text(sheet.cell(row=1, column=index).value)
        for index in range(1, max_column + 1)
    )
    last_header_index = next(
        (
            index
            for index in range(len(header_values), 0, -1)
            if header_values[index - 1]
        ),
        0,
    )
    effective_headers = header_values[:last_header_index]
    if effective_headers == PRODUCT_VARIANT_TEMPLATE_HEADERS:
        return ProductTemplateLayout(
            kind=TEMPLATE_LAYOUT_PRODUCT_VARIANTS,
            headers=PRODUCT_VARIANT_TEMPLATE_HEADERS,
            image_column_offset=10,
            image_column_count=1,
        )
    image_column_count = _product_image_column_count(sheet)
    if image_column_count is None:
        return None
    return ProductTemplateLayout(
        kind=TEMPLATE_LAYOUT_SKU_ROWS,
        headers=PRODUCT_TEMPLATE_BASE_HEADERS
        + tuple(
            f"商品图片{index}"
            for index in range(1, image_column_count + 1)
        ),
        image_column_offset=PRODUCT_IMAGE_COLUMN_OFFSET,
        image_column_count=image_column_count,
    )


def _product_sheet_has_data(
    sheet: object,
    layout: ProductTemplateLayout,
) -> bool:
    max_row = getattr(sheet, "max_row", None) or 1
    if max_row < 2:
        return False
    return any(
        any(_cell_text(value) for value in row)
        for row in sheet.iter_rows(
            min_row=2,
            max_row=max_row,
            max_col=len(layout.headers),
            values_only=True,
        )
    )


def _closest_product_sheet(
    workbook: object,
) -> tuple[object, ProductTemplateLayout] | None:
    """Find one near-compatible sheet so header errors stay field-specific."""

    scored_candidates: list[tuple[int, object, ProductTemplateLayout]] = []
    for sheet in workbook.worksheets:
        max_column = getattr(sheet, "max_column", None) or 1
        header_values = tuple(
            _cell_text(sheet.cell(row=1, column=index).value)
            for index in range(1, max_column + 1)
        )
        last_header_index = next(
            (
                index
                for index in range(len(header_values), 0, -1)
                if header_values[index - 1]
            ),
            0,
        )
        possible_image_count = last_header_index - len(PRODUCT_TEMPLATE_BASE_HEADERS)
        sku_image_count = (
            possible_image_count
            if PRODUCT_IMAGE_COLUMN_COUNT
            <= possible_image_count
            <= MAX_PRODUCT_IMAGE_COLUMN_COUNT
            else PRODUCT_IMAGE_COLUMN_COUNT
        )
        layouts = (
            ProductTemplateLayout(
                kind=TEMPLATE_LAYOUT_SKU_ROWS,
                headers=PRODUCT_TEMPLATE_BASE_HEADERS
                + tuple(
                    f"商品图片{index}"
                    for index in range(1, sku_image_count + 1)
                ),
                image_column_offset=PRODUCT_IMAGE_COLUMN_OFFSET,
                image_column_count=sku_image_count,
            ),
            ProductTemplateLayout(
                kind=TEMPLATE_LAYOUT_PRODUCT_VARIANTS,
                headers=PRODUCT_VARIANT_TEMPLATE_HEADERS,
                image_column_offset=10,
                image_column_count=1,
            ),
        )
        for layout in layouts:
            score = sum(
                actual == expected
                for actual, expected in zip(
                    header_values,
                    layout.headers,
                    strict=False,
                )
            )
            if score >= max(3, len(layout.headers) // 2):
                scored_candidates.append((score, sheet, layout))

    best_score = max((score for score, _sheet, _layout in scored_candidates), default=0)
    best_candidates = [
        (sheet, layout)
        for score, sheet, layout in scored_candidates
        if score == best_score
    ]
    if len(best_candidates) == 1:
        return best_candidates[0]
    return None


def _select_product_sheet(
    workbook: object,
) -> tuple[object, ProductTemplateLayout, str | None]:
    """Select a supported product worksheet by structure, never by title."""

    compatible_sheets = [
        (sheet, layout)
        for sheet in workbook.worksheets
        if (layout := _product_template_layout(sheet)) is not None
    ]
    populated_sheets = [
        (sheet, layout)
        for sheet, layout in compatible_sheets
        if _product_sheet_has_data(sheet, layout)
    ]
    candidates = populated_sheets or compatible_sheets
    if len(candidates) == 1:
        sheet, layout = candidates[0]
        behavior = (
            "相同商品的多行规格会合并为一个商品，并生成独立 SKU。"
            if layout.kind == TEMPLATE_LAYOUT_PRODUCT_VARIANTS
            else "每一行仍按一个 SKU 导入。"
        )
        sheet_warning = (
            None
            if sheet.title == PRODUCT_TEMPLATE_SHEET
            else f"已根据列结构识别工作表“{sheet.title}”；{behavior}"
        )
        return (
            sheet,
            layout,
            sheet_warning,
        )
    if len(candidates) > 1:
        sheet_names = "、".join(
            f"“{sheet.title}”" for sheet, _layout in candidates
        )
        issue = _issue(
            row_number=None,
            column="工作表",
            code="SHEET_AMBIGUOUS",
            message=f"发现多个符合商品列结构的数据页：{sheet_names}。",
            suggestion=(
                "请只保留一个商品数据页，或拆分为多个文件分别导入；"
                "工作表名称不限。"
            ),
        )
        raise ProductTemplateValidationError(issue.message, issues=(issue,))

    closest_sheet = _closest_product_sheet(workbook)
    if closest_sheet is not None:
        sheet, layout = closest_sheet
        return (
            sheet,
            layout,
            None,
        )

    available_sheets = "、".join(
        f"“{sheet_name}”" for sheet_name in workbook.sheetnames
    )
    issue = _issue(
        row_number=None,
        column="工作表",
        code="SHEET_MISSING",
        message=(
            f"未找到包含受支持商品列结构的数据页。现有工作表："
            f"{available_sheets or '无'}。"
        ),
        suggestion=(
            "工作表名称不限。请保留第一行的商品列名与顺序；支持标准 SKU "
            "列结构，或“商品名称、分类名称、商品型号、商品价格、商品描述、"
            "备注、是否是新品、一箱个数、规格名称、规格价格、商品图片”"
            "的多规格列结构。"
        ),
    )
    raise ProductTemplateValidationError(issue.message, issues=(issue,))


def _variant_product_key(
    *,
    name: str,
    category: str,
    product_model: str | None,
) -> str:
    if product_model:
        return f"MODEL:{_normalize_sku_code(product_model)}"
    normalized = "\x1f".join(
        (
            " ".join(name.split()).casefold(),
            category.casefold(),
        )
    )
    return f"AUTO:{hashlib.sha256(normalized.encode('utf-8')).hexdigest().upper()}"


def _variant_sku_code(
    *,
    product_key: str,
    product_model: str | None,
    specification: str | None,
) -> str:
    base = (
        _normalize_sku_code(product_model)
        if product_model
        else f"AUTO-{hashlib.sha256(product_key.encode('utf-8')).hexdigest()[:16].upper()}"
    )
    if not specification:
        return base[:160]
    specification_digest = hashlib.sha256(
        " ".join(specification.split()).casefold().encode("utf-8")
    ).hexdigest()[:10].upper()
    return f"{base[:149]}-{specification_digest}"


def _new_product_flag(
    value: object,
    *,
    row_number: int,
) -> bool:
    normalized = _cell_text(value).strip().casefold()
    if not normalized:
        return False
    if normalized in {"是", "true", "1", "yes", "y", "新品"}:
        return True
    if normalized in {"否", "false", "0", "no", "n"}:
        return False
    raise ProductTemplateValidationError(
        f"第 {row_number} 行的是否是新品应填写“是”或“否”。"
    )


def _is_variant_template_instruction_row(values: tuple[object, ...]) -> bool:
    name = _cell_text(values[0])
    if name.startswith("以上均为例子"):
        return True
    return (
        name == "商品名称"
        and any("可填" in _cell_text(value) for value in values[1:])
    )


def _parse_product_variant_rows(
    sheet: object,
    *,
    effective_headers: tuple[str, ...],
    embedded_images_by_row: dict[int, tuple[EmbeddedTemplateImage, ...]],
    embedded_image_warnings: tuple[str, ...],
    sheet_warning: str | None,
    progress_callback: Callable[[int, int], None] | None,
) -> ProductTemplateParseResult:
    candidates: list[ProductVariantTemplateCandidate] = []
    warnings: list[str] = [
        "已识别商品+规格模板：相同商品型号的多行会合并为一个商品。"
    ]
    if sheet_warning is not None:
        warnings.insert(0, sheet_warning)
    warnings.extend(embedded_image_warnings)
    issues: list[ProductTemplateIssue] = []
    skipped_rows = 0
    generated_model_rows = 0
    visited_rows: set[int] = set()
    effective_data_rows = 0
    total_rows = max(0, (sheet.max_row or 1) - 1)
    progress_interval = max(100, total_rows // 100) if total_rows else 100

    for row_number, raw_values in enumerate(
        sheet.iter_rows(
            min_row=2,
            max_col=len(effective_headers),
            values_only=True,
        ),
        start=2,
    ):
        values = tuple(raw_values)
        processed_rows = row_number - 1
        if progress_callback is not None and (
            processed_rows == 1
            or processed_rows % progress_interval == 0
            or processed_rows == total_rows
        ):
            progress_callback(processed_rows, total_rows)
        embedded_images = embedded_images_by_row.get(row_number, ())
        if not any(_cell_text(value) for value in values) and not embedded_images:
            continue
        visited_rows.add(row_number)
        if _is_variant_template_instruction_row(values):
            skipped_rows += 1
            continue
        effective_data_rows += 1
        if effective_data_rows > MAX_TEMPLATE_ROWS:
            raise _effective_row_limit_error(
                sheet_label="商品工作表",
                record_label="商品",
                excluded_rows="表头、说明行和空白行",
            )

        row_issues: list[ProductTemplateIssue] = []
        formula_indexes = {
            index
            for index, value in enumerate(values)
            if isinstance(value, str) and value.lstrip().startswith("=")
        }
        for index in sorted(formula_indexes):
            column = effective_headers[index]
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column=column,
                    code="FORMULA_NOT_ALLOWED",
                    message=f"第 {row_number} 行的{column}包含公式。",
                    value=values[index],
                    suggestion="请复制计算结果并粘贴为固定值后重新导入。",
                )
            )

        name = _cell_text(values[0])
        if not name and 0 not in formula_indexes:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="商品名称",
                    code="REQUIRED_VALUE_MISSING",
                    message=f"第 {row_number} 行缺少商品名称。",
                    suggestion="请填写商品名称；这是该模板唯一必填的商品字段。",
                )
            )
        elif len(name) > 500 and 0 not in formula_indexes:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="商品名称",
                    code="VALUE_TOO_LONG",
                    message=f"第 {row_number} 行的商品名称超过 500 个字符。",
                    value=values[0],
                    suggestion="请缩短商品名称，详细信息可填写在“商品描述”中。",
                )
            )

        category = UNCATEGORIZED_CATEGORY_NAME
        category_text = _cell_text(values[1])
        if category_text and 1 not in formula_indexes:
            try:
                category = "/".join(
                    _normalize_category_path(
                        category_text,
                        row_number=row_number,
                    )
                )
            except ProductTemplateValidationError as exc:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="分类名称",
                        code="CATEGORY_INVALID",
                        message=str(exc),
                        value=values[1],
                        suggestion="请填写“一级分类”或“一级分类/二级分类”，最多两级。",
                    )
                )

        raw_model = _cell_text(values[2])
        product_model = raw_model or None
        if len(raw_model) > 160 and 2 not in formula_indexes:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="商品型号",
                    code="VALUE_TOO_LONG",
                    message=f"第 {row_number} 行的商品型号超过 160 个字符。",
                    value=values[2],
                    suggestion="请将商品型号缩短至 160 个字符以内。",
                )
            )
        product_key = _variant_product_key(
            name=name,
            category=category,
            product_model=product_model,
        )
        if product_model is None:
            generated_model_rows += 1

        product_price = Decimal("0.00")
        if _cell_text(values[3]) and 3 not in formula_indexes:
            try:
                product_price = _decimal(
                    values[3],
                    field="商品价格",
                    row_number=row_number,
                )
            except ProductTemplateValidationError as exc:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="商品价格",
                        code="PRICE_INVALID",
                        message=str(exc),
                        value=values[3],
                        suggestion="价格可以留空（系统按 0 处理），或填写大于等于 0 的数字。",
                    )
                )

        is_new = False
        if 6 not in formula_indexes:
            try:
                is_new = _new_product_flag(values[6], row_number=row_number)
            except ProductTemplateValidationError as exc:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="是否是新品",
                        code="NEW_PRODUCT_FLAG_INVALID",
                        message=str(exc),
                        value=values[6],
                        suggestion="可以留空；需要标记新品时填写“是”，否则填写“否”。",
                    )
                )

        specification = _cell_text(values[8]) or None
        if specification is not None and len(specification) > 500:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="规格名称",
                    code="VALUE_TOO_LONG",
                    message=f"第 {row_number} 行的规格名称超过 500 个字符。",
                    value=values[8],
                    suggestion="请缩短规格名称。",
                )
            )
        specification_price: Decimal | None = None
        if _cell_text(values[9]) and 9 not in formula_indexes:
            try:
                specification_price = _decimal(
                    values[9],
                    field="规格价格",
                    row_number=row_number,
                )
            except ProductTemplateValidationError as exc:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="规格价格",
                        code="PRICE_INVALID",
                        message=str(exc),
                        value=values[9],
                        suggestion="规格价格可以留空，或填写大于等于 0 的数字。",
                    )
                )

        image_urls: list[str] = []
        image_url_columns: list[int] = []
        if _cell_text(values[10]) and 10 not in formula_indexes:
            image_url = _valid_image_url(values[10])
            if image_url is None:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="商品图片",
                        code="IMAGE_URL_INVALID",
                        message=f"第 {row_number} 行商品图片不是有效的 HTTP(S) 链接。",
                        value=values[10],
                        suggestion=(
                            "图片可以留空或直接插入该单元格位置；"
                            "填写文本时请使用可公开访问的 http:// 或 https:// 地址。"
                        ),
                    )
                )
            else:
                image_urls.append(image_url)
                image_url_columns.append(1)

        if row_issues:
            issues.extend(row_issues)
            continue
        candidates.append(
            ProductVariantTemplateCandidate(
                row_number=row_number,
                name=name,
                category=category,
                product_model=product_model,
                product_key=product_key,
                product_price=product_price,
                description=_cell_text(values[4]) or None,
                note=_cell_text(values[5]) or None,
                is_new=is_new,
                units_per_carton=_cell_text(values[7]) or None,
                specification=specification,
                specification_price=specification_price,
                image_urls=tuple(image_urls),
                image_url_columns=tuple(image_url_columns),
                embedded_images=embedded_images,
            )
        )

    grouped: dict[str, list[ProductVariantTemplateCandidate]] = defaultdict(list)
    for candidate in candidates:
        grouped[candidate.product_key].append(candidate)

    rows: list[ProductTemplateRow] = []
    first_row_by_sku: dict[str, int] = {}
    for group in grouped.values():
        first = group[0]
        normalized_names = {
            " ".join(candidate.name.split()).casefold() for candidate in group
        }
        normalized_categories = {
            candidate.category.casefold() for candidate in group
        }
        if len(normalized_names) > 1 or len(normalized_categories) > 1:
            for candidate in group[1:]:
                if (
                    " ".join(candidate.name.split()).casefold()
                    == " ".join(first.name.split()).casefold()
                    and candidate.category.casefold() == first.category.casefold()
                ):
                    continue
                issues.append(
                    _issue(
                        row_number=candidate.row_number,
                        column="商品型号",
                        code="PRODUCT_GROUP_CONFLICT",
                        message=(
                            f"第 {candidate.row_number} 行与第 {first.row_number} 行"
                            "使用相同商品型号，但商品名称或分类不同。"
                        ),
                        value=candidate.product_model,
                        suggestion="同一商品型号的所有规格行请使用相同商品名称和分类。",
                    )
                )
            continue

        seen_specs: dict[str, int] = {}
        if len(group) > 1:
            for candidate in group:
                if not candidate.specification:
                    issues.append(
                        _issue(
                            row_number=candidate.row_number,
                            column="规格名称",
                            code="SPECIFICATION_REQUIRED",
                            message=(
                                f"商品“{first.name}”包含多行规格，"
                                f"第 {candidate.row_number} 行缺少规格名称。"
                            ),
                            suggestion="多规格商品的每一行都需要填写唯一的规格名称。",
                        )
                    )
                    continue
                normalized_spec = " ".join(
                    candidate.specification.split()
                ).casefold()
                if normalized_spec in seen_specs:
                    issues.append(
                        _issue(
                            row_number=candidate.row_number,
                            column="规格名称",
                            code="SPECIFICATION_DUPLICATE",
                            message=(
                                f"第 {candidate.row_number} 行规格名称与第 "
                                f"{seen_specs[normalized_spec]} 行重复。"
                            ),
                            value=candidate.specification,
                            suggestion="同一商品下的规格名称必须唯一。",
                        )
                    )
                else:
                    seen_specs[normalized_spec] = candidate.row_number

        for candidate in group:
            sku_code = _variant_sku_code(
                product_key=candidate.product_key,
                product_model=candidate.product_model,
                specification=candidate.specification,
            )
            if sku_code in first_row_by_sku:
                issues.append(
                    _issue(
                        row_number=candidate.row_number,
                        column="规格名称",
                        code="SKU_IDENTITY_DUPLICATE",
                        message=(
                            f"第 {candidate.row_number} 行生成的 SKU 与第 "
                            f"{first_row_by_sku[sku_code]} 行重复。"
                        ),
                        value=candidate.specification,
                        suggestion="请确认商品型号和规格名称组合唯一。",
                    )
                )
                continue
            first_row_by_sku[sku_code] = candidate.row_number
            sku_name = (
                f"{candidate.name} · {candidate.specification}"
                if candidate.specification
                else candidate.name
            )
            rows.append(
                ProductTemplateRow(
                    row_number=candidate.row_number,
                    name=candidate.name,
                    category=candidate.category,
                    product_key=candidate.product_key,
                    product_model=candidate.product_model,
                    sku_code=sku_code,
                    sku_name=sku_name,
                    specification=candidate.specification,
                    units_per_carton=candidate.units_per_carton,
                    is_new=candidate.is_new,
                    schema_version=2,
                    supplier_name=None,
                    unit_price=(
                        candidate.specification_price
                        if candidate.specification_price is not None
                        else candidate.product_price
                    ),
                    description=candidate.description or first.description,
                    note=candidate.note,
                    tags=("新品",) if candidate.is_new else (),
                    variant_options=(),
                    default_moq=None,
                    gross_weight=None,
                    image_urls=candidate.image_urls,
                    image_url_columns=candidate.image_url_columns,
                    embedded_images=candidate.embedded_images,
                )
            )

    if issues:
        raise ProductTemplateValidationError(
            _validation_summary(issues),
            issues=tuple(issues),
        )
    if not rows:
        raise ProductTemplateValidationError("模版中没有可导入的有效商品。")
    if skipped_rows:
        warnings.append(f"已自动忽略 {skipped_rows} 行模板说明。")
    if generated_model_rows:
        warnings.append(
            f"有 {generated_model_rows} 行未填写商品型号，"
            "系统已根据商品名称和分类生成稳定的商品与 SKU 标识。"
        )
    unmatched_embedded_images = sum(
        len(images)
        for row_number, images in embedded_images_by_row.items()
        if row_number not in visited_rows
    )
    if unmatched_embedded_images:
        warnings.append(
            f"有 {unmatched_embedded_images} 张内嵌图片未对应到有效商品行，已忽略。"
        )
    return ProductTemplateParseResult(
        rows=tuple(rows),
        warnings=tuple(warnings),
        skipped_rows=skipped_rows,
    )


def _effective_sheet_headers(sheet: object) -> tuple[str, ...]:
    max_column = getattr(sheet, "max_column", None) or 1
    headers = tuple(
        _cell_text(sheet.cell(row=1, column=index).value)
        for index in range(1, max_column + 1)
    )
    last_header_index = next(
        (
            index
            for index in range(len(headers), 0, -1)
            if headers[index - 1]
        ),
        0,
    )
    return headers[:last_header_index]


def _resolve_read_only_sheet_dimensions(workbook: object) -> None:
    """Recover worksheet bounds when an XLSX omits cached dimension metadata.

    Some spreadsheet exporters write valid cells without a ``dimension`` entry.
    OpenPyXL leaves ``max_row`` and ``max_column`` unset for those worksheets in
    read-only mode, which previously made the header detector inspect only A1.
    Calculating the bounds keeps the memory-safe streaming reader while making
    these otherwise valid workbooks importable.
    """

    for sheet in workbook.worksheets:
        if (
            getattr(sheet, "max_row", None) is not None
            and getattr(sheet, "max_column", None) is not None
        ):
            continue

        # A completely empty auxiliary worksheet cannot be a supported data
        # sheet and OpenPyXL cannot calculate bounds for it reliably.
        if sheet.cell(row=1, column=1).value is None:
            continue

        calculate_dimension = getattr(sheet, "calculate_dimension", None)
        if not callable(calculate_dimension):
            continue
        try:
            calculate_dimension(force=True)
        except (
            OSError,
            TypeError,
            ValueError,
            UnboundLocalError,
            ET.ParseError,
        ) as exc:
            issue = _issue(
                row_number=None,
                column="工作表",
                code="SHEET_DIMENSION_INVALID",
                message=f"工作表“{sheet.title}”的有效数据范围无法识别。",
                suggestion="请使用 Excel 或 WPS 将文件另存为 XLSX 后重新导入。",
            )
            raise ProductTemplateValidationError(
                issue.message,
                issues=(issue,),
            ) from exc

        if (
            getattr(sheet, "max_row", None) is None
            or getattr(sheet, "max_column", None) is None
        ):
            issue = _issue(
                row_number=None,
                column="工作表",
                code="SHEET_DIMENSION_INVALID",
                message=f"工作表“{sheet.title}”的有效数据范围无法识别。",
                suggestion="请使用 Excel 或 WPS 将文件另存为 XLSX 后重新导入。",
            )
            raise ProductTemplateValidationError(
                issue.message,
                issues=(issue,),
            )


def _dual_sheet_header_issues(
    sheet: object,
    *,
    expected_headers: tuple[str, ...],
) -> list[ProductTemplateIssue]:
    actual_headers = _effective_sheet_headers(sheet)
    issues: list[ProductTemplateIssue] = []
    for index, expected in enumerate(expected_headers, start=1):
        actual = actual_headers[index - 1] if index <= len(actual_headers) else ""
        if actual == expected:
            continue
        issues.append(
            _issue(
                row_number=1,
                column=expected,
                code="HEADER_MISMATCH",
                message=(
                    f"工作表“{sheet.title}”第 {index} 列应为“{expected}”，"
                    f"实际为“{actual or '空白'}”。"
                ),
                value=actual,
                suggestion="请下载最新模板，不要修改第一行的列名或列顺序。",
            )
        )
    for index, actual in enumerate(
        actual_headers[len(expected_headers) :],
        start=len(expected_headers) + 1,
    ):
        if not actual:
            continue
        issues.append(
            _issue(
                row_number=1,
                column=f"第 {index} 列",
                code="UNEXPECTED_HEADER",
                message=f"工作表“{sheet.title}”包含模板之外的额外列“{actual}”。",
                value=actual,
                suggestion="请删除额外列；需要补充的内容可填写在备注或规格列。",
            )
        )
    return issues


def _select_product_sku_sheets(
    workbook: object,
) -> tuple[object, object, tuple[str, ...], int] | None:
    """Select a Product + SKU pair by structure while retaining v3 support."""

    contracts = (
        (5, PRODUCT_MASTER_TEMPLATE_HEADERS, SKU_DETAIL_TEMPLATE_HEADERS),
        (
            4,
            PRODUCT_MASTER_TEMPLATE_HEADERS_V4,
            SKU_DETAIL_TEMPLATE_HEADERS_V4,
        ),
        (3, PRODUCT_MASTER_TEMPLATE_HEADERS_V3, SKU_DETAIL_TEMPLATE_HEADERS_V3),
    )
    matches: list[tuple[int, object, object]] = []
    all_product_candidates: list[object] = []
    all_sku_candidates: list[object] = []
    for schema_version, product_headers, sku_headers in contracts:
        product_candidates = [
            sheet
            for sheet in workbook.worksheets
            if _effective_sheet_headers(sheet) == product_headers
        ]
        sku_candidates = [
            sheet
            for sheet in workbook.worksheets
            if _effective_sheet_headers(sheet) == sku_headers
        ]
        all_product_candidates.extend(product_candidates)
        all_sku_candidates.extend(sku_candidates)
        if len(product_candidates) > 1 or len(sku_candidates) > 1:
            names = "、".join(
                f"“{sheet.title}”"
                for sheet in (*product_candidates, *sku_candidates)
            )
            issue = _issue(
                row_number=None,
                column="工作表",
                code="SHEET_AMBIGUOUS",
                message=f"发现重复的 Product 或 SKU 数据页：{names}。",
                suggestion="请只保留一张商品主表和一张 SKU 明细表；工作表名称可以修改。",
            )
            raise ProductTemplateValidationError(issue.message, issues=(issue,))
        if len(product_candidates) == 1 and len(sku_candidates) == 1:
            matches.append(
                (schema_version, product_candidates[0], sku_candidates[0])
            )

    if len(matches) > 1:
        names = "、".join(
            f"“{sheet.title}”"
            for _schema, product_sheet, sku_sheet in matches
            for sheet in (product_sheet, sku_sheet)
        )
        issue = _issue(
            row_number=None,
            column="工作表",
            code="SHEET_AMBIGUOUS",
            message=f"发现重复的 Product 或 SKU 数据页：{names}。",
            suggestion="请只保留一张商品主表和一张 SKU 明细表；工作表名称可以修改。",
        )
        raise ProductTemplateValidationError(issue.message, issues=(issue,))
    if len(matches) == 1:
        schema_version, product_sheet, sku_sheet = matches[0]
        warnings: list[str] = []
        if product_sheet.title != PRODUCT_MASTER_TEMPLATE_SHEET:
            warnings.append(
                f"已根据列结构将工作表“{product_sheet.title}”识别为 Product。"
            )
        if sku_sheet.title != SKU_DETAIL_TEMPLATE_SHEET:
            warnings.append(
                f"已根据列结构将工作表“{sku_sheet.title}”识别为 SKU。"
            )
        if schema_version < 5:
            warnings.append("已按历史 Product + SKU 双表结构兼容导入。")
        return product_sheet, sku_sheet, tuple(warnings), schema_version

    product_named = next(
        (
            sheet
            for sheet in workbook.worksheets
            if sheet.title.casefold() == PRODUCT_MASTER_TEMPLATE_SHEET.casefold()
        ),
        None,
    )
    sku_named = next(
        (
            sheet
            for sheet in workbook.worksheets
            if sheet.title.casefold() == SKU_DETAIL_TEMPLATE_SHEET.casefold()
        ),
        None,
    )
    if (
        not all_product_candidates
        and not all_sku_candidates
        and product_named is None
        and sku_named is None
    ):
        return None

    candidate_headers = {
        _effective_sheet_headers(sheet)
        for sheet in (*all_product_candidates, *all_sku_candidates)
    }
    if candidate_headers.intersection(
        {PRODUCT_MASTER_TEMPLATE_HEADERS_V3, SKU_DETAIL_TEMPLATE_HEADERS_V3}
    ):
        schema_version = 3
        expected_product_headers = PRODUCT_MASTER_TEMPLATE_HEADERS_V3
        expected_sku_headers = SKU_DETAIL_TEMPLATE_HEADERS_V3
    elif SKU_DETAIL_TEMPLATE_HEADERS_V4 in candidate_headers:
        schema_version = 4
        expected_product_headers = PRODUCT_MASTER_TEMPLATE_HEADERS_V4
        expected_sku_headers = SKU_DETAIL_TEMPLATE_HEADERS_V4
    else:
        schema_version = 5
        expected_product_headers = PRODUCT_MASTER_TEMPLATE_HEADERS
        expected_sku_headers = SKU_DETAIL_TEMPLATE_HEADERS

    issues: list[ProductTemplateIssue] = []
    if all_product_candidates:
        product_sheet = all_product_candidates[0]
        if _effective_sheet_headers(product_sheet) != expected_product_headers:
            issues.extend(
                _dual_sheet_header_issues(
                    product_sheet,
                    expected_headers=expected_product_headers,
                )
            )
    elif product_named is not None:
        issues.extend(
            _dual_sheet_header_issues(
                product_named,
                expected_headers=expected_product_headers,
            )
        )
    else:
        issues.append(
            _issue(
                row_number=None,
                column="Product",
                code="SHEET_MISSING",
                message="缺少包含商品主数据列结构的 Product 数据页。",
                suggestion="请保留最新版模板中的 Product 与 SKU 两张数据页。",
            )
        )
    if all_sku_candidates:
        sku_sheet = all_sku_candidates[0]
        if _effective_sheet_headers(sku_sheet) != expected_sku_headers:
            issues.extend(
                _dual_sheet_header_issues(
                    sku_sheet,
                    expected_headers=expected_sku_headers,
                )
            )
    elif sku_named is not None:
        issues.extend(
            _dual_sheet_header_issues(
                sku_named,
                expected_headers=expected_sku_headers,
            )
        )
    else:
        issues.append(
            _issue(
                row_number=None,
                column="SKU",
                code="SHEET_MISSING",
                message="缺少包含 SKU 明细列结构的 SKU 数据页。",
                suggestion="请保留最新版模板中的 Product 与 SKU 两张数据页。",
            )
        )
    raise ProductTemplateValidationError(
        _validation_summary(issues),
        issues=tuple(issues),
    )


def _parse_product_sku_rows(
    path: Path,
    *,
    product_sheet: object,
    sku_sheet: object,
    sheet_warnings: tuple[str, ...],
    schema_version: int,
    progress_callback: Callable[[int, int], None] | None,
) -> ProductTemplateParseResult:
    if schema_version >= 5:
        product_headers = PRODUCT_MASTER_TEMPLATE_HEADERS
        sku_headers = SKU_DETAIL_TEMPLATE_HEADERS
    elif schema_version == 4:
        product_headers = PRODUCT_MASTER_TEMPLATE_HEADERS_V4
        sku_headers = SKU_DETAIL_TEMPLATE_HEADERS_V4
    else:
        product_headers = PRODUCT_MASTER_TEMPLATE_HEADERS_V3
        sku_headers = SKU_DETAIL_TEMPLATE_HEADERS_V3
    product_image_offset = 8 if schema_version >= 4 else 9
    product_rows_total = max(0, (product_sheet.max_row or 1) - 1)
    sku_rows_total = max(0, (sku_sheet.max_row or 1) - 1)

    embedded_images_by_row, embedded_image_warnings = (
        _extract_embedded_template_images(
            path,
            sheet_name=product_sheet.title,
            image_column_offset=product_image_offset,
            image_column_count=PRODUCT_IMAGE_COLUMN_COUNT,
        )
    )
    products: dict[str, ProductMasterTemplateCandidate] = {}
    issues: list[ProductTemplateIssue] = []
    visited_product_rows: set[int] = set()
    effective_product_rows = 0
    for row_number, raw_values in enumerate(
        product_sheet.iter_rows(
            min_row=2,
            max_col=len(product_headers),
            values_only=True,
        ),
        start=2,
    ):
        values = tuple(raw_values)
        embedded_images = embedded_images_by_row.get(row_number, ())
        if not any(_cell_text(value) for value in values) and not embedded_images:
            continue
        effective_product_rows += 1
        if effective_product_rows > MAX_TEMPLATE_ROWS:
            raise _effective_row_limit_error(
                sheet_label="Product 工作表",
                record_label="商品",
            )
        visited_product_rows.add(row_number)
        row_issues: list[ProductTemplateIssue] = []
        formula_indexes = {
            index
            for index, value in enumerate(values)
            if isinstance(value, str) and value.lstrip().startswith("=")
        }
        for index in sorted(formula_indexes):
            column = product_headers[index]
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column=column,
                    code="FORMULA_NOT_ALLOWED",
                    message=f"Product 第 {row_number} 行的{column}包含公式。",
                    value=values[index],
                    suggestion="请复制计算结果并粘贴为固定值后重新导入。",
                )
            )

        product_code = _normalize_sku_code(values[0])
        name = _cell_text(values[1])
        if not product_code and 0 not in formula_indexes:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="商品编码",
                    code="REQUIRED_VALUE_MISSING",
                    message=f"Product 第 {row_number} 行缺少商品编码。",
                    suggestion="请填写稳定且唯一的商品编码，SKU 表将使用它关联商品。",
                )
            )
        elif len(product_code) > 160 and 0 not in formula_indexes:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="商品编码",
                    code="VALUE_TOO_LONG",
                    message=f"Product 第 {row_number} 行的商品编码超过 160 个字符。",
                    value=values[0],
                    suggestion="请将商品编码缩短至 160 个字符以内。",
                )
            )
        if not name and 1 not in formula_indexes:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="商品名称",
                    code="REQUIRED_VALUE_MISSING",
                    message=f"Product 第 {row_number} 行缺少商品名称。",
                    suggestion="请填写面向客户展示的商品名称。",
                )
            )
        elif len(name) > 500 and 1 not in formula_indexes:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="商品名称",
                    code="VALUE_TOO_LONG",
                    message=f"Product 第 {row_number} 行的商品名称超过 500 个字符。",
                    value=values[1],
                    suggestion="请缩短商品名称，详细信息放入商品描述。",
                )
            )

        category = UNCATEGORIZED_CATEGORY_NAME
        if _cell_text(values[2]) and 2 not in formula_indexes:
            try:
                category = "/".join(
                    _normalize_category_path(values[2], row_number=row_number)
                )
            except ProductTemplateValidationError as exc:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="商品分类",
                        code="CATEGORY_INVALID",
                        message=str(exc),
                        value=values[2],
                        suggestion="请填写“一级分类”或“一级分类/二级分类”，最多两级。",
                    )
                )

        product_model = _cell_text(values[3]) or None
        if product_model and len(product_model) > 160 and 3 not in formula_indexes:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="商品型号",
                    code="VALUE_TOO_LONG",
                    message=f"Product 第 {row_number} 行的商品型号超过 160 个字符。",
                    value=values[3],
                    suggestion="请将商品型号缩短至 160 个字符以内。",
                )
            )

        product_price = Decimal("0.00")
        if _cell_text(values[4]) and 4 not in formula_indexes:
            try:
                product_price = _decimal(
                    values[4],
                    field="商品价格",
                    row_number=row_number,
                )
            except ProductTemplateValidationError as exc:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="商品价格",
                        code="PRICE_INVALID",
                        message=str(exc),
                        value=values[4],
                        suggestion="商品价格可以留空（按 0 处理），或填写大于等于 0 的数字。",
                    )
                )

        tags: tuple[str, ...] = ()
        if _cell_text(values[7]) and 7 not in formula_indexes:
            try:
                tags = _normalize_tags(values[7], row_number=row_number)
            except ProductTemplateValidationError as exc:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="标签",
                        code="TAGS_INVALID",
                        message=str(exc),
                        value=values[7],
                        suggestion=f"最多填写 {MAX_TAGS} 个标签，使用逗号分隔。",
                    )
                )

        is_new = False
        if schema_version == 3 and 8 not in formula_indexes:
            try:
                is_new = _new_product_flag(values[8], row_number=row_number)
            except ProductTemplateValidationError as exc:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="是否是新品",
                        code="NEW_PRODUCT_FLAG_INVALID",
                        message=str(exc),
                        value=values[8],
                        suggestion="可以留空；需要标记新品时填写“是”，否则填写“否”。",
                    )
                )

        image_urls: list[str] = []
        image_url_columns: list[int] = []
        for image_index, value in enumerate(
            values[product_image_offset:],
            start=1,
        ):
            value_index = image_index + product_image_offset - 1
            if not _cell_text(value) or value_index in formula_indexes:
                continue
            image_url = _valid_image_url(value)
            if image_url is None:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column=f"商品图片{image_index}",
                        code="IMAGE_URL_INVALID",
                        message=(
                            f"Product 第 {row_number} 行商品图片{image_index}"
                            "不是有效的 HTTP(S) 链接。"
                        ),
                        value=value,
                        suggestion=(
                            "图片可以留空或直接插入对应单元格；"
                            "填写文本时请使用可公开访问的 HTTP(S) 地址。"
                        ),
                    )
                )
            elif image_url not in image_urls:
                image_urls.append(image_url)
                image_url_columns.append(image_index)

        if product_code and product_code in products:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="商品编码",
                    code="PRODUCT_CODE_DUPLICATE",
                    message=(
                        f"Product 第 {row_number} 行商品编码“{product_code}”与第 "
                        f"{products[product_code].row_number} 行重复。"
                    ),
                    value=values[0],
                    suggestion="Product 表中的商品编码必须唯一。",
                )
            )
        if row_issues:
            issues.extend(row_issues)
            continue
        products[product_code] = ProductMasterTemplateCandidate(
            row_number=row_number,
            product_code=product_code,
            name=name,
            category=category,
            product_model=product_model,
            product_price=product_price,
            description=_cell_text(values[5]) or None,
            note=_cell_text(values[6]) or None,
            tags=tags,
            is_new=is_new,
            image_urls=tuple(image_urls),
            image_url_columns=tuple(image_url_columns),
            embedded_images=embedded_images,
        )

    rows: list[ProductTemplateRow] = []
    first_row_by_sku: dict[str, int] = {}
    first_row_by_sku_definition: dict[str, int] = {}
    expanded_definition_rows = 0
    generated_source_sku_count = 0
    referenced_product_codes: set[str] = set()
    effective_sku_rows = 0
    progress_interval = max(100, sku_rows_total // 100) if sku_rows_total else 100
    reserved_option_keys = {
        TEMPLATE_SOURCE_KEY.casefold(),
        "商品编码".casefold(),
        "商品型号".casefold(),
        "规格名称".casefold(),
        "一箱个数".casefold(),
        "装箱数".casefold(),
        "毛重".casefold(),
        "起定数".casefold(),
        "是否是新品".casefold(),
        "备注".casefold(),
    }
    for row_number, raw_values in enumerate(
        sku_sheet.iter_rows(
            min_row=2,
            max_col=len(sku_headers),
            values_only=True,
        ),
        start=2,
    ):
        values = tuple(raw_values)
        processed_rows = row_number - 1
        if progress_callback is not None and (
            processed_rows == 1
            or processed_rows % progress_interval == 0
            or processed_rows == sku_rows_total
        ):
            progress_callback(processed_rows, sku_rows_total)
        if not any(_cell_text(value) for value in values):
            continue
        effective_sku_rows += 1
        if effective_sku_rows > MAX_TEMPLATE_ROWS:
            raise _effective_row_limit_error(
                sheet_label="SKU 工作表",
                record_label=" SKU ",
            )

        row_issues: list[ProductTemplateIssue] = []
        formula_indexes = {
            index
            for index, value in enumerate(values)
            if isinstance(value, str) and value.lstrip().startswith("=")
        }
        for index in sorted(formula_indexes):
            column = sku_headers[index]
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column=column,
                    code="FORMULA_NOT_ALLOWED",
                    message=f"SKU 第 {row_number} 行的{column}包含公式。",
                    value=values[index],
                    suggestion="请复制计算结果并粘贴为固定值后重新导入。",
                )
            )

        product_code = _normalize_sku_code(values[0])
        sku_code = _normalize_sku_code(values[1])
        if not product_code and 0 not in formula_indexes:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="商品编码",
                    code="REQUIRED_VALUE_MISSING",
                    message=f"SKU 第 {row_number} 行缺少商品编码。",
                    suggestion="请填写 Product 表中已存在的商品编码。",
                )
            )
        product = products.get(product_code)
        if product_code and product is None and 0 not in formula_indexes:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="商品编码",
                    code="PRODUCT_REFERENCE_MISSING",
                    message=(
                        f"SKU 第 {row_number} 行引用的商品编码“{product_code}”"
                        "在 Product 表中不存在。"
                    ),
                    value=values[0],
                    suggestion="请先在 Product 表中新增相同商品编码，或修正拼写。",
                )
            )
        if len(sku_code) > 160 and 1 not in formula_indexes:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="SKU编号",
                    code="VALUE_TOO_LONG",
                    message=f"SKU 第 {row_number} 行的 SKU 编号超过 160 个字符。",
                    value=values[1],
                    suggestion="请将 SKU 编号缩短至 160 个字符以内。",
                )
            )
        identity_rows = (
            first_row_by_sku_definition
            if schema_version >= 5
            else first_row_by_sku
        )
        if sku_code and sku_code in identity_rows:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="SKU编号",
                    code="SKU_IDENTITY_DUPLICATE",
                    message=(
                        f"SKU 第 {row_number} 行编号“{sku_code}”与第 "
                        f"{identity_rows[sku_code]} 行重复。"
                    ),
                    value=values[1],
                    suggestion="SKU 表中的 SKU 编号必须唯一。",
                )
            )

        explicit_specification: str | None = None
        variant_combinations: list[tuple[tuple[str, str], ...]] = [()]
        seen_option_keys: set[str] = set()
        if schema_version >= 5:
            option_dimensions: list[tuple[str, tuple[str, ...]]] = []
            option_groups = (
                (1, 3, tuple(range(4, 9))),
                (2, 9, tuple(range(10, 15))),
                (3, 15, tuple(range(16, 21))),
            )
            for option_number, name_index, value_indexes in option_groups:
                option_name = _cell_text(values[name_index])
                option_values = [
                    _cell_text(values[value_index])
                    for value_index in value_indexes
                    if _cell_text(values[value_index])
                ]
                if not option_name and option_values:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column=f"规格{option_number}名称",
                            code="VARIANT_OPTION_NAME_MISSING",
                            message=(
                                f"SKU 第 {row_number} 行填写了规格{option_number}值，"
                                "但没有填写对应的规格名称。"
                            ),
                            suggestion="请填写规格名称，例如“尺寸”或“颜色”。",
                        )
                    )
                    continue
                if option_name and not option_values:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column=f"规格{option_number}值（1）",
                            code="VARIANT_OPTION_VALUE_MISSING",
                            message=(
                                f"SKU 第 {row_number} 行填写了规格{option_number}名称"
                                f"“{option_name}”，但没有填写任何候选值。"
                            ),
                            suggestion="请至少填写一个候选值，或将规格名称一并留空。",
                        )
                    )
                    continue
                if not option_name:
                    continue
                normalized_option_name = option_name.casefold()
                if normalized_option_name in reserved_option_keys:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column=f"规格{option_number}名称",
                            code="VARIANT_OPTION_RESERVED",
                            message=(
                                f"SKU 第 {row_number} 行规格名称“{option_name}”"
                                "与系统字段冲突。"
                            ),
                            value=option_name,
                            suggestion="请使用“颜色”“尺寸”“材质”等业务规格名称。",
                        )
                    )
                    continue
                if normalized_option_name in seen_option_keys:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column=f"规格{option_number}名称",
                            code="VARIANT_OPTION_DUPLICATE",
                            message=(
                                f"SKU 第 {row_number} 行重复填写规格名称“{option_name}”。"
                            ),
                            value=option_name,
                            suggestion="同一个 SKU 定义中的规格名称不能重复。",
                        )
                    )
                    continue
                normalized_values: set[str] = set()
                duplicate_value: str | None = None
                for option_value in option_values:
                    normalized_value = option_value.casefold()
                    if normalized_value in normalized_values:
                        duplicate_value = option_value
                        break
                    normalized_values.add(normalized_value)
                if duplicate_value is not None:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column=f"规格{option_number}值",
                            code="VARIANT_VALUE_DUPLICATE",
                            message=(
                                f"SKU 第 {row_number} 行规格“{option_name}”"
                                f"重复填写候选值“{duplicate_value}”。"
                            ),
                            value=duplicate_value,
                            suggestion="同一规格下的候选值不能重复。",
                        )
                    )
                    continue
                if len(option_name) > 100 or any(
                    len(option_value) > 500 for option_value in option_values
                ):
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column=f"规格{option_number}",
                            code="VALUE_TOO_LONG",
                            message=f"SKU 第 {row_number} 行的规格名称或规格值过长。",
                            suggestion="规格名称最多 100 个字符，规格值最多 500 个字符。",
                        )
                    )
                    continue
                seen_option_keys.add(normalized_option_name)
                option_dimensions.append((option_name, tuple(option_values)))
            if option_dimensions:
                variant_combinations = [
                    tuple(
                        (option_dimensions[index][0], option_value)
                        for index, option_value in enumerate(combination)
                    )
                    for combination in cartesian_product(
                        *(values for _name, values in option_dimensions)
                    )
                ]
        else:
            explicit_specification = _cell_text(values[3]) or None
            if explicit_specification and len(explicit_specification) > 500:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="规格名称",
                        code="VALUE_TOO_LONG",
                        message=f"SKU 第 {row_number} 行的规格名称超过 500 个字符。",
                        value=values[3],
                        suggestion="请缩短规格名称。",
                    )
                )
            variant_options: list[tuple[str, str]] = []
            option_indexes = (
                ((4, 5), (6, 7), (8, 9), (10, 11), (12, 13))
                if schema_version == 4
                else ((4, 5), (6, 7), (8, 9))
            )
            for option_number, (name_index, value_index) in enumerate(
                option_indexes,
                start=1,
            ):
                option_name = _cell_text(values[name_index])
                option_value = _cell_text(values[value_index])
                if bool(option_name) != bool(option_value):
                    missing = "值" if option_name else "名称"
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column=f"规格{option_number}{missing}",
                            code="VARIANT_OPTION_INCOMPLETE",
                            message=(
                                f"SKU 第 {row_number} 行的规格{option_number}"
                                "名称和值需要成对填写。"
                            ),
                            value=values[
                                value_index if option_name else name_index
                            ],
                            suggestion="请同时填写规格名称和规格值，或同时留空。",
                        )
                    )
                    continue
                if not option_name:
                    continue
                normalized_option_name = option_name.casefold()
                if normalized_option_name in reserved_option_keys:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column=f"规格{option_number}名称",
                            code="VARIANT_OPTION_RESERVED",
                            message=(
                                f"SKU 第 {row_number} 行规格名称“{option_name}”"
                                "与系统字段冲突。"
                            ),
                            value=option_name,
                            suggestion="请使用“颜色”“尺寸”“材质”等业务规格名称。",
                        )
                    )
                elif normalized_option_name in seen_option_keys:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column=f"规格{option_number}名称",
                            code="VARIANT_OPTION_DUPLICATE",
                            message=(
                                f"SKU 第 {row_number} 行重复填写规格名称“{option_name}”。"
                            ),
                            value=option_name,
                            suggestion="同一个 SKU 的规格名称不能重复。",
                        )
                    )
                elif len(option_name) > 100 or len(option_value) > 500:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column=f"规格{option_number}",
                            code="VALUE_TOO_LONG",
                            message=f"SKU 第 {row_number} 行的规格名称或规格值过长。",
                            suggestion="规格名称最多 100 个字符，规格值最多 500 个字符。",
                        )
                    )
                else:
                    seen_option_keys.add(normalized_option_name)
                    variant_options.append((option_name, option_value))
            variant_combinations = [tuple(variant_options)]

        if schema_version >= 5:
            supplier_index = 21
            price_index = 22
        elif schema_version == 4:
            supplier_index = 14
            price_index = 15
        else:
            supplier_index = 10
            price_index = 11
        supplier_name: str | None = None
        if (
            _cell_text(values[supplier_index])
            and supplier_index not in formula_indexes
        ):
            try:
                supplier_name = _normalize_supplier_name(
                    values[supplier_index], row_number=row_number
                )
            except ProductTemplateValidationError as exc:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="供应商",
                        code="SUPPLIER_INVALID",
                        message=str(exc),
                        value=values[supplier_index],
                        suggestion="供应商可以留空；填写时请使用简洁、稳定的名称。",
                    )
                )

        unit_price = product.product_price if product is not None else Decimal("0.00")
        if _cell_text(values[price_index]) and price_index not in formula_indexes:
            try:
                unit_price = _decimal(
                    values[price_index], field="SKU价格", row_number=row_number
                )
            except ProductTemplateValidationError as exc:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="SKU价格",
                        code="PRICE_INVALID",
                        message=str(exc),
                        value=values[price_index],
                        suggestion="SKU 价格可以留空并继承商品价格，或填写大于等于 0 的数字。",
                    )
                )

        gross_weight: Decimal | None = None
        default_moq: Decimal | None = None
        if schema_version >= 4:
            gross_weight_index = 23 if schema_version >= 5 else 16
            default_moq_index = 24 if schema_version >= 5 else 17
            units_per_carton_index = 25 if schema_version >= 5 else 18
            for value_index, field in (
                (gross_weight_index, "毛重"),
                (default_moq_index, "起定数"),
            ):
                if not _cell_text(values[value_index]) or value_index in formula_indexes:
                    continue
                try:
                    parsed_value = _sku_quantity_decimal(
                        values[value_index],
                        field=field,
                        row_number=row_number,
                    )
                    if field == "毛重":
                        gross_weight = parsed_value
                    else:
                        default_moq = parsed_value
                except ProductTemplateValidationError as exc:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column=field,
                            code="QUANTITY_INVALID",
                            message=str(exc),
                            value=values[value_index],
                            suggestion=f"{field}可以留空，或填写大于等于 0 的数字。",
                        )
                    )
            units_per_carton = None
            if (
                _cell_text(values[units_per_carton_index])
                and units_per_carton_index not in formula_indexes
            ):
                try:
                    units_per_carton = _decimal_option_text(
                        _sku_quantity_decimal(
                            values[units_per_carton_index],
                            field="装箱数",
                            row_number=row_number,
                        )
                    )
                except ProductTemplateValidationError as exc:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column="装箱数",
                            code="QUANTITY_INVALID",
                            message=str(exc),
                            value=values[units_per_carton_index],
                            suggestion="装箱数可以留空，或填写大于等于 0 的数字。",
                        )
                    )
        else:
            units_per_carton = _cell_text(values[12]) or None
            if units_per_carton and len(units_per_carton) > 100:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="一箱个数",
                        code="VALUE_TOO_LONG",
                        message=f"SKU 第 {row_number} 行的一箱个数超过 100 个字符。",
                        value=values[12],
                        suggestion="请填写简洁的装箱数量，例如 24。",
                    )
                )

        sku_name = _cell_text(values[2]) or None
        if sku_name and len(sku_name) > 500:
            row_issues.append(
                _issue(
                    row_number=row_number,
                    column="SKU名称",
                    code="VALUE_TOO_LONG",
                    message=f"SKU 第 {row_number} 行的 SKU 名称超过 500 个字符。",
                    value=values[2],
                    suggestion="请缩短 SKU 名称。",
                )
            )

        if row_issues:
            issues.extend(row_issues)
            continue
        assert product is not None
        tags = list(product.tags)
        if schema_version == 3 and product.is_new and "新品".casefold() not in {
            tag.casefold() for tag in tags
        }:
            tags.append("新品")
        if len(rows) + len(variant_combinations) > MAX_EXPANDED_SKUS:
            issues.append(
                _issue(
                    row_number=row_number,
                    column="规格候选值",
                    code="SKU_EXPANSION_LIMIT_EXCEEDED",
                    message=(
                        f"SKU 第 {row_number} 行展开后会使本次导入超过 "
                        f"{MAX_EXPANDED_SKUS} 个 SKU。"
                    ),
                    suggestion="请减少候选值数量，或拆分为多个文件分批导入。",
                )
            )
            continue
        prepared_rows: list[
            tuple[str, str, str | None, tuple[tuple[str, str], ...]]
        ] = []
        generated_codes: set[str] = set()
        base_sku_name = sku_name or product.name
        for variant_options in variant_combinations:
            specification = explicit_specification or (
                " / ".join(value for _name, value in variant_options) or None
            )
            generated_sku_code = _variant_sku_code(
                product_key=f"PRODUCT:{product.product_code}",
                product_model=sku_code or None,
                specification=(
                    specification if schema_version >= 5 else None
                ),
            )
            if generated_sku_code in generated_codes:
                issues.append(
                    _issue(
                        row_number=row_number,
                        column="规格候选值",
                        code="SKU_IDENTITY_DUPLICATE",
                        message=(
                            f"SKU 第 {row_number} 行的候选值生成了重复 SKU 编号"
                            f"“{generated_sku_code}”。"
                        ),
                        suggestion="请删除重复候选值或调整 SKU 编号前缀。",
                    )
                )
                continue
            if generated_sku_code in first_row_by_sku:
                issues.append(
                    _issue(
                        row_number=row_number,
                        column="SKU编号",
                        code="SKU_IDENTITY_DUPLICATE",
                        message=(
                            f"SKU 第 {row_number} 行生成的编号“{generated_sku_code}”"
                            f"与第 {first_row_by_sku[generated_sku_code]} 行重复。"
                        ),
                        suggestion="请使用不同的 SKU 编号前缀或候选值组合。",
                    )
                )
                continue
            generated_codes.add(generated_sku_code)
            final_sku_name = (
                f"{base_sku_name} · {specification}"
                if schema_version >= 5 and specification
                else (
                    sku_name
                    or (
                        f"{product.name} · {specification}"
                        if specification
                        else product.name
                    )
                )
            )
            prepared_rows.append(
                (
                    generated_sku_code,
                    final_sku_name,
                    specification,
                    tuple(variant_options),
                )
            )
        if len(prepared_rows) != len(variant_combinations):
            continue
        if not sku_code:
            generated_source_sku_count += len(prepared_rows)
        if schema_version >= 5 and sku_code:
            first_row_by_sku_definition[sku_code] = row_number
        if schema_version >= 5:
            if len(prepared_rows) > 1:
                expanded_definition_rows += 1
        referenced_product_codes.add(product_code)
        for (
            generated_sku_code,
            final_sku_name,
            specification,
            variant_options,
        ) in prepared_rows:
            first_row_by_sku[generated_sku_code] = row_number
            rows.append(
                ProductTemplateRow(
                    row_number=row_number,
                    name=product.name,
                    category=product.category,
                    product_key=f"PRODUCT:{product.product_code}",
                    product_model=product.product_model,
                    sku_code=generated_sku_code,
                    sku_name=final_sku_name,
                    specification=specification,
                    units_per_carton=units_per_carton,
                    is_new=product.is_new,
                    schema_version=schema_version,
                    supplier_name=supplier_name,
                    unit_price=unit_price,
                    description=product.description,
                    note=product.note,
                    tags=tuple(tags),
                    variant_options=variant_options,
                    default_moq=default_moq,
                    gross_weight=gross_weight,
                    image_urls=product.image_urls,
                    image_url_columns=product.image_url_columns,
                    embedded_images=product.embedded_images,
                )
            )

    if issues:
        raise ProductTemplateValidationError(
            _validation_summary(issues),
            issues=tuple(issues),
        )
    if not products:
        raise ProductTemplateValidationError("Product 表中没有可导入的有效商品。")

    sku_count = len(rows)
    unreferenced_product_codes = sorted(
        set(products) - referenced_product_codes
    )
    for product_code in unreferenced_product_codes:
        product = products[product_code]
        rows.append(
            ProductTemplateRow(
                row_number=product.row_number,
                name=product.name,
                category=product.category,
                product_key=f"PRODUCT:{product.product_code}",
                product_model=product.product_model,
                # Product-only rows create a generated no-specification SKU
                # during the apply phase. Keeping the source product code here
                # still gives image storage and planning a deterministic key.
                sku_code=product.product_code,
                sku_name=product.name,
                specification=None,
                units_per_carton=None,
                is_new=product.is_new,
                schema_version=schema_version,
                supplier_name=None,
                unit_price=product.product_price,
                description=product.description,
                note=product.note,
                tags=product.tags,
                variant_options=(),
                default_moq=None,
                gross_weight=None,
                image_urls=product.image_urls,
                image_url_columns=product.image_url_columns,
                embedded_images=product.embedded_images,
                product_only=True,
            )
        )

    warnings = [
        *sheet_warnings,
        *embedded_image_warnings,
        (
            f"已识别 Product + SKU 双表模板："
            f"{len(products)} 个商品，{sku_count} 个 SKU。"
        ),
    ]
    if expanded_definition_rows:
        warnings.append(
            f"已将 {expanded_definition_rows} 行规格候选值自动组合为具体 SKU。"
        )
    if generated_source_sku_count:
        warnings.append(
            f"有 {generated_source_sku_count} 个 SKU 未填写来源编号，"
            "系统已按商品与规格生成稳定的导入标识。"
        )
    if unreferenced_product_codes:
        warnings.append(
            f"Product 表中有 {len(unreferenced_product_codes)} 个商品没有 SKU，"
            "系统将为每个商品创建 1 个无规格基础 SKU，可继续按 SKU 管理。"
        )
    unmatched_embedded_images = sum(
        len(images)
        for row_number, images in embedded_images_by_row.items()
        if row_number not in visited_product_rows
    )
    if unmatched_embedded_images:
        warnings.append(
            f"有 {unmatched_embedded_images} 张内嵌图片未对应到有效 Product 行，已忽略。"
        )
    return ProductTemplateParseResult(
        rows=tuple(rows),
        warnings=tuple(warnings),
        skipped_rows=0,
    )


def parse_product_template(
    path: Path,
    *,
    progress_callback: Callable[[int, int], None] | None = None,
) -> ProductTemplateParseResult:
    if path.suffix.lower() != ".xlsx":
        raise ProductTemplateValidationError("只支持 .xlsx 商品文件。")
    _inspect_archive(path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except (OSError, ValueError, BadZipFile) as exc:
        raise ProductTemplateValidationError("商品模版无法读取，请重新导出为 XLSX。") from exc

    try:
        _resolve_read_only_sheet_dimensions(workbook)
        product_sku_sheets = _select_product_sku_sheets(workbook)
        if product_sku_sheets is not None:
            product_sheet, sku_sheet, sheet_warnings, schema_version = (
                product_sku_sheets
            )
            return _parse_product_sku_rows(
                path,
                product_sheet=product_sheet,
                sku_sheet=sku_sheet,
                sheet_warnings=sheet_warnings,
                schema_version=schema_version,
                progress_callback=progress_callback,
            )
        sheet, layout, sheet_warning = _select_product_sheet(workbook)
        effective_headers = layout.headers
        received_headers = tuple(
            _cell_text(sheet.cell(row=1, column=index).value)
            for index in range(1, len(effective_headers) + 1)
        )
        max_column = sheet.max_column or len(effective_headers)
        extra_headers = [
            (index, _cell_text(sheet.cell(row=1, column=index).value))
            for index in range(len(effective_headers) + 1, max_column + 1)
            if _cell_text(sheet.cell(row=1, column=index).value)
        ]
        if received_headers != effective_headers or extra_headers:
            header_issues: list[ProductTemplateIssue] = []
            for index, expected in enumerate(effective_headers, start=1):
                actual = received_headers[index - 1]
                if actual == expected:
                    continue
                header_issues.append(
                    _issue(
                        row_number=1,
                        column=expected,
                        code="HEADER_MISMATCH",
                        message=(
                            f"表头第 {index} 列应为“{expected}”，"
                            f"实际为“{actual or '空白'}”。"
                        ),
                        value=actual,
                        suggestion="请下载最新模板，不要修改第一行的列名或列顺序。",
                    )
                )
            for index, actual in extra_headers:
                header_issues.append(
                    _issue(
                        row_number=1,
                        column=f"第 {index} 列",
                        code="UNEXPECTED_HEADER",
                        message=f"表头包含模板之外的额外列“{actual}”。",
                        value=actual,
                        suggestion="请删除额外列，或把补充信息放入“备注”列。",
                    )
                )
            raise ProductTemplateValidationError(
                f"表头与固定商品模版不一致。{_validation_summary(header_issues)}",
                issues=tuple(header_issues),
            )
        embedded_images_by_row, embedded_image_warnings = (
            _extract_embedded_template_images(
                path,
                sheet_name=sheet.title,
                image_column_offset=layout.image_column_offset,
                image_column_count=layout.image_column_count,
            )
        )
        if layout.kind == TEMPLATE_LAYOUT_PRODUCT_VARIANTS:
            return _parse_product_variant_rows(
                sheet,
                effective_headers=effective_headers,
                embedded_images_by_row=embedded_images_by_row,
                embedded_image_warnings=embedded_image_warnings,
                sheet_warning=sheet_warning,
                progress_callback=progress_callback,
            )
        rows: list[ProductTemplateRow] = []
        warnings: list[str] = []
        if sheet_warning is not None:
            warnings.append(sheet_warning)
        warnings.extend(embedded_image_warnings)
        issues: list[ProductTemplateIssue] = []
        first_row_by_sku: dict[str, int] = {}
        generated_sku_occurrences: dict[str, int] = {}
        generated_sku_count = 0
        skipped_rows = 0
        visited_rows: set[int] = set()
        effective_data_rows = 0
        total_rows = max(0, (sheet.max_row or 1) - 1)
        progress_interval = max(100, total_rows // 100) if total_rows else 100
        for row_number, values in enumerate(
            sheet.iter_rows(
                min_row=2,
                max_col=len(effective_headers),
                values_only=True,
            ),
            start=2,
        ):
            processed_rows = row_number - 1
            if progress_callback is not None and (
                processed_rows == 1
                or processed_rows % progress_interval == 0
                or processed_rows == total_rows
            ):
                progress_callback(processed_rows, total_rows)
            embedded_images = embedded_images_by_row.get(row_number, ())
            if not any(_cell_text(value) for value in values) and not embedded_images:
                continue
            effective_data_rows += 1
            if effective_data_rows > MAX_TEMPLATE_ROWS:
                raise _effective_row_limit_error(
                    sheet_label="商品工作表",
                    record_label="商品",
                )
            visited_rows.add(row_number)

            row_issues: list[ProductTemplateIssue] = []
            formula_indexes = {
                index
                for index, value in enumerate(values)
                if isinstance(value, str) and value.lstrip().startswith("=")
            }
            for index in sorted(formula_indexes):
                column = effective_headers[index]
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column=column,
                        code="FORMULA_NOT_ALLOWED",
                        message=f"第 {row_number} 行的{column}包含公式。",
                        value=values[index],
                        suggestion="请复制计算结果并粘贴为固定值后重新导入。",
                    )
                )

            name = _cell_text(values[0])
            category_text = _cell_text(values[1])
            sku_code = _normalize_sku_code(values[2])
            if not name and 0 not in formula_indexes:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="商品名称",
                        code="REQUIRED_VALUE_MISSING",
                        message=f"第 {row_number} 行缺少商品名称。",
                        suggestion="请填写商品名称；这是商品导入唯一的必填项。",
                    )
                )

            if len(name) > 500 and 0 not in formula_indexes:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="商品名称",
                        code="VALUE_TOO_LONG",
                        message=f"第 {row_number} 行的商品名称超过 500 个字符。",
                        value=values[0],
                        suggestion="请缩短商品名称，详细信息可填写在“商品描述”中。",
                    )
                )
            if len(sku_code) > 160 and 2 not in formula_indexes:
                row_issues.append(
                    _issue(
                        row_number=row_number,
                        column="商品型号",
                        code="VALUE_TOO_LONG",
                        message=f"第 {row_number} 行的商品型号超过 160 个字符。",
                        value=values[2],
                        suggestion="请将商品型号缩短至 160 个字符以内。",
                    )
                )

            category = UNCATEGORIZED_CATEGORY_NAME
            if category_text and 1 not in formula_indexes:
                try:
                    category = "/".join(
                        _normalize_category_path(
                            category_text,
                            row_number=row_number,
                        )
                    )
                except ProductTemplateValidationError as exc:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column="商品分类",
                            code="CATEGORY_INVALID",
                            message=str(exc),
                            value=values[1],
                            suggestion="请填写“一级分类”或“一级分类/二级分类”，最多两级。",
                        )
                    )

            supplier_name: str | None = None
            if _cell_text(values[3]) and 3 not in formula_indexes:
                try:
                    supplier_name = _normalize_supplier_name(
                        values[3],
                        row_number=row_number,
                    )
                except ProductTemplateValidationError as exc:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column="供应商",
                            code="SUPPLIER_INVALID",
                            message=str(exc),
                            value=values[3],
                            suggestion=(
                                f"供应商可以留空；填写时请控制在 "
                                f"{MAX_SUPPLIER_NAME_LENGTH} 个字符以内。"
                            ),
                        )
                    )

            unit_price = Decimal("0.00")
            if _cell_text(values[4]) and 4 not in formula_indexes:
                try:
                    unit_price = _decimal(
                        values[4],
                        field="商品价格",
                        row_number=row_number,
                    )
                except ProductTemplateValidationError as exc:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column="商品价格",
                            code="PRICE_INVALID",
                            message=str(exc),
                            value=values[4],
                            suggestion="价格可以留空（系统按 0 处理），或填写大于等于 0 的数字。",
                        )
                    )

            tags: tuple[str, ...] = ()
            if _cell_text(values[7]) and 7 not in formula_indexes:
                try:
                    tags = _normalize_tags(values[7], row_number=row_number)
                except ProductTemplateValidationError as exc:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column="标签",
                            code="TAGS_INVALID",
                            message=str(exc),
                            value=values[7],
                            suggestion=(
                                f"标签可以留空；最多填写 {MAX_TAGS} 个，"
                                "使用逗号分隔。"
                            ),
                        )
                    )

            image_urls: list[str] = []
            image_url_columns: list[int] = []
            for image_index, value in enumerate(values[8:], start=1):
                value_index = image_index + 7
                if not _cell_text(value) or value_index in formula_indexes:
                    continue
                image_url = _valid_image_url(value)
                if image_url is None:
                    row_issues.append(
                        _issue(
                            row_number=row_number,
                            column=f"商品图片{image_index}",
                            code="IMAGE_URL_INVALID",
                            message=(
                                f"第 {row_number} 行商品图片{image_index}"
                                "不是有效的 HTTP(S) 链接。"
                            ),
                            value=value,
                            suggestion=(
                                "图片可以留空或直接插入该单元格位置；"
                                "填写文本时请使用可公开访问的 http:// 或 https:// 地址。"
                            ),
                        )
                    )
                    continue
                if image_url not in image_urls:
                    image_urls.append(image_url)
                    image_url_columns.append(image_index)

            if row_issues:
                issues.extend(row_issues)
                continue

            if not sku_code:
                generated_base = _generated_sku_code(
                    name=name,
                    category=category,
                    supplier_name=supplier_name,
                    occurrence=1,
                )
                occurrence = generated_sku_occurrences.get(generated_base, 0) + 1
                sku_code = _generated_sku_code(
                    name=name,
                    category=category,
                    supplier_name=supplier_name,
                    occurrence=occurrence,
                )
                while sku_code in first_row_by_sku:
                    occurrence += 1
                    sku_code = _generated_sku_code(
                        name=name,
                        category=category,
                        supplier_name=supplier_name,
                        occurrence=occurrence,
                    )
                generated_sku_occurrences[generated_base] = occurrence
                generated_sku_count += 1

            if sku_code in first_row_by_sku:
                warnings.append(
                    f"第 {row_number} 行商品型号“{sku_code}”与第 "
                    f"{first_row_by_sku[sku_code]} 行重复，已保留首次出现的记录。"
                )
                skipped_rows += 1
                continue

            note = _cell_text(values[6]) or None
            first_row_by_sku[sku_code] = row_number
            rows.append(
                ProductTemplateRow(
                    row_number=row_number,
                    name=name,
                    category=category,
                    product_key=sku_code,
                    product_model=sku_code,
                    sku_code=sku_code,
                    sku_name=name,
                    specification=None,
                    units_per_carton=None,
                    is_new=False,
                    schema_version=1,
                    supplier_name=supplier_name,
                    unit_price=unit_price,
                    description=_cell_text(values[5]) or None,
                    note=note,
                    tags=tags,
                    variant_options=(),
                    default_moq=None,
                    gross_weight=None,
                    image_urls=tuple(image_urls),
                    image_url_columns=tuple(image_url_columns),
                    embedded_images=embedded_images,
                )
            )
        if issues:
            raise ProductTemplateValidationError(
                _validation_summary(issues),
                issues=tuple(issues),
            )
        if not rows:
            raise ProductTemplateValidationError("模版中没有可导入的有效商品。")
        if generated_sku_count:
            warnings.insert(
                0,
                f"有 {generated_sku_count} 行未填写商品型号，"
                "系统已根据商品名称、分类和供应商生成临时型号。",
            )
        unmatched_embedded_images = sum(
            len(images)
            for row_number, images in embedded_images_by_row.items()
            if row_number not in visited_rows
        )
        if unmatched_embedded_images:
            warnings.append(
                f"有 {unmatched_embedded_images} 张内嵌图片未对应到有效商品行，已忽略。"
            )
        return ProductTemplateParseResult(
            rows=tuple(rows),
            warnings=tuple(warnings),
            skipped_rows=skipped_rows,
        )
    finally:
        workbook.close()


def _category_code(name: str) -> str:
    digest = hashlib.sha256(name.encode("utf-8")).hexdigest()[:12].upper()
    return f"TPL-{digest}"


def _product_code(product_key: str) -> str:
    # Product code is an internal, deterministic ownership marker. In the
    # original template the key equals the SKU; in the variant template all
    # rows of one product deliberately share the same key.
    digest = hashlib.sha256(product_key.encode("utf-8")).hexdigest()[:24].upper()
    return f"TPL-{digest}"


def _alternate_product_code(product_key: str) -> str:
    digest = hashlib.sha256(
        f"PRODUCT_TEMPLATE:{product_key}".encode("utf-8")
    ).hexdigest()
    return f"TPLX-{digest[:48].upper()}"


def _supplier_name_key(name: str) -> str:
    """Return the tenant-local identity used when matching supplier names.

    Supplier records belong to a merchant.  Whitespace differences in an
    uploaded workbook should not create a second record for that same
    merchant, so use the same normalization for planning and assignment.
    """

    return " ".join(name.split()).casefold()


def _supplier_identity(tenant_id: UUID, name: str) -> tuple[str, str]:
    """Build an auto-created supplier identity scoped to one merchant.

    ``suppliers.id`` is still a global primary key for legacy reasons, while
    the business identity is tenant-local.  Including the tenant in the
    digest means two merchants may both import a supplier with the same name
    without colliding.  Existing records are matched by name before this
    helper is called, so re-imports in the same merchant continue to append to
    the existing supplier.
    """

    identity = f"{tenant_id}:{_supplier_name_key(name)}"
    digest = hashlib.sha256(identity.encode("utf-8")).hexdigest().upper()
    return f"SUP-TPL-{digest[:24]}", f"TPL-{digest[:24]}"


def _template_option_values(
    row: ProductTemplateRow,
    *,
    existing: dict[str, object] | None = None,
    base_product: bool = False,
) -> dict[str, object]:
    # The template owns its internal marker and the free-text note, but it
    # must not erase variant attributes maintained elsewhere in the product
    # center (for example color or size).
    values: dict[str, object] = dict(existing or {})
    previous_marker = values.get(TEMPLATE_SOURCE_KEY)
    previous_variant_keys = (
        previous_marker.get("variant_option_keys", [])
        if isinstance(previous_marker, dict)
        else []
    )
    for key in previous_variant_keys:
        if isinstance(key, str):
            values.pop(key, None)
    marker: dict[str, object] = {
        "source": TEMPLATE_SOURCE_VALUE,
        "schema": row.schema_version,
    }
    if base_product:
        marker[TEMPLATE_BASE_PRODUCT_FLAG] = True
    if row.schema_version >= 3:
        marker["product_code"] = row.product_key.removeprefix("PRODUCT:")
        marker["variant_option_keys"] = [key for key, _value in row.variant_options]
        for key, value in row.variant_options:
            values[key] = value
    values.pop("商品编码", None)
    values[TEMPLATE_SOURCE_KEY] = marker
    if row.note:
        values["备注"] = row.note
    else:
        values.pop("备注", None)
    if row.schema_version >= 2:
        if row.product_model:
            values["商品型号"] = row.product_model
        else:
            values.pop("商品型号", None)
        if row.specification:
            values["规格名称"] = row.specification
        else:
            values.pop("规格名称", None)
        if row.schema_version >= 4:
            values.pop("一箱个数", None)
            values.pop("是否是新品", None)
            if row.units_per_carton:
                values["装箱数"] = row.units_per_carton
            else:
                values.pop("装箱数", None)
            if row.gross_weight is not None:
                values["毛重"] = _decimal_option_text(row.gross_weight)
            else:
                values.pop("毛重", None)
            if row.default_moq is not None:
                values["起定数"] = _decimal_option_text(row.default_moq)
            else:
                values.pop("起定数", None)
        else:
            values.pop("装箱数", None)
            values.pop("毛重", None)
            values.pop("起定数", None)
            if row.units_per_carton:
                values["一箱个数"] = row.units_per_carton
            else:
                values.pop("一箱个数", None)
            if row.is_new:
                values["是否是新品"] = True
            else:
                values.pop("是否是新品", None)
    else:
        # Keep the historical SKU-row template contract unchanged. These keys
        # belong to the product/variant template and must not be reintroduced
        # after a merchant edits ordinary SKU option values.
        values.pop("商品型号", None)
        values.pop("规格名称", None)
        values.pop("一箱个数", None)
        values.pop("装箱数", None)
        values.pop("毛重", None)
        values.pop("起定数", None)
        values.pop("是否是新品", None)
    return values


def _is_template_managed_sku(sku: SkuRow) -> bool:
    marker = sku.option_values.get(TEMPLATE_SOURCE_KEY)
    return (
        isinstance(marker, dict)
        and marker.get("source") == TEMPLATE_SOURCE_VALUE
    )


def _is_base_product_sku(sku: SkuRow) -> bool:
    marker = sku.option_values.get(TEMPLATE_SOURCE_KEY)
    return (
        isinstance(marker, dict)
        and marker.get("source") == TEMPLATE_SOURCE_VALUE
        and marker.get(TEMPLATE_BASE_PRODUCT_FLAG) is True
    )


def _record_import_progress(
    *,
    job_id: str,
    tenant_id: UUID,
    progress: int,
    stage: str,
    processed_rows: int = 0,
    total_rows: int = 0,
) -> None:
    """Publish observable progress without committing the catalog transaction."""

    publish_runtime_import_progress(
        job_id=job_id,
        tenant_id=tenant_id,
        progress=progress,
        stage=stage,
        processed_rows=processed_rows,
        total_rows=total_rows,
    )
    try:
        with SessionLocal() as progress_session:
            if (
                progress_session.bind is not None
                and progress_session.bind.dialect.name == "sqlite"
            ):
                # SQLite only permits one writer. The catalog transaction must
                # remain atomic, so local progress is served from the in-process
                # registry instead of waiting repeatedly on the same write lock.
                return
            set_request_context(
                progress_session,
                organization_id=UUID(int=0),
                tenant_id=tenant_id,
                user_id=UUID(int=0),
            )
            worker = progress_session.scalar(
                select(WorkerJobRow)
                .where(
                    WorkerJobRow.tenant_id == tenant_id,
                    WorkerJobRow.import_job_id == job_id,
                )
                .order_by(WorkerJobRow.created_at.desc())
                .limit(1)
            )
            if worker is None:
                return
            checkpoint = dict(worker.checkpoint)
            checkpoint.update(
                {
                    "import_progress": max(0, min(100, progress)),
                    "import_stage": stage,
                    "processed_rows": max(0, processed_rows),
                    "total_rows": max(0, total_rows),
                }
            )
            worker.checkpoint = checkpoint
            progress_session.commit()
    except Exception:
        # Progress reporting must never make an otherwise valid import fail.
        return


def _fail_import(
    session,
    *,
    job: ImportJobRow,
    message: str,
    issues: tuple[ProductTemplateIssue, ...] = (),
) -> ProductTemplateImportResult:
    job.status = "failed"
    job.progress = 100
    job.warnings_count = max(1, job.warnings_count)
    job.error_message = message
    job.completed_at = utcnow()
    session.commit()
    _record_import_progress(
        job_id=job.id,
        tenant_id=job.tenant_id,
        progress=100,
        stage="FAILED",
    )
    return ProductTemplateImportResult(
        status="failed",
        warnings=(message,),
        issues=issues,
        message=message,
    )


def _image_content_type(url: str) -> str:
    content_type, _encoding = mimetypes.guess_type(urlsplit(url).path)
    return content_type if content_type and content_type.startswith("image/") else "image/jpeg"


def _filename_from_url(url: str) -> str:
    filename = Path(urlsplit(url).path).name
    return filename[:500] or "product-image"


def _embedded_storage_provider() -> str:
    return (
        "S3"
        if os.getenv("OBJECT_STORAGE_BACKEND", "local").casefold().strip() == "s3"
        else "LOCAL"
    )


def _safe_embedded_image_suffix(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    return suffix if re.fullmatch(r"\.[a-z0-9]{1,10}", suffix) else ".img"


def _embedded_image_object_key(
    *,
    tenant_id: UUID,
    sku_code: str,
    image: EmbeddedTemplateImage,
) -> str:
    sku_digest = hashlib.sha256(sku_code.encode("utf-8")).hexdigest()[:24]
    suffix = _safe_embedded_image_suffix(image.original_filename)
    return (
        f"tenants/{tenant_id}/approved-media/product-template/"
        f"{sku_digest}/{image.image_column:02d}-{image.sha256}{suffix}"
    )


def _embedded_original_filename(
    *,
    source_filename: str,
    image: EmbeddedTemplateImage,
) -> str:
    source_stem = Path(source_filename).stem.strip() or "商品"
    suffix = _safe_embedded_image_suffix(image.original_filename)
    filename = (
        f"{source_stem}-第{image.row_number}行-"
        f"商品图片{image.image_column}{suffix}"
    )
    return filename[-500:]


def _template_image_specs(
    row: ProductTemplateRow,
    *,
    tenant_id: UUID,
    source_filename: str,
) -> tuple[StoredTemplateImage, ...]:
    specs: list[StoredTemplateImage] = []
    for sequence, (image_url, image_column) in enumerate(
        zip(row.image_urls, row.image_url_columns, strict=True),
        start=1,
    ):
        specs.append(
            StoredTemplateImage(
                image_column=image_column,
                sequence=sequence,
                object_key=image_url,
                original_filename=_filename_from_url(image_url),
                content_type=_image_content_type(image_url),
                byte_size=0,
                sha256=hashlib.sha256(image_url.encode("utf-8")).hexdigest(),
                storage_provider="EXTERNAL",
            )
        )
    for image in row.embedded_images:
        specs.append(
            StoredTemplateImage(
                image_column=image.image_column,
                sequence=image.sequence,
                object_key=_embedded_image_object_key(
                    tenant_id=tenant_id,
                    sku_code=row.sku_code,
                    image=image,
                ),
                original_filename=_embedded_original_filename(
                    source_filename=source_filename,
                    image=image,
                ),
                content_type=image.content_type,
                byte_size=image.byte_size,
                sha256=image.sha256,
                storage_provider=_embedded_storage_provider(),
                archive_path=image.archive_path,
            )
        )
    return tuple(
        sorted(
            specs,
            key=lambda spec: (
                spec.image_column,
                0 if spec.archive_path is None else 1,
                spec.sequence,
            ),
        )
    )


def _store_new_embedded_images(
    source_path: Path,
    *,
    specs: tuple[StoredTemplateImage, ...],
    existing_object_keys: set[str],
    progress_callback: Callable[[int, int], None] | None = None,
) -> None:
    unique_specs: dict[str, StoredTemplateImage] = {}
    for spec in specs:
        if spec.archive_path is not None:
            unique_specs.setdefault(spec.object_key, spec)
    pending = [
        spec
        for object_key, spec in unique_specs.items()
        if object_key not in existing_object_keys
    ]
    if not pending:
        return

    storage = get_object_storage()
    staging_root = Path(
        os.getenv("OBJECT_STORAGE_STAGING_DIR", tempfile.gettempdir())
    )
    staging_root.mkdir(parents=True, exist_ok=True)
    try:
        configured_concurrency = int(
            os.getenv("PRODUCT_TEMPLATE_IMAGE_UPLOAD_CONCURRENCY", "6")
        )
    except ValueError:
        configured_concurrency = 6
    concurrency = min(16, max(1, configured_concurrency), len(pending))

    # Extract once, then upload several small product images concurrently.
    # R2 and other S3-compatible stores are latency-bound for this workload;
    # sequential HEAD + PUT requests make image-heavy workbooks needlessly slow.
    # Object keys contain the content hash, so an idempotent PUT is safe after a
    # previously interrupted import and removes the extra HEAD request.
    with tempfile.TemporaryDirectory(
        prefix="atc-template-images-",
        dir=staging_root,
    ) as raw_staging_dir:
        staging_dir = Path(raw_staging_dir)
        prepared: list[tuple[StoredTemplateImage, Path]] = []
        with ZipFile(source_path) as archive:
            for index, spec in enumerate(pending, start=1):
                assert spec.archive_path is not None
                suffix = _safe_embedded_image_suffix(spec.original_filename)
                path = staging_dir / f"{index:06d}{suffix}"
                digest = hashlib.sha256()
                byte_size = 0
                with path.open("wb") as output, archive.open(
                    spec.archive_path
                ) as source:
                    while chunk := source.read(1024 * 1024):
                        digest.update(chunk)
                        byte_size += len(chunk)
                        output.write(chunk)
                if byte_size != spec.byte_size or digest.hexdigest() != spec.sha256:
                    raise RuntimeError("embedded product image changed during import")
                prepared.append((spec, path))

        def upload(prepared_image: tuple[StoredTemplateImage, Path]) -> None:
            spec, path = prepared_image
            storage.put_file(
                path,
                object_key=spec.object_key,
                content_type=spec.content_type,
            )

        completed = 0
        if concurrency == 1:
            for prepared_image in prepared:
                upload(prepared_image)
                completed += 1
                if progress_callback is not None:
                    progress_callback(completed, len(prepared))
            return

        with ThreadPoolExecutor(
            max_workers=concurrency,
            thread_name_prefix="product-image-upload",
        ) as executor:
            futures = [executor.submit(upload, item) for item in prepared]
            for future in as_completed(futures):
                future.result()
                completed += 1
                if progress_callback is not None:
                    progress_callback(completed, len(prepared))


def _load_image_map(
    session,
    *,
    tenant_id: UUID,
    image_urls: set[str],
) -> dict[str, ProductImageRow]:
    result: dict[str, ProductImageRow] = {}
    urls = list(image_urls)
    for start in range(0, len(urls), 500):
        chunk = urls[start : start + 500]
        rows = session.scalars(
            select(ProductImageRow)
            .where(
                ProductImageRow.tenant_id == tenant_id,
                ProductImageRow.object_key.in_(chunk),
            )
            .execution_options(include_deleted=True)
        ).all()
        result.update({row.object_key: row for row in rows})
    return result


def process_product_template_import(
    job_id: str,
    *,
    tenant_id: UUID,
    source_path: Path,
) -> ProductTemplateImportResult:
    with SessionLocal() as session:
        set_request_context(
            session,
            organization_id=UUID(int=0),
            tenant_id=tenant_id,
            user_id=UUID(int=0),
        )
        statement = (
            select(ImportJobRow)
            .options(selectinload(ImportJobRow.source_file))
            .where(
                ImportJobRow.id == job_id,
                ImportJobRow.tenant_id == tenant_id,
            )
        )
        job = session.scalar(statement)
        if job is None:
            raise RuntimeError("product template import job was not found")
        job.status = "parsing"
        job.progress = 25
        session.commit()
        _record_import_progress(
            job_id=job.id,
            tenant_id=tenant_id,
            progress=25,
            stage="READING_WORKBOOK",
        )

        def report_row_validation(processed_rows: int, total_rows: int) -> None:
            fraction = processed_rows / total_rows if total_rows else 1
            _record_import_progress(
                job_id=job.id,
                tenant_id=tenant_id,
                progress=25 + int(min(1, fraction) * 15),
                stage="VALIDATING_ROWS",
                processed_rows=processed_rows,
                total_rows=total_rows,
            )

        try:
            parsed = parse_product_template(
                source_path,
                progress_callback=report_row_validation,
            )
        except ProductTemplateValidationError as exc:
            session.rollback()
            job = session.scalar(statement)
            if job is None:
                raise RuntimeError("product template import job disappeared") from exc
            failure_issues = exc.issues or (
                _issue(
                    row_number=None,
                    column="商品文件",
                    code="FILE_VALIDATION_ERROR",
                    message=str(exc),
                    suggestion="请下载最新模板，确认工作表、表头和商品数据后重新导入。",
                ),
            )
            job.status = "failed"
            job.progress = 100
            job.warnings_count = max(1, len(failure_issues), job.warnings_count)
            job.error_message = str(exc)
            job.completed_at = utcnow()
            session.commit()
            _record_import_progress(
                job_id=job.id,
                tenant_id=tenant_id,
                progress=100,
                stage="VALIDATION_FAILED",
            )
            warning_messages = tuple(issue.message for issue in failure_issues)
            return ProductTemplateImportResult(
                status="failed",
                warnings=warning_messages,
                issues=failure_issues,
                message=str(exc),
            )

        job = session.scalar(statement)
        if job is None:
            raise RuntimeError("product template import job disappeared")
        job.progress = 40
        session.commit()
        _record_import_progress(
            job_id=job.id,
            tenant_id=tenant_id,
            progress=40,
            stage="VALIDATING_ROWS",
            total_rows=len(parsed.rows),
        )
        job = session.scalar(statement)
        if job is None:
            raise RuntimeError("product template import job disappeared")
        # Serialize import batches per tenant. PostgreSQL takes a row lock
        # here; SQLite already serializes writers. Once the lock is held, an
        # older retry is rejected if a newer import has already won.
        tenant = session.scalar(
            select(TenantRow)
            .where(TenantRow.id == tenant_id)
            .with_for_update()
        )
        if tenant is None:
            return _fail_import(
                session,
                job=job,
                message="当前租户不存在或已停用，未执行本次商品模版同步。",
            )
        if job.batch_id is not None:
            batch = session.scalar(
                select(CatalogImportBatchRow)
                .where(
                    CatalogImportBatchRow.tenant_id == tenant_id,
                    CatalogImportBatchRow.id == job.batch_id,
                    CatalogImportBatchRow.deleted_at.is_(None),
                )
                .execution_options(populate_existing=True)
                .with_for_update()
            )
            if batch is None or batch.status != "ACTIVE":
                return _fail_import(
                    session,
                    job=job,
                    message="导入批次已被撤回或不再可用，本文件未写入商品库。",
                )
        newer_published_job_id = session.scalar(
            select(ImportJobRow.id)
            .where(
                ImportJobRow.tenant_id == tenant_id,
                ImportJobRow.source_type == "PRODUCT_TEMPLATE",
                ImportJobRow.status == "published",
                ImportJobRow.id != job.id,
                or_(
                    ImportJobRow.created_at > job.created_at,
                    and_(
                        ImportJobRow.created_at == job.created_at,
                        ImportJobRow.id > job.id,
                    ),
                ),
            )
            .order_by(ImportJobRow.created_at.desc())
            .limit(1)
            .execution_options(include_deleted=True)
        )
        if newer_published_job_id is not None:
            message = (
                "本次商品模版早于已经生效的新版本，系统已跳过旧导入批次，"
                "不会覆盖当前商品库。"
            )
            job.status = "failed"
            job.progress = 100
            job.warnings_count = 1
            job.error_message = message
            job.completed_at = utcnow()
            session.commit()
            return ProductTemplateImportResult(
                status="superseded",
                warnings=(message,),
                message=message,
            )
        currency = tenant.default_currency
        media = (
            session.get(MediaObjectRow, job.source_file.media_object_id)
            if job.source_file.media_object_id
            else None
        )
        user_id = media.created_by_user_id if media else None
        now = utcnow()

        category_rows = list(
            session.scalars(
                select(ProductCategoryRow)
                .where(
                    ProductCategoryRow.tenant_id == tenant_id,
                )
                .execution_options(include_deleted=True)
            ).all()
        )
        categories = {row.code: row for row in category_rows}
        categories_by_parent_and_name: dict[
            tuple[UUID | None, str], ProductCategoryRow
        ] = {}
        # Category codes differ between manually created, category-template,
        # and product-template records. Name + parent is the human identity
        # shared by every import path, so prefer an active existing sibling
        # before considering a new product-template category.
        for row in sorted(
            category_rows,
            key=lambda item: (
                item.deleted_at is not None,
                item.status != "ACTIVE",
                item.sort_order,
                str(item.id),
            ),
        ):
            categories_by_parent_and_name.setdefault(
                (row.parent_id, category_name_key(row.name)),
                row,
            )
        product_rows = session.scalars(
            select(ProductRow)
            .where(
                ProductRow.tenant_id == tenant_id,
            )
            .execution_options(include_deleted=True)
        ).all()
        products = {
            row.product_code: row
            for row in product_rows
            if row.product_code
        }
        sku_rows = session.scalars(
            select(SkuRow)
            .where(
                SkuRow.tenant_id == tenant_id,
            )
            .execution_options(include_deleted=True)
        ).all()
        sku_rows_by_product: dict[UUID, list[SkuRow]] = defaultdict(list)
        for sku_row in sku_rows:
            sku_rows_by_product[sku_row.product_id].append(sku_row)
        sku_code_allocator = CatalogSkuCodeAllocator(
            tenant=tenant,
            products=product_rows,
            skus=sku_rows,
            issued_at=now,
        )
        sku_groups: dict[str, list[SkuRow]] = defaultdict(list)
        for sku_row in sku_rows:
            sku_groups[_source_sku_identity(sku_row)].append(sku_row)

        skus = {
            code: rows[0]
            for code, rows in sku_groups.items()
            if len(rows) == 1
        }
        current_sku_count = sum(1 for row in sku_rows if row.deleted_at is None)

        def product_only_consumes_capacity(template_row: ProductTemplateRow) -> bool:
            """Return whether importing a no-spec product creates an active SKU."""

            candidate_product = next(
                (
                    products.get(candidate_code)
                    for candidate_code in (
                        _product_code(template_row.product_key),
                        _alternate_product_code(template_row.product_key),
                    )
                    if products.get(candidate_code) is not None
                ),
                None,
            )
            if candidate_product is None:
                return True
            return not any(
                row.deleted_at is None
                for row in sku_rows_by_product.get(candidate_product.id, ())
            )

        additional_sku_count = sum(
            1
            for row in parsed.rows
            if (
                product_only_consumes_capacity(row)
                if row.product_only
                else (
                    (existing := skus.get(_normalize_sku_code(row.sku_code))) is None
                    or existing.deleted_at is not None
                )
            )
        )
        quota = sku_quota_snapshot(
            session,
            tenant_id=tenant_id,
            additional=additional_sku_count,
            current_count=current_sku_count,
            lock_tenant=False,
        )
        remaining_capacity = quota.remaining
        import_rows: list[ProductTemplateRow] = []
        quota_skipped_rows: list[ProductTemplateRow] = []
        for template_row in parsed.rows:
            if template_row.product_only:
                consumes_capacity = product_only_consumes_capacity(template_row)
            else:
                existing = skus.get(_normalize_sku_code(template_row.sku_code))
                consumes_capacity = existing is None or existing.deleted_at is not None
            if (
                consumes_capacity
                and remaining_capacity is not None
                and remaining_capacity <= 0
            ):
                quota_skipped_rows.append(template_row)
                continue
            import_rows.append(template_row)
            if consumes_capacity and remaining_capacity is not None:
                remaining_capacity -= 1

        accepted_rows = tuple(import_rows)
        quota_skipped_count = len(quota_skipped_rows)
        quota_warnings: list[str] = []
        quota_issues: tuple[ProductTemplateIssue, ...] = ()
        if quota_skipped_rows:
            assert quota.limit is not None
            first_skipped = quota_skipped_rows[0]
            accepted_new_count = additional_sku_count - quota_skipped_count
            warning = (
                f"当前等级最多可保留 {quota.limit} 个 SKU；导入前已使用 "
                f"{quota.current} 个，本次新增 {accepted_new_count} 个，"
                f"另有 {quota_skipped_count} 行超出额度未导入。"
            )
            quota_warnings.append(warning)
            sample_codes = "、".join(
                row.sku_code for row in quota_skipped_rows[:5]
            )
            if quota_skipped_count > 5:
                sample_codes += f" 等 {quota_skipped_count} 个 SKU"
            quota_issues = (
                _issue(
                    row_number=first_skipped.row_number,
                    column="SKU编号",
                    code="SKU_LIMIT_EXCEEDED",
                    message=(
                        f"共有 {quota_skipped_count} 行因超出 SKU 额度未导入，"
                        f"首个未导入项位于第 {first_skipped.row_number} 行；"
                        "额度内商品及已有 SKU 的更新已正常处理。"
                    ),
                    value=sample_codes,
                    suggestion=(
                        "删除不再使用的 SKU、提高商家额度后重新导入；"
                        "重新导入不会重复创建已成功写入的 SKU。"
                    ),
                ),
            )

        incoming_sku_codes = {
            row.sku_code for row in accepted_rows if not row.product_only
        }
        conflicting_codes = sorted(
            code
            for code in incoming_sku_codes
            if len(sku_groups.get(code, ())) > 1
        )
        if conflicting_codes:
            return _fail_import(
                session,
                job=job,
                message=(
                    "租户商品库存在忽略大小写或首尾空白后重复的 SKU："
                    f"{'、'.join(conflicting_codes[:5])}。请先合并重复记录后再导入。"
                ),
            )

        supplier_rows = session.scalars(
            select(SupplierRow)
            .where(SupplierRow.tenant_id == tenant_id)
            .execution_options(include_deleted=True)
        ).all()
        suppliers_by_name: dict[str, list[SupplierRow]] = defaultdict(list)
        suppliers_by_id = {row.id: row for row in supplier_rows}
        suppliers_by_code = {row.supplier_code: row for row in supplier_rows}
        for supplier_row in supplier_rows:
            suppliers_by_name[_supplier_name_key(supplier_row.name)].append(
                supplier_row
            )

        requested_supplier_names: dict[str, str] = {}
        for template_row in accepted_rows:
            if template_row.supplier_name:
                requested_supplier_names.setdefault(
                    _supplier_name_key(template_row.supplier_name),
                    template_row.supplier_name,
                )

        supplier_plan: dict[str, tuple[str, SupplierRow | None]] = {}
        for normalized_name, display_name in requested_supplier_names.items():
            matches = suppliers_by_name.get(normalized_name, [])
            live_matches = [row for row in matches if row.deleted_at is None]
            if len(live_matches) > 1 or (not live_matches and len(matches) > 1):
                return _fail_import(
                    session,
                    job=job,
                    message=(
                        f"供应商“{display_name}”在当前商家下存在多个同名记录，"
                        "请先合并供应商后再导入。"
                    ),
                )
            existing_supplier = (
                live_matches[0]
                if live_matches
                else matches[0] if matches else None
            )
            if existing_supplier is None:
                supplier_id, supplier_code = _supplier_identity(
                    tenant_id,
                    display_name,
                )
                id_collision = suppliers_by_id.get(supplier_id)
                code_collision = suppliers_by_code.get(supplier_code)
                if (
                    id_collision is not None
                    and _supplier_name_key(id_collision.name) != normalized_name
                ) or (
                    code_collision is not None
                    and _supplier_name_key(code_collision.name) != normalized_name
                ):
                    return _fail_import(
                        session,
                        job=job,
                        message=(
                            f"供应商“{display_name}”的自动编码与现有供应商冲突，"
                            "请先在供应商管理中创建该供应商。"
                        ),
                    )
            supplier_plan[normalized_name] = (display_name, existing_supplier)

        products_by_id = {row.id: row for row in product_rows}
        existing_sku_ids = [row.id for row in sku_rows]
        offers = (
            {
                row.sku_id: row
                for row in session.scalars(
                    select(PublicCatalogOfferRow)
                    .where(
                        PublicCatalogOfferRow.tenant_id == tenant_id,
                        PublicCatalogOfferRow.sku_id.in_(existing_sku_ids),
                    )
                    .execution_options(include_deleted=True)
                ).all()
            }
            if existing_sku_ids
            else {}
        )
        template_rows_by_product_key: dict[str, list[ProductTemplateRow]] = (
            defaultdict(list)
        )
        for template_row in accepted_rows:
            template_rows_by_product_key[template_row.product_key].append(
                template_row
            )
        image_specs_by_product_key: dict[
            str, tuple[StoredTemplateImage, ...]
        ] = {}
        for product_key, template_rows in template_rows_by_product_key.items():
            specs_by_content_and_column: dict[
                tuple[str, int], StoredTemplateImage
            ] = {}
            for template_row in template_rows:
                for spec in _template_image_specs(
                    template_row,
                    tenant_id=tenant_id,
                    source_filename=job.source_file.original_filename,
                ):
                    # The same image repeated in the same image column across
                    # several variant rows is one product image. The same
                    # bytes placed in two distinct image columns are two
                    # intentional gallery slots and must both be retained.
                    specs_by_content_and_column.setdefault(
                        (spec.sha256, spec.image_column),
                        spec,
                    )
            image_specs_by_product_key[product_key] = tuple(
                sorted(
                    specs_by_content_and_column.values(),
                    key=lambda spec: (
                        spec.image_column,
                        0 if spec.archive_path is None else 1,
                        spec.sequence,
                    ),
                )
            )
        all_image_specs = tuple(
            spec
            for specs in image_specs_by_product_key.values()
            for spec in specs
        )
        images = _load_image_map(
            session,
            tenant_id=tenant_id,
            image_urls={spec.object_key for spec in all_image_specs},
        )
        embedded_specs = tuple(
            spec for spec in all_image_specs if spec.archive_path is not None
        )
        _store_new_embedded_images(
            source_path,
            specs=embedded_specs,
            # A withdrawn batch leaves soft-deleted image rows for audit. The
            # corresponding R2 objects have already been removed, so only
            # active rows can suppress a fresh upload on a later re-import.
            existing_object_keys={
                object_key
                for object_key, image in images.items()
                if image.deleted_at is None
            },
            progress_callback=lambda processed, total: _record_import_progress(
                job_id=job.id,
                tenant_id=tenant_id,
                progress=55 + int((processed / total) * 9) if total else 64,
                stage="STORING_IMAGES",
                processed_rows=processed,
                total_rows=total,
            ),
        )
        template_images_by_product: dict[UUID, list[ProductImageRow]] = defaultdict(list)
        for image in session.scalars(
            select(ProductImageRow)
            .where(
                ProductImageRow.tenant_id == tenant_id,
                ProductImageRow.bucket == TEMPLATE_IMAGE_BUCKET,
            )
            .execution_options(include_deleted=True)
        ).all():
            template_images_by_product[image.product_id].append(image)

        _record_import_progress(
            job_id=job.id,
            tenant_id=tenant_id,
            progress=55,
            stage="LOADING_CATALOG",
            total_rows=len(accepted_rows),
        )

        # A product/variant template intentionally maps several SKU rows onto
        # one ProductRow. The original template uses a unique product_key per
        # SKU, so it retains its historical one-row-per-product behavior.
        planned_product_by_key: dict[str, ProductRow | None] = {}
        planned_product_code_by_key: dict[str, str | None] = {}
        for product_key, template_rows in template_rows_by_product_key.items():
            product_only = any(row.product_only for row in template_rows)
            incoming_codes = {
                row.sku_code for row in template_rows if not row.product_only
            }
            current_products = list(
                dict.fromkeys(
                    products_by_id[sku.product_id]
                    for row in template_rows
                    if not row.product_only
                    if (sku := skus.get(row.sku_code)) is not None
                    and sku.product_id in products_by_id
                )
            )
            selected_product: ProductRow | None = None
            selected_code: str | None = None
            candidate_codes = (
                _product_code(product_key),
                _alternate_product_code(product_key),
            )
            for candidate_code in candidate_codes:
                candidate = products.get(candidate_code)
                candidate_skus = (
                    sku_rows_by_product.get(candidate.id, ())
                    if candidate is not None
                    else ()
                )
                candidate_sku_codes = {
                    _source_sku_identity(row)
                    for row in candidate_skus
                    if not _is_base_product_sku(row)
                }
                if (
                    candidate is None
                    or product_only
                    or candidate in current_products
                    or not candidate_sku_codes
                    or bool(candidate_sku_codes & incoming_codes)
                ):
                    selected_product = candidate
                    selected_code = candidate_code
                    break
            if selected_code is None:
                reusable_product = next(
                    (
                        product
                        for product in current_products
                        if {
                            _source_sku_identity(row)
                            for row in sku_rows_by_product.get(product.id, ())
                            if not _is_base_product_sku(row)
                        }.issubset(incoming_codes)
                    ),
                    None,
                )
                if reusable_product is not None:
                    selected_product = reusable_product
                else:
                    return _fail_import(
                        session,
                        job=job,
                        message=(
                            f"商品“{template_rows[0].name}”无法分配商品记录；"
                            "保留的模板商品编码已被其他商品占用。"
                        ),
                    )
            planned_product_by_key[product_key] = selected_product
            planned_product_code_by_key[product_key] = (
                selected_code if selected_product is None else None
            )

        _record_import_progress(
            job_id=job.id,
            tenant_id=tenant_id,
            progress=65,
            stage="PLANNING_CHANGES",
            total_rows=len(accepted_rows),
        )

        suppliers_for_template: dict[str, SupplierRow] = {}
        new_supplier_ids: set[str] = set()
        for normalized_name, (display_name, existing_supplier) in supplier_plan.items():
            supplier = existing_supplier
            if supplier is None:
                supplier_id, supplier_code = _supplier_identity(
                    tenant_id,
                    display_name,
                )
                supplier = SupplierRow(
                    id=supplier_id,
                    tenant_id=tenant_id,
                    supplier_code=supplier_code,
                    name=display_name,
                    category="商品模版",
                    category_summary="由商品模版自动创建",
                    status="ACTIVE",
                    risk_level="UNKNOWN",
                    active_skus=0,
                    health="good",
                )
                session.add(supplier)
                supplier_rows.append(supplier)
                suppliers_by_id[supplier.id] = supplier
                suppliers_by_code[supplier.supplier_code] = supplier
                new_supplier_ids.add(supplier.id)
            elif supplier.deleted_at is not None:
                supplier.deleted_at = None
                supplier.status = "ACTIVE"
                supplier.version += 1
            suppliers_for_template[normalized_name] = supplier
        if suppliers_for_template:
            # Establish auto-created suppliers before inserting SKU rows that
            # reference them through the tenant-scoped composite foreign key.
            session.flush()

        created = 0
        updated = 0
        unchanged = 0
        dirty_product_ids: set[UUID] = set()
        imported_base_product_ids: set[UUID] = set()
        touched_supplier_ids: set[str] = set()
        synced_image_product_keys: set[str] = set()
        moved_from_product_ids: set[UUID] = set()
        runtime_warnings = [*parsed.warnings, *quota_warnings]

        def sync_product_images(
            product: ProductRow,
            template_row: ProductTemplateRow,
        ) -> bool:
            if template_row.product_key in synced_image_product_keys:
                return False
            synced_image_product_keys.add(template_row.product_key)
            image_changed = False
            product_image_specs = image_specs_by_product_key[
                template_row.product_key
            ]
            desired_image_keys = {
                spec.object_key for spec in product_image_specs
            }
            for old_image in template_images_by_product.get(product.id, ()):
                if (
                    old_image.object_key not in desired_image_keys
                    and old_image.deleted_at is None
                ):
                    old_image.deleted_at = now
                    image_changed = True

            for image_index, image_spec in enumerate(product_image_specs):
                image = images.get(image_spec.object_key)
                if image is not None and image.product_id != product.id:
                    runtime_warnings.append(
                        f"商品“{template_row.name}”的图片"
                        f"{image_spec.image_column}已被其他商品使用，已跳过该图片。"
                    )
                    continue
                if image is None:
                    image = ProductImageRow(
                        id=uuid4(),
                        tenant_id=tenant_id,
                        product_id=product.id,
                        storage_provider=image_spec.storage_provider,
                        bucket=TEMPLATE_IMAGE_BUCKET,
                        object_key=image_spec.object_key,
                        original_filename=image_spec.original_filename,
                        content_type=image_spec.content_type,
                        byte_size=image_spec.byte_size,
                        sha256=image_spec.sha256,
                        image_role="MAIN" if image_index == 0 else "GALLERY",
                        sort_order=image_index,
                        approval_status="APPROVED",
                        alt_text=template_row.name,
                        created_by=user_id,
                    )
                    session.add(image)
                    images[image_spec.object_key] = image
                    template_images_by_product[product.id].append(image)
                    image_changed = True
                    continue

                image_values = {
                    "storage_provider": image_spec.storage_provider,
                    "bucket": TEMPLATE_IMAGE_BUCKET,
                    "original_filename": image_spec.original_filename,
                    "content_type": image_spec.content_type,
                    "byte_size": image_spec.byte_size,
                    "sha256": image_spec.sha256,
                    "deleted_at": None,
                    "image_role": "MAIN" if image_index == 0 else "GALLERY",
                    "sort_order": image_index,
                    "approval_status": "APPROVED",
                    "alt_text": template_row.name,
                }
                if any(
                    getattr(image, key) != value
                    for key, value in image_values.items()
                ):
                    for key, value in image_values.items():
                        setattr(image, key, value)
                    image_changed = True
                if image not in template_images_by_product[product.id]:
                    template_images_by_product[product.id].append(image)
            return image_changed

        progress_interval = max(1, len(accepted_rows) // 100)
        for row_index, template_row in enumerate(accepted_rows, start=1):
            if row_index == 1 or row_index % progress_interval == 0:
                progress = 70 + int((row_index / len(accepted_rows)) * 22)
                _record_import_progress(
                    job_id=job.id,
                    tenant_id=tenant_id,
                    progress=progress,
                    stage="APPLYING_PRODUCTS",
                    processed_rows=row_index,
                    total_rows=len(accepted_rows),
                )
            changed = False
            supplier = (
                suppliers_for_template.get(
                    _supplier_name_key(template_row.supplier_name)
                )
                if template_row.supplier_name
                else None
            )
            category_parts = tuple(template_row.category.split("/"))
            parent: ProductCategoryRow | None = None
            category: ProductCategoryRow | None = None
            for depth, category_name in enumerate(category_parts, start=1):
                imported_category_path = "/".join(category_parts[:depth])
                category_code = _category_code(imported_category_path)
                parent_id = parent.id if parent is not None else None
                category_identity = (
                    parent_id,
                    category_name_key(category_name),
                )
                category = categories_by_parent_and_name.get(category_identity)
                if category is None:
                    category = categories.get(category_code)
                effective_name = (
                    category.name if category is not None else category_name
                )
                category_path = (
                    f"{parent.path or parent.name}/{effective_name}"
                    if parent is not None
                    else effective_name
                )
                category_values = {
                    "parent_id": parent_id,
                    "name": effective_name,
                    "path": category_path,
                    "status": "ACTIVE",
                    "deleted_at": None,
                }
                if category is None:
                    category = ProductCategoryRow(
                        id=uuid4(),
                        tenant_id=tenant_id,
                        code=category_code,
                        **category_values,
                    )
                    session.add(category)
                    categories[category_code] = category
                    categories_by_parent_and_name[category_identity] = category
                    session.flush()
                    changed = True
                elif any(
                    getattr(category, key) != value
                    for key, value in category_values.items()
                ):
                    for key, value in category_values.items():
                        setattr(category, key, value)
                    category.version += 1
                    changed = True
                # Keep both indexes hot for later rows in this file. An
                # existing category may use a manual or category-template
                # code, while this import uses a deterministic TPL code.
                categories[category_code] = category
                categories_by_parent_and_name[category_identity] = category
                parent = category
            assert category is not None

            sku = (
                None
                if template_row.product_only
                else skus.get(template_row.sku_code)
            )
            if sku is not None and sku.supplier_id is not None:
                touched_supplier_ids.add(sku.supplier_id)
            product = planned_product_by_key[template_row.product_key]
            product_code = planned_product_code_by_key[template_row.product_key]
            is_new = product is None if template_row.product_only else sku is None
            if product is None:
                assert product_code is not None
                product = ProductRow(
                    id=uuid4(),
                    tenant_id=tenant_id,
                    product_code=product_code,
                    name=template_row.name,
                    description=template_row.description,
                    category_id=category.id,
                    status="ACTIVE",
                    default_unit="piece",
                    created_by=user_id,
                    updated_by=user_id,
                )
                session.add(product)
                sku_code_allocator.ensure_product(product)
                products[product_code] = product
                products_by_id[product.id] = product
                planned_product_by_key[template_row.product_key] = product
                # Product and category must exist before the composite SKU
                # foreign key is evaluated (notably by SQLite executemany).
                session.flush()
                changed = True
            else:
                product_values = {
                    "name": template_row.name,
                    "description": template_row.description,
                    "category_id": category.id,
                    "status": "ACTIVE",
                    "default_unit": product.default_unit or "piece",
                    "archived_at": None,
                    "deleted_at": None,
                }
                if any(getattr(product, key) != value for key, value in product_values.items()):
                    for key, value in product_values.items():
                        setattr(product, key, value)
                    product.current_version += 1
                    product.updated_by = user_id
                    changed = True

            if template_row.product_only:
                # A Product row without any explicit SKU is still a concrete
                # catalog SKU: one generated, no-specification base SKU. Reuse
                # the marked row on subsequent imports so this operation is
                # idempotent and does not consume another SKU sequence.
                live_skus = [
                    row
                    for row in sku_rows_by_product.get(product.id, ())
                    if (
                        row.deleted_at is None
                        and row.status != "ARCHIVED"
                        and not _is_base_product_sku(row)
                    )
                ]
                base_sku = next(
                    (
                        row
                        for row in sku_rows_by_product.get(product.id, ())
                        if _is_base_product_sku(row)
                    ),
                    None,
                )
                if not live_skus:
                    if base_sku is None:
                        system_sku_code, sku_sequence = sku_code_allocator.issue(product)
                        base_sku = SkuRow(
                            id=uuid4(),
                            tenant_id=tenant_id,
                            product_id=product.id,
                            supplier_id=None,
                            latest_import_job_id=job.id,
                            rollback_owner_batch_id=job.batch_id,
                            sku_code=system_sku_code,
                            source_sku_code=None,
                            sku_sequence=sku_sequence,
                            name=template_row.sku_name or template_row.name,
                            option_values=_template_option_values(
                                template_row,
                                base_product=True,
                            ),
                            default_moq=template_row.default_moq,
                            moq_unit=None,
                            weight=template_row.gross_weight,
                            weight_unit=(
                                "kg" if template_row.gross_weight is not None else None
                            ),
                            status="ACTIVE",
                            created_by_user_id=user_id,
                            updated_by_user_id=user_id,
                        )
                        session.add(base_sku)
                        sku_rows.append(base_sku)
                        sku_rows_by_product[product.id].append(base_sku)
                        # Public offers reference the SKU through a composite
                        # tenant foreign key, so establish it first.
                        session.flush()
                        changed = True
                    else:
                        base_values = {
                            "product_id": product.id,
                            "supplier_id": None,
                            "source_sku_code": None,
                            "name": template_row.sku_name or template_row.name,
                            "option_values": _template_option_values(
                                template_row,
                                existing=base_sku.option_values,
                                base_product=True,
                            ),
                            "default_moq": template_row.default_moq,
                            "moq_unit": None,
                            "weight": template_row.gross_weight,
                            "weight_unit": (
                                "kg" if template_row.gross_weight is not None else None
                            ),
                            "status": "ACTIVE",
                            "deleted_at": None,
                        }
                        if any(
                            getattr(base_sku, key) != value
                            for key, value in base_values.items()
                        ):
                            for key, value in base_values.items():
                                setattr(base_sku, key, value)
                            base_sku.version += 1
                            base_sku.updated_by_user_id = user_id
                            changed = True
                        base_sku.latest_import_job_id = job.id
                        if base_sku.rollback_owner_batch_id != job.batch_id:
                            base_sku.rollback_owner_batch_id = None
                    sku = base_sku
                    imported_base_product_ids.add(product.id)
                else:
                    # A product that already has a concrete SKU does not need
                    # another base row; this product-only row still syncs the
                    # product master data and gallery below.
                    sku = None

                if sku is None:
                    if sync_product_images(product, template_row):
                        changed = True
                    if is_new:
                        created += 1
                    elif changed:
                        updated += 1
                    else:
                        unchanged += 1
                    if changed and product.status == "ACTIVE":
                        product.search_document_version = 0
                        dirty_product_ids.add(product.id)
                    if row_index % IMPORT_FLUSH_BATCH_SIZE == 0:
                        session.flush()
                    continue

            if not template_row.product_only:
                # A generated base SKU is synthetic and should be retired when
                # real variant rows arrive for the same product. It remains in
                # history (soft-deleted) but no longer appears as an extra
                # active variant.
                for base_sku in list(sku_rows_by_product.get(product.id, ())):
                    if not _is_base_product_sku(base_sku) or base_sku.id == getattr(sku, "id", None):
                        continue
                    if base_sku.deleted_at is None and base_sku.status != "ARCHIVED":
                        base_sku.status = "ARCHIVED"
                        base_sku.deleted_at = now
                        base_sku.version += 1
                        base_sku.updated_by_user_id = user_id
                        base_offer = offers.get(base_sku.id)
                        if base_offer is not None:
                            base_offer.publication_status = "SUSPENDED"
                            base_offer.deleted_at = now
                        changed = True

            if sku is None:
                system_sku_code, sku_sequence = sku_code_allocator.issue(product)
                sku = SkuRow(
                    id=uuid4(),
                    tenant_id=tenant_id,
                    product_id=product.id,
                    supplier_id=supplier.id if supplier is not None else None,
                    latest_import_job_id=job.id,
                    rollback_owner_batch_id=job.batch_id,
                    sku_code=system_sku_code,
                    source_sku_code=template_row.sku_code,
                    sku_sequence=sku_sequence,
                    name=template_row.sku_name,
                    option_values=_template_option_values(template_row),
                    default_moq=template_row.default_moq,
                    moq_unit=None,
                    weight=template_row.gross_weight,
                    weight_unit=(
                        "kg" if template_row.gross_weight is not None else None
                    ),
                    status="ACTIVE",
                    created_by_user_id=user_id,
                    updated_by_user_id=user_id,
                )
                session.add(sku)
                skus[template_row.sku_code] = sku
                sku_rows.append(sku)
                sku_rows_by_product[product.id].append(sku)
                # Public offers reference the SKU through a composite tenant
                # foreign key, so establish the SKU before staging its offer.
                session.flush()
                changed = True
            elif not template_row.product_only:
                old_product_id = sku.product_id
                sku_values = {
                    "product_id": product.id,
                    "supplier_id": supplier.id if supplier is not None else None,
                    "source_sku_code": template_row.sku_code,
                    "name": template_row.sku_name,
                    "option_values": _template_option_values(
                        template_row,
                        existing=sku.option_values,
                    ),
                    "default_moq": template_row.default_moq,
                    "moq_unit": None,
                    "weight": template_row.gross_weight,
                    "weight_unit": (
                        "kg" if template_row.gross_weight is not None else None
                    ),
                    "status": "ACTIVE",
                    "deleted_at": None,
                }
                if any(getattr(sku, key) != value for key, value in sku_values.items()):
                    for key, value in sku_values.items():
                        setattr(sku, key, value)
                    sku.version += 1
                    sku.updated_by_user_id = user_id
                    changed = True
                if old_product_id != product.id:
                    moved_from_product_ids.add(old_product_id)
                    sku_rows_by_product[old_product_id] = [
                        row
                        for row in sku_rows_by_product[old_product_id]
                        if row.id != sku.id
                    ]
                    sku_rows_by_product[product.id].append(sku)
                # Provenance follows the most recent successful import even
                # when the row's business fields were unchanged. It is kept
                # out of ``changed`` so the import summary and AI index only
                # report actual catalog-content changes.
                sku.latest_import_job_id = job.id
                # A later catalog import takes over a pre-existing row but
                # never gains permission to delete it. Only the batch that
                # originally created the SKU may retain rollback ownership.
                if sku.rollback_owner_batch_id != job.batch_id:
                    sku.rollback_owner_batch_id = None
            if supplier is not None:
                touched_supplier_ids.add(supplier.id)

            offer = offers.get(sku.id)
            offer_tags = list(template_row.tags)
            if offer is None:
                offer = PublicCatalogOfferRow(
                    id=uuid4(),
                    tenant_id=tenant_id,
                    sku_id=sku.id,
                    unit_price=template_row.unit_price,
                    currency=currency,
                    tags=offer_tags,
                    display_tag=offer_tags[0] if offer_tags else None,
                    publication_status="PUBLISHED",
                    published_at=now,
                )
                session.add(offer)
                offers[sku.id] = offer
                changed = True
            else:
                offer_values = {
                    "unit_price": template_row.unit_price,
                    "currency": currency,
                    "tags": offer_tags,
                    "display_tag": (
                        offer.display_tag
                        if offer.display_tag
                        and offer.display_tag.casefold()
                        in {tag.casefold() for tag in offer_tags}
                        else offer_tags[0] if offer_tags else None
                    ),
                    "publication_status": "PUBLISHED",
                    "deleted_at": None,
                }
                if any(getattr(offer, key) != value for key, value in offer_values.items()):
                    for key, value in offer_values.items():
                        setattr(offer, key, value)
                    offer.published_at = now
                    changed = True

            if sync_product_images(product, template_row):
                changed = True

            if is_new:
                created += 1
            elif changed:
                updated += 1
            else:
                unchanged += 1
            if changed and product.status == "ACTIVE":
                product.search_document_version = 0
                dirty_product_ids.add(product.id)

            # Keep one atomic transaction for the whole import, but send
            # accumulated SQL to the database in bounded batches. Without
            # this flush, SQLAlchemy retains the change history for every
            # imported SKU until the final commit; large catalogs can then
            # exceed the API container's memory limit during finalization.
            if row_index % IMPORT_FLUSH_BATCH_SIZE == 0:
                session.flush()

        for previous_product_id in moved_from_product_ids:
            previous_product = products_by_id.get(previous_product_id)
            if previous_product is None:
                continue
            still_active = any(
                row.status == "ACTIVE" and row.deleted_at is None
                for row in sku_rows_by_product.get(previous_product_id, ())
            )
            if not still_active and previous_product.status != "ARCHIVED":
                previous_product.status = "ARCHIVED"
                previous_product.archived_at = now
                previous_product.current_version += 1
                previous_product.updated_by = user_id
                previous_product.search_document_version = 0

        _record_import_progress(
            job_id=job.id,
            tenant_id=tenant_id,
            progress=94,
            stage="FINALIZING",
            processed_rows=len(parsed.rows),
            total_rows=len(parsed.rows),
        )

        # Imports are incremental merges. A missing row is not evidence that a
        # merchant intended to delete a product, so previously imported SKUs
        # remain untouched until an explicit archive/delete action is used.
        preserved = sum(
            1
            for sku in sku_rows
            if _is_template_managed_sku(sku)
            and not (
                _is_base_product_sku(sku)
                and sku.product_id in imported_base_product_ids
            )
            and _source_sku_identity(sku) not in incoming_sku_codes
            and sku.status == "ACTIVE"
            and sku.deleted_at is None
        )

        if touched_supplier_ids:
            session.flush()
            for supplier_id in touched_supplier_ids:
                supplier = suppliers_by_id.get(supplier_id)
                if supplier is None:
                    continue
                active_skus = int(
                    session.scalar(
                        select(func.count(SkuRow.id)).where(
                            SkuRow.tenant_id == tenant_id,
                            SkuRow.supplier_id == supplier_id,
                            SkuRow.status == "ACTIVE",
                            SkuRow.deleted_at.is_(None),
                        )
                    )
                    or 0
                )
                if supplier.active_skus != active_skus:
                    supplier.active_skus = active_skus
                    if supplier.id not in new_supplier_ids:
                        supplier.version += 1

        imported = created + updated + unchanged
        skipped = parsed.skipped_rows + quota_skipped_count
        warning_summary = "；".join(runtime_warnings[:3])
        index_summary = (
            f"，待更新智能索引 {len(dirty_product_ids)}"
            if dirty_product_ids
            else ""
        )
        result_summary = (
            f"商品导入完成：新建 {created}，更新 {updated}，"
            f"未变化 {unchanged}，保留未包含商品 {preserved}，"
            f"跳过 {skipped}"
            f"{index_summary}。"
        )
        job.products_count = imported
        job.warnings_count = len(runtime_warnings)
        job.progress = 100
        job.status = "published"
        job.error_message = (
            f"{result_summary}{' ' + warning_summary if warning_summary else ''}"
        )
        job.completed_at = now
        session.commit()
        _record_import_progress(
            job_id=job.id,
            tenant_id=tenant_id,
            progress=100,
            stage="COMPLETED",
            processed_rows=len(parsed.rows),
            total_rows=len(parsed.rows),
        )
        return ProductTemplateImportResult(
            status="published",
            imported=imported,
            created=created,
            updated=updated,
            unchanged=unchanged,
            skipped=skipped,
            warnings=tuple(runtime_warnings),
            issues=quota_issues,
            message=result_summary,
        )
