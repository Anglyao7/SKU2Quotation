from __future__ import annotations

import ipaddress
import os
import re
import socket
import unicodedata
from copy import copy
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Callable
from urllib.parse import urljoin, urlsplit
from xml.sax.saxutils import escape

import httpx
import reportlab
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PillowImage
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image as ReportLabImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from ..public_catalog_schemas import PUBLIC_QUOTE_PDF_MAX_COLUMNS, PublicQuoteDocument
from .public_catalog_privacy import public_specification
from .quote_localization import (
    localize_quote_unit,
    quote_field_label,
    quote_headers,
    quote_is_rtl,
    quote_locale,
    quote_text,
)


QuoteImageLoader = Callable[[str], bytes | None]
MAX_QUOTE_IMAGE_BYTES = 8 * 1024 * 1024
MAX_QUOTE_IMAGE_EDGE = 320
DEFAULT_QUOTE_HEADERS = quote_headers("zh-CN")
DEFAULT_QUOTE_WIDTHS = (
    8,
    15,
    20,
    36,
    12,
    12,
    14,
    22,
    14,
    14,
    15,
    17,
    17,
    18,
    38,
    32,
    26,
    28,
    18,
)

_LOGISTICS_OPTION_ALIASES: dict[str, tuple[str, ...]] = {
    "packing_quantity": (
        "装箱数量",
        "装箱数",
        "装箱量",
        "一箱个数",
        "每箱数量",
        "每箱个数",
        "qtyctn",
        "pcsctn",
        "packingquantity",
    ),
    "carton_dimensions": (
        "装箱尺寸",
        "外箱尺寸",
        "纸箱尺寸",
        "箱规",
        "cartonsize",
        "cartondimensions",
    ),
    "gross_weight": (
        "毛重",
        "箱毛重",
        "整箱毛重",
        "grossweight",
        "gw",
    ),
    "carton_volume": (
        "立方",
        "单箱立方",
        "箱体积",
        "cbm",
        "cartonvolume",
    ),
    "minimum_order_quantity": (
        "起订数",
        "起订量",
        "最低起订量",
        "最小起订量",
        "moq",
        "minimumorderquantity",
        "minimumquantity",
    ),
}


def _xlsx_text(value: object | None) -> str:
    """Force untrusted spreadsheet text to remain text, never a formula."""

    text = str(value or "")
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def _xlsx_value(value: object | None) -> object:
    return _xlsx_text(value) if isinstance(value, str) else value


def _normalized_option_key(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).casefold().strip()
    return re.sub(r"[\s\-_/\\:：,.，。()（）\[\]【】]+", "", text)


def _option_value(item: object, field: str) -> object | None:
    values = getattr(item, "option_values_snapshot", None) or {}
    marker = values.get("_sku2quotation")
    source_values = (
        marker.get("quote_source_option_values")
        if isinstance(marker, dict)
        else None
    )
    if isinstance(source_values, dict):
        values = {**values, **source_values}
    normalized = {
        _normalized_option_key(key): value
        for key, value in values.items()
        if _normalized_option_key(key)
    }
    for alias in _LOGISTICS_OPTION_ALIASES[field]:
        value = normalized.get(_normalized_option_key(alias))
        if value not in (None, "", [], {}):
            return value
    return None


def _positive_decimal(value: object | None) -> Decimal | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value if value > 0 else None
    if isinstance(value, (int, float)):
        number = Decimal(str(value))
        return number if number.is_finite() and number > 0 else None
    match = re.search(r"\d+(?:\.\d+)?", str(value).replace(",", ""))
    if match is None:
        return None
    try:
        number = Decimal(match.group(0))
    except InvalidOperation:
        return None
    return number if number > 0 else None


def _gross_weight_kg(value: object | None) -> Decimal | None:
    number = _positive_decimal(value)
    if number is None:
        return None
    text = unicodedata.normalize("NFKC", str(value or "")).casefold()
    if "lb" in text or "磅" in text:
        return number * Decimal("0.45359237")
    if ("g" in text and "kg" not in text) or (
        "克" in text and "千克" not in text and "公斤" not in text
    ):
        return number / Decimal("1000")
    return number


def _carton_volume_m3(value: object | None) -> Decimal | None:
    number = _positive_decimal(value)
    if number is None:
        return None
    text = unicodedata.normalize("NFKC", str(value or "")).casefold()
    if any(unit in text for unit in ("cm3", "cm³", "立方厘米")):
        return number / Decimal("1000000")
    if any(unit in text for unit in ("dm3", "dm³", "升")):
        return number / Decimal("1000")
    return number


def _dimensions_volume_m3(value: object | None) -> Decimal | None:
    if value in (None, ""):
        return None
    text = unicodedata.normalize("NFKC", str(value)).casefold()
    numbers = re.findall(r"\d+(?:\.\d+)?", text.replace(",", ""))
    if len(numbers) < 3:
        return None
    dimensions = [Decimal(number) for number in numbers[:3]]
    volume = dimensions[0] * dimensions[1] * dimensions[2]
    if "mm" in text or "毫米" in text:
        return volume / Decimal("1000000000")
    if re.search(r"(?:^|[^c])m(?:$|[^a-z])", text) or (
        "米" in text and "厘米" not in text
    ):
        return volume
    # Carton dimensions are conventionally supplied in centimetres when the
    # unit is omitted.
    return volume / Decimal("1000000")


def _logistics_values(item: object) -> dict[str, object | None]:
    packing_raw = _option_value(item, "packing_quantity")
    dimensions_raw = _option_value(item, "carton_dimensions")
    gross_raw = _option_value(item, "gross_weight")
    volume_raw = _option_value(item, "carton_volume")
    moq_raw = _option_value(item, "minimum_order_quantity")
    packing = _positive_decimal(packing_raw)
    gross_weight = _gross_weight_kg(gross_raw)
    carton_volume = _carton_volume_m3(volume_raw) or _dimensions_volume_m3(
        dimensions_raw
    )
    quantity = Decimal(str(getattr(item, "quantity", 0) or 0))
    carton_factor = quantity / packing if packing and quantity >= 0 else None
    return {
        "packing_quantity": (
            float(packing) if packing is not None else _xlsx_value(packing_raw)
        ),
        "carton_dimensions": _xlsx_value(dimensions_raw),
        "gross_weight": (
            float(gross_weight)
            if gross_weight is not None
            else _xlsx_value(gross_raw)
        ),
        "carton_volume": (
            float(carton_volume)
            if carton_volume is not None
            else _xlsx_value(volume_raw)
        ),
        "total_volume": (
            float(carton_volume * carton_factor)
            if carton_volume is not None and carton_factor is not None
            else None
        ),
        "total_gross_weight": (
            float(gross_weight * carton_factor)
            if gross_weight is not None and carton_factor is not None
            else None
        ),
        "minimum_order_quantity": (
            float(moq)
            if (moq := _positive_decimal(moq_raw)) is not None
            else _xlsx_value(moq_raw)
        ),
    }


def _validate_public_image_url(url: str) -> None:
    parsed = urlsplit(url)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise ValueError("unsupported quote image URL")
    if parsed.username or parsed.password:
        raise ValueError("credentials are not allowed in quote image URLs")
    port = parsed.port or (443 if parsed.scheme == "https" else 80)
    if port not in {80, 443}:
        raise ValueError("non-standard quote image ports are not allowed")
    addresses = socket.getaddrinfo(parsed.hostname, port, type=socket.SOCK_STREAM)
    if not addresses:
        raise ValueError("quote image host did not resolve")
    for address in addresses:
        ip = ipaddress.ip_address(address[4][0])
        if not ip.is_global:
            raise ValueError("quote image host is not public")


def fetch_remote_quote_image(url: str) -> bytes:
    """Fetch one public image with redirect, SSRF and size safeguards."""

    current = url
    timeout = httpx.Timeout(5.0, connect=3.0)
    with httpx.Client(timeout=timeout, follow_redirects=False, trust_env=False) as client:
        for _redirect in range(4):
            _validate_public_image_url(current)
            with client.stream(
                "GET",
                current,
                headers={"Accept": "image/*", "User-Agent": "AITradeCloud-Quote/1.0"},
            ) as response:
                if response.status_code in {301, 302, 303, 307, 308}:
                    location = response.headers.get("location")
                    if not location:
                        raise ValueError("quote image redirect is missing a target")
                    current = urljoin(current, location)
                    continue
                response.raise_for_status()
                content_type = response.headers.get("content-type", "").casefold()
                if content_type and not content_type.startswith("image/"):
                    raise ValueError("quote image URL did not return an image")
                declared_size = int(response.headers.get("content-length") or 0)
                if declared_size > MAX_QUOTE_IMAGE_BYTES:
                    raise ValueError("quote image exceeds the size limit")
                chunks: list[bytes] = []
                total = 0
                for chunk in response.iter_bytes():
                    total += len(chunk)
                    if total > MAX_QUOTE_IMAGE_BYTES:
                        raise ValueError("quote image exceeds the size limit")
                    chunks.append(chunk)
                return b"".join(chunks)
    raise ValueError("quote image exceeded the redirect limit")


def _normalized_quote_image(content: bytes) -> bytes:
    source_buffer = BytesIO(content)
    output = BytesIO()
    with PillowImage.open(source_buffer) as image:
        image.load()
        image.thumbnail((MAX_QUOTE_IMAGE_EDGE, MAX_QUOTE_IMAGE_EDGE))
        if image.mode in {"RGBA", "LA"} or (
            image.mode == "P" and "transparency" in image.info
        ):
            image.convert("RGBA").save(output, format="PNG", optimize=True)
        else:
            image.convert("RGB").save(output, format="JPEG", quality=82, optimize=True)
    return output.getvalue()


def _place_quote_image(
    sheet: object,
    *,
    row_number: int,
    column_number: int,
    image_url: str | None,
    image_loader: QuoteImageLoader | None,
) -> bool:
    if not image_url or image_loader is None:
        return False
    try:
        content = image_loader(image_url)
        if not content:
            return False
        image = OpenpyxlImage(BytesIO(_normalized_quote_image(content)))
    except Exception:
        return False
    scale = min(86 / max(image.width, 1), 62 / max(image.height, 1))
    image.width = max(1, int(image.width * scale))
    image.height = max(1, int(image.height * scale))
    cell = sheet.cell(row_number, column_number)
    cell.value = None
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[row_number].height = max(
        sheet.row_dimensions[row_number].height or 0,
        52,
    )
    sheet.add_image(image, cell.coordinate)
    return True


def _configure_default_quote_printing(
    sheet: object,
    *,
    header_row: int,
    last_row: int,
) -> None:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35
    sheet.print_options.horizontalCentered = True
    sheet.print_title_rows = f"1:{header_row}"
    sheet.print_area = f"A1:{get_column_letter(max(sheet.max_column, 1))}{last_row}"


def _template_item_value(field: str, document: PublicQuoteDocument, item) -> object:
    quote = document.quote
    locale = quote_locale(quote.locale)
    logistics = _logistics_values(item)
    values: dict[str, object | None] = {
        "serial_number": item.position,
        "sku_code": item.sku_code_snapshot,
        "product_name": item.name_snapshot,
        "description": item.description_snapshot,
        "specification": public_specification(item.specification_snapshot),
        "category": item.category_snapshot,
        "tags": quote_text(locale, "separator").join(item.tags_snapshot or []),
        "product_image": None,
        "quantity": float(item.quantity),
        "unit_code": localize_quote_unit(locale, item.unit_code_snapshot),
        **logistics,
        "unit_price": float(item.unit_price_snapshot),
        "line_total": float(item.line_total),
        "currency": item.currency_snapshot,
        "quote_number": quote.quote_number,
        "quote_date": quote.created_at.date(),
        "customer_name": quote.customer_name,
        "customer_company": quote.customer_company,
        "customer_email": quote.customer_email,
        "customer_phone": quote.customer_phone,
        "notes": quote.notes,
    }
    return _xlsx_value(values.get(field))


def _append_quote_extra_information(sheet: object, quote: object) -> None:
    """Append merchant-authored key/value notes without changing item rows.

    Custom quote templates are intentionally left intact above this section;
    the small block at the bottom gives merchants a safe place for delivery,
    payment, or lead-time notes without requiring another mapped column.
    """

    entries = getattr(quote, "extra_information", None) or []
    normalized: list[tuple[str, str]] = []
    for entry in entries:
        title = str(getattr(entry, "title", "") or "").strip()
        content = str(getattr(entry, "content", "") or "").strip()
        if title and content:
            normalized.append((title, content))
    if not normalized:
        return
    last_column = max(int(getattr(sheet, "max_column", 2) or 2), 2)
    sheet.append([])
    for title, content in normalized:
        sheet.append(
            [_xlsx_text(title), _xlsx_text(content)]
            + [None] * max(0, last_column - 2)
        )
        row_number = sheet.max_row
        if last_column > 2:
            sheet.merge_cells(
                start_row=row_number,
                start_column=2,
                end_row=row_number,
                end_column=last_column,
            )
        sheet.cell(row_number, 1).font = Font(bold=True)
        for cell in sheet[row_number]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _copy_quote_template_cell_format(source: object, target: object) -> None:
    if getattr(source, "has_style", False):
        target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format


def _merge_quote_value(
    sheet: object,
    *,
    row: int,
    start_column: int,
    end_column: int,
    value: object,
) -> None:
    cell = sheet.cell(row, start_column)
    cell.value = _xlsx_value(value)
    if end_column > start_column:
        sheet.merge_cells(
            start_row=row,
            start_column=start_column,
            end_row=row,
            end_column=end_column,
        )


def _compose_system_quote_header(
    sheet: object,
    document: PublicQuoteDocument,
    *,
    column_count: int,
) -> int:
    """Render the system-owned document header above any product template."""

    quote = document.quote
    locale = quote_locale(quote.locale)
    column_count = max(4, column_count)
    last_column = get_column_letter(column_count)
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = quote_text(locale, "document_title")
    sheet["A1"].font = Font(size=18, bold=True, color="172033")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 32

    split_column = max(2, column_count // 2)
    right_label_column = split_column + 1
    right_value_column = min(column_count, right_label_column + 1)
    rows = (
        (
            quote_text(locale, "merchant"),
            document.tenant_name,
            quote_text(locale, "quote_number"),
            quote.quote_number,
        ),
        (
            quote_text(locale, "customer"),
            quote.customer_name,
            quote_text(locale, "date"),
            quote.created_at.date(),
        ),
        (
            quote_text(locale, "company"),
            quote.customer_company or "-",
            quote_text(locale, "currency"),
            quote.currency,
        ),
        (
            quote_text(locale, "email"),
            quote.customer_email or "-",
            quote_text(locale, "phone"),
            quote.customer_phone or "-",
        ),
    )
    for row_number, (left_label, left_value, right_label, right_value) in enumerate(
        rows,
        start=2,
    ):
        left_label_cell = sheet.cell(row_number, 1, left_label)
        left_label_cell.font = Font(bold=True, color="475569")
        left_label_cell.fill = PatternFill("solid", fgColor="F1F5F9")
        left_label_cell.alignment = Alignment(vertical="center")
        _merge_quote_value(
            sheet,
            row=row_number,
            start_column=2,
            end_column=split_column,
            value=left_value,
        )
        right_label_cell = sheet.cell(row_number, right_label_column, right_label)
        right_label_cell.font = Font(bold=True, color="475569")
        right_label_cell.fill = PatternFill("solid", fgColor="F1F5F9")
        right_label_cell.alignment = Alignment(vertical="center")
        _merge_quote_value(
            sheet,
            row=row_number,
            start_column=right_value_column,
            end_column=column_count,
            value=right_value,
        )
        for cell in sheet[row_number]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[row_number].height = 24
    sheet.cell(3, right_value_column).number_format = "yyyy-mm-dd"
    # One quiet spacer row separates document metadata from the product area.
    sheet.row_dimensions[6].height = 9
    return 7


def _copy_single_row_merges(
    source_sheet: object,
    target_sheet: object,
    *,
    source_row: int,
    target_row: int,
    source_to_target_columns: dict[int, int],
) -> None:
    for merged in source_sheet.merged_cells.ranges:
        if merged.min_row != source_row or merged.max_row != source_row:
            continue
        mapped_columns = [
            source_to_target_columns.get(column)
            for column in range(merged.min_col, merged.max_col + 1)
        ]
        if not mapped_columns or any(column is None for column in mapped_columns):
            continue
        target_sheet.merge_cells(
            start_row=target_row,
            start_column=min(mapped_columns),
            end_row=target_row,
            end_column=max(mapped_columns),
        )


def _render_custom_quote_xlsx(
    document: PublicQuoteDocument,
    *,
    template_path: Path,
    image_loader: QuoteImageLoader | None,
) -> bytes:
    spec = document.excel_template
    if spec is None:
        raise ValueError("custom quote template configuration is missing")
    source_workbook = load_workbook(
        template_path,
        data_only=False,
        keep_links=False,
    )
    try:
        if spec.sheet_name not in source_workbook.sheetnames:
            raise ValueError("configured quote worksheet is missing")
        source_sheet = source_workbook[spec.sheet_name]
        columns = sorted(spec.columns, key=lambda column: column.index)
        if not columns:
            raise ValueError("configured quote template has no product columns")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = quote_text(
            quote_locale(document.quote.locale),
            "sheet_name",
        )[:31]
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.rightToLeft = quote_is_rtl(document.quote.locale)
        product_column_count = len(columns)
        sheet_column_count = max(4, product_column_count)
        product_header_row = _compose_system_quote_header(
            sheet,
            document,
            column_count=sheet_column_count,
        )
        source_to_target_columns = {
            column.index: target_index
            for target_index, column in enumerate(columns, start=1)
        }

        for target_index, column in enumerate(columns, start=1):
            source_letter = get_column_letter(column.index)
            target_letter = get_column_letter(target_index)
            source_dimension = source_sheet.column_dimensions[source_letter]
            target_dimension = sheet.column_dimensions[target_letter]
            target_dimension.width = source_dimension.width or 14
            target_dimension.hidden = source_dimension.hidden
            target_dimension.bestFit = source_dimension.bestFit

            source_header = source_sheet.cell(spec.header_row, column.index)
            target_header = sheet.cell(product_header_row, target_index)
            _copy_quote_template_cell_format(source_header, target_header)
            field = spec.column_mappings.get(column.key)
            target_header.value = (
                quote_field_label(document.quote.locale, field)
                if field
                else column.header
            )
            if not source_header.has_style:
                target_header.fill = PatternFill("solid", fgColor="172033")
                target_header.font = Font(color="FFFFFF", bold=True)
            target_header.alignment = copy(source_header.alignment)
            target_header.alignment = Alignment(
                horizontal=target_header.alignment.horizontal or "center",
                vertical=target_header.alignment.vertical or "center",
                wrap_text=True,
            )
        source_header_dimension = source_sheet.row_dimensions[spec.header_row]
        sheet.row_dimensions[product_header_row].height = (
            source_header_dimension.height or 28
        )
        _copy_single_row_merges(
            source_sheet,
            sheet,
            source_row=spec.header_row,
            target_row=product_header_row,
            source_to_target_columns=source_to_target_columns,
        )

        data_start_row = product_header_row + 1
        source_data_dimension = source_sheet.row_dimensions[spec.data_start_row]
        total_volume = Decimal("0")
        total_gross_weight = Decimal("0")
        has_total_volume = False
        has_total_gross_weight = False
        item_count = max(1, len(document.quote.items))
        for offset in range(item_count):
            item = document.quote.items[offset] if document.quote.items else None
            row_number = data_start_row + offset
            sheet.row_dimensions[row_number].height = source_data_dimension.height
            for target_index, column in enumerate(columns, start=1):
                source_cell = source_sheet.cell(spec.data_start_row, column.index)
                target_cell = sheet.cell(row_number, target_index)
                _copy_quote_template_cell_format(source_cell, target_cell)
                field = spec.column_mappings.get(column.key)
                if field == "product_image" and item is not None:
                    target_cell.value = None
                    _place_quote_image(
                        sheet,
                        row_number=row_number,
                        column_number=target_index,
                        image_url=item.image_url_snapshot,
                        image_loader=image_loader,
                    )
                elif field and item is not None:
                    target_cell.value = _template_item_value(field, document, item)
                    if field in {
                        "unit_price",
                        "line_total",
                        "gross_weight",
                        "carton_volume",
                        "total_volume",
                        "total_gross_weight",
                    } and target_cell.number_format == "General":
                        target_cell.number_format = "#,##0.00"
                else:
                    target_cell.value = None
                target_cell.alignment = Alignment(
                    horizontal=target_cell.alignment.horizontal,
                    vertical=target_cell.alignment.vertical or "center",
                    wrap_text=True,
                )
            _copy_single_row_merges(
                source_sheet,
                sheet,
                source_row=spec.data_start_row,
                target_row=row_number,
                source_to_target_columns=source_to_target_columns,
            )
            if item is not None:
                logistics = _logistics_values(item)
                if isinstance(logistics["total_volume"], (int, float)):
                    total_volume += Decimal(str(logistics["total_volume"]))
                    has_total_volume = True
                if isinstance(logistics["total_gross_weight"], (int, float)):
                    total_gross_weight += Decimal(
                        str(logistics["total_gross_weight"])
                    )
                    has_total_gross_weight = True

        total_row = data_start_row + item_count
        field_columns = {
            spec.column_mappings.get(column.key): target_index
            for target_index, column in enumerate(columns, start=1)
            if spec.column_mappings.get(column.key)
        }
        for column_number in range(1, product_column_count + 1):
            cell = sheet.cell(total_row, column_number)
            cell.fill = PatternFill("solid", fgColor="EEF2F7")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        total_value_column = field_columns.get("line_total", product_column_count)
        label_column = max(1, total_value_column - 1)
        sheet.cell(total_row, label_column).value = quote_text(
            document.quote.locale,
            "total",
        )
        sheet.cell(total_row, total_value_column).value = float(document.quote.total)
        sheet.cell(total_row, total_value_column).number_format = "#,##0.00"
        if has_total_volume and (column := field_columns.get("total_volume")):
            sheet.cell(total_row, column).value = float(total_volume)
            sheet.cell(total_row, column).number_format = "#,##0.00####"
        if has_total_gross_weight and (
            column := field_columns.get("total_gross_weight")
        ):
            sheet.cell(total_row, column).value = float(total_gross_weight)
            sheet.cell(total_row, column).number_format = "#,##0.00####"

        if document.quote.notes:
            sheet.append([])
            sheet.append(
                [
                    quote_text(document.quote.locale, "notes"),
                    _xlsx_text(document.quote.notes),
                ]
            )
            if sheet_column_count > 2:
                sheet.merge_cells(
                    start_row=sheet.max_row,
                    start_column=2,
                    end_row=sheet.max_row,
                    end_column=sheet_column_count,
                )
        _append_quote_extra_information(sheet, document.quote)
        product_last_column = get_column_letter(product_column_count)
        sheet.freeze_panes = f"A{data_start_row}"
        sheet.auto_filter.ref = (
            f"A{product_header_row}:{product_last_column}{data_start_row + item_count - 1}"
        )
        _configure_default_quote_printing(
            sheet,
            header_row=product_header_row,
            last_row=sheet.max_row,
        )
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        return buffer.getvalue()
    finally:
        source_workbook.close()


def _register_quote_pdf_font(locale: str) -> str:
    cid_fonts = {
        "zh-CN": "STSong-Light",
        "ja": "HeiseiMin-W3",
        "ko": "HYSMyeongJo-Medium",
    }
    cid_font = cid_fonts.get(locale)
    if cid_font:
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(cid_font))
        except KeyError:
            pass
        return cid_font

    if locale in {"ar", "fa"}:
        configured = os.environ.get("QUOTE_ARABIC_FONT_PATH", "").strip()
        candidates = [
            *((Path(configured),) if configured else ()),
            Path("/usr/share/fonts/truetype/noto/NotoSansArabic-Regular.ttf"),
            Path("/usr/share/fonts/opentype/noto/NotoSansArabic-Regular.ttf"),
            Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
        ]
        for candidate in candidates:
            if not candidate.is_file():
                continue
            try:
                pdfmetrics.registerFont(TTFont("QuoteArabic", str(candidate)))
            except KeyError:
                pass
            return "QuoteArabic"

    vera = Path(reportlab.__file__).resolve().parent / "fonts" / "Vera.ttf"
    try:
        pdfmetrics.registerFont(TTFont("QuoteSans", str(vera)))
    except KeyError:
        pass
    return "QuoteSans"


def _pdf_localized_text(value: object | None, locale: str) -> str:
    text = str(value or "")
    if locale in {"ar", "fa"} and re.search(r"[\u0600-\u06ff]", text):
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display

            text = get_display(arabic_reshaper.reshape(text))
        except ImportError:
            # Production installs the shaping helpers. This fallback keeps
            # development downloads functional before dependencies are synced.
            pass
    return escape(text)


_PUBLIC_QUOTE_TABLE_FIELDS = frozenset(
    {
        "serial_number",
        "sku_code",
        "product_name",
        "description",
        "specification",
        "category",
        "tags",
        "product_image",
        "quantity",
        "unit_code",
        "packing_quantity",
        "carton_dimensions",
        "gross_weight",
        "carton_volume",
        "minimum_order_quantity",
        "unit_price",
        "line_total",
        "total_volume",
        "total_gross_weight",
        "currency",
    }
)
_PUBLIC_QUOTE_DEFAULT_FIELDS = (
    "product_image",
    "product_name",
    "quantity",
    "unit_price",
    "line_total",
)
_PUBLIC_QUOTE_COLUMN_WIDTHS_MM = {
    "serial_number": 10,
    "sku_code": 32,
    "product_name": 42,
    "description": 38,
    "specification": 32,
    "category": 26,
    "tags": 28,
    "product_image": 20,
    "quantity": 18,
    "unit_code": 18,
    "packing_quantity": 22,
    "carton_dimensions": 30,
    "gross_weight": 22,
    "carton_volume": 22,
    "minimum_order_quantity": 18,
    "unit_price": 25,
    "line_total": 28,
    "total_volume": 24,
    "total_gross_weight": 26,
    "currency": 18,
}
def _clip_quote_table_text(value: object | None, field: str) -> str:
    # Paragraph cells wrap naturally in the PDF.  Keep the complete imported
    # value instead of clipping it, otherwise logistics/specification details
    # disappear from an otherwise valid quotation.
    return str(value if value is not None else "").strip()


def _public_quote_table_fields(document: PublicQuoteDocument) -> list[str]:
    quote = document.quote
    configured = [
        str(field).strip()
        for field in (getattr(quote, "visible_columns", None) or [])
        if str(field).strip()
    ]
    template_fields: list[str] = []
    if document.excel_template is not None:
        for column in document.excel_template.columns:
            field = document.excel_template.column_mappings.get(column.key)
            if field:
                template_fields.append(str(field))
    candidates = configured or template_fields or list(_PUBLIC_QUOTE_DEFAULT_FIELDS)
    # The merchant controls the visible columns. SKU and logistics fields are
    # ordinary catalog data and are available when the default template is used.
    fields: list[str] = []
    for field in candidates:
        if field not in _PUBLIC_QUOTE_TABLE_FIELDS:
            continue
        if field not in fields:
            fields.append(field)
    return (fields or list(_PUBLIC_QUOTE_DEFAULT_FIELDS))[:PUBLIC_QUOTE_PDF_MAX_COLUMNS]


def _public_quote_table_headers(
    document: PublicQuoteDocument,
    fields: list[str],
    locale: str,
) -> list[str]:
    custom_headers: dict[str, str] = {}
    if document.excel_template is not None:
        for column in document.excel_template.columns:
            field = document.excel_template.column_mappings.get(column.key)
            header = str(column.header or "").strip()
            if field and header and field not in custom_headers:
                custom_headers[str(field)] = header
    headers: list[str] = []
    for field in fields:
        # System fields have one canonical translation dictionary.  Keeping a
        # merchant's source-language header here would make a PDF disagree
        # with the workbench and the generated/custom Excel document after the
        # quote language is changed.  Unknown fields (if a future template
        # exposes one) can still retain their custom header.
        localized = quote_field_label(locale, field)
        headers.append(localized or custom_headers.get(field) or field)
    return headers


def _quote_table_value(field: str, document: PublicQuoteDocument, item: object) -> str:
    value = _template_item_value(field, document, item)
    if value in (None, ""):
        return ""
    if field in {"quantity", "packing_quantity", "minimum_order_quantity"}:
        try:
            return f"{Decimal(str(value)):f}".rstrip("0").rstrip(".")
        except (InvalidOperation, ValueError):
            return _clip_quote_table_text(value, field)
    if field in {
        "unit_price",
        "line_total",
        "gross_weight",
        "carton_volume",
        "total_volume",
        "total_gross_weight",
    }:
        try:
            return f"{Decimal(str(value)):,.2f}"
        except (InvalidOperation, ValueError):
            return _clip_quote_table_text(value, field)
    return _clip_quote_table_text(value, field)


def _quote_table_widths(fields: list[str]) -> list[float]:
    available = 178.0  # A4 width minus the 16 mm margins used below.
    widths = [float(_PUBLIC_QUOTE_COLUMN_WIDTHS_MM.get(field, 24)) for field in fields]
    total = sum(widths)
    if total <= available:
        return [width * mm for width in widths]
    scale = available / total
    scaled = [max(8.0, width * scale) for width in widths]
    overflow = sum(scaled) - available
    if overflow > 0:
        shrinkable = sum(max(width - 8.0, 0.0) for width in scaled)
        if shrinkable:
            scaled = [
                width - overflow * max(width - 8.0, 0.0) / shrinkable
                for width in scaled
            ]
    return [width * mm for width in scaled]


def _quote_pdf_image(
    image_url: str | None,
    image_loader: QuoteImageLoader | None,
    *,
    max_width: float,
    max_height: float,
) -> ReportLabImage | str:
    """Build a proportional thumbnail flowable for a PDF table cell.

    Images are loaded through the route-provided callback so private catalog
    media can be resolved with the same authorization as the quote download.
    The normalized in-memory copy keeps generated PDFs small and prevents a
    large source image from changing the table layout.
    """

    if not image_url or image_loader is None:
        return ""
    try:
        content = image_loader(image_url)
        if not content:
            return ""
        normalized = _normalized_quote_image(content)
        image_buffer = BytesIO(normalized)
        with PillowImage.open(image_buffer) as image_source:
            source_width, source_height = image_source.size
        if source_width <= 0 or source_height <= 0:
            return ""
        scale = min(
            max_width / source_width,
            max_height / source_height,
        )
        image_buffer.seek(0)
        image = ReportLabImage(
            image_buffer,
            width=max(1.0, source_width * scale),
            height=max(1.0, source_height * scale),
            hAlign="CENTER",
        )
        # Keep the BytesIO alive until reportlab has finished building the PDF.
        image._quote_image_buffer = image_buffer
        return image
    except Exception:
        return ""


def render_public_quote_draft_pdf(
    document: PublicQuoteDocument,
    *,
    image_loader: QuoteImageLoader | None = None,
) -> bytes:
    quote = document.quote
    locale = quote_locale(quote.locale)
    font_name = _register_quote_pdf_font(locale)
    style_palette = {
        "indigo": {"accent": "#314B9B", "soft": "#EEF2FF", "border": "#CBD5E1"},
        "emerald": {"accent": "#087F5B", "soft": "#E8F7F0", "border": "#B7E4D2"},
        "gold": {"accent": "#8A6418", "soft": "#FBF3D7", "border": "#E4CF8A"},
        "slate": {"accent": "#334155", "soft": "#F1F5F9", "border": "#CBD5E1"},
        "rose": {"accent": "#9F3B5B", "soft": "#FFF0F4", "border": "#F0C4D2"},
    }
    palette = style_palette.get(document.style, style_palette["indigo"])
    buffer = BytesIO()
    styles = getSampleStyleSheet()
    rtl = quote_is_rtl(locale)
    title_style = ParagraphStyle(
        "DraftLocalizedTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=18,
        leading=24,
        textColor=colors.HexColor(palette["accent"]),
        alignment=TA_RIGHT if rtl else styles["Title"].alignment,
    )
    body_style = ParagraphStyle(
        "DraftLocalizedBody",
        parent=styles["BodyText"],
        fontName=font_name,
        fontSize=9,
        leading=13,
        alignment=TA_RIGHT if rtl else styles["BodyText"].alignment,
    )
    right_style = ParagraphStyle("DraftRight", parent=body_style, alignment=TA_RIGHT)
    pdf = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=f"{quote_text(locale, 'document_title')} {quote.quote_number}",
    )
    story = [
        Paragraph(_pdf_localized_text(quote_text(locale, "document_title"), locale), title_style),
        Spacer(1, 6 * mm),
    ]
    metadata = [
        [
            Paragraph(_pdf_localized_text(f"{quote_text(locale, 'merchant')}: {document.tenant_name}", locale), body_style),
            Paragraph(_pdf_localized_text(f"{quote_text(locale, 'quote_number')}: {quote.quote_number}", locale), right_style),
        ],
        [
            Paragraph(_pdf_localized_text(f"{quote_text(locale, 'customer')}: {quote.customer_name}", locale), body_style),
            Paragraph("", right_style),
        ],
        [
            Paragraph(_pdf_localized_text(f"{quote_text(locale, 'company')}: {quote.customer_company or '-'}", locale), body_style),
            Paragraph(_pdf_localized_text(f"{quote_text(locale, 'date')}: {quote.created_at:%Y-%m-%d}", locale), right_style),
        ],
        [
            Paragraph(_pdf_localized_text(f"{quote_text(locale, 'email')}: {quote.customer_email or '-'}", locale), body_style),
            Paragraph(_pdf_localized_text(f"{quote_text(locale, 'currency')}: {quote.currency}", locale), right_style),
        ],
    ]
    meta_table = Table(metadata, colWidths=[90 * mm, 72 * mm])
    meta_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.extend([meta_table, Spacer(1, 7 * mm)])

    table_fields = _public_quote_table_fields(document)
    table_body_style = ParagraphStyle(
        "DraftTableBody",
        parent=body_style,
        fontSize=7.8,
        leading=10,
        wordWrap="CJK",
        splitLongWords=1,
    )
    table_header_style = ParagraphStyle(
        "DraftTableHeader",
        parent=table_body_style,
        textColor=colors.white,
        fontName=font_name,
        fontSize=7.8,
        leading=9,
        alignment=TA_CENTER,
    )
    table_headers = _public_quote_table_headers(document, table_fields, locale)
    table_widths = _quote_table_widths(table_fields)
    rows: list[list[object]] = [[
        Paragraph(_pdf_localized_text(header, locale), table_header_style)
        for header in table_headers
    ]]
    for item in quote.items:
        row: list[object] = []
        for index, field in enumerate(table_fields):
            if field == "product_image":
                row.append(
                    _quote_pdf_image(
                        item.image_url_snapshot,
                        image_loader,
                        max_width=max(4 * mm, table_widths[index] - 2 * mm),
                        max_height=18 * mm,
                    )
                )
            else:
                row.append(
                    Paragraph(
                        _pdf_localized_text(
                            _quote_table_value(field, document, item),
                            locale,
                        ),
                        table_body_style,
                    )
                )
        rows.append(row)
    total_row = [""] * len(table_fields)
    total_row[max(0, len(table_fields) - 2)] = Paragraph(
        _pdf_localized_text(quote_text(locale, "total"), locale),
        table_body_style,
    )
    total_row[-1] = Paragraph(
        _pdf_localized_text(f"{quote.currency} {quote.total:,.2f}", locale),
        table_body_style,
    )
    rows.append(total_row)
    numeric_indexes = {
        index for index, field in enumerate(table_fields)
        if field in {
            "quantity",
            "packing_quantity",
            "gross_weight",
            "carton_volume",
            "minimum_order_quantity",
            "unit_price",
            "line_total",
            "total_volume",
            "total_gross_weight",
        }
    }
    numeric_commands = [
        ("ALIGN", (index, 1), (index, -1), "RIGHT")
        for index in sorted(numeric_indexes)
    ]
    serial_index = table_fields.index("serial_number") if "serial_number" in table_fields else None
    if serial_index is not None:
        numeric_commands.append(("ALIGN", (serial_index, 1), (serial_index, -1), "CENTER"))
    table = Table(
        rows,
        repeatRows=1,
        colWidths=table_widths,
    )
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 8.2),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(palette["accent"])),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -2), 0.35, colors.HexColor(palette["border"])),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(palette["soft"])),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                *numeric_commands,
            ]
        )
    )
    story.append(table)
    if quote.notes:
        story.extend(
            [Spacer(1, 7 * mm), Paragraph(_pdf_localized_text(f"{quote_text(locale, 'notes')}: {quote.notes}", locale), body_style)]
        )
    for entry in getattr(quote, "extra_information", None) or []:
        title = str(getattr(entry, "title", "") or "").strip()
        content = str(getattr(entry, "content", "") or "").strip()
        if not title or not content:
            continue
        story.extend(
            [
                Spacer(1, 3 * mm),
                Paragraph(
                    _pdf_localized_text(f"{title}: {content}", locale),
                    body_style,
                ),
            ]
        )
    story.extend(
        [
            Spacer(1, 8 * mm),
            Paragraph(
                _pdf_localized_text(
                    f"{quote_text(locale, 'merchant_contact')}: "
                    f"{document.contact_email or '-'}  {document.contact_phone or ''}",
                    locale,
                ),
                body_style,
            ),
        ]
    )
    pdf.build(story)
    return buffer.getvalue()


def render_public_quote_draft_xlsx(
    document: PublicQuoteDocument,
    *,
    template_path: Path | None = None,
    image_loader: QuoteImageLoader | None = None,
) -> bytes:
    if document.excel_template is not None and template_path is not None:
        return _render_custom_quote_xlsx(
            document,
            template_path=template_path,
            image_loader=image_loader,
        )
    quote = document.quote
    locale = quote_locale(quote.locale)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = quote_text(locale, "sheet_name")[:31]
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.rightToLeft = quote_is_rtl(locale)
    dark_fill = PatternFill("solid", fgColor="172033")
    light_fill = PatternFill("solid", fgColor="EEF2F7")
    white_font = Font(color="FFFFFF", bold=True)

    column_count = len(DEFAULT_QUOTE_HEADERS)
    last_column = get_column_letter(column_count)
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = quote_text(locale, "document_title")
    sheet["A1"].font = Font(size=18, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.row_dimensions[1].height = 30

    sheet.append(
        [
            quote_text(locale, "merchant"),
            _xlsx_text(document.tenant_name),
            "",
            "",
            quote_text(locale, "quote_number"),
            _xlsx_text(quote.quote_number),
            "",
            "",
            quote_text(locale, "currency"),
            quote.currency,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    sheet.append(
        [
            quote_text(locale, "customer"),
            _xlsx_text(quote.customer_name),
            "",
            "",
            quote_text(locale, "company"),
            _xlsx_text(quote.customer_company),
            "",
            "",
            quote_text(locale, "date"),
            quote.created_at.date(),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    sheet.append(
        [
            quote_text(locale, "email"),
            _xlsx_text(quote.customer_email),
            "",
            "",
            quote_text(locale, "phone"),
            _xlsx_text(quote.customer_phone),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    for row_number in (2, 3, 4):
        sheet.merge_cells(start_row=row_number, start_column=2, end_row=row_number, end_column=4)
        sheet.merge_cells(start_row=row_number, start_column=6, end_row=row_number, end_column=8)
        sheet.merge_cells(start_row=row_number, start_column=10, end_row=row_number, end_column=column_count)
    sheet.cell(row=3, column=10).number_format = "yyyy-mm-dd"
    sheet.append([])
    sheet.append(list(quote_headers(locale)))
    header_row = sheet.max_row
    for cell in sheet[header_row]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 28

    total_volume = Decimal("0")
    total_gross_weight = Decimal("0")
    has_total_volume = False
    has_total_gross_weight = False
    for item in quote.items:
        logistics = _logistics_values(item)
        sheet.append(
            [
                item.position,
                None,
                _xlsx_text(item.sku_code_snapshot),
                _xlsx_text(item.name_snapshot),
                float(item.quantity),
                _xlsx_text(localize_quote_unit(locale, item.unit_code_snapshot)),
                logistics["packing_quantity"],
                logistics["carton_dimensions"],
                logistics["gross_weight"],
                logistics["carton_volume"],
                float(item.unit_price_snapshot),
                float(item.line_total),
                logistics["total_volume"],
                logistics["total_gross_weight"],
                _xlsx_text(item.description_snapshot),
                _xlsx_text(item.specification_snapshot),
                _xlsx_text(item.category_snapshot),
                _xlsx_text(quote_text(locale, "separator").join(item.tags_snapshot or [])),
                logistics["minimum_order_quantity"],
            ]
        )
        row_number = sheet.max_row
        _place_quote_image(
            sheet,
            row_number=row_number,
            column_number=2,
            image_url=item.image_url_snapshot,
            image_loader=image_loader,
        )
        for cell in sheet[row_number]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        if isinstance(logistics["total_volume"], (int, float)):
            total_volume += Decimal(str(logistics["total_volume"]))
            has_total_volume = True
        if isinstance(logistics["total_gross_weight"], (int, float)):
            total_gross_weight += Decimal(str(logistics["total_gross_weight"]))
            has_total_gross_weight = True
    sheet.append(
        [
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            quote_text(locale, "total"),
            float(quote.total),
            float(total_volume) if has_total_volume else None,
            float(total_gross_weight) if has_total_gross_weight else None,
            "",
            "",
            "",
            "",
            "",
        ]
    )
    total_row = sheet.max_row
    for cell in sheet[total_row]:
        cell.fill = light_fill
        cell.font = Font(bold=True)
    if quote.notes:
        sheet.append([])
        sheet.append([quote_text(locale, "notes"), _xlsx_text(quote.notes)])
        sheet.merge_cells(
            start_row=sheet.max_row,
            start_column=2,
            end_row=sheet.max_row,
            end_column=column_count,
        )

    _append_quote_extra_information(sheet, quote)

    for index, width in enumerate(DEFAULT_QUOTE_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{max(header_row, total_row - 1)}"
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=total_row):
        row[4].number_format = "#,##0.######"
        row[6].number_format = "#,##0.######"
        for column_index in (8, 9, 10, 11, 12, 13, 18):
            row[column_index].number_format = "#,##0.00####"
    _configure_default_quote_printing(
        sheet,
        header_row=header_row,
        last_row=sheet.max_row,
    )

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def render_default_quote_template_xlsx() -> bytes:
    """Return a product-region template merchants can map or customize.

    The final quotation header is system-owned and therefore intentionally
    absent from this workbook.  Uploading this file (or a merchant variant)
    only changes the columns and styling of the product-detail table.
    """

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品明细模板"
    sheet.sheet_view.showGridLines = False
    dark_fill = PatternFill("solid", fgColor="172033")
    light_fill = PatternFill("solid", fgColor="EEF2F7")
    column_count = len(DEFAULT_QUOTE_HEADERS)
    last_column = get_column_letter(column_count)
    sheet.append(list(DEFAULT_QUOTE_HEADERS))
    header_row = 1
    for cell in sheet[header_row]:
        cell.fill = dark_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 28
    sheet.append([None] * len(DEFAULT_QUOTE_HEADERS))
    sheet.row_dimensions[sheet.max_row].height = 52
    sheet.append([])
    sheet.append(
        [
            "说明",
            "该模板只决定报价单的商品明细区域；商家、单号、客户、日期和币种等顶部信息由系统统一生成。未映射列会保留为空。",
            *([None] * (column_count - 2)),
        ]
    )
    sheet.merge_cells(
        start_row=sheet.max_row,
        start_column=2,
        end_row=sheet.max_row,
        end_column=column_count,
    )
    for cell in sheet[sheet.max_row]:
        cell.fill = light_fill
    for index, width in enumerate(DEFAULT_QUOTE_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{header_row + 1}"
    _configure_default_quote_printing(
        sheet,
        header_row=header_row,
        last_row=sheet.max_row,
    )
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()
