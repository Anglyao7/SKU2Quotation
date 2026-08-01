from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook


CATEGORY_TEMPLATE_SHEET = "分类模板"
CATEGORY_TEMPLATE_HEADERS = ("一级分类", "二级分类")
CATEGORY_TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent / "assets" / "category-template.xlsx"
)
CATEGORY_TEMPLATE_MAX_FILE_BYTES = 50 * 1024 * 1024
CATEGORY_TEMPLATE_MAX_ROWS = 20_000
CATEGORY_TEMPLATE_MAX_ARCHIVE_ENTRIES = 2_000
CATEGORY_TEMPLATE_MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
CATEGORY_NAME_MAX_LENGTH = 200
_CONTROL_CHARACTER_PATTERN = re.compile(r"[\x00-\x1f\x7f]")


@dataclass(frozen=True, slots=True)
class CategoryTemplateIssue:
    row_number: int | None
    column: str
    code: str
    message: str
    value: str | None = None

    def as_dict(self) -> dict[str, object]:
        return {
            "row_number": self.row_number,
            "column": self.column,
            "code": self.code,
            "message": self.message,
            "value": self.value,
        }


class CategoryTemplateValidationError(ValueError):
    def __init__(
        self,
        message: str,
        *,
        issues: tuple[CategoryTemplateIssue, ...] = (),
    ) -> None:
        super().__init__(message)
        self.issues = issues


@dataclass(frozen=True, slots=True)
class CategoryTemplateGroup:
    primary_name: str
    secondary_names: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class CategoryTemplateParseResult:
    groups: tuple[CategoryTemplateGroup, ...]
    processed_rows: int
    blank_rows_ignored: int
    duplicate_rows_ignored: int


def category_name_key(value: str) -> str:
    return unicodedata.normalize("NFKC", value).casefold().strip()


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format(value, ".15g")
    return str(value).strip()


def _issue_value(value: object) -> str | None:
    text = _cell_text(value)
    if not text:
        return None
    return f"{text[:157]}…" if len(text) > 160 else text


def _validate_archive(content: bytes) -> None:
    if not content.startswith(b"PK"):
        raise CategoryTemplateValidationError(
            "文件不是有效的 XLSX 分类模板，请重新下载模板后填写。"
        )
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > CATEGORY_TEMPLATE_MAX_ARCHIVE_ENTRIES:
                raise CategoryTemplateValidationError(
                    "分类模板包含过多内部文件，无法安全解析。"
                )
            if sum(item.file_size for item in entries) > CATEGORY_TEMPLATE_MAX_UNCOMPRESSED_BYTES:
                raise CategoryTemplateValidationError(
                    "分类模板解压后体积过大，无法安全解析。"
                )
    except BadZipFile as exc:
        raise CategoryTemplateValidationError(
            "文件不是有效的 XLSX 分类模板，请重新下载模板后填写。"
        ) from exc


def _validate_name(
    value: object,
    *,
    row_number: int,
    column: str,
) -> tuple[str, CategoryTemplateIssue | None]:
    name = unicodedata.normalize("NFKC", _cell_text(value)).strip()
    if not name:
        return "", None
    if len(name) > CATEGORY_NAME_MAX_LENGTH:
        return "", CategoryTemplateIssue(
            row_number=row_number,
            column=column,
            code="CATEGORY_NAME_TOO_LONG",
            message=f"{column}最多填写 {CATEGORY_NAME_MAX_LENGTH} 个字符。",
            value=_issue_value(value),
        )
    if "/" in name or "／" in name:
        return "", CategoryTemplateIssue(
            row_number=row_number,
            column=column,
            code="CATEGORY_NAME_CONTAINS_SEPARATOR",
            message=f"{column}不需要填写“/”；一级、二级分类请分别放在 A、B 列。",
            value=_issue_value(value),
        )
    if _CONTROL_CHARACTER_PATTERN.search(name):
        return "", CategoryTemplateIssue(
            row_number=row_number,
            column=column,
            code="CATEGORY_NAME_CONTROL_CHARACTER",
            message=f"{column}包含换行或控制字符，请改为单行文字。",
            value=_issue_value(value),
        )
    return name, None


def parse_category_template(content: bytes) -> CategoryTemplateParseResult:
    if not content:
        raise CategoryTemplateValidationError("分类模板为空，请重新选择文件。")
    if len(content) > CATEGORY_TEMPLATE_MAX_FILE_BYTES:
        raise CategoryTemplateValidationError(
            "分类模板超过 50 MB，请拆分后分别导入。"
        )
    _validate_archive(content)

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=False,
        )
    except Exception as exc:
        raise CategoryTemplateValidationError(
            "无法读取分类模板，请确认文件未损坏且格式为 .xlsx。"
        ) from exc

    try:
        if CATEGORY_TEMPLATE_SHEET not in workbook.sheetnames:
            raise CategoryTemplateValidationError(
                f"缺少工作表“{CATEGORY_TEMPLATE_SHEET}”，请使用下载的分类模板导入。"
            )
        sheet = workbook[CATEGORY_TEMPLATE_SHEET]
        if sheet.max_row is None or sheet.max_column is None:
            sheet.calculate_dimension(force=True)
        max_row = sheet.max_row or 1
        actual_headers = tuple(
            _cell_text(sheet.cell(row=1, column=index).value)
            for index in range(1, len(CATEGORY_TEMPLATE_HEADERS) + 1)
        )
        if actual_headers != CATEGORY_TEMPLATE_HEADERS:
            issues = tuple(
                CategoryTemplateIssue(
                    row_number=1,
                    column=f"第 {index} 列",
                    code="CATEGORY_TEMPLATE_HEADER_INVALID",
                    message=f"表头应为“{expected}”。",
                    value=actual or None,
                )
                for index, (actual, expected) in enumerate(
                    zip(actual_headers, CATEGORY_TEMPLATE_HEADERS, strict=True),
                    start=1,
                )
                if actual != expected
            )
            raise CategoryTemplateValidationError(
                "分类模板表头不正确，请保留 A 列“一级分类”和 B 列“二级分类”。",
                issues=issues,
            )
        if max_row - 1 > CATEGORY_TEMPLATE_MAX_ROWS:
            raise CategoryTemplateValidationError(
                f"单次最多导入 {CATEGORY_TEMPLATE_MAX_ROWS} 行分类，请拆分文件后重试。"
            )

        issues: list[CategoryTemplateIssue] = []
        group_names: dict[str, str] = {}
        secondary_names: dict[str, list[str]] = {}
        secondary_keys: dict[str, set[str]] = {}
        seen_rows: set[tuple[str, str | None]] = set()
        processed_rows = 0
        blank_rows_ignored = 0
        duplicate_rows_ignored = 0

        for row_number, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                max_row=max_row,
                min_col=1,
                max_col=2,
            ),
            start=2,
        ):
            primary_cell, secondary_cell = row
            if primary_cell.data_type == "f" or secondary_cell.data_type == "f":
                formula_column = (
                    "一级分类" if primary_cell.data_type == "f" else "二级分类"
                )
                issues.append(
                    CategoryTemplateIssue(
                        row_number=row_number,
                        column=formula_column,
                        code="CATEGORY_TEMPLATE_FORMULA_NOT_ALLOWED",
                        message=f"{formula_column}不支持公式，请粘贴为普通文字。",
                    )
                )
                if len(issues) >= 100:
                    break
                continue

            primary_name, primary_issue = _validate_name(
                primary_cell.value,
                row_number=row_number,
                column="一级分类",
            )
            secondary_name, secondary_issue = _validate_name(
                secondary_cell.value,
                row_number=row_number,
                column="二级分类",
            )
            if primary_issue:
                issues.append(primary_issue)
            if secondary_issue:
                issues.append(secondary_issue)
            if primary_issue or secondary_issue:
                if len(issues) >= 100:
                    break
                continue
            if not primary_name and not secondary_name:
                blank_rows_ignored += 1
                continue
            if not primary_name:
                issues.append(
                    CategoryTemplateIssue(
                        row_number=row_number,
                        column="一级分类",
                        code="PRIMARY_CATEGORY_REQUIRED",
                        message="填写二级分类时，A 列一级分类不能为空。",
                        value=None,
                    )
                )
                if len(issues) >= 100:
                    break
                continue

            primary_key = category_name_key(primary_name)
            secondary_key = (
                category_name_key(secondary_name) if secondary_name else None
            )
            row_key = (primary_key, secondary_key)
            if row_key in seen_rows:
                duplicate_rows_ignored += 1
                continue
            seen_rows.add(row_key)
            processed_rows += 1
            group_names.setdefault(primary_key, primary_name)
            secondary_names.setdefault(primary_key, [])
            secondary_keys.setdefault(primary_key, set())
            if secondary_name and secondary_key not in secondary_keys[primary_key]:
                secondary_names[primary_key].append(secondary_name)
                secondary_keys[primary_key].add(secondary_key)

        if issues:
            affected_rows = len(
                {issue.row_number for issue in issues if issue.row_number is not None}
            )
            raise CategoryTemplateValidationError(
                f"发现 {len(issues)} 个分类数据问题，涉及 {affected_rows} 行，未执行导入。",
                issues=tuple(issues),
            )
        if not processed_rows:
            raise CategoryTemplateValidationError(
                "分类模板没有可导入的数据，请从第 2 行开始填写分类。"
            )

        return CategoryTemplateParseResult(
            groups=tuple(
                CategoryTemplateGroup(
                    primary_name=group_names[primary_key],
                    secondary_names=tuple(secondary_names[primary_key]),
                )
                for primary_key in group_names
            ),
            processed_rows=processed_rows,
            blank_rows_ignored=blank_rows_ignored,
            duplicate_rows_ignored=duplicate_rows_ignored,
        )
    finally:
        workbook.close()
