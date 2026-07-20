import csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Sequence
from zipfile import BadZipFile, ZipFile

import xlrd
from docx import Document
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..models import FileDetectionResponse


MAX_ROWS_PER_SHEET = 20_000
MAX_COLUMNS = 64
HEADER_SCAN_ROWS = 80
CSV_SAMPLE_BYTES = 64 * 1024
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030")

FIELD_LABELS = {
    "name": "产品名称",
    "model": "型号/货号",
    "category": "产品分类",
    "price": "采购价格",
    "moq": "MOQ",
    "specification": "规格/尺寸",
    "color": "颜色",
    "material": "材质",
    "packing": "包装/箱规",
    "weight": "重量",
    "description": "产品描述",
}
FIELD_ORDER = tuple(FIELD_LABELS)

HEADER_ALIASES = {
    "name": ("产品名称", "商品名称", "货品名称", "品名", "productname", "product name", "名称"),
    "model": ("产品型号", "商品编码", "产品编号", "货号", "款号", "型号", "itemno", "item no", "item#", "sku", "编号"),
    "category": ("产品分类", "商品分类", "品类", "category", "类别", "大类"),
    "price": ("人民币单价", "含税单价", "出厂价", "拿样价", "批发价", "最低报价", "采购价", "单价", "price", "exw", "fob", "rmb"),
    "moq": ("最小起订量", "最低起订量", "起订量", "moq"),
    "specification": ("产品规格", "产品尺寸", "商品规格", "尺寸", "规格", "specification", "size"),
    "color": ("产品颜色", "可选颜色", "颜色", "color"),
    "material": ("产品材质", "主要材质", "材质", "material"),
    "packing": ("独立包装箱规", "叠装箱规", "装箱规格", "包装方式", "包装规格", "装箱资料", "装箱数", "箱规", "包装", "packing", "carton"),
    "weight": ("产品重量", "单品重量", "毛重", "净重", "重量", "weight", "g.w", "n.w"),
    "description": ("产品描述", "商品描述", "description", "remarks", "remark", "备注"),
    "image": ("产品图片", "商品图片", "图片", "picture", "photo", "image"),
}


@dataclass(slots=True)
class ParsedRecord:
    name: str
    model: str
    category: str
    location: str
    fields: list[dict[str, Any]]
    raw_payload: dict[str, Any]


@dataclass(slots=True)
class ParseResult:
    records: list[ParsedRecord] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    source_image_count: int = 0
    supported: bool = True


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def _normalized(value: Any) -> str:
    return re.sub(r"[\s\-_./\\:：()（）\[\]【】]+", "", _text(value).casefold())


def _canonical_header(value: Any) -> str | None:
    raw = _normalized(value)
    if not raw:
        return None
    for key, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            token = _normalized(alias)
            if raw == token or (len(token) >= 3 and token in raw):
                return key
    return None


def _header_map(rows: Sequence[Sequence[Any]]) -> tuple[int, dict[str, int]] | None:
    best: tuple[int, int, dict[str, int]] | None = None
    for row_index in range(min(len(rows), HEADER_SCAN_ROWS)):
        mapping: dict[str, int] = {}
        for column in range(min(len(rows[row_index]), MAX_COLUMNS)):
            current_value = rows[row_index][column] if column < len(rows[row_index]) else None
            canonical = _canonical_header(current_value)
            pieces = []
            if canonical is None:
                for candidate_row in range(max(0, row_index - 2), row_index + 1):
                    if column < len(rows[candidate_row]):
                        value = _text(rows[candidate_row][column])
                        if value:
                            pieces.append(value)
                canonical = _canonical_header(" ".join(pieces))
            if canonical and canonical not in mapping:
                mapping[canonical] = column
        semantic_fields = set(mapping) - {"image"}
        score = len(semantic_fields) + (1 if "name" in mapping or "model" in mapping else 0)
        if score >= 3 and (best is None or score > best[0]):
            best = (score, row_index, mapping)
    return (best[1], best[2]) if best else None


def _field(
    key: str,
    value: str,
    confidence: float | None = None,
    source_location: str | None = None,
) -> dict[str, Any]:
    field = {
        "key": key,
        "label": FIELD_LABELS[key],
        "source": value,
        "normalized": value,
        "confidence": confidence if confidence is not None else (0.95 if key == "name" else 0.9),
    }
    if source_location:
        field["source_location"] = source_location
    return field


def _records_from_rows(rows: Sequence[Sequence[Any]], sheet_name: str) -> tuple[list[ParsedRecord], list[str]]:
    header = _header_map(rows)
    if header is None:
        return [], [f"{sheet_name}：未识别到包含产品名称/型号的表头"]

    header_row, mapping = header
    records: list[ParsedRecord] = []
    warnings: list[str] = []
    current: ParsedRecord | None = None

    for row_index, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        values = {key: _text(row[column]) if column < len(row) else "" for key, column in mapping.items()}
        values.pop("image", None)
        non_empty = {key: value for key, value in values.items() if value}
        if not non_empty:
            continue

        name = values.get("name", "")
        model = values.get("model", "")
        identity_present = bool(name or model)

        if identity_present:
            resolved_name = name or model
            field_locations = {
                key: f"{sheet_name}!{get_column_letter(mapping[key] + 1)}{row_index}"
                for key in non_empty
                if key in mapping
            }
            fields = [
                _field(
                    key,
                    non_empty[key],
                    0.94 if key in {"name", "model"} else 0.86,
                    field_locations.get(key),
                )
                for key in FIELD_ORDER
                if key in non_empty
            ]
            if not name:
                warnings.append(f"{sheet_name} 第 {row_index} 行：缺少产品名称，暂以型号作为名称")
                fields.insert(
                    0,
                    _field("name", resolved_name, 0.55, field_locations.get("model")),
                )
            current = ParsedRecord(
                name=resolved_name,
                model=model,
                category=values.get("category", "待分类") or "待分类",
                location=f"{sheet_name} · 第 {row_index} 行",
                fields=fields,
                raw_payload={
                    "sheet": sheet_name,
                    "row": row_index,
                    "values": non_empty,
                    "field_locations": field_locations,
                },
            )
            records.append(current)
            continue

        if current and any(key in non_empty for key in {"price", "specification", "color", "material", "packing", "weight", "moq"}):
            variant_text = "；".join(f"{FIELD_LABELS.get(key, key)}：{value}" for key, value in non_empty.items())
            existing = next((item for item in current.fields if item["key"] == "variants"), None)
            if existing:
                existing["source"] += f" | 第 {row_index} 行 {variant_text}"
                existing["normalized"] = existing["source"]
            else:
                current.fields.append({
                    "key": "variants",
                    "label": "规格/价格阶梯",
                    "source": f"第 {row_index} 行 {variant_text}",
                    "normalized": f"第 {row_index} 行 {variant_text}",
                    "confidence": 0.72,
                })
            current.raw_payload.setdefault("continuation_rows", []).append({"row": row_index, "values": non_empty})

    return records, warnings


def _count_zip_images(path: Path, prefix: str) -> int:
    try:
        with ZipFile(path) as archive:
            return sum(1 for name in archive.namelist() if name.startswith(prefix) and not name.endswith("/"))
    except (BadZipFile, OSError):
        return 0


def _parse_xlsx(path: Path) -> ParseResult:
    result = ParseResult(source_image_count=_count_zip_images(path, "xl/media/"))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            max_row = min(worksheet.max_row or 1, MAX_ROWS_PER_SHEET)
            max_column = min(worksheet.max_column or 1, MAX_COLUMNS)
            rows = list(worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column, values_only=True))
            records, warnings = _records_from_rows(rows, worksheet.title)
            result.records.extend(records)
            result.warnings.extend(warnings)
            if (worksheet.max_row or 0) > MAX_ROWS_PER_SHEET:
                result.warnings.append(f"{worksheet.title}：超过 {MAX_ROWS_PER_SHEET} 行，当前任务仅解析前半部分")
    finally:
        workbook.close()
    if result.source_image_count:
        result.warnings.append(f"发现 {result.source_image_count} 张来源图片；图片与产品的关联及纯净度需人工复核")
    return result


def _parse_xls(path: Path) -> ParseResult:
    result = ParseResult()
    workbook = xlrd.open_workbook(path, on_demand=True)
    try:
        for sheet in workbook.sheets():
            row_count = min(sheet.nrows, MAX_ROWS_PER_SHEET)
            column_count = min(sheet.ncols, MAX_COLUMNS)
            rows = [[sheet.cell_value(row, column) for column in range(column_count)] for row in range(row_count)]
            records, warnings = _records_from_rows(rows, sheet.name)
            result.records.extend(records)
            result.warnings.extend(warnings)
            if sheet.nrows > MAX_ROWS_PER_SHEET:
                result.warnings.append(f"{sheet.name}：超过 {MAX_ROWS_PER_SHEET} 行，当前任务仅解析前半部分")
    finally:
        workbook.release_resources()
    return result


def _csv_encoding_and_dialect(path: Path) -> tuple[str, csv.Dialect]:
    sample = path.read_bytes()[:CSV_SAMPLE_BYTES]
    if b"\x00" in sample:
        raise ValueError("CSV contains NUL bytes and cannot be treated as a text file")

    decoded: str | None = None
    encoding: str | None = None
    for candidate in CSV_ENCODINGS:
        try:
            decoded = sample.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    if decoded is None or encoding is None:
        raise ValueError("CSV encoding must be UTF-8, UTF-8 BOM, or GB18030")

    try:
        dialect = csv.Sniffer().sniff(decoded, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.get_dialect("excel")
    return encoding, dialect


def _parse_csv(path: Path, source_name: str | None = None) -> ParseResult:
    result = ParseResult()
    encoding, dialect = _csv_encoding_and_dialect(path)
    rows: list[list[str]] = []
    truncated = False
    with path.open("r", encoding=encoding, errors="strict", newline="") as source:
        reader = csv.reader(source, dialect)
        for row_index, row in enumerate(reader, start=1):
            if row_index > MAX_ROWS_PER_SHEET:
                truncated = True
                break
            rows.append(row[:MAX_COLUMNS])
            if len(row) > MAX_COLUMNS:
                result.warnings.append(
                    f"CSV 第 {row_index} 行超过 {MAX_COLUMNS} 列，超出部分未解析"
                )

    sheet_name = Path(source_name or path.name).stem[:80] or "CSV"
    records, warnings = _records_from_rows(rows, sheet_name)
    result.records.extend(records)
    result.warnings.extend(warnings)
    if truncated:
        result.warnings.append(f"CSV 超过 {MAX_ROWS_PER_SHEET} 行，当前任务仅解析前半部分")
    return result


def _vertical_docx_record(rows: Sequence[Sequence[Any]], table_index: int) -> ParsedRecord | None:
    values: dict[str, str] = {}
    for row in rows:
        pairs = zip(row[0::2], row[1::2])
        for key_cell, value_cell in pairs:
            key = _canonical_header(key_cell)
            value = _text(value_cell)
            if key and key != "image" and value:
                values[key] = value
    if not (values.get("name") or values.get("model")):
        return None
    name = values.get("name") or values.get("model") or "待命名产品"
    return ParsedRecord(
        name=name,
        model=values.get("model", ""),
        category=values.get("category", "待分类"),
        location=f"表格 {table_index}",
        fields=[_field(key, values[key], 0.9 if key in {"name", "model"} else 0.84) for key in FIELD_ORDER if key in values],
        raw_payload={"table": table_index, "values": values},
    )


def _parse_docx(path: Path) -> ParseResult:
    result = ParseResult(source_image_count=_count_zip_images(path, "word/media/"))
    document = Document(path)
    for table_index, table in enumerate(document.tables, start=1):
        rows = [[cell.text for cell in row.cells] for row in table.rows]
        if not rows:
            continue
        header = _header_map(rows)
        if header:
            records, warnings = _records_from_rows(rows, f"表格 {table_index}")
            result.records.extend(records)
            result.warnings.extend(warnings)
        else:
            record = _vertical_docx_record(rows, table_index)
            if record:
                result.records.append(record)
    if not result.records:
        result.warnings.append("DOCX 中未识别到结构化产品表格")
    if result.source_image_count:
        result.warnings.append(f"发现 {result.source_image_count} 张来源图片；图片与产品的关联及纯净度需人工复核")
    return result


def parse_document(path: Path, detection: FileDetectionResponse) -> ParseResult:
    if detection.parser == "openpyxl":
        return _parse_xlsx(path)
    if detection.parser == "xlrd":
        return _parse_xls(path)
    if detection.parser == "python-csv":
        return _parse_csv(path, detection.filename)
    if detection.parser == "python-docx":
        return _parse_docx(path)
    return ParseResult(
        warnings=[f"当前版本尚未启用 {detection.detected_type} 解析器，文件已安全保存并等待人工处理"],
        supported=False,
    )
