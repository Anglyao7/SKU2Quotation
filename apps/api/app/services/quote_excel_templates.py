from __future__ import annotations

import math
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


MAX_INSPECT_COLUMNS = 200
MAX_HEADER_SCAN_ROWS = 60
MAX_DATA_SCAN_ROWS = 500


class QuoteTemplateParseError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class InspectedQuoteColumn:
    key: str
    index: int
    header: str
    samples: list[str]
    suggested_field: str | None


@dataclass(frozen=True, slots=True)
class QuoteTemplateInspection:
    sheet_names: list[str]
    sheet_name: str
    header_row: int
    data_start_row: int
    data_end_row: int
    columns: list[InspectedQuoteColumn]
    score: float


FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "serial_number": ("序号", "行号", "no", "number", "#"),
    "sku_code": (
        "sku",
        "skuno",
        "skucode",
        "sku编号",
        "sku编码",
        "sku代码",
        "商品编号",
        "商品编码",
        "产品编号",
        "产品编码",
    ),
    "product_name": (
        "商品名称",
        "产品名称",
        "商品名",
        "产品名",
        "品名",
        "productname",
        "itemname",
        "name",
    ),
    "description": (
        "商品描述",
        "产品描述",
        "描述",
        "详情",
        "description",
        "details",
    ),
    "specification": (
        "规格",
        "产品规格",
        "商品规格",
        "型号规格",
        "specification",
        "spec",
        "variant",
    ),
    "category": ("商品分类", "产品分类", "分类", "category"),
    "tags": ("商品标签", "产品标签", "标签", "tags", "tag"),
    "quantity": ("数量", "采购数量", "报价数量", "quantity", "qty"),
    "unit_code": ("计量单位", "单位", "unit", "uom"),
    "unit_price": (
        "报价单价",
        "销售单价",
        "产品单价",
        "商品单价",
        "单价",
        "unitprice",
        "price",
    ),
    "line_total": (
        "金额小计",
        "行小计",
        "小计",
        "合计金额",
        "金额",
        "linetotal",
        "subtotal",
        "amount",
    ),
    "currency": ("币种", "货币", "currency"),
    "quote_number": (
        "报价单号",
        "报价编号",
        "询价单号",
        "quotenumber",
        "quotationnumber",
        "quoteno",
    ),
    "quote_date": ("报价日期", "日期", "quotedate", "date"),
    "customer_name": ("客户姓名", "联系人", "customername", "contact"),
    "customer_company": (
        "客户公司",
        "公司名称",
        "客户名称",
        "customercompany",
        "company",
    ),
    "customer_email": ("客户邮箱", "邮箱", "email"),
    "customer_phone": ("客户电话", "联系电话", "电话", "phone", "tel"),
    "notes": ("报价备注", "备注", "说明", "notes", "remark"),
}


TOTAL_MARKERS = {
    "合计",
    "总计",
    "总金额",
    "合计金额",
    "subtotal",
    "grandtotal",
    "total",
    "备注",
    "说明",
    "terms",
}


def _normalize(value: object | None) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).casefold().strip()
    return re.sub(r"[\s\-_/\\:：,.，。()（）\[\]【】]+", "", text)


NORMALIZED_ALIASES = {
    field: tuple(_normalize(alias) for alias in aliases)
    for field, aliases in FIELD_ALIASES.items()
}


def _suggest_field(header: object | None) -> str | None:
    normalized = _normalize(header)
    if not normalized:
        return None
    matches: list[tuple[int, str]] = []
    for field, aliases in NORMALIZED_ALIASES.items():
        for alias in aliases:
            if not alias:
                continue
            if normalized == alias:
                matches.append((10_000 + len(alias), field))
            elif len(alias) >= 2 and (
                normalized.startswith(alias)
                or normalized.endswith(alias)
                or alias in normalized
            ):
                matches.append((len(alias), field))
    return max(matches, default=(0, ""))[1] or None


def _display_value(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        return str(int(value)) if value.is_integer() else f"{value:g}"
    return str(value).strip()


def _nonempty_indices(row: Iterable[object | None]) -> list[int]:
    return [index for index, value in enumerate(row, 1) if _display_value(value)]


def _header_score(row: tuple[object | None, ...]) -> float:
    values = [_display_value(value) for value in row]
    nonempty = [value for value in values if value]
    if not nonempty:
        return -1.0
    text_count = sum(not value.replace(".", "", 1).isdigit() for value in nonempty)
    unique_count = len({_normalize(value) for value in nonempty})
    alias_hits = sum(_suggest_field(value) is not None for value in nonempty)
    density = min(len(nonempty), 20)
    return alias_hits * 10 + text_count * 1.2 + unique_count * 0.4 + density


def _looks_like_total(row: tuple[object | None, ...]) -> bool:
    normalized = [_normalize(value) for value in row if _display_value(value)]
    if not normalized:
        return False
    return any(
        value in TOTAL_MARKERS
        or value.startswith("合计")
        or value.startswith("总计")
        or value.endswith("total")
        for value in normalized
    )


def _inspect_sheet(
    worksheet,
    *,
    sheet_names: list[str],
    forced_header_row: int | None,
) -> QuoteTemplateInspection:
    max_column = max(1, min(int(worksheet.max_column or 1), MAX_INSPECT_COLUMNS))
    max_row = max(1, int(worksheet.max_row or 1))
    scan_end = min(
        max_row,
        max(
            MAX_HEADER_SCAN_ROWS,
            (forced_header_row or 1) + MAX_DATA_SCAN_ROWS,
        ),
    )
    rows = list(
        worksheet.iter_rows(
            min_row=1,
            max_row=scan_end,
            min_col=1,
            max_col=max_column,
            values_only=True,
        )
    )
    if not rows:
        raise QuoteTemplateParseError(f"工作表“{worksheet.title}”没有可解析内容")

    if forced_header_row is not None:
        if forced_header_row > len(rows):
            raise QuoteTemplateParseError(
                f"表头行 {forced_header_row} 超出工作表“{worksheet.title}”的有效范围"
            )
        header_row = forced_header_row
        score = _header_score(rows[header_row - 1])
    else:
        candidates = [
            (_header_score(row), index)
            for index, row in enumerate(rows[:MAX_HEADER_SCAN_ROWS], 1)
        ]
        score, header_row = max(candidates, default=(-1.0, 1))
        if score < 0:
            raise QuoteTemplateParseError(f"工作表“{worksheet.title}”没有识别到表头")

    header = rows[header_row - 1]
    data_start_row = header_row + 1
    for row_index in range(header_row + 1, len(rows) + 1):
        candidate = rows[row_index - 1]
        if _nonempty_indices(candidate) and not _looks_like_total(candidate):
            data_start_row = row_index
            break

    data_end_row = data_start_row
    saw_data = False
    for row_index in range(data_start_row, len(rows) + 1):
        candidate = rows[row_index - 1]
        if _looks_like_total(candidate):
            break
        if not _nonempty_indices(candidate):
            if saw_data:
                break
            continue
        saw_data = True
        data_end_row = row_index
    if not saw_data:
        data_end_row = data_start_row

    sampled_rows = rows[
        max(0, data_start_row - 1) : min(len(rows), data_start_row + 5)
    ]
    used_indices = set(_nonempty_indices(header))
    for row in sampled_rows:
        used_indices.update(_nonempty_indices(row))
    if not used_indices:
        raise QuoteTemplateParseError(f"工作表“{worksheet.title}”没有可映射列")
    first_column = min(used_indices)
    last_column = max(used_indices)

    used_suggestions: set[str] = set()
    columns: list[InspectedQuoteColumn] = []
    for index in range(first_column, last_column + 1):
        raw_header = _display_value(header[index - 1])
        samples = []
        for row in sampled_rows:
            sample = _display_value(row[index - 1])
            if sample and sample not in samples:
                samples.append(sample[:240])
            if len(samples) >= 3:
                break
        suggestion = _suggest_field(raw_header)
        if suggestion in used_suggestions:
            suggestion = None
        if suggestion:
            used_suggestions.add(suggestion)
        columns.append(
            InspectedQuoteColumn(
                key=get_column_letter(index),
                index=index,
                header=raw_header or f"第 {index} 列",
                samples=samples,
                suggested_field=suggestion,
            )
        )
    return QuoteTemplateInspection(
        sheet_names=sheet_names,
        sheet_name=worksheet.title,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        columns=columns,
        score=score,
    )


def inspect_quote_excel_template(
    path: Path,
    *,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> QuoteTemplateInspection:
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise QuoteTemplateParseError("无法读取该 Excel，请确认文件未损坏且未加密") from exc
    try:
        sheet_names = list(workbook.sheetnames)
        if not sheet_names:
            raise QuoteTemplateParseError("Excel 中没有工作表")
        if sheet_name is not None:
            if sheet_name not in workbook.sheetnames:
                raise QuoteTemplateParseError(f"工作表“{sheet_name}”不存在")
            return _inspect_sheet(
                workbook[sheet_name],
                sheet_names=sheet_names,
                forced_header_row=header_row,
            )

        inspections: list[QuoteTemplateInspection] = []
        for candidate_name in sheet_names[:30]:
            try:
                inspections.append(
                    _inspect_sheet(
                        workbook[candidate_name],
                        sheet_names=sheet_names,
                        forced_header_row=None,
                    )
                )
            except QuoteTemplateParseError:
                continue
        if not inspections:
            raise QuoteTemplateParseError("没有找到包含商品明细表头的工作表")
        return max(
            inspections,
            key=lambda item: (item.score, len(item.columns), -item.header_row),
        )
    finally:
        workbook.close()
