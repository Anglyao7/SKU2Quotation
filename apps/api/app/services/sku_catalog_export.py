from __future__ import annotations

from collections import Counter
from collections.abc import Mapping, Sequence
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..product_center_models import SKU_TEMPLATE_SOURCE_OPTION_KEY
from ..product_supplier_models import ProductImageRow
from ..repositories.product_center_repository import SkuListRow
from .product_template_import import (
    MAX_PRODUCT_IMAGE_COLUMN_COUNT,
    PRODUCT_MASTER_TEMPLATE_HEADERS,
    PRODUCT_MASTER_TEMPLATE_SHEET,
    SKU_DETAIL_TEMPLATE_HEADERS,
    SKU_DETAIL_TEMPLATE_SHEET,
)

# Export the same workbook contract that the current importer and download
# template use. The previous exporter used a report-shaped schema and packed
# every option into one "规格" cell, which made an exported workbook difficult
# to edit and impossible to round-trip for multi-variant products.
PRODUCT_HEADERS = PRODUCT_MASTER_TEMPLATE_HEADERS
SKU_HEADERS = SKU_DETAIL_TEMPLATE_HEADERS


_RESERVED_OPTION_KEYS = frozenset(
    {
        SKU_TEMPLATE_SOURCE_OPTION_KEY.casefold(),
        "商品编码",
        "商品型号",
        "规格名称",
        "一箱个数",
        "装箱数",
        "毛重",
        "起定数",
        "是否是新品",
        "备注",
        "标签",
        "packing_quantity",
        "units_per_carton",
        "gross_weight",
        "default_moq",
        "moq_unit",
        "weight",
        "weight_unit",
        "product_model",
        "specification",
    }
)


def _category_name(row: SkuListRow) -> str:
    category = row.category
    if category is None:
        return "未分类"
    path = str(category.path or "").strip()
    if path and path.casefold() != category.code.casefold():
        return path
    return category.name


def _text(value: Any) -> str:
    """Return a safe, human-editable value for a template cell."""

    if value is None or isinstance(value, (Mapping, list, tuple, set)):
        return ""
    return str(value).strip()


def _option_value(values: Mapping[str, Any], name: str) -> Any:
    if name in values:
        return values[name]
    normalized = name.casefold()
    for key, value in values.items():
        if _text(key).casefold() == normalized:
            return value
    return None


def _variant_options(values: Mapping[str, Any]) -> tuple[tuple[str, str], ...]:
    """Extract ordered variant dimensions without leaking internal metadata."""

    ordered_keys: list[str] = []
    marker = values.get(SKU_TEMPLATE_SOURCE_OPTION_KEY)
    if isinstance(marker, Mapping):
        marker_keys = marker.get("variant_option_keys")
        if isinstance(marker_keys, Sequence) and not isinstance(
            marker_keys, (str, bytes)
        ):
            ordered_keys.extend(_text(key) for key in marker_keys if _text(key))

    for key in values:
        text_key = _text(key)
        if text_key and text_key not in ordered_keys:
            ordered_keys.append(text_key)

    options: list[tuple[str, str]] = []
    seen: set[str] = set()
    for key in ordered_keys:
        normalized_key = key.casefold()
        if normalized_key in _RESERVED_OPTION_KEYS or normalized_key in seen:
            continue
        value = _text(_option_value(values, key))
        if not value:
            continue
        options.append((key, value))
        seen.add(normalized_key)
        if len(options) == 3:
            break
    return tuple(options)


def _units_per_carton(values: Mapping[str, Any]) -> Any:
    for key in ("装箱数", "一箱个数", "packing_quantity", "units_per_carton"):
        value = values.get(key)
        if value not in (None, ""):
            return value
    return None


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _first_option_text(rows: Sequence[SkuListRow], name: str) -> str:
    for row in rows:
        value = _text(_option_value(row.sku.option_values, name))
        if value:
            return value
    return ""


def _product_tags(rows: Sequence[SkuListRow]) -> str:
    tags: list[str] = []
    seen: set[str] = set()
    for row in rows:
        offer_tags = row.public_offer.tags if row.public_offer is not None else ()
        if isinstance(offer_tags, Sequence) and not isinstance(offer_tags, (str, bytes)):
            raw_tags: Sequence[Any] = offer_tags
        else:
            raw_tags = (_text(offer_tags),)
        for raw_tag in raw_tags:
            tag = _text(raw_tag)
            if not tag:
                continue
            normalized = tag.casefold()
            if normalized not in seen:
                tags.append(tag)
                seen.add(normalized)

        option_tags = _text(_option_value(row.sku.option_values, "标签"))
        if option_tags:
            for tag in option_tags.replace("，", ",").replace("；", ",").split(","):
                tag = tag.strip()
                if tag and tag.casefold() not in seen:
                    tags.append(tag)
                    seen.add(tag.casefold())
    return "，".join(tags)


def _product_price(
    rows: Sequence[SkuListRow],
    *,
    public_price_overrides: Mapping[UUID, float] | None = None,
) -> float:
    for row in rows:
        if public_price_overrides is not None and row.sku.id in public_price_overrides:
            return float(public_price_overrides[row.sku.id])
        if row.public_offer is not None:
            price = _number(row.public_offer.unit_price)
            if price is not None:
                return price
        price = _number(_option_value(row.sku.option_values, "商品价格"))
        if price is not None:
            return price
    return 0.0


def _export_sku_identifier(
    sku: Any,
    *,
    source_is_unique: bool,
    include_source_sku_codes: bool = True,
) -> str:
    """Use source codes for ordinary SKUs and system codes for variants.

    A source code identifies an import definition and can intentionally be
    shared by several expanded variants. In that case the system SKU code is
    the only unambiguous value that can be exported and imported again.
    """

    system_code = _text(getattr(sku, "sku_code", None))
    source_code = _text(getattr(sku, "source_sku_code", None))
    return (
        source_code
        if (
            include_source_sku_codes
            and source_code
            and source_is_unique
            and not _variant_options(sku.option_values)
        )
        else system_code
    )


def _style_sheet(sheet: object, *, headers: Sequence[str], widths: Sequence[int]) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="2D1B69")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90


def build_sku_catalog_workbook(
    *,
    rows: Sequence[SkuListRow],
    images_by_product: Mapping[UUID, Sequence[ProductImageRow]],
    image_urls: Mapping[UUID, str],
    supplier_names: Mapping[str, str],
    public_price_overrides: Mapping[UUID, float] | None = None,
    include_source_sku_codes: bool = True,
) -> bytes:
    """Build a Product + SKU workbook with editable variant columns.

    Product image cells intentionally contain the public object-storage URL
    (and a hyperlink) instead of embedding binary image objects. This keeps
    exports small and preserves the R2 address when the workbook is edited.
    """

    workbook = Workbook()
    product_sheet = workbook.active
    product_sheet.title = PRODUCT_MASTER_TEMPLATE_SHEET
    sku_sheet = workbook.create_sheet(SKU_DETAIL_TEMPLATE_SHEET)
    product_sheet.append(list(PRODUCT_HEADERS))
    sku_sheet.append(list(SKU_HEADERS))

    _style_sheet(
        product_sheet,
        headers=PRODUCT_HEADERS,
        widths=(
            20,
            30,
            24,
            20,
            14,
            46,
            28,
            28,
            *([18] * MAX_PRODUCT_IMAGE_COLUMN_COUNT),
        ),
    )
    _style_sheet(
        sku_sheet,
        headers=SKU_HEADERS,
        widths=(
            20,
            20,
            32,
            16,
            *([16] * 5),
            16,
            *([16] * 5),
            16,
            *([16] * 5),
            24,
            14,
            14,
            14,
            14,
        ),
    )
    product_sheet.sheet_properties.tabColor = "D4AF37"
    sku_sheet.sheet_properties.tabColor = "42A58B"

    product_rows: dict[UUID, list[SkuListRow]] = {}
    for row in rows:
        product_rows.setdefault(row.product.id, []).append(row)

    image_offset = PRODUCT_HEADERS.index("商品图片1") + 1
    for row_number, product_rows_for_product in enumerate(product_rows.values(), start=2):
        row = product_rows_for_product[0]
        product = row.product
        images = list(images_by_product.get(product.id, ()))[:MAX_PRODUCT_IMAGE_COLUMN_COUNT]
        urls = [image_urls.get(image.id, "") for image in images]
        product_sheet.append(
            [
                product.product_code or "",
                product.name or "",
                _category_name(row),
                _first_option_text(product_rows_for_product, "商品型号"),
                _product_price(
                    product_rows_for_product,
                    public_price_overrides=public_price_overrides,
                ),
                product.description or "",
                _first_option_text(product_rows_for_product, "备注"),
                _product_tags(product_rows_for_product),
                *urls,
                *("" for _ in range(MAX_PRODUCT_IMAGE_COLUMN_COUNT - len(urls))),
            ]
        )
        for offset, url in enumerate(urls, start=image_offset):
            cell = product_sheet.cell(row=row_number, column=offset)
            if url.startswith(("https://", "http://")):
                cell.hyperlink = url
                cell.style = "Hyperlink"

    source_counts = Counter(
        _text(row.sku.source_sku_code)
        for row in rows
        if _text(row.sku.source_sku_code)
        and not _variant_options(row.sku.option_values)
    )
    for row in rows:
        sku = row.sku
        offer = row.public_offer
        options = _variant_options(sku.option_values)
        source_code = _text(sku.source_sku_code)
        sku_values: list[Any] = [
            row.product.product_code or "",
            _export_sku_identifier(
                sku,
                source_is_unique=bool(source_code and source_counts[source_code] == 1),
                include_source_sku_codes=include_source_sku_codes,
            ),
            sku.name or row.product.name,
        ]
        for dimension in range(3):
            if dimension < len(options):
                option_name, option_value = options[dimension]
                sku_values.extend([option_name, option_value, "", "", "", ""])
            else:
                sku_values.extend(["", "", "", "", "", ""])
        sku_values.extend(
            [
                supplier_names.get(sku.supplier_id or "", ""),
                (
                    float(public_price_overrides[sku.id])
                    if public_price_overrides is not None and sku.id in public_price_overrides
                    else _number(offer.unit_price) if offer is not None else 0.0
                ),
                _number(sku.weight),
                _number(sku.default_moq),
                _number(_units_per_carton(sku.option_values)),
            ]
        )
        sku_sheet.append(sku_values)

    if product_sheet.max_row >= 2:
        product_sheet.auto_filter.ref = f"A1:{get_column_letter(len(PRODUCT_HEADERS))}{product_sheet.max_row}"
        for row in product_sheet.iter_rows(min_row=2, max_col=len(PRODUCT_HEADERS)):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=cell.column in {2, 3, 4, 6, 7, 8},
                )
            row[0].number_format = "@"
            row[4].number_format = "0.00"
            product_sheet.row_dimensions[row[0].row].height = 36

    if sku_sheet.max_row >= 2:
        sku_sheet.auto_filter.ref = f"A1:{get_column_letter(len(SKU_HEADERS))}{sku_sheet.max_row}"
        number_formats = {
            "SKU价格": "0.00",
            "毛重": "0.######",
            "起定数": "0.######",
            "装箱数": "0.######",
        }
        for header, number_format in number_formats.items():
            column = SKU_HEADERS.index(header) + 1
            for cells in sku_sheet.iter_cols(
                min_col=column,
                max_col=column,
                min_row=2,
                max_row=sku_sheet.max_row,
            ):
                for cell in cells:
                    cell.number_format = number_format
        for row in sku_sheet.iter_rows(min_row=2, max_col=len(SKU_HEADERS)):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=2 <= cell.column <= 22,
                )
            sku_sheet.row_dimensions[row[0].row].height = 30

    workbook.calculation.fullCalcOnLoad = True
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
