from __future__ import annotations

import os
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

import psycopg
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import and_, or_, select
from sqlalchemy.orm import Session

from ..database import SessionLocal, set_request_context
from ..data import PRODUCTS
from ..domain.errors import ApplicationError
from ..models import (
    FileDetectionResponse,
    ImportJob,
    PriceCalculationRequest,
    PriceCalculationResponse,
    Product,
    ReviewApprovalResponse,
    ReviewItem,
    ReviewItemUpdate,
    Supplier,
    SupplierFileImportResponse,
)
from ..repositories.legacy_repository import (
    find_supplier,
    get_review_item,
)
from ..repositories.file_security_repository import add_file_security_records
from ..adapters.object_storage import get_object_storage
from ..file_security_models import MediaObjectRow, WorkerJobRow
from ..db_models import ImportJobRow, SourceFileRow
from ..model_mixins import utcnow
from ..services.file_detection import detect_file_path, detect_file_type
from ..services.import_processing import new_id
from ..services.pricing import calculate_price
from ..services.product_template_import import (
    PRODUCT_MASTER_TEMPLATE_HEADERS,
    PRODUCT_MASTER_TEMPLATE_SHEET,
    SKU_DETAIL_TEMPLATE_HEADERS,
    SKU_DETAIL_TEMPLATE_SHEET,
)
from ..services.repository import (
    get_import_job,
    import_job_model,
    list_import_job_models,
    list_review_item_models,
    review_item_model,
    supplier_models,
)
from ..services.storage import UploadTooLargeError, store_upload
from ..workers.file_processing import inline_worker_enabled, process_file_worker_job


_deferred_import_executor = ThreadPoolExecutor(
    max_workers=1,
    thread_name_prefix="local-import-worker",
)
_ZERO_IDENTITY = UUID(int=0)


def _require_permission(permissions: frozenset[str], code: str) -> None:
    if code not in permissions:
        raise ApplicationError(
            "PERMISSION_DENIED",
            f"Permission is required: {code}",
            kind="forbidden",
        )


def list_suppliers(
    session: Session,
    *,
    tenant_id: UUID,
    permissions: frozenset[str],
) -> list[Supplier]:
    _require_permission(permissions, "supplier.view")
    return supplier_models(session, tenant_id=tenant_id)


def list_imports(
    session: Session,
    *,
    tenant_id: UUID,
    permissions: frozenset[str],
    limit: int,
) -> list[ImportJob]:
    _require_permission(permissions, "product.import")
    return list_import_job_models(session, tenant_id=tenant_id, limit=limit)


def get_import(
    session: Session,
    *,
    tenant_id: UUID,
    permissions: frozenset[str],
    job_id: str,
) -> ImportJob:
    _require_permission(permissions, "product.import")
    row = get_import_job(session, job_id, tenant_id=tenant_id)
    if row is None:
        raise ApplicationError("IMPORT_NOT_FOUND", "Import job was not found.", kind="not_found")
    return import_job_model(row)


def detect_upload(filename: str, header: bytes) -> FileDetectionResponse:
    return detect_file_type(filename, header)


async def create_import(
    session: Session,
    *,
    upload: Any,
    supplier_id: str | None,
    source_type: str,
    tenant_id: UUID,
    user_id: UUID,
    permissions: frozenset[str],
    defer_inline_worker: bool = False,
) -> SupplierFileImportResponse:
    _require_permission(permissions, "product.import")
    original_filename = Path(upload.filename or "unnamed").name
    normalized_source_type = source_type.strip().upper()[:40] or "UNKNOWN"
    if normalized_source_type == "PRODUCT_TEMPLATE":
        # A fixed-template import writes authoritative products and may publish
        # public offers. Requiring all three capabilities prevents a scoped
        # importer from escalating into product editing or catalog publishing.
        _require_permission(permissions, "product.edit")
        _require_permission(permissions, "catalog.publish")
    if (
        normalized_source_type == "PRODUCT_TEMPLATE"
        and Path(original_filename).suffix.lower() != ".xlsx"
    ):
        raise ApplicationError(
            "PRODUCT_TEMPLATE_FORMAT_INVALID",
            "商品库只接受 .xlsx 商品文件。",
        )
    supplier = (
        find_supplier(session, tenant_id=tenant_id, supplier_id=supplier_id)
        if supplier_id
        else None
    )
    if supplier_id and supplier is None:
        raise ApplicationError("SUPPLIER_NOT_FOUND", "Supplier was not found.", kind="not_found")
    if normalized_source_type in {"SUPPLIER_CATALOG", "SUPPLIER_QUOTE"} and supplier is None:
        raise ApplicationError(
            "SUPPLIER_REQUIRED",
            "Select a supplier before importing a supplier catalog or quote.",
        )
    source_id = new_id("SRC")
    job_id = new_id("JOB")
    storage = get_object_storage()
    try:
        stored = await store_upload(
            upload,
            source_id,
            tenant_id=tenant_id,
            storage=storage,
        )
    except UploadTooLargeError as exc:
        raise ApplicationError("UPLOAD_TOO_LARGE", str(exc), kind="too_large") from exc

    with storage.materialize(stored.object_key) as quarantine_path:
        detection = detect_file_path(quarantine_path, original_filename)
    if normalized_source_type == "PRODUCT_TEMPLATE" and (
        detection.detected_type != "OOXML / XLSX" or not detection.extension_matches
    ):
        try:
            storage.delete(stored.object_key)
        except Exception:
            pass
        raise ApplicationError(
            "PRODUCT_TEMPLATE_FORMAT_INVALID",
            "文件不是有效的 XLSX 商品模版，请使用根目录约定的商品模版格式。",
        )
    now = utcnow()
    media_id = uuid4()
    worker_job_id = uuid4()
    media = MediaObjectRow(
        id=media_id,
        tenant_id=tenant_id,
        object_key=stored.object_key,
        zone="QUARANTINE",
        original_filename=original_filename,
        sha256=stored.sha256,
        byte_size=stored.byte_size,
        declared_media_type=stored.declared_media_type,
        detected_media_type=detection.detected_type,
        status="QUARANTINED",
        scan_status="PENDING",
        scan_result={},
        created_by_user_id=user_id,
    )
    local_quarantine_path = storage.local_path(stored.object_key)
    source = SourceFileRow(
        id=source_id,
        tenant_id=tenant_id,
        media_object_id=media_id,
        security_status="PENDING_SCAN",
        original_filename=original_filename,
        stored_filename=stored.stored_filename,
        local_path=str(local_quarantine_path) if local_quarantine_path else "",
        sha256=stored.sha256,
        byte_size=stored.byte_size,
        extension=detection.extension,
        detected_type=detection.detected_type,
        extension_matches=detection.extension_matches,
        parser=detection.parser,
        warning=detection.warning,
    )
    import_job = ImportJobRow(
        id=job_id,
        tenant_id=tenant_id,
        source_file_id=source_id,
        supplier_id=supplier.id if supplier else None,
        supplier_name=(
            supplier.name
            if supplier
            else "商品模版"
            if normalized_source_type == "PRODUCT_TEMPLATE"
            else "Supplier pending selection"
        ),
        source_type=normalized_source_type,
        status="scanning",
        progress=5,
        warnings_count=1 if detection.warning else 0,
    )
    worker_job = WorkerJobRow(
        id=worker_job_id,
        tenant_id=tenant_id,
        job_type="FILE_SCAN_AND_PARSE",
        media_object_id=media_id,
        source_file_id=source_id,
        import_job_id=job_id,
        status="PENDING",
        attempt_count=0,
        max_attempts=int(os.getenv("FILE_WORKER_MAX_ATTEMPTS", "3")),
        available_at=now,
        checkpoint={},
        idempotency_key=f"file-scan-parse:{source_id}:{stored.sha256}",
    )
    add_file_security_records(
        session,
        media=media,
        source=source,
        import_job=import_job,
        worker_job=worker_job,
    )
    session.commit()

    worker_result = None
    dialect = session.bind.dialect.name if session.bind is not None else "unknown"
    if inline_worker_enabled(database_dialect=dialect) and not defer_inline_worker:
        worker_result = process_file_worker_job(
            session,
            tenant_id=tenant_id,
            job_id=worker_job_id,
            worker_id="inline-api-worker",
            storage=storage,
        )
    session.expire_all()
    completed = get_import_job(session, job_id, tenant_id=tenant_id)
    if completed is None:
        raise ApplicationError(
            "IMPORT_CREATION_FAILED",
            "Import job could not be loaded after processing.",
            kind="internal",
        )
    return SupplierFileImportResponse(
        **import_job_model(completed).model_dump(),
        ai_task_id=str(worker_result.ai_task_id) if worker_result and worker_result.ai_task_id else None,
        candidate_fields=worker_result.candidate_fields if worker_result else 0,
        candidate_status=worker_result.candidate_status if worker_result else None,
        candidate_idempotent=worker_result.candidate_idempotent if worker_result else False,
    )


def inline_import_worker_enabled(session: Session) -> bool:
    dialect = session.bind.dialect.name if session.bind is not None else "unknown"
    return inline_worker_enabled(database_dialect=dialect)


def _inline_import_tenant_ids(session: Session) -> tuple[UUID, ...]:
    """Discover tenants without weakening tenant-scoped worker access."""

    dialect = session.bind.dialect.name if session.bind is not None else "unknown"
    if dialect != "postgresql":
        return tuple(
            session.scalars(select(WorkerJobRow.tenant_id).distinct()).all()
        )

    directory_url = os.getenv("TENANT_DIRECTORY_DATABASE_URL", "").strip()
    if not directory_url:
        return ()
    psycopg_url = directory_url.replace(
        "postgresql+psycopg://",
        "postgresql://",
        1,
    )
    with psycopg.connect(psycopg_url, connect_timeout=5) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT id FROM tenants "
                "WHERE status = 'active' AND deleted_at IS NULL ORDER BY id"
            )
            return tuple(UUID(str(row[0])) for row in cursor.fetchall())


def process_deferred_import(*, tenant_id: UUID, import_job_id: str) -> None:
    """Run an inline-profile job after the HTTP response has been sent."""

    with SessionLocal() as session:
        set_request_context(
            session,
            organization_id=UUID(int=0),
            tenant_id=tenant_id,
            user_id=UUID(int=0),
        )
        now = utcnow()
        worker_job_id = session.scalar(
            select(WorkerJobRow.id)
            .where(
                WorkerJobRow.tenant_id == tenant_id,
                WorkerJobRow.import_job_id == import_job_id,
                or_(
                    and_(
                        WorkerJobRow.status.in_(("PENDING", "RETRY")),
                        WorkerJobRow.available_at <= now,
                    ),
                    and_(
                        WorkerJobRow.status == "RUNNING",
                        WorkerJobRow.lease_expires_at.is_not(None),
                        WorkerJobRow.lease_expires_at <= now,
                    ),
                ),
            )
            .order_by(WorkerJobRow.created_at.desc())
            .limit(1)
        )
        session.rollback()
        if worker_job_id is None:
            return
        process_file_worker_job(
            session,
            tenant_id=tenant_id,
            job_id=worker_job_id,
            worker_id="deferred-api-worker",
            storage=get_object_storage(),
        )


def resume_deferred_imports() -> int:
    """Resume interrupted inline-profile imports after a local API restart."""

    with SessionLocal() as session:
        if not inline_import_worker_enabled(session):
            return 0
        tenant_ids = _inline_import_tenant_ids(session)
        session.rollback()

    queued: list[tuple[UUID, str]] = []
    for tenant_id in tenant_ids:
        with SessionLocal() as session:
            set_request_context(
                session,
                organization_id=_ZERO_IDENTITY,
                tenant_id=tenant_id,
                user_id=_ZERO_IDENTITY,
            )
            now = utcnow()

            # In the inline profile, all import workers live inside this API
            # process. A RUNNING lease found during startup therefore belongs
            # to the process that just stopped (for example after an OOM
            # restart), even when its short lease has not expired yet.
            interrupted = session.scalars(
                select(WorkerJobRow).where(
                    WorkerJobRow.tenant_id == tenant_id,
                    WorkerJobRow.import_job_id.is_not(None),
                    WorkerJobRow.status == "RUNNING",
                )
            ).all()
            for worker_job in interrupted:
                worker_job.lease_expires_at = now
            if interrupted:
                session.flush()

            tenant_queue = session.execute(
                select(WorkerJobRow.tenant_id, WorkerJobRow.import_job_id)
                .where(
                    WorkerJobRow.tenant_id == tenant_id,
                    WorkerJobRow.import_job_id.is_not(None),
                    or_(
                        and_(
                            WorkerJobRow.status.in_(("PENDING", "RETRY")),
                            WorkerJobRow.available_at <= now,
                        ),
                        and_(
                            WorkerJobRow.status == "RUNNING",
                            WorkerJobRow.lease_expires_at.is_not(None),
                            WorkerJobRow.lease_expires_at <= now,
                        ),
                    ),
                )
                .order_by(WorkerJobRow.available_at, WorkerJobRow.created_at)
            ).all()
            queued.extend(
                (queued_tenant_id, import_job_id)
                for queued_tenant_id, import_job_id in tenant_queue
                if import_job_id is not None
            )
            session.commit()

    for tenant_id, import_job_id in queued:
        _deferred_import_executor.submit(
            process_deferred_import,
            tenant_id=tenant_id,
            import_job_id=import_job_id,
        )
    return len(queued)


def build_product_template_workbook() -> bytes:
    """Build the canonical Product + SKU workbook used by new imports."""

    workbook = Workbook()
    product_sheet = workbook.active
    product_sheet.title = PRODUCT_MASTER_TEMPLATE_SHEET
    sku_sheet = workbook.create_sheet(SKU_DETAIL_TEMPLATE_SHEET)
    product_sheet.append(list(PRODUCT_MASTER_TEMPLATE_HEADERS))
    sku_sheet.append(list(SKU_DETAIL_TEMPLATE_HEADERS))

    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_border = Border(bottom=Side(style="medium", color="D4AF37"))
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    product_instructions = {
        "商品编码": "必填且唯一。用于连接 Product 与 SKU 两张表；后续增量更新时请保持不变。",
        "商品名称": "必填。面向客户展示的商品名称；一个商品可以关联多个 SKU。",
        "商品分类": "选填。填写“一级分类”或“一级分类/二级分类”，最多两级；留空归入“未分类”且不进入智能索引。",
        "商品型号": "选填。商品级型号，会写入该商品下各 SKU 的商品信息。",
        "商品价格": "选填。作为该商品下 SKU 的默认价格；SKU 表填写 SKU 价格时以 SKU 价格为准，均留空按 0.00 处理。",
        "商品描述": "选填。展示在商品详情页，并参与智能搜索内容构建。",
        "备注": "选填。商品级补充说明。",
        "标签": "选填。多个标签使用中文或英文逗号分隔，最多 20 个，并应用到该商品下全部 SKU。",
    }
    sku_instructions = {
        "商品编码": "必填。必须与 Product 表中的商品编码一致，用于确定该 SKU 属于哪个商品。",
        "SKU编号": "必填且全表唯一。没有候选规格时直接作为 SKU 编号；填写多个候选值时作为稳定编号前缀，系统会为每种组合生成独立 SKU。",
        "SKU名称": "选填。作为生成后 SKU 的名称前缀；每个具体 SKU 会自动附加所选规格值。留空时使用商品名称。",
        "供应商": "选填。用于关联进销存；同名供应商自动复用，不存在时自动创建。",
        "SKU价格": "选填。填写时覆盖 Product 表的商品价格；留空时继承商品价格。",
        "毛重": "选填。该 SKU 的毛重，单位为 kg；填写大于等于 0 的数字。",
        "起定数": "选填。该 SKU 的最小起订数量；填写大于等于 0 的数字。",
        "装箱数": "选填。每箱装入的 SKU 数量，例如 24。",
    }
    option_examples = (
        ("尺寸", "小号"),
        ("颜色", "红色"),
        ("材质", "不锈钢"),
    )
    for option_number, (option_name, option_value) in enumerate(
        option_examples,
        start=1,
    ):
        sku_instructions[f"规格{option_number}名称"] = (
            f"选填。该组候选值的规格名称，例如“{option_name}”。填写名称后至少填写一个候选值。"
        )
        for value_number in range(1, 6):
            sku_instructions[f"规格{option_number}值（{value_number}）"] = (
                f"选填。规格“{option_name}”的第 {value_number} 个候选值"
                f"；例如“{option_value}”。系统会与其他规格组自动组合。"
            )
    image_instruction = (
        "选填。可把真实图片直接插入对应单元格位置，"
        "也可填写可公开访问的 HTTP(S) 商品图片地址。图片属于 Product，"
        "会供该商品下全部 SKU 共用。"
    )

    def style_sheet(
        sheet: object,
        *,
        headers: tuple[str, ...],
        widths: tuple[int, ...],
        instructions: dict[str, str],
        header_color: str,
        tab_color: str,
    ) -> None:
        header_fill = PatternFill(fill_type="solid", fgColor=header_color)
        for index, (header, width) in enumerate(
            zip(headers, widths, strict=True),
            start=1,
        ):
            cell = sheet.cell(row=1, column=index)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = header_alignment
            cell.comment = Comment(
                instructions.get(header, image_instruction),
                "智贸云",
            )
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.row_dimensions[1].height = 36
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.zoomScale = 90
        sheet.sheet_properties.tabColor = tab_color
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_title_rows = "1:1"

    style_sheet(
        product_sheet,
        headers=PRODUCT_MASTER_TEMPLATE_HEADERS,
        widths=(18, 28, 22, 20, 14, 44, 28, 26, *([38] * 10)),
        instructions=product_instructions,
        header_color="2D1B69",
        tab_color="D4AF37",
    )
    style_sheet(
        sku_sheet,
        headers=SKU_DETAIL_TEMPLATE_HEADERS,
        widths=(
            18, 22, 30,
            18, 16, 16, 16, 16, 16,
            18, 16, 16, 16, 16, 16,
            18, 16, 16, 16, 16, 16,
            28, 14, 14, 14, 14,
        ),
        instructions=sku_instructions,
        header_color="23453B",
        tab_color="42A58B",
    )
    sku_sheet.freeze_panes = "D2"
    for start_column, end_column, color in (
        (4, 9, "4B3A79"),
        (10, 15, "285D61"),
        (16, 21, "755A28"),
    ):
        group_fill = PatternFill(fill_type="solid", fgColor=color)
        for column in range(start_column, end_column + 1):
            sku_sheet.cell(row=1, column=column).fill = group_fill

    for sheet, cell_range, title in (
        (product_sheet, f"E2:E{20_001}", "商品价格格式错误"),
        (sku_sheet, f"W2:W{20_001}", "SKU 价格格式错误"),
    ):
        price_validation = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
        )
        price_validation.error = "价格必须是大于或等于 0 的数字，也可以留空。"
        price_validation.errorTitle = title
        price_validation.showErrorMessage = True
        sheet.add_data_validation(price_validation)
        price_validation.add(cell_range)

    for cell_range, title in (
        (f"X2:X{20_001}", "毛重格式错误"),
        (f"Y2:Y{20_001}", "起定数格式错误"),
        (f"Z2:Z{20_001}", "装箱数格式错误"),
    ):
        quantity_validation = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
        )
        quantity_validation.error = "请填写大于或等于 0 的数字，也可以留空。"
        quantity_validation.errorTitle = title
        quantity_validation.showErrorMessage = True
        sku_sheet.add_data_validation(quantity_validation)
        quantity_validation.add(cell_range)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def list_review_items(
    session: Session,
    *,
    tenant_id: UUID,
    permissions: frozenset[str],
    job_id: str | None,
    review_status: str | None,
    limit: int,
) -> list[ReviewItem]:
    _require_permission(permissions, "product.review")
    return list_review_item_models(
        session,
        tenant_id=tenant_id,
        job_id=job_id,
        status=review_status,
        limit=limit,
    )


def update_review_item(
    session: Session,
    *,
    tenant_id: UUID,
    permissions: frozenset[str],
    item_id: str,
    request: ReviewItemUpdate,
) -> ReviewItem:
    _require_permission(permissions, "product.review")
    row = get_review_item(session, tenant_id=tenant_id, item_id=item_id)
    if row is None:
        raise ApplicationError("REVIEW_ITEM_NOT_FOUND", "Review item was not found.", kind="not_found")
    fields = [dict(field) for field in row.fields]
    valid_keys = {field.get("key") for field in fields}
    unknown_keys = set(request.normalized_values) - valid_keys
    if unknown_keys:
        raise ApplicationError(
            "UNKNOWN_REVIEW_FIELD",
            f"Unknown fields: {', '.join(sorted(unknown_keys))}",
        )
    for field in fields:
        key = field.get("key")
        if key in request.normalized_values:
            field["normalized"] = request.normalized_values[key].strip()
    row.fields = fields
    row.status = "pending"
    session.commit()
    session.refresh(row)
    return review_item_model(row)


def approve_review_item(
    session: Session,
    *,
    tenant_id: UUID,
    permissions: frozenset[str],
    item_id: str,
) -> ReviewApprovalResponse:
    _require_permission(permissions, "product.review")
    row = get_review_item(session, tenant_id=tenant_id, item_id=item_id)
    if row is None:
        raise ApplicationError("REVIEW_ITEM_NOT_FOUND", "Review item was not found.", kind="not_found")
    row.status = "approved"
    session.commit()
    return ReviewApprovalResponse(id=row.id, status=row.status, image_status=row.image_status)


def list_products(
    *,
    query: str,
    supplier: str | None,
    approved_images_only: bool,
) -> list[Product]:
    normalized_query = query.casefold().strip()
    rows = PRODUCTS
    if normalized_query:
        rows = [
            row
            for row in rows
            if normalized_query
            in " ".join(
                [row.name, row.model, row.category, row.supplier, *row.tags]
            ).casefold()
        ]
    if supplier:
        rows = [row for row in rows if row.supplier == supplier]
    if approved_images_only:
        rows = [row for row in rows if row.image_status == "APPROVED"]
    return rows


def calculate_pricing(request: PriceCalculationRequest) -> PriceCalculationResponse:
    return calculate_price(request)
