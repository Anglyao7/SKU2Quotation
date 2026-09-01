import hashlib
import re
from decimal import Decimal
from pathlib import Path
from uuid import UUID, uuid4
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from app.services import product_template_import as product_template_import_service
from app.services.product_template_import import (
    PRODUCT_MASTER_TEMPLATE_HEADERS,
    PRODUCT_MASTER_TEMPLATE_HEADERS_V3,
    PRODUCT_MASTER_TEMPLATE_SHEET,
    PRODUCT_TEMPLATE_HEADERS,
    PRODUCT_TEMPLATE_SHEET,
    PRODUCT_VARIANT_TEMPLATE_HEADERS,
    SKU_CATALOG_EXPORT_PRODUCT_HEADERS,
    SKU_CATALOG_EXPORT_SKU_HEADERS,
    SKU_DETAIL_TEMPLATE_HEADERS,
    SKU_DETAIL_TEMPLATE_HEADERS_V3,
    SKU_DETAIL_TEMPLATE_HEADERS_V4,
    SKU_DETAIL_TEMPLATE_SHEET,
    ProductTemplateValidationError,
    parse_product_template,
)
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import PatternFill


def _write_workbook(path: Path, rows: list[list[object]], *, headers=None) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = PRODUCT_TEMPLATE_SHEET
    sheet.append(list(headers or PRODUCT_TEMPLATE_HEADERS))
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def _write_product_sku_workbook(
    path: Path,
    *,
    product_rows: list[list[object]],
    sku_rows: list[list[object]],
    product_sheet_name: str = PRODUCT_MASTER_TEMPLATE_SHEET,
    sku_sheet_name: str = SKU_DETAIL_TEMPLATE_SHEET,
) -> None:
    workbook = Workbook()
    product_sheet = workbook.active
    product_sheet.title = product_sheet_name
    product_sheet.append(list(PRODUCT_MASTER_TEMPLATE_HEADERS))
    for row in product_rows:
        product_sheet.append(row)
    sku_sheet = workbook.create_sheet(sku_sheet_name)
    sku_sheet.append(list(SKU_DETAIL_TEMPLATE_HEADERS))
    for row in sku_rows:
        sku_sheet.append(row)
    workbook.save(path)
    workbook.close()


def test_supplier_identity_is_scoped_to_merchant_and_normalizes_name() -> None:
    first_tenant = UUID("11111111-1111-1111-1111-111111111111")
    second_tenant = UUID("22222222-2222-2222-2222-222222222222")

    first_identity = product_template_import_service._supplier_identity(
        first_tenant,
        "  同名供应商  ",
    )
    normalized_identity = product_template_import_service._supplier_identity(
        first_tenant,
        "同名供应商",
    )
    other_tenant_identity = product_template_import_service._supplier_identity(
        second_tenant,
        "同名供应商",
    )

    assert first_identity == normalized_identity
    assert first_identity != other_tenant_identity


def test_product_sku_template_allows_limit_effective_rows_and_ignores_formatted_blanks(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(product_template_import_service, "MAX_TEMPLATE_ROWS", 2)
    path = tmp_path / "有效行边界.xlsx"
    _write_product_sku_workbook(
        path,
        product_rows=[
            [
                "PRODUCT-1",
                "商品一",
                *([None] * (len(PRODUCT_MASTER_TEMPLATE_HEADERS) - 2)),
            ],
            [
                "PRODUCT-2",
                "商品二",
                *([None] * (len(PRODUCT_MASTER_TEMPLATE_HEADERS) - 2)),
            ],
        ],
        sku_rows=[
            [
                "PRODUCT-1",
                "SKU-1",
                *([None] * (len(SKU_DETAIL_TEMPLATE_HEADERS) - 2)),
            ],
            [
                "PRODUCT-2",
                "SKU-2",
                *([None] * (len(SKU_DETAIL_TEMPLATE_HEADERS) - 2)),
            ],
        ],
    )
    workbook = load_workbook(path)
    blank_style = PatternFill(fill_type="solid", fgColor="FFF4CC")
    workbook[PRODUCT_MASTER_TEMPLATE_SHEET]["A50"].fill = blank_style
    workbook[SKU_DETAIL_TEMPLATE_SHEET]["A60"].fill = blank_style
    workbook.save(path)
    workbook.close()

    result = parse_product_template(path)

    assert [row.sku_code for row in result.rows] == ["SKU-1", "SKU-2"]


def test_product_sku_template_rejects_only_after_effective_row_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(product_template_import_service, "MAX_TEMPLATE_ROWS", 2)
    path = tmp_path / "有效行超限.xlsx"
    _write_product_sku_workbook(
        path,
        product_rows=[[
            "PRODUCT-1",
            "商品一",
            *([None] * (len(PRODUCT_MASTER_TEMPLATE_HEADERS) - 2)),
        ]],
        sku_rows=[
            [
                "PRODUCT-1",
                f"SKU-{index}",
                *([None] * (len(SKU_DETAIL_TEMPLATE_HEADERS) - 2)),
            ]
            for index in range(1, 4)
        ],
    )

    with pytest.raises(
        ProductTemplateValidationError,
        match="SKU 工作表最多允许 2 条有效 SKU 数据",
    ) as captured:
        parse_product_template(path)

    assert captured.value.issues[0].code == "ROW_LIMIT_EXCEEDED"
    assert captured.value.issues[0].suggestion == (
        "请保留不超过 2 条有效数据；更多数据可拆分为多个文件分批导入。"
    )


def test_product_sku_template_groups_multiple_skus_under_one_product(
    tmp_path: Path,
) -> None:
    path = tmp_path / "商品导入模板.xlsx"
    _write_product_sku_workbook(
        path,
        product_rows=[[
            "PET-BOWL",
            "可调节宠物碗",
            "宠物用品/食具",
            "BOWL-2026",
            18,
            "高度可调节，适合猫犬。",
            "支持定制包装",
            "新品, 防滑",
            "https://img.example.com/pet-bowl.jpg",
            *([None] * 9),
        ]],
        sku_rows=[[
            "PET-BOWL",
            "PET-BOWL",
            "宠物碗",
            "尺寸",
            "小号",
            "中号",
            *([None] * 3),
            "颜色",
            "红色",
            "绿色",
            *([None] * 3),
            "材质",
            "不锈钢",
            *([None] * 4),
            "宠物用品供应商",
            20,
            "1.25",
            12,
            24,
        ]],
    )

    result = parse_product_template(path)

    assert len(result.rows) == 4
    assert {row.product_key for row in result.rows} == {"PRODUCT:PET-BOWL"}
    assert {row.direct_product_code for row in result.rows} == {"PET-BOWL"}
    assert len({row.sku_code for row in result.rows}) == 4
    assert all(row.sku_code.startswith("PET-BOWL-") for row in result.rows)
    assert {row.specification for row in result.rows} == {
        "小号 / 红色 / 不锈钢",
        "小号 / 绿色 / 不锈钢",
        "中号 / 红色 / 不锈钢",
        "中号 / 绿色 / 不锈钢",
    }
    assert {row.sku_name for row in result.rows} == {
        f"宠物碗 · {specification}"
        for specification in (
            "小号 / 红色 / 不锈钢",
            "小号 / 绿色 / 不锈钢",
            "中号 / 红色 / 不锈钢",
            "中号 / 绿色 / 不锈钢",
        )
    }
    assert {
        row.variant_options for row in result.rows
    } == {
        (("尺寸", size), ("颜色", color), ("材质", "不锈钢"))
        for size in ("小号", "中号")
        for color in ("红色", "绿色")
    }
    assert all(row.unit_price == Decimal("20.00") for row in result.rows)
    assert all(row.gross_weight == Decimal("1.250000") for row in result.rows)
    assert all(row.default_moq == Decimal("12.000000") for row in result.rows)
    assert all(row.supplier_name == "宠物用品供应商" for row in result.rows)
    assert all(row.units_per_carton == "24" for row in result.rows)
    assert result.rows[0].tags == ("新品", "防滑")
    assert result.rows[0].image_urls == (
        "https://img.example.com/pet-bowl.jpg",
    )
    assert result.rows[0].schema_version == 5
    assert any("1 个商品，4 个 SKU" in warning for warning in result.warnings)
    assert result.warnings[-1] == "已将 1 行规格候选值自动组合为具体 SKU。"
    repeated = parse_product_template(path)
    assert [row.sku_code for row in repeated.rows] == [
        row.sku_code for row in result.rows
    ]


def test_product_sku_template_allows_an_empty_sku_sheet(tmp_path: Path) -> None:
    path = tmp_path / "无SKU商品.xlsx"
    _write_product_sku_workbook(
        path,
        product_rows=[[
            "PRODUCT-ONLY-001",
            "暂未维护 SKU 的商品",
            "待完善商品",
            None,
            None,
            "商品主数据应当可以先于 SKU 导入。",
            None,
            None,
            *([None] * 10),
        ]],
        sku_rows=[],
    )

    result = parse_product_template(path)

    assert len(result.rows) == 1
    assert result.rows[0].product_key == "PRODUCT:PRODUCT-ONLY-001"
    assert result.rows[0].product_only is True
    assert result.rows[0].sku_code == "PRODUCT-ONLY-001"
    assert result.warnings == (
        "已识别 Product + SKU 双表模板：1 个商品，0 个 SKU。",
        "Product 表中有 1 个商品没有 SKU，系统将为每个商品创建 1 个无规格基础 SKU，可继续按 SKU 管理。",
    )


def test_product_sku_template_keeps_products_unreferenced_by_partial_sku_sheet(
    tmp_path: Path,
) -> None:
    path = tmp_path / "部分商品无SKU.xlsx"
    _write_product_sku_workbook(
        path,
        product_rows=[
            [
                "PRODUCT-WITH-SKU",
                "已有 SKU 商品",
                *([None] * (len(PRODUCT_MASTER_TEMPLATE_HEADERS) - 2)),
            ],
            [
                "PRODUCT-WITHOUT-SKU",
                "暂无 SKU 商品",
                *([None] * (len(PRODUCT_MASTER_TEMPLATE_HEADERS) - 2)),
            ],
        ],
        sku_rows=[[
            "PRODUCT-WITH-SKU",
            "SKU-001",
            *([None] * (len(SKU_DETAIL_TEMPLATE_HEADERS) - 2)),
        ]],
    )

    result = parse_product_template(path)
    rows_by_product = {row.product_key: row for row in result.rows}

    assert set(rows_by_product) == {
        "PRODUCT:PRODUCT-WITH-SKU",
        "PRODUCT:PRODUCT-WITHOUT-SKU",
    }
    assert rows_by_product["PRODUCT:PRODUCT-WITH-SKU"].product_only is False
    assert rows_by_product["PRODUCT:PRODUCT-WITHOUT-SKU"].product_only is True
    assert result.warnings[-1] == (
        "Product 表中有 1 个商品没有 SKU，系统将为每个商品创建 1 个无规格基础 SKU，可继续按 SKU 管理。"
    )


def test_product_sku_template_is_detected_by_headers_after_sheet_rename(
    tmp_path: Path,
) -> None:
    path = tmp_path / "客户双表.xlsx"
    _write_product_sku_workbook(
        path,
        product_sheet_name="商品主数据",
        sku_sheet_name="规格明细",
        product_rows=[[
            "RENAMED-001",
            "改名工作表商品",
            None,
            None,
            None,
            None,
            None,
            None,
            *([None] * 10),
        ]],
        sku_rows=[[
            "RENAMED-001",
            "RENAMED-SKU-001",
            *([None] * (len(SKU_DETAIL_TEMPLATE_HEADERS) - 2)),
        ]],
    )

    result = parse_product_template(path)

    assert len(result.rows) == 1
    assert result.rows[0].sku_code == "RENAMED-SKU-001"
    assert result.warnings[0] == "已根据列结构将工作表“商品主数据”识别为 Product。"
    assert result.warnings[1] == "已根据列结构将工作表“规格明细”识别为 SKU。"


def test_product_sku_template_without_cached_dimensions_is_imported(
    tmp_path: Path,
) -> None:
    source_path = tmp_path / "有尺寸信息.xlsx"
    path = tmp_path / "缺少尺寸信息.xlsx"
    _write_product_sku_workbook(
        source_path,
        product_rows=[[
            "NO-DIMENSION-PRODUCT",
            "缺少缓存尺寸的商品",
            "测试分类",
            *([None] * (len(PRODUCT_MASTER_TEMPLATE_HEADERS) - 3)),
        ]],
        sku_rows=[[
            "NO-DIMENSION-PRODUCT",
            "NO-DIMENSION-SKU",
            *([None] * (len(SKU_DETAIL_TEMPLATE_HEADERS) - 2)),
        ]],
    )

    dimension_pattern = re.compile(br"<dimension\s+[^>]*/>")
    with ZipFile(source_path, "r") as source_archive, ZipFile(
        path,
        "w",
        compression=ZIP_DEFLATED,
    ) as target_archive:
        for archive_entry in source_archive.infolist():
            content = source_archive.read(archive_entry.filename)
            if archive_entry.filename in {
                "xl/worksheets/sheet1.xml",
                "xl/worksheets/sheet2.xml",
            }:
                content, replacements = dimension_pattern.subn(
                    b"",
                    content,
                    count=1,
                )
                assert replacements == 1
            target_archive.writestr(archive_entry, content)

    unsized_workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        assert unsized_workbook[PRODUCT_MASTER_TEMPLATE_SHEET].max_row is None
        assert unsized_workbook[PRODUCT_MASTER_TEMPLATE_SHEET].max_column is None
        assert unsized_workbook[SKU_DETAIL_TEMPLATE_SHEET].max_row is None
        assert unsized_workbook[SKU_DETAIL_TEMPLATE_SHEET].max_column is None
    finally:
        unsized_workbook.close()

    result = parse_product_template(path)

    assert len(result.rows) == 1
    assert result.rows[0].product_key == "PRODUCT:NO-DIMENSION-PRODUCT"
    assert result.rows[0].sku_code == "NO-DIMENSION-SKU"


def test_previous_product_sku_template_remains_compatible(tmp_path: Path) -> None:
    path = tmp_path / "上一版双表.xlsx"
    workbook = Workbook()
    product_sheet = workbook.active
    product_sheet.title = "旧商品表"
    product_sheet.append(list(PRODUCT_MASTER_TEMPLATE_HEADERS_V3))
    product_sheet.append([
        "LEGACY-PRODUCT",
        "上一版商品",
        "历史分类",
        None,
        8,
        None,
        None,
        "经典",
        "是",
        *([None] * 10),
    ])
    sku_sheet = workbook.create_sheet("旧规格表")
    sku_sheet.append(list(SKU_DETAIL_TEMPLATE_HEADERS_V3))
    sku_sheet.append([
        "LEGACY-PRODUCT",
        "LEGACY-SKU",
        None,
        None,
        "颜色",
        "蓝色",
        *([None] * 4),
        None,
        None,
        24,
    ])
    workbook.save(path)
    workbook.close()

    result = parse_product_template(path)

    assert result.rows[0].schema_version == 3
    assert result.rows[0].units_per_carton == "24"
    assert result.rows[0].tags == ("经典", "新品")
    assert result.rows[0].gross_weight is None
    assert "历史 Product + SKU" in result.warnings[-2]


def test_product_sku_template_rejects_negative_logistics_values(
    tmp_path: Path,
) -> None:
    path = tmp_path / "错误物流数据.xlsx"
    _write_product_sku_workbook(
        path,
        product_rows=[[
            "PRODUCT-NEGATIVE",
            "错误物流商品",
            *([None] * (len(PRODUCT_MASTER_TEMPLATE_HEADERS) - 2)),
        ]],
        sku_rows=[[
            "PRODUCT-NEGATIVE",
            "SKU-NEGATIVE",
            *([None] * 21),
            -1,
            -2,
            -3,
        ]],
    )

    with pytest.raises(ProductTemplateValidationError) as captured:
        parse_product_template(path)

    assert {
        (issue.column, issue.code)
        for issue in captured.value.issues
    } == {
        ("毛重", "QUANTITY_INVALID"),
        ("起定数", "QUANTITY_INVALID"),
        ("装箱数", "QUANTITY_INVALID"),
    }


def test_v4_five_dimension_template_remains_compatible(tmp_path: Path) -> None:
    path = tmp_path / "v4双表.xlsx"
    workbook = Workbook()
    product_sheet = workbook.active
    product_sheet.title = PRODUCT_MASTER_TEMPLATE_SHEET
    product_sheet.append(list(PRODUCT_MASTER_TEMPLATE_HEADERS))
    product_sheet.append([
        "V4-PRODUCT",
        "V4 兼容商品",
        "历史分类",
        None,
        9,
        None,
        None,
        "兼容",
        *([None] * 10),
    ])
    sku_sheet = workbook.create_sheet(SKU_DETAIL_TEMPLATE_SHEET)
    sku_sheet.append(list(SKU_DETAIL_TEMPLATE_HEADERS_V4))
    sku_sheet.append([
        "V4-PRODUCT",
        "V4-SKU",
        None,
        None,
        "尺寸",
        "小号",
        *([None] * 8),
        None,
        None,
        1.2,
        6,
        24,
    ])
    workbook.save(path)
    workbook.close()

    result = parse_product_template(path)

    assert result.rows[0].schema_version == 4
    assert result.rows[0].sku_code == "V4-SKU"
    assert result.rows[0].variant_options == (("尺寸", "小号"),)
    assert result.rows[0].gross_weight == Decimal("1.200000")
    assert result.rows[0].default_moq == Decimal("6.000000")
    assert result.rows[0].units_per_carton == "24"
    assert any("历史 Product + SKU" in warning for warning in result.warnings)


def test_product_sku_template_rejects_missing_product_reference(
    tmp_path: Path,
) -> None:
    path = tmp_path / "错误关联.xlsx"
    _write_product_sku_workbook(
        path,
        product_rows=[[
            "PRODUCT-001",
            "已有商品",
            None,
            None,
            None,
            None,
            None,
            None,
            *([None] * 10),
        ]],
        sku_rows=[[
            "NOT-FOUND",
            "SKU-001",
            *([None] * (len(SKU_DETAIL_TEMPLATE_HEADERS) - 2)),
        ]],
    )

    with pytest.raises(ProductTemplateValidationError) as captured:
        parse_product_template(path)

    assert captured.value.issues[0].code == "PRODUCT_REFERENCE_MISSING"
    assert captured.value.issues[0].column == "商品编码"


def test_fixed_template_keeps_note_without_creating_a_moq(tmp_path: Path) -> None:
    path = tmp_path / "商品模版.xlsx"
    _write_workbook(
        path,
        [[
            "柔雾唇釉",
            "唇彩",
            "SKU-001",
            " 青湾 供应链 ",
            "20",
            "轻盈柔雾质地",
            "12",
            "新品，热卖, 新品",
            "https://img.example.com/sku-001.jpg",
            *([None] * 9),
        ]],
    )

    result = parse_product_template(path)

    assert len(result.rows) == 1
    assert result.rows[0].sku_code == "SKU-001"
    assert result.rows[0].supplier_name == "青湾 供应链"
    assert str(result.rows[0].unit_price) == "20.00"
    assert result.rows[0].note == "12"
    assert result.rows[0].tags == ("新品", "热卖")
    assert result.rows[0].default_moq is None
    assert result.rows[0].image_urls == ("https://img.example.com/sku-001.jpg",)
    assert result.rows[0].image_url_columns == (1,)
    assert result.rows[0].embedded_images == ()
    assert result.warnings == ()


def test_embedded_images_are_mapped_to_product_rows_and_image_columns(
    tmp_path: Path,
) -> None:
    path = tmp_path / "内嵌图片商品.xlsx"
    image_path = tmp_path / "product.png"
    image_bytes = (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
        b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00"
        b"\x1f\x15\xc4\x89\x00\x00\x00\rIDAT\x08\xd7c\xf8\xcf\xc0"
        b"\xf0\x1f\x00\x05\x00\x01\xff\x89\x99=\x1d\x00\x00\x00"
        b"\x00IEND\xaeB`\x82"
    )
    image_path.write_bytes(image_bytes)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = PRODUCT_TEMPLATE_SHEET
    sheet.append(list(PRODUCT_TEMPLATE_HEADERS))
    sheet.append([
        "内嵌图片商品",
        "配件",
        "EMBEDDED-001",
        None,
        None,
        "包含两张真实图片",
        None,
        None,
        *([None] * 10),
    ])
    first = OpenpyxlImage(image_path)
    first.anchor = "I2"
    sheet.add_image(first)
    second = OpenpyxlImage(image_path)
    second.anchor = "J2"
    sheet.add_image(second)
    workbook.save(path)
    workbook.close()

    result = parse_product_template(path)

    assert len(result.rows) == 1
    embedded = result.rows[0].embedded_images
    assert len(embedded) == 2
    assert [image.image_column for image in embedded] == [1, 2]
    assert all(image.row_number == 2 for image in embedded)
    assert all(image.content_type == "image/png" for image in embedded)
    assert all(image.byte_size == len(image_bytes) for image in embedded)
    assert all(
        image.sha256 == hashlib.sha256(image_bytes).hexdigest()
        for image in embedded
    )


def test_product_sku_template_maps_embedded_product_images_to_each_sku(
    tmp_path: Path,
) -> None:
    path = tmp_path / "双表内嵌图片.xlsx"
    image_path = tmp_path / "product-cover.png"
    image_bytes = (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
        b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00"
        b"\x1f\x15\xc4\x89\x00\x00\x00\rIDAT\x08\xd7c\xf8\xcf\xc0"
        b"\xf0\x1f\x00\x05\x00\x01\xff\x89\x99=\x1d\x00\x00\x00"
        b"\x00IEND\xaeB`\x82"
    )
    image_path.write_bytes(image_bytes)
    workbook = Workbook()
    product_sheet = workbook.active
    product_sheet.title = PRODUCT_MASTER_TEMPLATE_SHEET
    product_sheet.append(list(PRODUCT_MASTER_TEMPLATE_HEADERS))
    product_sheet.append([
        "PRODUCT-WITH-IMAGE",
        "双表内嵌图片商品",
        "配件/测试",
        None,
        12,
        "图片只在 Product 表保存一次",
        None,
        None,
        *([None] * 10),
    ])
    image = OpenpyxlImage(image_path)
    image.anchor = "I2"
    product_sheet.add_image(image)
    sku_sheet = workbook.create_sheet(SKU_DETAIL_TEMPLATE_SHEET)
    sku_sheet.append(list(SKU_DETAIL_TEMPLATE_HEADERS))
    for sku_code, sku_name in (
        ("SKU-WITH-IMAGE-A", "内嵌图片商品 A"),
        ("SKU-WITH-IMAGE-B", "内嵌图片商品 B"),
    ):
        sku_sheet.append([
            "PRODUCT-WITH-IMAGE",
            sku_code,
            sku_name,
            *([None] * (len(SKU_DETAIL_TEMPLATE_HEADERS) - 3)),
        ])
    workbook.save(path)
    workbook.close()

    result = parse_product_template(path)

    assert [row.sku_code for row in result.rows] == [
        "SKU-WITH-IMAGE-A",
        "SKU-WITH-IMAGE-B",
    ]
    assert all(len(row.embedded_images) == 1 for row in result.rows)
    assert {
        row.embedded_images[0].sha256 for row in result.rows
    } == {hashlib.sha256(image_bytes).hexdigest()}
    assert {
        row.embedded_images[0].archive_path for row in result.rows
    } == {"xl/media/image1.png"}


def test_embedded_image_storage_uploads_without_head_requests(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source_path = tmp_path / "images.xlsx"
    first_bytes = b"first-image"
    second_bytes = b"second-image"
    with ZipFile(source_path, "w", ZIP_DEFLATED) as archive:
        archive.writestr("xl/media/image1.png", first_bytes)
        archive.writestr("xl/media/image2.jpg", second_bytes)

    uploaded: dict[str, bytes] = {}

    class FakeStorage:
        backend_name = "s3"

        def put_file(
            self,
            source: Path,
            *,
            object_key: str,
            content_type: str | None,
        ) -> None:
            del content_type
            uploaded[object_key] = source.read_bytes()

        def exists(self, object_key: str) -> bool:
            raise AssertionError(f"unexpected HEAD request for {object_key}")

    monkeypatch.setattr(
        product_template_import_service,
        "get_object_storage",
        lambda: FakeStorage(),
    )
    monkeypatch.setenv("PRODUCT_TEMPLATE_IMAGE_UPLOAD_CONCURRENCY", "2")
    specs = (
        product_template_import_service.StoredTemplateImage(
            image_column=1,
            sequence=1,
            object_key="images/first.png",
            original_filename="image1.png",
            content_type="image/png",
            byte_size=len(first_bytes),
            sha256=hashlib.sha256(first_bytes).hexdigest(),
            storage_provider="S3",
            archive_path="xl/media/image1.png",
        ),
        product_template_import_service.StoredTemplateImage(
            image_column=2,
            sequence=2,
            object_key="images/second.jpg",
            original_filename="image2.jpg",
            content_type="image/jpeg",
            byte_size=len(second_bytes),
            sha256=hashlib.sha256(second_bytes).hexdigest(),
            storage_provider="S3",
            archive_path="xl/media/image2.jpg",
        ),
    )
    progress: list[tuple[int, int]] = []

    product_template_import_service._store_new_embedded_images(
        source_path,
        specs=specs,
        existing_object_keys=set(),
        progress_callback=lambda processed, total: progress.append(
            (processed, total)
        ),
    )

    assert uploaded == {
        "images/first.png": first_bytes,
        "images/second.jpg": second_bytes,
    }
    assert progress[-1] == (2, 2)


def test_alternate_sheet_name_is_auto_detected_without_grouping_skus(
    tmp_path: Path,
) -> None:
    path = tmp_path / "供应商原始图册.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(list(PRODUCT_TEMPLATE_HEADERS))
    sheet.append([
        "同名商品",
        "服饰/套装",
        "STYLE-001",
        None,
        None,
        None,
        None,
        None,
        *([None] * 10),
    ])
    sheet.append([
        "同名商品",
        "服饰/套装",
        "STYLE-002",
        None,
        None,
        None,
        None,
        None,
        *([None] * 10),
    ])
    workbook.create_sheet("Sheet2")
    workbook.save(path)
    workbook.close()

    result = parse_product_template(path)

    assert [row.sku_code for row in result.rows] == ["STYLE-001", "STYLE-002"]
    assert [row.name for row in result.rows] == ["同名商品", "同名商品"]
    assert result.warnings == (
        "已根据列结构识别工作表“Sheet1”；每一行仍按一个 SKU 导入。",
    )


def test_product_sheet_selection_does_not_prefer_a_reserved_title(
    tmp_path: Path,
) -> None:
    path = tmp_path / "两个商品页.xlsx"
    workbook = Workbook()
    named_sheet = workbook.active
    named_sheet.title = PRODUCT_TEMPLATE_SHEET
    named_sheet.append(list(PRODUCT_TEMPLATE_HEADERS))
    named_sheet.append([
        "命名页商品",
        "分类",
        "NAMED-001",
        None,
        None,
        None,
        None,
        None,
        *([None] * 10),
    ])
    alternate_sheet = workbook.create_sheet("客户数据")
    alternate_sheet.append(list(PRODUCT_TEMPLATE_HEADERS))
    alternate_sheet.append([
        "客户页商品",
        "分类",
        "CUSTOM-001",
        None,
        None,
        None,
        None,
        None,
        *([None] * 10),
    ])
    workbook.save(path)
    workbook.close()

    with pytest.raises(ProductTemplateValidationError, match="多个符合商品列结构") as caught:
        parse_product_template(path)

    assert caught.value.issues[0].code == "SHEET_AMBIGUOUS"
    assert "工作表名称不限" in (caught.value.issues[0].suggestion or "")


def test_product_sheet_selection_ignores_an_empty_compatible_copy(
    tmp_path: Path,
) -> None:
    path = tmp_path / "带空模板副本.xlsx"
    workbook = Workbook()
    empty_sheet = workbook.active
    empty_sheet.title = PRODUCT_TEMPLATE_SHEET
    empty_sheet.append(list(PRODUCT_TEMPLATE_HEADERS))
    data_sheet = workbook.create_sheet("本次商品")
    data_sheet.append(list(PRODUCT_TEMPLATE_HEADERS))
    data_sheet.append([
        "客户页商品",
        "分类",
        "CUSTOM-002",
        None,
        None,
        None,
        None,
        None,
        *([None] * 10),
    ])
    workbook.save(path)
    workbook.close()

    result = parse_product_template(path)

    assert [row.sku_code for row in result.rows] == ["CUSTOM-002"]
    assert result.warnings == (
        "已根据列结构识别工作表“本次商品”；每一行仍按一个 SKU 导入。",
    )


def test_product_variant_template_groups_rows_and_generates_stable_skus(
    tmp_path: Path,
) -> None:
    path = tmp_path / "商品图册模板（更新了商品规格分类）.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(list(PRODUCT_VARIANT_TEMPLATE_HEADERS))
    sheet.append([
        "双目云台摄像机",
        "数码/摄像头",
        "CAMERA-0389",
        None,
        "室外双目联动摄像机",
        "支持定制包装",
        "是",
        12,
        "普通款",
        10,
        None,
    ])
    sheet.append([
        "双目云台摄像机",
        "数码/摄像头",
        "CAMERA-0389",
        None,
        "室外双目联动摄像机",
        None,
        "否",
        12,
        "蓝牙款",
        20,
        None,
    ])
    sheet.append([
        "单规格摄像头",
        "数码",
        "CAMERA-3333",
        600,
        "室内摄像头",
        None,
        None,
        24,
        None,
        None,
        None,
    ])
    workbook.save(path)
    workbook.close()

    first = parse_product_template(path)
    repeated = parse_product_template(path)

    assert len(first.rows) == 3
    assert first.rows[0].product_key == first.rows[1].product_key
    assert first.rows[0].product_key != first.rows[2].product_key
    assert first.rows[0].sku_code != first.rows[1].sku_code
    assert [row.sku_code for row in first.rows] == [
        row.sku_code for row in repeated.rows
    ]
    assert [row.specification for row in first.rows] == [
        "普通款",
        "蓝牙款",
        None,
    ]
    assert [row.sku_name for row in first.rows] == [
        "双目云台摄像机 · 普通款",
        "双目云台摄像机 · 蓝牙款",
        "单规格摄像头",
    ]
    assert [row.unit_price for row in first.rows] == [
        Decimal("10.00"),
        Decimal("20.00"),
        Decimal("600.00"),
    ]
    assert first.rows[0].tags == ("新品",)
    assert first.rows[0].units_per_carton == "12"
    assert first.rows[2].sku_code == "CAMERA-3333"
    assert first.warnings[0].startswith("已根据列结构识别工作表")
    assert "商品+规格模板" in first.warnings[1]


def test_product_variant_template_reads_embedded_image_from_column_k(
    tmp_path: Path,
) -> None:
    path = tmp_path / "多规格内嵌图片.xlsx"
    image_path = tmp_path / "variant.png"
    image_bytes = (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
        b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00"
        b"\x1f\x15\xc4\x89\x00\x00\x00\rIDAT\x08\xd7c\xf8\xcf\xc0"
        b"\xf0\x1f\x00\x05\x00\x01\xff\x89\x99=\x1d\x00\x00\x00"
        b"\x00IEND\xaeB`\x82"
    )
    image_path.write_bytes(image_bytes)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(list(PRODUCT_VARIANT_TEMPLATE_HEADERS))
    sheet.append([
        "带图商品",
        "配件",
        "IMAGE-PRODUCT",
        88,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    ])
    image = OpenpyxlImage(image_path)
    image.anchor = "K2"
    sheet.add_image(image)
    workbook.save(path)
    workbook.close()

    result = parse_product_template(path)

    assert len(result.rows) == 1
    assert len(result.rows[0].embedded_images) == 1
    assert result.rows[0].embedded_images[0].image_column == 1
    assert result.rows[0].embedded_images[0].sha256 == hashlib.sha256(
        image_bytes
    ).hexdigest()


def test_product_variant_template_requires_unique_specs_for_multi_row_product(
    tmp_path: Path,
) -> None:
    path = tmp_path / "重复规格.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(list(PRODUCT_VARIANT_TEMPLATE_HEADERS))
    for price in (10, 20):
        sheet.append([
            "重复规格商品",
            "数码",
            "DUP-SPEC",
            None,
            None,
            None,
            None,
            None,
            "标准款",
            price,
            None,
        ])
    workbook.save(path)
    workbook.close()

    with pytest.raises(
        ProductTemplateValidationError,
        match="规格名称|重复",
    ):
        parse_product_template(path)


def test_sequential_image_columns_can_extend_beyond_ten(
    tmp_path: Path,
) -> None:
    path = tmp_path / "扩展图片列.xlsx"
    image_path = tmp_path / "extended-product.png"
    image_bytes = (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
        b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00"
        b"\x1f\x15\xc4\x89\x00\x00\x00\rIDAT\x08\xd7c\xf8\xcf\xc0"
        b"\xf0\x1f\x00\x05\x00\x01\xff\x89\x99=\x1d\x00\x00\x00"
        b"\x00IEND\xaeB`\x82"
    )
    image_path.write_bytes(image_bytes)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    extended_headers = [
        *PRODUCT_TEMPLATE_HEADERS,
        *(f"商品图片{index}" for index in range(11, 19)),
    ]
    sheet.append(extended_headers)
    sheet.append([
        "十八图商品",
        "服饰/套装",
        "IMAGES-018",
        None,
        None,
        None,
        None,
        None,
        *([None] * 17),
        "https://img.example.com/image-18.jpg",
    ])
    embedded = OpenpyxlImage(image_path)
    embedded.anchor = "Y2"
    sheet.add_image(embedded)
    workbook.save(path)
    workbook.close()

    result = parse_product_template(path)

    assert len(result.rows) == 1
    assert result.rows[0].image_urls == (
        "https://img.example.com/image-18.jpg",
    )
    assert result.rows[0].image_url_columns == (18,)
    assert [image.image_column for image in result.rows[0].embedded_images] == [17]
    assert result.warnings == (
        "已根据列结构识别工作表“Sheet1”；每一行仍按一个 SKU 导入。",
    )


def test_product_sku_image_columns_can_extend_beyond_ten(
    tmp_path: Path,
) -> None:
    path = tmp_path / "双表扩展图片列.xlsx"
    image_path = tmp_path / "dual-extended-product.png"
    image_bytes = (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
        b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00"
        b"\x1f\x15\xc4\x89\x00\x00\x00\rIDAT\x08\xd7c\xf8\xcf\xc0"
        b"\xf0\x1f\x00\x05\x00\x01\xff\x89\x99=\x1d\x00\x00\x00"
        b"\x00IEND\xaeB`\x82"
    )
    image_path.write_bytes(image_bytes)
    workbook = Workbook()
    product_sheet = workbook.active
    product_sheet.title = PRODUCT_MASTER_TEMPLATE_SHEET
    product_headers = PRODUCT_MASTER_TEMPLATE_HEADERS[:8] + tuple(
        f"商品图片{index}" for index in range(1, 19)
    )
    product_sheet.append(list(product_headers))
    product_sheet.append([
        "DUAL-EXTENDED-PRODUCT",
        "双表扩展图片商品",
        "测试分类",
        None,
        10,
        None,
        None,
        None,
        *([None] * 17),
        "https://img.example.com/image-18.jpg",
    ])
    embedded = OpenpyxlImage(image_path)
    embedded.anchor = "Z2"  # Product column 26 = 商品图片18
    product_sheet.add_image(embedded)
    sku_sheet = workbook.create_sheet(SKU_DETAIL_TEMPLATE_SHEET)
    sku_sheet.append(list(SKU_DETAIL_TEMPLATE_HEADERS))
    sku_sheet.append([
        "DUAL-EXTENDED-PRODUCT",
        "DUAL-EXTENDED-SKU",
        "双表扩展图片 SKU",
        *([None] * (len(SKU_DETAIL_TEMPLATE_HEADERS) - 3)),
    ])
    workbook.save(path)
    workbook.close()

    result = parse_product_template(path)

    assert result.rows[0].image_urls == (
        "https://img.example.com/image-18.jpg",
    )
    assert result.rows[0].image_url_columns == (18,)
    assert [image.image_column for image in result.rows[0].embedded_images] == [18]


def test_duplicate_sku_is_case_and_whitespace_insensitive(tmp_path: Path) -> None:
    path = tmp_path / "商品模版.xlsx"
    _write_workbook(
        path,
        [
            ["商品 A", "分类", " sku-dup ", None, 10, None, 1, None, *([None] * 10)],
            ["商品 B", "分类", "SkU-DuP", None, 12, None, 1, None, *([None] * 10)],
        ],
    )

    result = parse_product_template(path)

    assert [row.name for row in result.rows] == ["商品 A"]
    assert [row.sku_code for row in result.rows] == ["SKU-DUP"]
    assert result.skipped_rows == 1
    assert "第 3 行" in result.warnings[0]
    assert "第 2 行" in result.warnings[0]


def test_category_path_normalizes_two_levels(tmp_path: Path) -> None:
    path = tmp_path / "商品模版.xlsx"
    _write_workbook(
        path,
        [["商品 A", " 办公用品 ／ 纸品 ", "SKU-CATEGORY", None, 10, None, None, None, *([None] * 10)]],
    )

    result = parse_product_template(path)

    assert result.rows[0].category == "办公用品/纸品"


@pytest.mark.parametrize("category", ["办公用品/纸品/A4", "/纸品", "办公用品/"])
def test_category_path_rejects_empty_or_third_level(
    tmp_path: Path, category: str
) -> None:
    path = tmp_path / "商品模版.xlsx"
    _write_workbook(
        path,
        [["商品 A", category, "SKU-BAD-CATEGORY", None, 10, None, None, None, *([None] * 10)]],
    )

    with pytest.raises(ProductTemplateValidationError, match="商品分类|两级"):
        parse_product_template(path)


def test_price_is_half_up_quantized_to_two_decimal_places(tmp_path: Path) -> None:
    path = tmp_path / "商品模版.xlsx"
    _write_workbook(
        path,
        [
            ["商品 A", "分类", "price-a", None, "12.344", None, None, None, *([None] * 10)],
            ["商品 B", "分类", "price-b", None, "12.345", None, None, None, *([None] * 10)],
        ],
    )

    result = parse_product_template(path)

    assert [row.sku_code for row in result.rows] == ["PRICE-A", "PRICE-B"]
    assert [str(row.unit_price) for row in result.rows] == ["12.34", "12.35"]
    assert result.warnings == ()


def test_optional_columns_can_be_blank_and_blank_price_defaults_to_zero(
    tmp_path: Path,
) -> None:
    path = tmp_path / "商品模版.xlsx"
    _write_workbook(
        path,
        [[
            "无选填信息商品",
            "基础分类",
            "OPTIONAL-BLANK",
            None,
            None,
            None,
            None,
            None,
            *([None] * 10),
        ]],
    )

    result = parse_product_template(path)

    assert len(result.rows) == 1
    assert result.rows[0].supplier_name is None
    assert result.rows[0].unit_price == Decimal("0.00")
    assert result.rows[0].description is None
    assert result.rows[0].note is None
    assert result.rows[0].tags == ()
    assert result.rows[0].image_urls == ()


def test_blank_category_defaults_to_uncategorized(tmp_path: Path) -> None:
    path = tmp_path / "商品模版.xlsx"
    _write_workbook(
        path,
        [[
            "待分类商品",
            None,
            "UNCATEGORIZED-001",
            None,
            None,
            None,
            None,
            None,
            *([None] * 10),
        ]],
    )

    result = parse_product_template(path)

    assert len(result.rows) == 1
    assert result.rows[0].category == "未分类"
    assert result.rows[0].unit_price == Decimal("0.00")


def test_blank_model_generates_a_stable_temporary_sku(tmp_path: Path) -> None:
    path = tmp_path / "商品模版.xlsx"
    _write_workbook(
        path,
        [
            [
                "无型号商品",
                "基础分类",
                None,
                "测试供应商",
                "10",
                "第一件",
                None,
                None,
                *([None] * 10),
            ],
            [
                "无型号商品",
                "基础分类",
                None,
                "测试供应商",
                "20",
                "第二件",
                None,
                None,
                *([None] * 10),
            ],
        ],
    )

    first = parse_product_template(path)
    repeated = parse_product_template(path)

    assert len(first.rows) == 2
    assert first.rows[0].sku_code.startswith("AUTO-")
    assert first.rows[1].sku_code == f"{first.rows[0].sku_code}-2"
    assert [row.sku_code for row in repeated.rows] == [
        row.sku_code for row in first.rows
    ]
    assert first.skipped_rows == 0
    assert first.warnings == (
        "有 2 行未填写商品型号，系统已根据商品名称、分类和供应商生成临时型号。",
    )


def test_parser_reports_determinate_row_progress(tmp_path: Path) -> None:
    path = tmp_path / "商品模版.xlsx"
    _write_workbook(
        path,
        [
            [
                f"进度商品 {index}",
                "进度分类",
                f"PROGRESS-{index:03d}",
                None,
                None,
                None,
                None,
                None,
                *([None] * 10),
            ]
            for index in range(1, 251)
        ],
    )
    progress: list[tuple[int, int]] = []

    result = parse_product_template(
        path,
        progress_callback=lambda processed, total: progress.append((processed, total)),
    )

    assert len(result.rows) == 250
    assert progress[0] == (1, 250)
    assert progress[-1] == (250, 250)
    assert all(total == 250 for _, total in progress)
    assert [processed for processed, _ in progress] == sorted(
        processed for processed, _ in progress
    )


def test_price_respects_numeric_20_2_boundary(tmp_path: Path) -> None:
    path = tmp_path / "商品模版.xlsx"
    _write_workbook(
        path,
        [
            [
                "最大价格",
                "分类",
                "MAX-PRICE",
                None,
                "999999999999999999.99",
                None,
                None,
                None,
                *([None] * 10),
            ],
            [
                "溢出价格",
                "分类",
                "OVERFLOW-PRICE",
                None,
                "999999999999999999.995",
                None,
                None,
                None,
                *([None] * 10),
            ],
        ],
    )

    with pytest.raises(ProductTemplateValidationError, match=r"Numeric\(20,2\)"):
        parse_product_template(path)


@pytest.mark.parametrize(
    "row, message",
    [
        (
            ["", "分类", "INVALID-NAME", None, 10, None, None, None, *([None] * 10)],
            "缺少商品名称",
        ),
        (
            ["商品", "分类", "INVALID-PRICE", None, "not-a-price", None, None, None, *([None] * 10)],
            "商品价格不是有效数字",
        ),
        (
            ["商品", "分类", "INVALID-IMAGE", None, 10, None, None, None, "javascript:alert(1)", *([None] * 9)],
            "商品图片1不是有效",
        ),
        (
            ["商品", "分类", "FORMULA", None, "=1+1", None, None, None, *([None] * 10)],
            "包含公式",
        ),
    ],
)
def test_invalid_row_rejects_the_full_snapshot(
    tmp_path: Path,
    row: list[object],
    message: str,
) -> None:
    path = tmp_path / "商品模版.xlsx"
    _write_workbook(path, [row])

    with pytest.raises(ProductTemplateValidationError, match=message):
        parse_product_template(path)


def test_validation_collects_complete_row_and_field_details(tmp_path: Path) -> None:
    path = tmp_path / "商品模版.xlsx"
    _write_workbook(
        path,
        [
            [
                "",
                "分类",
                "INVALID-MULTI-A",
                None,
                "not-a-price",
                None,
                None,
                None,
                "javascript:alert(1)",
                *([None] * 9),
            ],
            [
                "商品 B",
                "一级/二级/三级",
                "INVALID-MULTI-B",
                None,
                None,
                None,
                None,
                "x" * 81,
                *([None] * 10),
            ],
        ],
    )

    with pytest.raises(ProductTemplateValidationError) as captured:
        parse_product_template(path)

    issues = captured.value.issues
    assert len(issues) == 5
    assert {issue.row_number for issue in issues} == {2, 3}
    assert {
        (issue.row_number, issue.column, issue.code)
        for issue in issues
    } == {
        (2, "商品名称", "REQUIRED_VALUE_MISSING"),
        (2, "商品价格", "PRICE_INVALID"),
        (2, "商品图片1", "IMAGE_URL_INVALID"),
        (3, "商品分类", "CATEGORY_INVALID"),
        (3, "标签", "TAGS_INVALID"),
    }
    assert all(issue.suggestion for issue in issues)
    assert "发现 5 个数据问题" in str(captured.value)


def test_wrong_headers_are_rejected_before_rows_are_read(tmp_path: Path) -> None:
    path = tmp_path / "商品模版.xlsx"
    wrong_headers = list(PRODUCT_TEMPLATE_HEADERS)
    wrong_headers[2] = "SKU"
    _write_workbook(path, [], headers=wrong_headers)

    with pytest.raises(ProductTemplateValidationError, match="表头") as captured:
        parse_product_template(path)

    assert captured.value.issues[0].row_number == 1
    assert captured.value.issues[0].column == "商品型号"
    assert captured.value.issues[0].code == "HEADER_MISMATCH"


def test_current_root_template_matches_contract() -> None:
    template_path = Path(__file__).resolve().parents[3] / "商品模版.xlsx"
    if not template_path.exists():
        pytest.skip("root product template is not present")

    result = parse_product_template(template_path)

    assert len(result.rows) == 600
    assert result.skipped_rows == 8
    assert sum(row.unit_price == Decimal("0.00") for row in result.rows) == 103
    assert len(result.warnings) == 8
    assert sum("重复" in warning for warning in result.warnings) == 8


def test_catalog_export_can_be_imported_without_creating_new_identity(
    tmp_path: Path,
) -> None:
    product_id = uuid4()
    sku_id = uuid4()
    workbook = Workbook()
    product_sheet = workbook.active
    product_sheet.title = "商品"
    sku_sheet = workbook.create_sheet("SKU")
    product_sheet.append(list(SKU_CATALOG_EXPORT_PRODUCT_HEADERS))
    sku_sheet.append(list(SKU_CATALOG_EXPORT_SKU_HEADERS))
    product_row = [
        str(product_id),
        "P-EXPORT-1",
        "可回导商品",
        "家纺/床品",
        "导出描述",
        "piece",
        "ACTIVE",
        *([None] * 50),
        None,
    ]
    product_sheet.append(product_row)
    sku_sheet.append(
        [
            str(sku_id),
            str(product_id),
            "P-EXPORT-1",
            "SYSTEM-1",
            "SOURCE-1",
            "可回导 SKU",
            "商品型号: MODEL-1；颜色: 红；规格名称: 红",
            "",
            "供应商A",
            12.5,
            "CNY",
            "新品，热卖",
            2,
            "piece",
            1.2,
            "kg",
            10,
            "ACTIVE",
            "catalog.xlsx",
            None,
            None,
        ]
    )
    path = tmp_path / "catalog-export.xlsx"
    workbook.save(path)
    workbook.close()

    result = parse_product_template(path)

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row.existing_product_id == product_id
    assert row.existing_sku_id == sku_id
    assert row.direct_product_code == "P-EXPORT-1"
    assert row.sku_code == "SOURCE-1"
    assert row.variant_options == (("颜色", "红"),)
    assert row.product_model == "MODEL-1"
    assert row.specification == "红"
    assert row.default_unit == "piece"
    assert row.default_moq == Decimal("2.000000")
    assert row.gross_weight == Decimal("1.200000")


def test_catalog_export_round_trip_does_not_repeat_sku_specification(
    tmp_path: Path,
) -> None:
    path = tmp_path / "catalog-round-trip.xlsx"
    _write_product_sku_workbook(
        path,
        product_rows=[[
            "P-ROUND-TRIP",
            "软冰宠物冰垫",
            "宠物用品",
            *([None] * (len(PRODUCT_MASTER_TEMPLATE_HEADERS) - 3)),
        ]],
        sku_rows=[[
            "P-ROUND-TRIP",
            "SOURCE-ROUND-TRIP",
            "软冰宠物冰垫 · 灰色 / L · 灰色 / L · 灰色 / L",
            "颜色",
            "灰色",
            *([None] * 4),
            "尺寸",
            "L",
            *([None] * 4),
            *([None] * (len(SKU_DETAIL_TEMPLATE_HEADERS) - 15)),
        ]],
    )

    result = parse_product_template(path)

    assert len(result.rows) == 1
    assert result.rows[0].sku_name == "软冰宠物冰垫 · 灰色 / L"


def test_catalog_export_ignores_malformed_logistics_annotation_when_column_empty(
    tmp_path: Path,
) -> None:
    """Free-form export text must not replace an empty logistics column."""

    product_id = uuid4()
    sku_id = uuid4()
    workbook = Workbook()
    product_sheet = workbook.active
    product_sheet.title = "商品"
    sku_sheet = workbook.create_sheet("SKU")
    product_sheet.append(list(SKU_CATALOG_EXPORT_PRODUCT_HEADERS))
    sku_sheet.append(list(SKU_CATALOG_EXPORT_SKU_HEADERS))
    product_sheet.append(
        [
            str(product_id),
            "P-EXPORT-ANNOTATION",
            "带物流描述的导出商品",
            "家纺",
            None,
            "piece",
            "ACTIVE",
            *([None] * 50),
            None,
        ]
    )
    sku_sheet.append(
        [
            str(sku_id),
            str(product_id),
            "P-EXPORT-ANNOTATION",
            "SYSTEM-ANNOTATION",
            "SOURCE-ANNOTATION",
            "带物流描述的 SKU",
            "装箱数：24.单个含包装重量：0.283kg",
            None,
            None,
            10,
            "CNY",
            None,
            None,
            "piece",
            None,
            "kg",
            None,
            "ACTIVE",
            "现货产品报价单.xlsx",
            None,
            None,
        ]
    )
    path = tmp_path / "现货产品报价单.xlsx"
    workbook.save(path)
    workbook.close()

    result = parse_product_template(path)

    assert len(result.rows) == 1
    assert result.rows[0].units_per_carton is None
    assert result.rows[0].gross_weight is None
