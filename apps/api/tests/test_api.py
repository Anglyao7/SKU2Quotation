import atexit
import hashlib
import io
import os
import re
import shutil
import tempfile
from concurrent.futures import ThreadPoolExecutor
from dataclasses import replace
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from threading import Event, Lock
from time import monotonic, sleep
from types import SimpleNamespace
from uuid import UUID, uuid4

import pytest
from alembic import command
from alembic.config import Config
from fastapi import Request, Response
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill
from PIL import Image
from sqlalchemy import MetaData, create_engine, delete, func, inspect, select
from sqlalchemy.dialects import postgresql
from sqlalchemy.exc import IntegrityError


TEST_RUNTIME = Path(tempfile.mkdtemp(prefix="mercator-tests-"))
os.environ["DATABASE_URL"] = f"sqlite:///{(TEST_RUNTIME / 'test.db').as_posix()}"
os.environ["UPLOAD_DIR"] = str(TEST_RUNTIME / "uploads")
os.environ["APP_ENV"] = "test"
os.environ["AUTH_PROFILE"] = "local_fake"
os.environ["AUTH_TEST_BYPASS"] = "true"

from app.main import app
from app.domain.errors import ApplicationError
from app.routers.auth import _set_refresh_cookie
from app.database import API_ROOT, SessionLocal, engine
from app.services.import_progress import (
    clear_runtime_import_progress,
    publish_runtime_import_progress,
)
from app.services.repository import import_job_model
from app.identity_models import (
    AuthRefreshTokenRow,
    AuthSessionRow,
    LocalAccountCredentialRow,
    MembershipRoleRow,
    MembershipRow,
    OrganizationRow,
    PermissionRow,
    RolePermissionRow,
    RoleRow,
    TenantRow,
    UserRow,
)
from app.inventory_models import WarehouseRow
from app.ai_data_models import AIProviderRouteRow, AISourceEvidenceRow, AITaskRow
from app.knowledge_embedding_models import EmbeddingRow, KnowledgeChunkRow, KnowledgeDocumentRow
from app.embedding_management_models import (
    EmbeddingProviderSettingsRow,
    KnowledgeIndexJobRow,
)
from app.translation_management_models import TranslationProviderSettingsRow
from app.catalog_translation_models import (
    CatalogLanguagePackRow,
    CatalogSkuTranslationRow,
    CatalogTextTranslationRow,
    CatalogTranslationJobRow,
)
from app.storefront_analytics_models import (
    StorefrontProductViewDailyRow,
    StorefrontProductViewEventRow,
)
from app.announcement_models import StorefrontAnnouncementRow
from app.image_intelligence_models import ImageEmbeddingRow, ImageSearchRow, VisionObservationRow
from app.db_models import ImportJobRow, ReviewItemRow, SourceFileRow, SupplierRow
from app.file_security_models import MediaObjectRow, WorkerJobRow
from app.product_center_models import (
    AttributeDefinitionRow,
    ProductAuditEventRow,
    SkuRow,
    SupplierPriceRow,
)
from app.product_intelligence_models import (
    AIRunRow,
    AITaskStepRow,
    InboxEventRow,
    OutboxEventRow,
    ProductCandidateDecisionRow,
    ProductFieldCandidateRow,
)
from app.product_supplier_models import (
    ProductAttributeRow,
    ProductCategoryRow,
    ProductImageRow,
    ProductRow,
    ProductVersionRow,
    SupplierProductRow,
    SupplierScoreRow,
)
from app.models import PriceCalculationRequest
from app.repositories import public_catalog_repository
from app.saas_seed import (
    DEFAULT_MEMBERSHIP_ID,
    DEFAULT_ORGANIZATION_ID,
    DEFAULT_OWNER_USER_ID,
    DEFAULT_TENANT_ID,
    PERMISSION_SEEDS,
    seed_saas_foundation,
)
from app.services.file_detection import OLE_SIGNATURE, detect_file_path, detect_file_type
from app.services.embedding import DeterministicFeatureHashEmbedding, validate_vectors
from app.services.embedding_configuration import decrypt_api_key
from app.services.translation_configuration import (
    decrypt_translation_api_key,
    resolved_catalog_translator,
    translation_provider_is_configured,
)
from app.services.catalog_translation import catalog_translation_source
from app.services.translation import TranslationIdentity, TranslationProviderError
from app.services.storefront_analytics import (
    request_country_code,
    request_visitor_ip,
)
from app.services import translation_memory as translation_memory_service
from app.services.hybrid_search import (
    _retrieval_tokens,
    _score_tag_relevance,
    hybrid_product_search,
)
from app.services.knowledge import (
    KnowledgeIndexExcludedError,
    build_product_chunks,
    indexed_product_ids,
    project_product_knowledge,
    update_knowledge_index,
)
from app.services.parsers import parse_document
from app.services.pricing import calculate_price
from app.services.product_intelligence.fake_parser import FakeProductParserAdapter
from app.services.product_intelligence.native_parser import NativeSupplierFileParserAdapter
from app.services.product_intelligence.workflow import (
    ProductWorkflowNotFound,
    run_product_draft_workflow,
)
from app.services.product_intelligence.adoption import (
    ProductAdoptionError,
    approve_candidate_group,
    dispatch_product_committed_event,
    reject_candidate_group,
)
from app.services.product_template_import import (
    PRODUCT_MASTER_TEMPLATE_HEADERS,
    PRODUCT_MASTER_TEMPLATE_SHEET,
    PRODUCT_TEMPLATE_HEADERS,
    PRODUCT_TEMPLATE_SHEET,
    PRODUCT_VARIANT_TEMPLATE_HEADERS,
    SKU_DETAIL_TEMPLATE_HEADERS,
    SKU_DETAIL_TEMPLATE_SHEET,
    parse_product_template,
)
from app.services.category_template_import import (
    CATEGORY_TEMPLATE_HEADERS,
    CATEGORY_TEMPLATE_SHEET,
)
from app.services.product_intelligence.normalization import normalize_product_field
from app.services.rbac import has_permission, list_permissions
from app.services.auth.tokens import (
    ACCESS_TTL_SECONDS,
    REFRESH_COOKIE_NAME,
    REFRESH_TTL_SECONDS,
    SESSION_TTL_SECONDS,
    hash_secret,
)
from app.services.auth.contracts import IdentityClaim, IdentityProviderError
from app.services.auth.oidc_provider import OidcIdentityProviderAdapter
from app.services.auth.local_credentials import new_local_password_material
from app.services.auth.service import (
    AuthError,
    _validate_new_password,
    verify_current_user_password,
)
from app.production_bootstrap import bootstrap_production_owner
from app.product_center_seed import seed_product_center_demo
from app.tenant_slugs import RESERVED_TENANT_SLUGS, storefront_slug_from_name
from app.model_mixins import mark_deleted, restore_deleted
from app.adapters.file_scanner import (
    DeterministicDevelopmentScanner,
    get_file_scanner,
)
from app.adapters.object_storage import get_object_storage
from app.ports.file_scanner import FileScanResult
from app.adapters.image_intelligence import get_image_intelligence_provider
from app.adapters.outbox_publisher import InMemoryOutboxPublisher, get_outbox_publisher
from app.services.outbox_consumer import consume_product_committed_message
from app.workers.file_processing import process_file_worker_job
import app.workers.file_processing as file_processing_worker
import app.use_cases.legacy_operations as legacy_operations_use_cases
from app.workers.outbox_relay import relay_one_outbox_event
from app.use_cases.product_center import list_products as list_authoritative_products
from app.use_cases.product_center import delete_all_products as delete_all_products_use_case
from app.use_cases.product_center import list_skus as list_authoritative_skus
from app.use_cases.product_center import upsert_public_offer as upsert_public_offer_use_case
from app.product_center_schemas import PublicCatalogOfferUpsertRequest
from app.use_cases.workspace import create_supplier as create_supplier_use_case
from app.use_cases import catalog_translations as catalog_translation_use_cases
from app.use_cases import public_catalog as public_catalog_use_cases
from app.use_cases import storefront_analytics as storefront_analytics_use_cases
from app.workspace_schemas import SupplierCreateRequest
from app.trade_flow_models import InquiryMatchResultRow, InquiryRow, QuotationApprovalRow, QuotationItemRow, QuotationRow, QuotationVersionRow
from app.public_catalog_models import (
    PublicCatalogOfferRow,
    PublicQuoteDownloadTokenRow,
    PublicQuoteDraftItemRow,
    PublicQuoteDraftRow,
    TenantPublicProfileRow,
)


client = TestClient(app)


def _product_template_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = PRODUCT_TEMPLATE_SHEET
    sheet.append(list(PRODUCT_TEMPLATE_HEADERS))
    for row in rows:
        values = list(row)
        if len(values) == len(PRODUCT_TEMPLATE_HEADERS) - 2:
            values.insert(6, None)
            values.insert(3, None)
        elif len(values) == len(PRODUCT_TEMPLATE_HEADERS) - 1:
            values.insert(3, None)
        sheet.append(values)
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    return content.getvalue()


def _cleanup_template_test_records(
    *,
    import_job_ids: list[str],
    sku_codes: list[str],
    category_names: list[str],
) -> None:
    object_keys: list[str] = []
    with SessionLocal() as session:
        sku_rows = session.scalars(
            select(SkuRow)
            .where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.sku_code.in_(sku_codes),
            )
            .execution_options(include_deleted=True)
        ).all()
        sku_ids = [row.id for row in sku_rows]
        product_ids = [row.product_id for row in sku_rows]
        if sku_ids:
            session.execute(
                delete(PublicCatalogOfferRow).where(
                    PublicCatalogOfferRow.tenant_id == DEFAULT_TENANT_ID,
                    PublicCatalogOfferRow.sku_id.in_(sku_ids),
                )
            )
        if product_ids:
            image_rows = session.scalars(
                select(ProductImageRow)
                .where(
                    ProductImageRow.tenant_id == DEFAULT_TENANT_ID,
                    ProductImageRow.product_id.in_(product_ids),
                )
                .execution_options(include_deleted=True)
            ).all()
            object_keys.extend(
                row.object_key
                for row in image_rows
                if not row.object_key.startswith(("http://", "https://"))
            )
            session.execute(
                delete(ProductImageRow).where(
                    ProductImageRow.tenant_id == DEFAULT_TENANT_ID,
                    ProductImageRow.product_id.in_(product_ids),
                )
            )
        if sku_ids:
            session.execute(
                delete(SkuRow).where(
                    SkuRow.tenant_id == DEFAULT_TENANT_ID,
                    SkuRow.id.in_(sku_ids),
                )
            )
        if product_ids:
            session.execute(
                delete(ProductAuditEventRow).where(
                    ProductAuditEventRow.tenant_id == DEFAULT_TENANT_ID,
                    ProductAuditEventRow.product_id.in_(product_ids),
                )
            )
            session.execute(
                delete(ProductRow).where(
                    ProductRow.tenant_id == DEFAULT_TENANT_ID,
                    ProductRow.id.in_(product_ids),
                )
            )

        if import_job_ids:
            import_rows = session.scalars(
                select(ImportJobRow)
                .where(
                    ImportJobRow.tenant_id == DEFAULT_TENANT_ID,
                    ImportJobRow.id.in_(import_job_ids),
                )
                .execution_options(include_deleted=True)
            ).all()
            worker_rows = session.scalars(
                select(WorkerJobRow).where(
                    WorkerJobRow.tenant_id == DEFAULT_TENANT_ID,
                    WorkerJobRow.import_job_id.in_(import_job_ids),
                )
            ).all()
            source_file_ids = {
                *(row.source_file_id for row in import_rows),
                *(row.source_file_id for row in worker_rows),
            }
            source_rows = (
                session.scalars(
                    select(SourceFileRow)
                    .where(
                        SourceFileRow.tenant_id == DEFAULT_TENANT_ID,
                        SourceFileRow.id.in_(source_file_ids),
                    )
                    .execution_options(include_deleted=True)
                ).all()
                if source_file_ids
                else []
            )
            media_object_ids = {
                *(row.media_object_id for row in worker_rows if row.media_object_id),
                *(row.media_object_id for row in source_rows if row.media_object_id),
            }
            media_rows = (
                session.scalars(
                    select(MediaObjectRow)
                    .where(
                        MediaObjectRow.tenant_id == DEFAULT_TENANT_ID,
                        MediaObjectRow.id.in_(media_object_ids),
                    )
                    .execution_options(include_deleted=True)
                ).all()
                if media_object_ids
                else []
            )
            object_keys.extend(row.object_key for row in media_rows)
            session.execute(
                delete(WorkerJobRow).where(
                    WorkerJobRow.tenant_id == DEFAULT_TENANT_ID,
                    WorkerJobRow.import_job_id.in_(import_job_ids),
                )
            )
            session.execute(
                delete(ImportJobRow).where(
                    ImportJobRow.tenant_id == DEFAULT_TENANT_ID,
                    ImportJobRow.id.in_(import_job_ids),
                )
            )
            if source_file_ids:
                session.execute(
                    delete(SourceFileRow).where(
                        SourceFileRow.tenant_id == DEFAULT_TENANT_ID,
                        SourceFileRow.id.in_(source_file_ids),
                    )
                )
            if media_object_ids:
                session.execute(
                    delete(MediaObjectRow).where(
                        MediaObjectRow.tenant_id == DEFAULT_TENANT_ID,
                        MediaObjectRow.id.in_(media_object_ids),
                    )
                )
        if category_names:
            session.execute(
                delete(ProductCategoryRow).where(
                    ProductCategoryRow.tenant_id == DEFAULT_TENANT_ID,
                    ProductCategoryRow.name.in_(category_names),
                )
            )
        session.commit()

    storage = get_object_storage()
    for object_key in object_keys:
        storage.delete(object_key)
        if "/source/" in object_key:
            storage.delete(object_key.replace("/source/", "/quarantine/", 1))
        elif "/quarantine/" in object_key:
            storage.delete(object_key.replace("/quarantine/", "/source/", 1))


def _create_pending_product_event(tmp_path: Path, *, suffix: str) -> UUID:
    workbook_path = tmp_path / f"outbox-{suffix}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(["Product Name", "Item No", "Material", "MOQ"])
    sheet.append([f"Outbox Product {suffix}", f"OB-{suffix}", "TPR", 100])
    workbook.save(workbook_path)
    source_id = f"SRC-{uuid4().hex[:12].upper()}"
    job_id = f"JOB-{uuid4().hex[:12].upper()}"
    with SessionLocal() as session:
        supplier = session.scalar(
            select(SupplierRow).where(
                SupplierRow.tenant_id == DEFAULT_TENANT_ID,
                SupplierRow.status == "ACTIVE",
            )
        )
        assert supplier is not None
        session.add(
            SourceFileRow(
                id=source_id,
                tenant_id=DEFAULT_TENANT_ID,
                original_filename=workbook_path.name,
                stored_filename=workbook_path.name,
                local_path=str(workbook_path),
                sha256=hashlib.sha256(workbook_path.read_bytes()).hexdigest(),
                byte_size=workbook_path.stat().st_size,
                extension=".xlsx",
                detected_type="OOXML / XLSX",
                extension_matches=True,
                parser="openpyxl",
                security_status="LEGACY_ACCEPTED",
            )
        )
        session.add(
            ImportJobRow(
                id=job_id,
                tenant_id=DEFAULT_TENANT_ID,
                source_file_id=source_id,
                supplier_id=supplier.id,
                supplier_name=supplier.name,
                source_type="SUPPLIER_CATALOG",
                status="needs_review",
                progress=100,
                products_count=1,
            )
        )
        session.commit()
        workflow = run_product_draft_workflow(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            source_file_id=source_id,
            parser=NativeSupplierFileParserAdapter(),
            idempotency_context=f"supplier:{supplier.id}:{suffix}",
        )
        session.commit()
        candidates = session.scalars(
            select(ProductFieldCandidateRow).where(
                ProductFieldCandidateRow.tenant_id == DEFAULT_TENANT_ID,
                ProductFieldCandidateRow.ai_task_id == workflow.task_id,
            )
        ).all()
        assert candidates
        group_key = candidates[0].candidate_group_key
        values = {
            candidate.field_key: candidate.raw_value
            for candidate in candidates
            if candidate.candidate_group_key == group_key
        }
        result = approve_candidate_group(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            task_id=workflow.task_id,
            candidate_group_key=group_key,
            reviewer_membership_id=DEFAULT_MEMBERSHIP_ID,
            idempotency_key=f"outbox-approve-{uuid4()}",
            confirmed_values=values,
            activate=True,
            product_code=f"ATC-OB-{uuid4().hex[:10].upper()}",
            change_reason="ACG-008 relay verification",
        )
        session.commit()
        return result.outbox_event_id


def cleanup_test_runtime() -> None:
    engine.dispose()
    shutil.rmtree(TEST_RUNTIME, ignore_errors=True)


atexit.register(cleanup_test_runtime)


def test_health() -> None:
    response = client.get("/api/v1/health")
    assert response.status_code == 200
    assert response.json()["mode"] == "local_persistence"
    assert response.json()["config_version"] == "local"
    assert client.get("/api/v1/health/live").status_code == 200

    ready = client.get("/api/v1/health/ready")
    assert ready.status_code == 200
    assert ready.json()["status"] == "ready"
    assert ready.json()["dependencies"]["database"]["status"] == "ready"


def test_sqlite_runtime_uses_wal_and_waits_for_short_lock_contention() -> None:
    with engine.connect() as connection:
        journal_mode = connection.exec_driver_sql("PRAGMA journal_mode").scalar_one()
        busy_timeout = connection.exec_driver_sql("PRAGMA busy_timeout").scalar_one()
        synchronous = connection.exec_driver_sql("PRAGMA synchronous").scalar_one()
        foreign_keys = connection.exec_driver_sql("PRAGMA foreign_keys").scalar_one()

    assert str(journal_mode).casefold() == "wal"
    assert int(busy_timeout) >= 30_000
    assert int(synchronous) == 1
    assert int(foreign_keys) == 1


def test_import_job_model_exposes_runtime_progress_without_database_writes() -> None:
    job_id = f"JOB-RUNTIME-{uuid4().hex[:8].upper()}"
    row = SimpleNamespace(
        id=job_id,
        tenant_id=DEFAULT_TENANT_ID,
        worker_jobs=[],
        source_file=SimpleNamespace(
            original_filename="大批量商品.xlsx",
            detected_type="OOXML / XLSX",
            parser="openpyxl",
            extension_matches=True,
        ),
        supplier_name="商品模版",
        source_type="PRODUCT_TEMPLATE",
        status="parsing",
        progress=40,
        products_count=0,
        warnings_count=0,
        created_at=datetime.now(UTC),
        error_message=None,
    )
    publish_runtime_import_progress(
        tenant_id=DEFAULT_TENANT_ID,
        job_id=job_id,
        progress=78,
        stage="APPLYING_PRODUCTS",
        processed_rows=13_000,
        total_rows=26_019,
    )
    try:
        model = import_job_model(row)
    finally:
        clear_runtime_import_progress(
            tenant_id=DEFAULT_TENANT_ID,
            job_id=job_id,
        )

    assert model.progress == 78
    assert model.result_details["import_stage"] == "APPLYING_PRODUCTS"
    assert model.result_details["processed_rows"] == 13_000
    assert model.result_details["total_rows"] == 26_019


def test_readiness_fails_closed_on_migration_mismatch(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("ATC_MIGRATION_HEAD", "20990101_9999")
    response = client.get("/api/v1/health/ready")
    assert response.status_code == 503
    assert response.json()["status"] == "not_ready"
    assert response.json()["dependencies"]["database"]["reason"] == "MIGRATION_HEAD_MISMATCH"


def test_merchant_name_updates_storefront_path_and_preserves_old_link() -> None:
    new_name = f"智贸云测试商家{uuid4().hex[:6]}"
    with SessionLocal() as session:
        tenant = session.get(TenantRow, DEFAULT_TENANT_ID)
        profile = session.get(TenantPublicProfileRow, DEFAULT_TENANT_ID)
        assert tenant is not None and profile is not None
        original_name = tenant.name
        original_tenant_slug = tenant.slug
        original_profile_slug = profile.slug
        original_aliases = list(profile.legacy_slugs or [])

    try:
        response = client.patch(
            "/api/v1/me/merchant",
            json={"name": new_name},
        )
        assert response.status_code == 200, response.text
        updated = response.json()
        assert updated["name"] == new_name
        assert updated["slug"] == new_name.casefold()
        assert updated["storefront_path"] == f"/{new_name.casefold()}"

        canonical = client.get(f"/api/store/{updated['slug']}")
        assert canonical.status_code == 200
        assert canonical.json()["name"] == new_name
        assert canonical.json()["slug"] == updated["slug"]

        legacy = client.get(f"/api/store/{original_tenant_slug}")
        assert legacy.status_code == 200
        assert legacy.json()["slug"] == updated["slug"]

        me = client.get("/api/v1/me")
        assert me.status_code == 200
        assert me.json()["context"]["tenant_name"] == new_name
        assert me.json()["context"]["tenant_slug"] == updated["slug"]
    finally:
        with SessionLocal() as session:
            tenant = session.get(TenantRow, DEFAULT_TENANT_ID)
            profile = session.get(TenantPublicProfileRow, DEFAULT_TENANT_ID)
            assert tenant is not None and profile is not None
            tenant.name = original_name
            tenant.slug = original_tenant_slug
            profile.slug = original_profile_slug
            profile.legacy_slugs = original_aliases
            session.commit()


def test_merchant_name_collision_receives_unique_storefront_path() -> None:
    target_name = f"Shared Merchant {uuid4().hex[:8]}"
    base_slug = storefront_slug_from_name(target_name)
    occupied = client.post(
        "/api/admin/tenants",
        json={"name": target_name, "active": True},
    )
    assert occupied.status_code == 201, occupied.text
    assert occupied.json()["slug"] == base_slug

    with SessionLocal() as session:
        tenant = session.get(TenantRow, DEFAULT_TENANT_ID)
        profile = session.get(TenantPublicProfileRow, DEFAULT_TENANT_ID)
        assert tenant is not None and profile is not None
        original_name = tenant.name
        original_tenant_slug = tenant.slug
        original_profile_slug = profile.slug
        original_aliases = list(profile.legacy_slugs or [])

    try:
        response = client.patch(
            "/api/v1/me/merchant",
            json={"name": target_name},
        )
        assert response.status_code == 200, response.text
        assert response.json()["slug"] == f"{base_slug}-2"
        assert client.get(f"/api/store/{base_slug}-2").status_code == 200
    finally:
        with SessionLocal() as session:
            tenant = session.get(TenantRow, DEFAULT_TENANT_ID)
            profile = session.get(TenantPublicProfileRow, DEFAULT_TENANT_ID)
            assert tenant is not None and profile is not None
            tenant.name = original_name
            tenant.slug = original_tenant_slug
            profile.slug = original_profile_slug
            profile.legacy_slugs = original_aliases
            session.commit()


def test_merchant_business_mode_switches_default_currency_safely() -> None:
    with SessionLocal() as session:
        tenant = session.get(TenantRow, DEFAULT_TENANT_ID)
        assert tenant is not None
        original_currency = tenant.default_currency
        original_warehouses = {
            row.id: (row.is_default, row.version)
            for row in session.scalars(
                select(WarehouseRow).where(
                    WarehouseRow.tenant_id == DEFAULT_TENANT_ID
                )
            ).all()
        }

    try:
        export_response = client.patch(
            "/api/v1/me/merchant",
            json={"business_mode": "EXPORT"},
        )
        assert export_response.status_code == 200, export_response.text
        assert export_response.json()["business_mode"] == "EXPORT"
        assert export_response.json()["default_currency"] == "USD"

        with SessionLocal() as session:
            tenant = session.get(TenantRow, DEFAULT_TENANT_ID)
            assert tenant is not None and tenant.default_currency == "USD"
            defaults = session.scalars(
                select(WarehouseRow).where(
                    WarehouseRow.tenant_id == DEFAULT_TENANT_ID,
                    WarehouseRow.is_default.is_(True),
                )
            ).all()
            assert len(defaults) == 1
            assert defaults[0].currency == "USD"

        me_response = client.get("/api/v1/me")
        assert me_response.status_code == 200
        assert me_response.json()["context"]["business_mode"] == "EXPORT"
        assert me_response.json()["context"]["default_currency"] == "USD"

        domestic_response = client.patch(
            "/api/v1/me/merchant",
            json={"business_mode": "DOMESTIC"},
        )
        assert domestic_response.status_code == 200, domestic_response.text
        assert domestic_response.json()["business_mode"] == "DOMESTIC"
        assert domestic_response.json()["default_currency"] == "CNY"
    finally:
        with SessionLocal() as session:
            tenant = session.get(TenantRow, DEFAULT_TENANT_ID)
            assert tenant is not None
            tenant.default_currency = original_currency
            current = session.scalars(
                select(WarehouseRow).where(
                    WarehouseRow.tenant_id == DEFAULT_TENANT_ID
                )
            ).all()
            for warehouse in current:
                warehouse.is_default = False
            session.flush()
            for warehouse in current:
                if warehouse.id in original_warehouses:
                    is_default, version = original_warehouses[warehouse.id]
                    warehouse.is_default = is_default
                    warehouse.version = version
                else:
                    session.delete(warehouse)
            session.commit()


def test_merchant_controls_languages_visible_on_the_storefront(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        public_catalog_use_cases,
        "catalog_translation_is_configured",
        lambda: True,
    )
    with SessionLocal() as session:
        profile = session.get(TenantPublicProfileRow, DEFAULT_TENANT_ID)
        assert profile is not None
        original_locales = list(profile.storefront_locales or [])

    try:
        response = client.patch(
            "/api/v1/me/merchant",
            json={"storefront_locales": ["es", "ar", "ja", "pt"]},
        )
        assert response.status_code == 200, response.text
        assert response.json()["storefront_locales"] == [
            "zh-CN",
            "es",
            "ar",
            "ja",
            "pt",
        ]

        settings = client.get("/api/v1/me/merchant")
        assert settings.status_code == 200, settings.text
        assert settings.json()["storefront_locales"] == [
            "zh-CN",
            "es",
            "ar",
            "ja",
            "pt",
        ]

        store = client.get("/api/store/demo", params={"locale": "pt-BR"})
        assert store.status_code == 200, store.text
        assert store.json()["locale"] == "pt"
        assert store.json()["available_locales"] == [
            "zh-CN",
            "es",
            "ar",
            "ja",
            "pt",
        ]

        disabled = client.get("/api/store/demo", params={"locale": "ko"})
        assert disabled.status_code == 422
        assert disabled.json()["detail"]["code"] == "PUBLIC_LOCALE_DISABLED"

        invalid = client.patch(
            "/api/v1/me/merchant",
            json={"storefront_locales": ["zh-CN", "de"]},
        )
        assert invalid.status_code == 422
    finally:
        with SessionLocal() as session:
            profile = session.get(TenantPublicProfileRow, DEFAULT_TENANT_ID)
            assert profile is not None
            profile.storefront_locales = original_locales
            session.commit()


def test_merchant_controls_hot_product_merchandising() -> None:
    with SessionLocal() as session:
        profile = session.get(TenantPublicProfileRow, DEFAULT_TENANT_ID)
        assert profile is not None
        original = bool(profile.hot_products_enabled)

    try:
        response = client.patch(
            "/api/v1/me/merchant",
            json={"hot_products_enabled": not original},
        )
        assert response.status_code == 200, response.text
        assert response.json()["hot_products_enabled"] is (not original)

        settings = client.get("/api/v1/me/merchant")
        assert settings.status_code == 200, settings.text
        assert settings.json()["hot_products_enabled"] is (not original)

        store = client.get("/api/store/demo")
        assert store.status_code == 200, store.text
        assert store.json()["hot_products_enabled"] is (not original)
    finally:
        with SessionLocal() as session:
            profile = session.get(TenantPublicProfileRow, DEFAULT_TENANT_ID)
            assert profile is not None
            profile.hot_products_enabled = original
            session.commit()


def test_user_can_persist_console_locale_preference() -> None:
    with SessionLocal() as session:
        user = session.get(UserRow, DEFAULT_OWNER_USER_ID)
        assert user is not None
        original_locale = user.locale
    try:
        response = client.patch(
            "/api/v1/me/preferences",
            json={"locale": "en-US"},
        )
        assert response.status_code == 200, response.text
        assert response.json() == {"locale": "en-US"}
        me_response = client.get("/api/v1/me")
        assert me_response.status_code == 200
        assert me_response.json()["user"]["locale"] == "en-US"
        invalid = client.patch(
            "/api/v1/me/preferences",
            json={"locale": "fr-FR"},
        )
        assert invalid.status_code == 422
    finally:
        with SessionLocal() as session:
            user = session.get(UserRow, DEFAULT_OWNER_USER_ID)
            assert user is not None
            user.locale = original_locale
            session.commit()


def test_merchant_can_schedule_ticker_and_safe_rich_popup_announcements() -> None:
    starts_at = datetime.now(UTC) - timedelta(minutes=1)
    with SessionLocal() as session:
        def published_sku() -> SkuRow | None:
            return session.scalar(
                select(SkuRow)
                .join(
                    PublicCatalogOfferRow,
                    (PublicCatalogOfferRow.tenant_id == SkuRow.tenant_id)
                    & (PublicCatalogOfferRow.sku_id == SkuRow.id),
                )
                .join(
                    ProductRow,
                    (ProductRow.tenant_id == SkuRow.tenant_id)
                    & (ProductRow.id == SkuRow.product_id),
                )
                .where(
                    SkuRow.tenant_id == DEFAULT_TENANT_ID,
                    SkuRow.status == "ACTIVE",
                    SkuRow.deleted_at.is_(None),
                    ProductRow.status == "ACTIVE",
                    ProductRow.deleted_at.is_(None),
                    PublicCatalogOfferRow.publication_status == "PUBLISHED",
                    PublicCatalogOfferRow.deleted_at.is_(None),
                )
                .limit(1)
            )

        related_sku = published_sku()
        if related_sku is None:
            seed_product_center_demo(session)
            demo_sku = session.scalar(
                select(SkuRow)
                .where(
                    SkuRow.tenant_id == DEFAULT_TENANT_ID,
                    SkuRow.sku_code == "AQ-320S",
                )
                .execution_options(include_deleted=True)
            )
            assert demo_sku is not None
            restore_deleted(demo_sku)
            demo_sku.status = "ACTIVE"
            demo_product = session.get(
                ProductRow,
                demo_sku.product_id,
                execution_options={"include_deleted": True},
            )
            assert demo_product is not None
            restore_deleted(demo_product)
            demo_product.status = "ACTIVE"
            demo_offer = session.scalar(
                select(PublicCatalogOfferRow)
                .where(
                    PublicCatalogOfferRow.tenant_id == DEFAULT_TENANT_ID,
                    PublicCatalogOfferRow.sku_id == demo_sku.id,
                )
                .execution_options(include_deleted=True)
            )
            assert demo_offer is not None
            restore_deleted(demo_offer)
            demo_offer.publication_status = "PUBLISHED"
            session.commit()
            related_sku = published_sku()
        assert related_sku is not None
        related_sku_id = related_sku.id
    ticker = client.post(
        "/api/v1/announcements",
        json={
            "display_type": "TICKER",
            "ticker_text": "Orders placed this week ship on Monday.",
            "content_blocks": [],
            "related_sku_ids": [str(related_sku_id)],
            "starts_at": starts_at.isoformat(),
            "duration_days": 2,
            "ticker_speed_px_per_second": 85,
            "publication_status": "PUBLISHED",
        },
    )
    assert ticker.status_code == 201, ticker.text
    ticker_data = ticker.json()
    assert ticker_data["is_active"] is True
    assert ticker_data["title"] is None
    assert ticker_data["ticker_speed_px_per_second"] == 85
    assert ticker_data["related_skus"][0]["id"] == str(related_sku_id)

    popup = client.post(
        "/api/v1/announcements",
        json={
            "title": "New collection",
            "display_type": "MODAL",
            "content_blocks": [
                {"type": "heading", "text": "Explore the new collection"},
                {"type": "paragraph", "text": "New products are now available."},
                {
                    "type": "image",
                    "url": "https://cdn.example.test/new.jpg",
                    "alt": "New collection",
                },
                {
                    "type": "video",
                    "url": "https://cdn.example.test/new.mp4",
                    "caption": "Product walkthrough",
                },
            ],
            "starts_at": starts_at.isoformat(),
            "duration_days": 3,
            "publication_status": "PUBLISHED",
        },
    )
    assert popup.status_code == 201, popup.text
    popup_data = popup.json()
    assert popup_data["content_blocks"][2]["type"] == "image"

    listed = client.get("/api/v1/announcements")
    assert listed.status_code == 200, listed.text
    listed_ids = {row["id"] for row in listed.json()["items"]}
    assert {ticker_data["id"], popup_data["id"]} <= listed_ids

    storefront = client.get("/api/store/demo")
    assert storefront.status_code == 200, storefront.text
    public_rows = {
        row["id"]: row for row in storefront.json()["announcements"]
    }
    assert public_rows[ticker_data["id"]]["ticker_text"].startswith("Orders")
    assert public_rows[ticker_data["id"]]["related_skus"][0]["id"] == str(related_sku_id)
    assert public_rows[ticker_data["id"]]["ticker_speed_px_per_second"] == 85
    assert public_rows[popup_data["id"]]["ticker_speed_px_per_second"] == 60

    missing_sku = client.post(
        "/api/v1/announcements",
        json={
            "display_type": "TICKER",
            "ticker_text": "Unavailable product",
            "related_sku_ids": [str(uuid4())],
            "starts_at": starts_at.isoformat(),
            "duration_days": 1,
            "publication_status": "PUBLISHED",
        },
    )
    assert missing_sku.status_code == 422

    unsafe = client.post(
        "/api/v1/announcements",
        json={
            "title": "Unsafe",
            "display_type": "MODAL",
            "content_blocks": [
                {"type": "image", "url": "javascript:alert(1)"}
            ],
            "starts_at": starts_at.isoformat(),
            "duration_days": 1,
            "publication_status": "PUBLISHED",
        },
    )
    assert unsafe.status_code == 422

    for announcement_id in (ticker_data["id"], popup_data["id"]):
        removed = client.delete(f"/api/v1/announcements/{announcement_id}")
        assert removed.status_code == 204
    with SessionLocal() as session:
        assert session.scalars(
            select(StorefrontAnnouncementRow).where(
                StorefrontAnnouncementRow.id.in_(
                    (UUID(ticker_data["id"]), UUID(popup_data["id"]))
                )
            )
        ).all() == []


def test_storefront_support_settings_and_human_conversation_flow() -> None:
    settings = client.get("/api/v1/support/settings")
    assert settings.status_code == 200, settings.text
    assert [row["slot"] for row in settings.json()["custom_actions"]] == [2, 3]

    action_image = BytesIO()
    Image.new("RGBA", (24, 24), (20, 180, 140, 255)).save(action_image, "PNG")
    uploaded = client.post(
        "/api/v1/support/settings/actions/2/image",
        files={"image": ("contact.png", action_image.getvalue(), "image/png")},
    )
    assert uploaded.status_code == 200, uploaded.text
    assert uploaded.json()["custom_actions"][0]["has_uploaded_image"] is True

    saved = client.patch(
        "/api/v1/support/settings",
        json={
            "welcome_message": "Tell us which product you are looking for.",
            "custom_actions": [
                {
                    "slot": 2,
                    "visible": True,
                    "label": "WhatsApp",
                },
                {"slot": 3, "visible": False},
            ],
        },
    )
    assert saved.status_code == 200, saved.text
    assert saved.json()["custom_actions"][0]["visible"] is True
    assert saved.json()["custom_actions"][0]["has_uploaded_image"] is True

    public_image = client.get("/api/store/demo/support/actions/2/image")
    assert public_image.status_code == 200, public_image.text
    assert public_image.headers["content-type"] == "image/webp"

    storefront = client.get("/api/store/demo")
    assert storefront.status_code == 200, storefront.text
    widget = storefront.json()["support_widget"]
    assert widget["ai_enabled"] is False
    assert widget["welcome_message"].startswith("Tell us")
    assert [row["slot"] for row in widget["custom_actions"]] == [2]

    created = client.post(
        "/api/store/demo/support/conversations",
        json={
            "message": "Do you have this item in blue?",
            "client_message_id": str(uuid4()),
            "locale": "en-US",
        },
    )
    assert created.status_code == 201, created.text
    conversation_id = created.json()["id"]
    support_token = created.json()["access_token"]
    assert support_token
    assert created.json()["messages"][0]["sender_type"] == "VISITOR"

    second_message_id = str(uuid4())
    for _ in range(2):
        sent = client.post(
            "/api/store/demo/support/conversations/current/messages",
            headers={"X-Support-Token": support_token},
            json={
                "message": "A medium size, please.",
                "client_message_id": second_message_id,
            },
        )
        assert sent.status_code == 200, sent.text
    assert len(sent.json()["messages"]) == 2

    listed = client.get("/api/v1/support/conversations")
    assert listed.status_code == 200, listed.text
    assert conversation_id in {row["id"] for row in listed.json()["items"]}

    detail = client.get(f"/api/v1/support/conversations/{conversation_id}")
    assert detail.status_code == 200, detail.text
    assert detail.json()["unread"] is False
    assert detail.json()["automation_state"] == "AI_ACTIVE"

    replied = client.post(
        f"/api/v1/support/conversations/{conversation_id}/messages",
        json={"message": "Yes, blue is available."},
    )
    assert replied.status_code == 200, replied.text
    assert replied.json()["messages"][-1]["sender_type"] == "MERCHANT"
    assert replied.json()["automation_state"] == "HUMAN_TAKEOVER"

    resumed = client.patch(
        f"/api/v1/support/conversations/{conversation_id}/automation",
        json={"automation_state": "AI_ACTIVE"},
    )
    assert resumed.status_code == 200, resumed.text
    assert resumed.json()["automation_state"] == "AI_ACTIVE"

    public_detail = client.get(
        "/api/store/demo/support/conversations/current",
        headers={"X-Support-Token": support_token},
    )
    assert public_detail.status_code == 200, public_detail.text
    assert public_detail.json()["messages"][-1]["body"].startswith("Yes")

    closed = client.patch(
        f"/api/v1/support/conversations/{conversation_id}",
        json={"status": "CLOSED"},
    )
    assert closed.status_code == 200, closed.text
    rejected = client.post(
        "/api/store/demo/support/conversations/current/messages",
        headers={"X-Support-Token": support_token},
        json={"message": "One more question"},
    )
    assert rejected.status_code == 409, rejected.text


def test_support_ai_configuration_and_file_knowledge_lifecycle(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from app.services.chat_generation import (
        ChatGenerationIdentity,
        ChatGenerationResult,
        OpenAICompatibleChatGeneration,
    )
    from app.services import support_ai_orchestrator
    from app.support_ai_models import (
        SupportAIEvidenceUseRow,
        SupportAIKnowledgeSourceRow,
        SupportAIProviderSettingsRow,
        SupportAIRunRow,
        SupportAISettingsRow,
    )
    from app.support_models import (
        StorefrontChatConversationRow,
        StorefrontChatMessageRow,
    )

    def unexpected_generation_request(*_args: object, **_kwargs: object) -> object:
        raise AssertionError("saving model settings must not test provider connectivity")

    monkeypatch.setenv("SUPPORT_AI_SETTINGS_MASTER_KEY", "test-master-key-" * 3)
    monkeypatch.setattr(
        OpenAICompatibleChatGeneration,
        "_request",
        unexpected_generation_request,
    )
    source_id: UUID | None = None
    run_id: UUID | None = None
    task_id: UUID | None = None
    conversation_ids: set[UUID] = set()
    object_key: str | None = None
    try:
        configured = client.put(
            "/api/v1/system/ai-generation/settings",
            json={
                "enabled": True,
                "base_url": "https://generation.invalid/v1",
                "model_name": "support-json-model",
                "api_key": "sk-support-test-secret",
                "timeout_seconds": 30,
                "max_output_tokens": 1024,
                "temperature": 0.1,
            },
        )
        assert configured.status_code == 200, configured.text
        provider_payload = configured.json()
        assert provider_payload["source"] == "database"
        assert provider_payload["api_key_configured"] is True
        assert provider_payload["api_key_hint"] == "••••cret"
        assert "api_key" not in provider_payload

        settings = client.patch(
            "/api/v1/support/ai/settings",
            json={"mode": "DRAFT"},
        )
        assert settings.status_code == 200, settings.text
        assert settings.json()["mode"] == "DRAFT"
        assert settings.json()["provider_configured"] is True

        uploaded = client.post(
            "/api/v1/support/ai/knowledge/sources/upload",
            data={
                "title": "Brand shipping policy",
                "classification": "CUSTOMER_APPROVED",
                "language": "en-US",
            },
            files={
                "file": (
                    "brand-policy.txt",
                    b"Standard dispatch takes three business days.",
                    "text/plain",
                )
            },
        )
        assert uploaded.status_code == 202, uploaded.text
        source_id = UUID(uploaded.json()["source"]["id"])
        job_id = uploaded.json()["job"]["id"]

        job = client.get(f"/api/v1/support/ai/knowledge/jobs/{job_id}")
        assert job.status_code == 200, job.text
        assert job.json()["status"] == "SUCCEEDED"
        assert job.json()["chunks_written"] >= 1

        sources = client.get("/api/v1/support/ai/knowledge/sources")
        assert sources.status_code == 200, sources.text
        source = next(
            row for row in sources.json() if row["id"] == str(source_id)
        )
        assert source["status"] == "READY"
        assert source["chunk_count"] >= 1

        approved = client.post(
            f"/api/v1/support/ai/knowledge/sources/{source_id}/approve"
        )
        assert approved.status_code == 200, approved.text
        assert approved.json()["status"] == "APPROVED"

        class FakeSupportGenerationProvider:
            identity = ChatGenerationIdentity(
                provider="fake-support",
                model_name="grounded-test-model",
            )

            def generate_json(
                self,
                *,
                messages: list[dict[str, str]],
                temperature: float | None = None,
                max_output_tokens: int | None = None,
            ) -> ChatGenerationResult:
                del messages, temperature, max_output_tokens
                data = {
                    "detected_language": "en-US",
                    "answer": "Standard dispatch takes three business days. [1]",
                    "confidence": 0.92,
                    "citations": [1],
                    "handoff": False,
                    "handoff_reason": None,
                }
                return ChatGenerationResult(
                    content="",
                    data=data,
                    finish_reason="stop",
                    usage={"total_tokens": 42},
                )

        monkeypatch.setattr(
            support_ai_orchestrator,
            "resolved_support_ai_provider",
            lambda _session: FakeSupportGenerationProvider(),
        )
        test_run = client.post(
            "/api/v1/support/ai/test-runs",
            json={
                "question": "Standard dispatch takes three business days.",
                "locale": "en-US",
            },
        )
        assert test_run.status_code == 200, test_run.text
        run_payload = test_run.json()
        run_id = UUID(run_payload["id"])
        task_id = UUID(run_payload["ai_task_id"])
        assert run_payload["status"] == "SUCCEEDED"
        assert run_payload["detected_language"] == "en-US"
        assert run_payload["answer"].endswith("[1]")
        assert run_payload["evidence"][0]["source_entity_id"] == str(source_id)
        assert run_payload["evidence"][0]["classification"] == "CUSTOMER_APPROVED"
        assert run_payload["decision_trace"]["citations_valid"] is True

        auto_settings = client.patch(
            "/api/v1/support/ai/settings",
            json={"mode": "AUTO_LIMITED", "daily_auto_reply_limit": 1},
        )
        assert auto_settings.status_code == 200, auto_settings.text

        first_chat = client.post(
            "/api/store/demo/support/conversations",
            json={
                "message": "Standard dispatch takes three business days.",
                "client_message_id": str(uuid4()),
                "locale": "en-US",
            },
        )
        assert first_chat.status_code == 201, first_chat.text
        conversation_ids.add(UUID(first_chat.json()["id"]))
        first_public = client.get(
            "/api/store/demo/support/conversations/current",
            headers={"X-Support-Token": first_chat.json()["access_token"]},
        )
        assert first_public.status_code == 200, first_public.text
        first_ai_message = first_public.json()["messages"][-1]
        assert first_ai_message["sender_type"] == "AI"
        assert first_ai_message["citations"][0]["source_entity_id"] == str(source_id)

        limited_chat = client.post(
            "/api/store/demo/support/conversations",
            json={
                "message": "Standard dispatch takes three business days.",
                "client_message_id": str(uuid4()),
                "locale": "en-US",
            },
        )
        assert limited_chat.status_code == 201, limited_chat.text
        conversation_ids.add(UUID(limited_chat.json()["id"]))
        assert limited_chat.json()["automation_state"] == "HUMAN_TAKEOVER"
        assert limited_chat.json()["messages"][-1]["sender_type"] == "SYSTEM"

        with SessionLocal() as session:
            limited_run = session.scalar(
                select(SupportAIRunRow).where(
                    SupportAIRunRow.conversation_id == UUID(limited_chat.json()["id"])
                )
            )
            assert limited_run is not None
            assert limited_run.status == "SKIPPED"
            assert limited_run.handoff_reason == "DAILY_AUTO_REPLY_LIMIT_REACHED"
            assert limited_run.decision_trace["model_called"] is False

        revoked = client.delete(
            f"/api/v1/support/ai/knowledge/sources/{source_id}"
        )
        assert revoked.status_code == 200, revoked.text
        assert revoked.json()["status"] == "REVOKED"
    finally:
        with SessionLocal() as session:
            run_rows = session.scalars(
                select(SupportAIRunRow).where(
                    SupportAIRunRow.id == run_id
                    if not conversation_ids
                    else (
                        (SupportAIRunRow.id == run_id)
                        | SupportAIRunRow.conversation_id.in_(conversation_ids)
                    )
                )
            ).all()
            cleanup_run_ids = [row.id for row in run_rows]
            cleanup_task_ids = [row.ai_task_id for row in run_rows]
            if cleanup_run_ids:
                session.execute(
                    delete(SupportAIEvidenceUseRow).where(
                        SupportAIEvidenceUseRow.run_id.in_(cleanup_run_ids)
                    )
                )
                session.execute(
                    delete(SupportAIRunRow).where(
                        SupportAIRunRow.id.in_(cleanup_run_ids)
                    )
                )
            if task_id is not None and task_id not in cleanup_task_ids:
                cleanup_task_ids.append(task_id)
            if cleanup_task_ids:
                session.execute(
                    delete(AITaskRow).where(AITaskRow.id.in_(cleanup_task_ids))
                )
            if conversation_ids:
                session.execute(
                    delete(StorefrontChatMessageRow).where(
                        StorefrontChatMessageRow.conversation_id.in_(conversation_ids)
                    )
                )
                session.execute(
                    delete(StorefrontChatConversationRow).where(
                        StorefrontChatConversationRow.id.in_(conversation_ids)
                    )
                )
            if source_id is not None:
                source_row = session.get(SupportAIKnowledgeSourceRow, source_id)
                if source_row is not None:
                    media_row = session.get(MediaObjectRow, source_row.media_object_id)
                    object_key = media_row.object_key if media_row is not None else None
                    session.delete(source_row)
                    session.flush()
                    if media_row is not None:
                        session.delete(media_row)
            provider_row = session.get(
                SupportAIProviderSettingsRow,
                "SUPPORT_AI_GENERATION",
            )
            if provider_row is not None:
                session.delete(provider_row)
            tenant_settings = session.get(SupportAISettingsRow, DEFAULT_TENANT_ID)
            if tenant_settings is not None:
                session.delete(tenant_settings)
            session.commit()
        if object_key:
            get_object_storage().delete(object_key)


class _SupportTranslationTestProvider:
    identity = TranslationIdentity(provider="support-translation-test", version="v1")

    def translate(
        self,
        text: str,
        *,
        source_locale: str,
        target_locale: str,
    ) -> str:
        if source_locale == "auto" and target_locale == "zh-CN":
            replacements = {
                "¿Tienen este producto en azul?": "这个商品有蓝色吗？",
            }
        elif source_locale == "auto" and target_locale == "en-US":
            replacements = {
                "Este produto está disponível?": "Is this product available?",
            }
        else:
            replacements = {
                ("zh-CN", "es"): {
                    "蓝色有库存，可以立即报价。":
                    "El azul está disponible y podemos cotizarlo ahora.",
                },
            }.get((source_locale, target_locale))
        if replacements is None:
            raise AssertionError(
                f"unexpected support translation: {source_locale} -> {target_locale}"
            )
        translated = text
        for source, target in replacements.items():
            translated = translated.replace(source, target)
        return translated


def test_support_conversation_bidirectional_translation(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from app.use_cases import support as support_use_cases

    provider = _SupportTranslationTestProvider()
    translation_memory_service._reset_translation_memory_for_tests()
    monkeypatch.setattr(
        support_use_cases,
        "catalog_translation_is_configured",
        lambda: True,
    )
    monkeypatch.setattr(
        support_use_cases,
        "configured_catalog_translator",
        lambda: provider,
    )
    with SessionLocal() as session:
        tenant = session.get(TenantRow, DEFAULT_TENANT_ID)
        assert tenant is not None
        original_currency = tenant.default_currency
        tenant.default_currency = "CNY"
        session.execute(
            delete(CatalogTextTranslationRow).where(
                CatalogTextTranslationRow.tenant_id == DEFAULT_TENANT_ID,
                CatalogTextTranslationRow.provider == provider.identity.provider,
            )
        )
        session.commit()

    try:
        created = client.post(
            "/api/store/demo/support/conversations",
            json={
                "message": "¿Tienen este producto en azul?",
                "client_message_id": str(uuid4()),
                "locale": "es",
            },
        )
        assert created.status_code == 201, created.text
        public_message = created.json()["messages"][0]
        assert public_message["body"] == "¿Tienen este producto en azul?"
        assert "translated_body" not in public_message

        conversation_id = created.json()["id"]
        support_token = created.json()["access_token"]
        detail = client.get(f"/api/v1/support/conversations/{conversation_id}")
        assert detail.status_code == 200, detail.text
        visitor_message = detail.json()["messages"][0]
        assert visitor_message["translated_body"] == "这个商品有蓝色吗？"
        assert visitor_message["translation_source_locale"] == "es"
        assert visitor_message["translation_target_locale"] == "zh-CN"
        assert visitor_message["translation_status"] == "READY"

        preview = client.post(
            f"/api/v1/support/conversations/{conversation_id}/translation-preview",
            json={
                "message": "蓝色有库存，可以立即报价。",
                "target_locale": "es",
            },
        )
        assert preview.status_code == 200, preview.text
        assert preview.json()["translated_message"].startswith("El azul")

        final_reply = "El azul está disponible; podemos cotizarlo hoy."
        replied = client.post(
            f"/api/v1/support/conversations/{conversation_id}/messages",
            json={
                "message": final_reply,
                "draft_message": "蓝色有库存，可以立即报价。",
                "source_locale": "zh-CN",
                "target_locale": "es",
            },
        )
        assert replied.status_code == 200, replied.text
        merchant_message = replied.json()["messages"][-1]
        assert merchant_message["body"] == final_reply
        assert merchant_message["draft_body"] == "蓝色有库存，可以立即报价。"

        public_detail = client.get(
            "/api/store/demo/support/conversations/current",
            headers={"X-Support-Token": support_token},
        )
        assert public_detail.status_code == 200, public_detail.text
        assert public_detail.json()["messages"][-1]["body"] == final_reply
        assert "draft_body" not in public_detail.json()["messages"][-1]

        with SessionLocal() as session:
            tenant = session.get(TenantRow, DEFAULT_TENANT_ID)
            assert tenant is not None
            tenant.default_currency = "USD"
            session.commit()
        export_conversation = client.post(
            "/api/store/demo/support/conversations",
            json={
                "message": "Este produto está disponível?",
                "client_message_id": str(uuid4()),
                "locale": "pt",
            },
        )
        assert export_conversation.status_code == 201, export_conversation.text
        export_detail = client.get(
            f"/api/v1/support/conversations/{export_conversation.json()['id']}"
        )
        export_message = export_detail.json()["messages"][0]
        assert export_message["translated_body"] == "Is this product available?"
        assert export_message["translation_target_locale"] == "en-US"
    finally:
        with SessionLocal() as session:
            tenant = session.get(TenantRow, DEFAULT_TENANT_ID)
            assert tenant is not None
            tenant.default_currency = original_currency
            session.commit()


def test_trusted_auth_session_membership_refresh_and_logout(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("AUTH_TEST_BYPASS", "false")
    assert client.get("/api/v1/suppliers").status_code == 401

    login_response = client.post(
        "/api/v1/auth/login",
        json={
            "provider": "local_fake",
            "authorization_code": f"fake:{DEFAULT_OWNER_USER_ID}",
            "code_verifier": "A" * 43,
            "redirect_uri": "http://127.0.0.1:5173/login/callback",
            "device_label": "pytest-browser",
        },
    )
    assert login_response.status_code == 200, login_response.text
    token_data = login_response.json()["data"]
    assert token_data["memberships"] == [
        {
            "id": str(DEFAULT_MEMBERSHIP_ID),
            "tenant_id": str(DEFAULT_TENANT_ID),
            "tenant_name": "Local Demo Company",
            "tenant_slug": "demo",
            "status": "active",
        }
    ]
    assert token_data["permission_version"] == 1
    assert "product.view" in token_data["permissions"]
    assert "announcement.manage" in token_data["permissions"]
    assert ACCESS_TTL_SECONDS == 10 * 60
    assert token_data["expires_in"] == ACCESS_TTL_SECONDS
    access_token = token_data["access_token"]
    csrf_token = token_data["csrf_token"]
    raw_refresh = login_response.cookies.get(REFRESH_COOKIE_NAME)
    assert raw_refresh
    assert "httponly" in login_response.headers["set-cookie"].lower()
    assert raw_refresh not in login_response.text

    with SessionLocal() as session:
        auth_session = session.get(AuthSessionRow, UUID(token_data["session_id"]))
        assert auth_session is not None
        assert auth_session.active_membership_id == DEFAULT_MEMBERSHIP_ID
        stored_refresh = session.scalar(
            select(AuthRefreshTokenRow).where(
                AuthRefreshTokenRow.auth_session_id == auth_session.id,
                AuthRefreshTokenRow.sequence_number == 0,
            )
        )
        assert stored_refresh is not None
        assert SESSION_TTL_SECONDS == 7 * 24 * 60 * 60
        assert REFRESH_TTL_SECONDS == 7 * 24 * 60 * 60
        assert auth_session.expires_at - auth_session.issued_at == timedelta(days=7)
        assert stored_refresh.expires_at - stored_refresh.issued_at == timedelta(days=7)
        assert stored_refresh.token_hash == hash_secret(raw_refresh)
        assert stored_refresh.token_hash != raw_refresh

    bearer_headers = {
        "Authorization": f"Bearer {access_token}",
        "X-Tenant-ID": str(uuid4()),
    }
    me_response = client.get("/api/v1/me", headers=bearer_headers)
    assert me_response.status_code == 200, me_response.text
    assert me_response.json()["context"]["tenant_id"] == str(DEFAULT_TENANT_ID)
    assert me_response.json()["context"]["tenant_slug"] == "demo"
    assert me_response.json()["user"]["is_platform_admin"] is True
    permissions_response = client.get(
        "/api/v1/me/permissions", headers=bearer_headers
    )
    assert permissions_response.status_code == 200
    assert "product.view" in permissions_response.json()["permissions"]
    bootstrap_response = client.get(
        "/api/v1/auth/bootstrap", headers=bearer_headers
    )
    assert bootstrap_response.status_code == 200
    bootstrap = bootstrap_response.json()
    assert bootstrap["profile"] == me_response.json()
    assert bootstrap["permissions"] == permissions_response.json()
    assert bootstrap_response.headers["cache-control"] == "no-store"
    memberships_response = client.get(
        "/api/v1/auth/memberships", headers=bearer_headers
    )
    assert memberships_response.status_code == 200
    assert memberships_response.json() == [
        {
            "id": str(DEFAULT_MEMBERSHIP_ID),
            "tenant_id": str(DEFAULT_TENANT_ID),
            "tenant_name": "Local Demo Company",
            "tenant_slug": "demo",
            "status": "active",
        }
    ]

    with SessionLocal() as session:
        membership = session.get(MembershipRow, DEFAULT_MEMBERSHIP_ID)
        assert membership is not None
        membership.status = "suspended"
        session.commit()
    suspended_response = client.get("/api/v1/me", headers=bearer_headers)
    assert suspended_response.status_code == 403
    with SessionLocal() as session:
        membership = session.get(MembershipRow, DEFAULT_MEMBERSHIP_ID)
        assert membership is not None
        membership.status = "active"
        session.commit()

    # A stale tab can temporarily pair the current shared refresh cookie with
    # its previous CSRF token. This is recoverable and must not erase the valid
    # cookie; the browser can retry after reading the CSRF value shared by the
    # tab that completed rotation.
    invalid_csrf_response = client.post(
        "/api/v1/auth/refresh",
        headers={"X-CSRF-Token": "stale-tab-csrf-token"},
    )
    assert invalid_csrf_response.status_code == 403
    assert invalid_csrf_response.json()["detail"]["code"] == "AUTH_CSRF_INVALID"
    assert REFRESH_COOKIE_NAME not in invalid_csrf_response.headers.get(
        "set-cookie", ""
    )
    assert client.cookies.get(REFRESH_COOKIE_NAME) == raw_refresh

    refresh_response = client.post(
        "/api/v1/auth/refresh",
        headers={"X-CSRF-Token": csrf_token},
    )
    assert refresh_response.status_code == 200, refresh_response.text
    rotated_data = refresh_response.json()["data"]
    rotated_refresh = refresh_response.cookies.get(REFRESH_COOKIE_NAME)
    assert rotated_refresh

    # A second tab can already have sent the same cookie/CSRF pair before the
    # first response updates the shared cookie jar. The row lock serializes
    # that loser behind the first rotation; the bounded retry must return the
    # same successor without advancing or revoking the token family.
    with TestClient(app) as concurrent_client:
        concurrent_client.cookies.set(
            REFRESH_COOKIE_NAME,
            raw_refresh,
            path="/api/v1/auth",
        )
        concurrent_response = concurrent_client.post(
            "/api/v1/auth/refresh",
            headers={"X-CSRF-Token": csrf_token},
        )
    assert concurrent_response.status_code == 200, concurrent_response.text
    assert concurrent_response.cookies.get(REFRESH_COOKIE_NAME) == rotated_refresh
    assert concurrent_response.json()["data"]["csrf_token"] == rotated_data["csrf_token"]
    with SessionLocal() as session:
        auth_session = session.get(AuthSessionRow, UUID(token_data["session_id"]))
        assert auth_session is not None
        assert auth_session.revoked_at is None
        assert auth_session.rotation_counter == 1
        stored_tokens = session.scalars(
            select(AuthRefreshTokenRow)
            .where(AuthRefreshTokenRow.auth_session_id == auth_session.id)
            .order_by(AuthRefreshTokenRow.sequence_number)
        ).all()
        assert len(stored_tokens) == 2
        assert stored_tokens[0].rotation_request_hash
        assert stored_tokens[0].retry_grace_expires_at
        assert stored_tokens[1].token_hash == hash_secret(rotated_refresh)
        assert rotated_refresh not in {
            stored_tokens[0].token_hash,
            stored_tokens[0].rotation_request_hash,
            stored_tokens[1].token_hash,
        }
        # Move beyond the explicit grace window to model a genuine replay,
        # without slowing the suite down.
        stored_tokens[0].retry_grace_expires_at = datetime.now(UTC) - timedelta(
            seconds=1
        )
        session.commit()

    with TestClient(app) as replay_client:
        replay_client.cookies.set(REFRESH_COOKIE_NAME, raw_refresh, path="/api/v1/auth")
        replay_response = replay_client.post(
            "/api/v1/auth/refresh",
            headers={"X-CSRF-Token": csrf_token},
        )
    assert replay_response.status_code == 401
    assert replay_response.json()["detail"]["code"] == "AUTH_REFRESH_REUSE_DETECTED"
    replay_cookie = replay_response.headers.get("set-cookie", "").lower()
    assert REFRESH_COOKIE_NAME in replay_cookie
    assert "max-age=0" in replay_cookie
    assert client.get(
        "/api/v1/me",
        headers={"Authorization": f"Bearer {rotated_data['access_token']}"},
    ).status_code == 401

    second_login = client.post(
        "/api/v1/auth/login",
        json={
            "provider": "local_fake",
            "authorization_code": f"fake:{DEFAULT_OWNER_USER_ID}",
            "code_verifier": "B" * 43,
            "redirect_uri": "http://localhost:5173/login/callback",
        },
    )
    assert second_login.status_code == 200
    second_access = second_login.json()["data"]["access_token"]
    logout_response = client.post(
        "/api/v1/auth/logout",
        headers={"Authorization": f"Bearer {second_access}"},
    )
    assert logout_response.status_code == 204
    assert client.get(
        "/api/v1/me",
        headers={"Authorization": f"Bearer {second_access}"},
    ).status_code == 401


def test_public_auth_config_never_exposes_oidc_secrets(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    local = client.get("/api/v1/auth/config")
    assert local.status_code == 200
    assert local.json() == {
        "provider": "local_fake",
        "client_id": None,
        "authorization_endpoint": None,
        "end_session_endpoint": None,
        "post_logout_redirect_uri": None,
        "scopes": [],
        "code_challenge_method": "S256",
    }
    assert local.headers["cache-control"] == "no-store"

    monkeypatch.setenv("AUTH_PROFILE", "enterprise_oidc")
    monkeypatch.setattr(
        "app.routers.auth.public_oidc_config",
        lambda: {
            "provider": "enterprise_oidc",
            "client_id": "public-web-client",
            "authorization_endpoint": "https://identity.example.test/authorize",
            "end_session_endpoint": "https://identity.example.test/logout",
            "post_logout_redirect_uri": "https://app.example.test/login",
            "scopes": ["openid", "profile", "email"],
            "code_challenge_method": "S256",
        },
    )
    configured = client.get("/api/v1/auth/config")
    assert configured.status_code == 200
    assert configured.json()["client_id"] == "public-web-client"
    assert configured.json()["post_logout_redirect_uri"] == "https://app.example.test/login"
    serialized = configured.text.lower()
    assert "secret" not in serialized
    assert "token_endpoint" not in serialized
    assert "jwks" not in serialized


def test_local_password_login_uses_the_same_browser_contract(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("APP_ENV", "development")
    monkeypatch.setenv("AUTH_PROFILE", "local_fake")
    monkeypatch.setenv("LOCAL_LOGIN_ACCOUNT", "owner")
    monkeypatch.setenv("LOCAL_LOGIN_PASSWORD", "localpass123")

    with TestClient(app) as password_client:
        response = password_client.post(
            "/api/v1/auth/login",
            json={
                "grant_type": "password",
                "identifier": "owner",
                "password": "localpass123",
                "device_label": "pytest-local-password-browser",
            },
        )

    assert response.status_code == 200, response.text
    assert response.json()["data"]["context"]["membership_id"] == str(
        DEFAULT_MEMBERSHIP_ID
    )
    assert "localpass123" not in response.text
    assert "httponly" in response.headers["set-cookie"].lower()


def test_password_login_uses_keycloak_and_reuses_hardened_session(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    user_id = uuid4()
    membership_id = uuid4()
    provider_key = f"oidc:{'c' * 32}"
    subject = f"password-subject-{uuid4()}"
    email = f"password-{uuid4().hex}@example.test"
    with SessionLocal() as session:
        session.add(
            UserRow(
                id=user_id,
                email_normalized=email,
                display_name="Password User",
                identity_provider=provider_key,
                identity_subject=subject,
                status="active",
            )
        )
        session.add(
            MembershipRow(
                id=membership_id,
                tenant_id=DEFAULT_TENANT_ID,
                user_id=user_id,
                status="active",
            )
        )
        session.commit()

    captured_provider: dict[str, str] = {}
    captured_limit: dict[str, object] = {}

    def authenticate(
        _self: OidcIdentityProviderAdapter,
        *,
        identifier: str,
        password: str,
    ) -> IdentityClaim:
        captured_provider.update(identifier=identifier, password=password)
        return IdentityClaim(
            provider=provider_key,
            subject=subject,
            email_normalized=email,
            email_verified=True,
            display_name="Password User",
        )

    def capture_limit(_request: object, **kwargs: object) -> None:
        captured_limit.update(kwargs)

    monkeypatch.setenv("AUTH_PROFILE", "enterprise_oidc")
    monkeypatch.setattr(
        OidcIdentityProviderAdapter,
        "authenticate_password",
        authenticate,
    )
    monkeypatch.setattr("app.routers.auth.enforce_rate_limit", capture_limit)

    with TestClient(app) as password_client:
        response = password_client.post(
            "/api/v1/auth/login",
            json={
                "grant_type": "password",
                "identifier": "+8613800138000",
                "password": "not-stored-password",
                "device_label": "pytest-password-browser",
            },
        )

    assert response.status_code == 200, response.text
    assert response.headers["cache-control"] == "no-store"
    assert "httponly" in response.headers["set-cookie"].lower()
    assert "not-stored-password" not in response.text
    assert captured_provider == {
        "identifier": "+8613800138000",
        "password": "not-stored-password",
    }
    assert captured_limit["scope"] == "auth-password-login"
    assert captured_limit["additional_subjects"] == (
        ("account", "+8613800138000"),
    )
    assert "token" not in captured_limit
    token_data = response.json()["data"]
    assert token_data["context"]["membership_id"] == str(membership_id)
    with SessionLocal() as session:
        auth_session = session.get(AuthSessionRow, UUID(token_data["session_id"]))
        assert auth_session is not None
        assert auth_session.user_id == user_id
        assert auth_session.device_label == "pytest-password-browser"


def test_password_login_returns_only_generic_authentication_errors(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def reject_password(
        _self: OidcIdentityProviderAdapter,
        *,
        identifier: str,
        password: str,
    ) -> IdentityClaim:
        assert identifier == "missing@example.test"
        assert password == "wrong-password"
        raise IdentityProviderError("upstream account does not exist")

    monkeypatch.setenv("AUTH_PROFILE", "enterprise_oidc")
    monkeypatch.setattr(
        OidcIdentityProviderAdapter,
        "authenticate_password",
        reject_password,
    )
    with TestClient(app) as password_client:
        response = password_client.post(
            "/api/v1/auth/login",
            json={
                "grant_type": "password",
                "identifier": "missing@example.test",
                "password": "wrong-password",
            },
        )

    assert response.status_code == 401
    assert response.headers["cache-control"] == "no-store"
    assert response.json() == {
        "detail": {
            "code": "AUTH_INVALID_CREDENTIALS",
            "message": "authentication failed",
        }
    }
    serialized = response.text.lower()
    assert "missing@example.test" not in serialized
    assert "wrong-password" not in serialized
    assert "does not exist" not in serialized


def test_password_change_verifies_current_secret_and_revokes_peer_sessions(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    user_id = uuid4()
    membership_id = uuid4()
    provider_key = f"oidc:{'d' * 32}"
    subject = f"password-change-subject-{uuid4()}"
    email = f"password-change-{uuid4().hex}@example.test"
    with SessionLocal() as session:
        session.add(
            UserRow(
                id=user_id,
                email_normalized=email,
                display_name="Password Change User",
                identity_provider=provider_key,
                identity_subject=subject,
                status="active",
            )
        )
        session.add(
            MembershipRow(
                id=membership_id,
                tenant_id=DEFAULT_TENANT_ID,
                user_id=user_id,
                status="active",
            )
        )
        session.commit()

    authenticated_passwords: list[tuple[str, str]] = []
    changed_passwords: list[tuple[str, str]] = []
    limits: list[dict[str, object]] = []

    def authenticate(
        _self: OidcIdentityProviderAdapter,
        *,
        identifier: str,
        password: str,
    ) -> IdentityClaim:
        authenticated_passwords.append((identifier, password))
        return IdentityClaim(
            provider=provider_key,
            subject=subject,
            email_normalized=email,
            email_verified=True,
            display_name="Password Change User",
        )

    def update(
        _self: OidcIdentityProviderAdapter,
        *,
        subject: str,
        new_password: str,
    ) -> None:
        changed_passwords.append((subject, new_password))

    monkeypatch.setenv("AUTH_PROFILE", "enterprise_oidc")
    monkeypatch.setattr(
        OidcIdentityProviderAdapter,
        "authenticate_password",
        authenticate,
    )
    monkeypatch.setattr(
        OidcIdentityProviderAdapter,
        "change_password",
        update,
    )
    monkeypatch.setattr(
        "app.routers.auth.enforce_rate_limit",
        lambda _request, **kwargs: limits.append(kwargs),
    )

    with TestClient(app) as password_client:
        current_login = password_client.post(
            "/api/v1/auth/login",
            json={
                "grant_type": "password",
                "identifier": email,
                "password": "InitialPass!123",
            },
        )
        peer_login = password_client.post(
            "/api/v1/auth/login",
            json={
                "grant_type": "password",
                "identifier": email,
                "password": "InitialPass!123",
            },
        )
        assert current_login.status_code == 200, current_login.text
        assert peer_login.status_code == 200, peer_login.text
        current_data = current_login.json()["data"]
        peer_data = peer_login.json()["data"]
        limits.clear()
        response = password_client.put(
            "/api/v1/auth/password",
            headers={
                "Authorization": f"Bearer {current_data['access_token']}",
                "X-CSRF-Token": current_data["csrf_token"],
            },
            json={
                "current_password": "InitialPass!123",
                "new_password": "Simple42",
            },
        )

    assert response.status_code == 204, response.text
    assert response.content == b""
    assert response.headers["cache-control"] == "no-store"
    assert authenticated_passwords[-1] == (email, "InitialPass!123")
    assert changed_passwords == [(subject, "Simple42")]
    assert limits == [
        {
            "scope": "auth-password-change",
            "limit": 5,
            "window_seconds": 900,
            "token": current_data["access_token"],
        }
    ]
    with SessionLocal() as session:
        current_auth_session = session.get(
            AuthSessionRow,
            UUID(current_data["session_id"]),
        )
        peer_auth_session = session.get(
            AuthSessionRow,
            UUID(peer_data["session_id"]),
        )
        assert current_auth_session is not None
        assert current_auth_session.revoked_at is None
        assert peer_auth_session is not None
        assert peer_auth_session.revoked_at is not None
        assert peer_auth_session.revocation_reason == "PASSWORD_CHANGED"
        peer_tokens = session.scalars(
            select(AuthRefreshTokenRow).where(
                AuthRefreshTokenRow.auth_session_id == peer_auth_session.id
            )
        ).all()
        assert peer_tokens
        assert all(token.revoked_at is not None for token in peer_tokens)
        current_tokens = session.scalars(
            select(AuthRefreshTokenRow).where(
                AuthRefreshTokenRow.auth_session_id == current_auth_session.id
            )
        ).all()
        assert current_tokens
        assert all(token.revoked_at is None for token in current_tokens)


@pytest.mark.parametrize(
    ("new_password", "expected_code"),
    [
        ("short1", "PASSWORD_POLICY_VIOLATION"),
        ("12345678", "PASSWORD_POLICY_VIOLATION"),
        ("abcdefgh", "PASSWORD_POLICY_VIOLATION"),
        ("Abcd 123", "PASSWORD_POLICY_VIOLATION"),
        ("A1" + "b" * 127, "PASSWORD_POLICY_VIOLATION"),
        ("InitialPass!123", "PASSWORD_POLICY_VIOLATION"),
    ],
)
def test_password_change_rejects_weak_or_reused_password_before_provider_update(
    monkeypatch: pytest.MonkeyPatch,
    new_password: str,
    expected_code: str,
) -> None:
    user_id = uuid4()
    provider_key = f"oidc:{'e' * 32}"
    subject = f"password-policy-subject-{uuid4()}"
    email = f"password-policy-{uuid4().hex}@example.test"
    with SessionLocal() as session:
        session.add(
            UserRow(
                id=user_id,
                email_normalized=email,
                display_name="Password Policy User",
                identity_provider=provider_key,
                identity_subject=subject,
                status="active",
            )
        )
        session.add(
            MembershipRow(
                id=uuid4(),
                tenant_id=DEFAULT_TENANT_ID,
                user_id=user_id,
                status="active",
            )
        )
        session.commit()

    provider_calls: list[str] = []

    def authenticate(
        _self: OidcIdentityProviderAdapter,
        *,
        identifier: str,
        password: str,
    ) -> IdentityClaim:
        provider_calls.append(password)
        return IdentityClaim(
            provider=provider_key,
            subject=subject,
            email_normalized=email,
            email_verified=True,
        )

    monkeypatch.setenv("AUTH_PROFILE", "enterprise_oidc")
    monkeypatch.setattr(
        OidcIdentityProviderAdapter,
        "authenticate_password",
        authenticate,
    )
    monkeypatch.setattr(
        OidcIdentityProviderAdapter,
        "change_password",
        lambda *_args, **_kwargs: pytest.fail("provider update must not run"),
    )
    monkeypatch.setattr("app.routers.auth.enforce_rate_limit", lambda *_args, **_kwargs: None)

    with TestClient(app) as password_client:
        login_response = password_client.post(
            "/api/v1/auth/login",
            json={
                "grant_type": "password",
                "identifier": email,
                "password": "InitialPass!123",
            },
        )
        data = login_response.json()["data"]
        provider_calls.clear()
        response = password_client.put(
            "/api/v1/auth/password",
            headers={
                "Authorization": f"Bearer {data['access_token']}",
                "X-CSRF-Token": data["csrf_token"],
            },
            json={
                "current_password": "InitialPass!123",
                "new_password": new_password,
            },
        )

    assert response.status_code == 422
    assert response.json()["detail"]["code"] == expected_code
    assert "InitialPass!123" not in response.text
    assert new_password not in response.text
    assert provider_calls == []


def test_new_password_policy_accepts_letters_digits_and_optional_symbols() -> None:
    user = UserRow(
        id=uuid4(),
        email_normalized="merchant42@example.test",
        display_name="Merchant42",
        identity_provider=f"oidc:{'9' * 32}",
        identity_subject=f"password-policy-unit-{uuid4()}",
        status="active",
    )

    for password in ("Simple42", "ABCDEFG1", "abcdefg1", "Abcd!234"):
        _validate_new_password(
            current_password="Current1",
            new_password=password,
            user=user,
        )

    for password in (
        "short1",
        "12345678",
        "abcdefgh",
        "Abcd 123",
        "A1" + "b" * 127,
        "Current1",
        "merchant42@example.test",
        "merchant42",
        "Merchant42",
    ):
        with pytest.raises(AuthError) as exc_info:
            _validate_new_password(
                current_password="Current1",
                new_password=password,
                user=user,
            )
        assert exc_info.value.code == "PASSWORD_POLICY_VIOLATION"


def test_password_change_hides_wrong_password_and_cross_account_claims(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    user_id = uuid4()
    provider_key = f"oidc:{'f' * 32}"
    subject = f"password-current-subject-{uuid4()}"
    email = f"password-current-{uuid4().hex}@example.test"
    with SessionLocal() as session:
        session.add(
            UserRow(
                id=user_id,
                email_normalized=email,
                display_name="Current Password User",
                identity_provider=provider_key,
                identity_subject=subject,
                status="active",
            )
        )
        session.add(
            MembershipRow(
                id=uuid4(),
                tenant_id=DEFAULT_TENANT_ID,
                user_id=user_id,
                status="active",
            )
        )
        session.commit()

    def authenticate(
        _self: OidcIdentityProviderAdapter,
        *,
        identifier: str,
        password: str,
    ) -> IdentityClaim:
        if password == "InitialPass!123":
            return IdentityClaim(
                provider=provider_key,
                subject=subject,
                email_normalized=email,
                email_verified=True,
            )
        if password == "CrossAccount!123":
            return IdentityClaim(
                provider=provider_key,
                subject="another-keycloak-user",
                email_normalized="another@example.test",
                email_verified=True,
            )
        raise IdentityProviderError("upstream current password detail")

    monkeypatch.setenv("AUTH_PROFILE", "enterprise_oidc")
    monkeypatch.setattr(
        OidcIdentityProviderAdapter,
        "authenticate_password",
        authenticate,
    )
    monkeypatch.setattr(
        OidcIdentityProviderAdapter,
        "change_password",
        lambda *_args, **_kwargs: pytest.fail("provider update must not run"),
    )
    monkeypatch.setattr("app.routers.auth.enforce_rate_limit", lambda *_args, **_kwargs: None)

    with TestClient(app) as password_client:
        login_response = password_client.post(
            "/api/v1/auth/login",
            json={
                "grant_type": "password",
                "identifier": email,
                "password": "InitialPass!123",
            },
        )
        data = login_response.json()["data"]
        for candidate in ("WrongCurrent!123", "CrossAccount!123"):
            response = password_client.put(
                "/api/v1/auth/password",
                headers={
                    "Authorization": f"Bearer {data['access_token']}",
                    "X-CSRF-Token": data["csrf_token"],
                },
                json={
                    "current_password": candidate,
                    "new_password": "UpdatedPass!456",
                },
            )
            assert response.status_code == 401
            assert response.json() == {
                "detail": {
                    "code": "CURRENT_PASSWORD_INVALID",
                    "message": "current password is invalid",
                }
            }
            assert candidate not in response.text
            assert "upstream" not in response.text


def test_password_change_requires_current_session_csrf(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    user_id = uuid4()
    provider_key = f"oidc:{'1' * 32}"
    subject = f"password-csrf-subject-{uuid4()}"
    email = f"password-csrf-{uuid4().hex}@example.test"
    with SessionLocal() as session:
        session.add(
            UserRow(
                id=user_id,
                email_normalized=email,
                display_name="Password CSRF User",
                identity_provider=provider_key,
                identity_subject=subject,
                status="active",
            )
        )
        session.add(
            MembershipRow(
                id=uuid4(),
                tenant_id=DEFAULT_TENANT_ID,
                user_id=user_id,
                status="active",
            )
        )
        session.commit()

    provider_calls: list[str] = []

    def authenticate(
        _self: OidcIdentityProviderAdapter,
        *,
        identifier: str,
        password: str,
    ) -> IdentityClaim:
        provider_calls.append(password)
        return IdentityClaim(
            provider=provider_key,
            subject=subject,
            email_normalized=email,
            email_verified=True,
        )

    monkeypatch.setenv("AUTH_PROFILE", "enterprise_oidc")
    monkeypatch.setattr(
        OidcIdentityProviderAdapter,
        "authenticate_password",
        authenticate,
    )
    monkeypatch.setattr("app.routers.auth.enforce_rate_limit", lambda *_args, **_kwargs: None)

    with TestClient(app) as password_client:
        login_response = password_client.post(
            "/api/v1/auth/login",
            json={
                "grant_type": "password",
                "identifier": email,
                "password": "InitialPass!123",
            },
        )
        data = login_response.json()["data"]
        provider_calls.clear()
        response = password_client.put(
            "/api/v1/auth/password",
            headers={
                "Authorization": f"Bearer {data['access_token']}",
                "X-CSRF-Token": "incorrect-csrf-token",
            },
            json={
                "current_password": "InitialPass!123",
                "new_password": "UpdatedPass!456",
            },
        )

    assert response.status_code == 403
    assert response.json()["detail"]["code"] == "AUTH_CSRF_INVALID"
    assert provider_calls == []


def test_enterprise_oidc_binds_only_verified_pending_invitation(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    email = f"verified-{uuid4().hex}@example.test"
    invited_user_id = uuid4()
    invited_membership_id = uuid4()
    with SessionLocal() as session:
        session.add(
            UserRow(
                id=invited_user_id,
                email_normalized=email,
                display_name="Pending Owner",
                identity_provider="pending_oidc",
                identity_subject=f"pending:{invited_user_id}",
                status="invited",
            )
        )
        session.add(
            MembershipRow(
                id=invited_membership_id,
                tenant_id=DEFAULT_TENANT_ID,
                user_id=invited_user_id,
                status="invited",
            )
        )
        session.commit()

    provider_key = f"oidc:{'a' * 32}"

    def verified_exchange(
        _self: OidcIdentityProviderAdapter,
        *,
        authorization_code: str,
        code_verifier: str,
        redirect_uri: str,
        nonce: str | None = None,
    ) -> IdentityClaim:
        assert authorization_code == "one-time-code"
        assert len(code_verifier) >= 43
        assert redirect_uri == "https://app.example.test/login/callback"
        assert nonce == "N" * 43
        return IdentityClaim(
            provider=provider_key,
            subject="enterprise-subject-1",
            email_normalized=email,
            email_verified=True,
            display_name="Verified Owner",
        )

    monkeypatch.setenv("AUTH_PROFILE", "enterprise_oidc")
    monkeypatch.setattr(
        OidcIdentityProviderAdapter,
        "exchange_authorization_code",
        verified_exchange,
    )
    with TestClient(app) as oidc_client:
        response = oidc_client.post(
            "/api/v1/auth/login",
            json={
                "provider": "enterprise_oidc",
                "authorization_code": "one-time-code",
                "code_verifier": "V" * 64,
                "redirect_uri": "https://app.example.test/login/callback",
                "nonce": "N" * 43,
            },
        )
    assert response.status_code == 200, response.text
    assert response.headers["cache-control"] == "no-store"
    with SessionLocal() as session:
        user = session.get(UserRow, invited_user_id)
        membership = session.get(MembershipRow, invited_membership_id)
        assert user is not None and membership is not None
        assert (user.identity_provider, user.identity_subject, user.status) == (
            provider_key,
            "enterprise-subject-1",
            "active",
        )
        assert user.display_name == "Verified Owner"
        assert membership.status == "active"
        assert membership.joined_at is not None


def test_enterprise_oidc_rejects_unverified_email_and_jit_registration(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("AUTH_PROFILE", "enterprise_oidc")
    unknown_email = f"unknown-{uuid4().hex}@example.test"

    def claim(email_verified: bool) -> IdentityClaim:
        return IdentityClaim(
            provider=f"oidc:{'b' * 32}",
            subject=f"subject-{uuid4()}",
            email_normalized=unknown_email,
            email_verified=email_verified,
            display_name="Unknown User",
        )

    monkeypatch.setattr(
        OidcIdentityProviderAdapter,
        "exchange_authorization_code",
        lambda _self, **_kwargs: claim(False),
    )
    payload = {
        "provider": "enterprise_oidc",
        "authorization_code": "unverified",
        "code_verifier": "V" * 64,
        "redirect_uri": "https://app.example.test/login/callback",
        "nonce": "N" * 43,
    }
    with TestClient(app) as oidc_client:
        unverified = oidc_client.post("/api/v1/auth/login", json=payload)
    assert unverified.status_code == 401

    monkeypatch.setattr(
        OidcIdentityProviderAdapter,
        "exchange_authorization_code",
        lambda _self, **_kwargs: claim(True),
    )
    with TestClient(app) as oidc_client:
        no_invite = oidc_client.post("/api/v1/auth/login", json=payload)
    assert no_invite.status_code == 401
    with SessionLocal() as session:
        assert (
            session.scalar(
                select(func.count(UserRow.id)).where(
                    UserRow.email_normalized == unknown_email
                )
            )
            == 0
        )


def test_production_bootstrap_is_idempotent_and_leaves_owner_pending_oidc() -> None:
    unique = uuid4().hex[:12]
    tenant_name = f"Production Tenant {unique}"
    parameters = {
        "organization_code": f"ORG{unique.upper()}",
        "organization_name": "Production Organization",
        "tenant_slug": f"production-{unique}",
        "tenant_name": tenant_name,
        "owner_email": f"owner-{unique}@example.test",
        "owner_display_name": "Production Owner",
        "platform_admin": True,
    }
    with SessionLocal() as session:
        first = bootstrap_production_owner(session, **parameters)
        tenant = session.get(TenantRow, first.tenant_id)
        assert tenant is not None
        assert tenant.slug == storefront_slug_from_name(tenant_name)
        tenant.name = "自定义商家名"
        tenant.slug = "自定义商家名"
        profile = session.get(TenantPublicProfileRow, tenant.id)
        assert profile is not None
        profile.slug = tenant.slug
        session.commit()
    with SessionLocal() as session:
        second = bootstrap_production_owner(session, **parameters)
        user = session.get(UserRow, first.user_id)
        membership = session.get(MembershipRow, first.membership_id)
        tenant = session.get(TenantRow, first.tenant_id)
        profile = session.get(TenantPublicProfileRow, first.tenant_id)
        owner_role = session.scalar(
            select(RoleRow).where(
                RoleRow.tenant_id == first.tenant_id,
                RoleRow.code == "OWNER",
            )
        )
        assert (
            user is not None
            and membership is not None
            and owner_role is not None
            and tenant is not None
            and profile is not None
        )
        assert tenant.name == "自定义商家名"
        assert tenant.slug == "自定义商家名"
        assert profile.slug == "自定义商家名"
        assert user.identity_provider == "pending_oidc"
        assert user.status == "invited"
        assert user.is_platform_admin is True
        assert membership.status == "invited"
        assert session.scalar(
            select(func.count(MembershipRoleRow.id)).where(
                MembershipRoleRow.tenant_id == first.tenant_id,
                MembershipRoleRow.membership_id == first.membership_id,
                MembershipRoleRow.role_id == owner_role.id,
            )
        ) == 1
    assert first == second


@pytest.mark.parametrize("slug", sorted(RESERVED_TENANT_SLUGS))
def test_production_bootstrap_rejects_reserved_tenant_slug_before_writes(
    slug: str,
) -> None:
    unique = uuid4().hex[:12]
    with SessionLocal() as session:
        organization_count = session.scalar(select(func.count(OrganizationRow.id)))
        tenant_count = session.scalar(select(func.count(TenantRow.id)))
        with pytest.raises(ValueError, match="reserved by the platform"):
            bootstrap_production_owner(
                session,
                organization_code=f"RSV{unique.upper()}",
                organization_name="Must Not Be Created",
                tenant_slug=slug,
                tenant_name="Reserved Storefront",
                owner_email=f"reserved-{unique}@example.test",
                owner_display_name="Reserved Owner",
            )
        assert session.scalar(select(func.count(OrganizationRow.id))) == organization_count
        assert session.scalar(select(func.count(TenantRow.id))) == tenant_count


def test_platform_admin_rejects_every_reserved_storefront_slug() -> None:
    for slug in sorted(RESERVED_TENANT_SLUGS):
        response = client.post(
            "/api/admin/tenants",
            json={"name": f"Reserved {slug}", "slug": slug},
        )
        assert response.status_code == 422, (slug, response.text)
    with SessionLocal() as session:
        assert (
            session.scalar(
                select(func.count(TenantRow.id)).where(
                    TenantRow.slug.in_(RESERVED_TENANT_SLUGS)
                )
            )
            == 0
        )


def test_platform_admin_auto_allocates_duplicate_name_storefront_paths() -> None:
    suffix = uuid4().hex[:10]
    base_name = f"YoYo {suffix}"
    expected_base = storefront_slug_from_name(base_name)

    created = []
    for name in (base_name, f"{base_name}~", f"{base_name}!!!"):
        response = client.post(
            "/api/admin/tenants",
            json={"name": name, "active": True},
        )
        assert response.status_code == 201, response.text
        created.append(response.json())

    assert [tenant["slug"] for tenant in created] == [
        expected_base,
        f"{expected_base}-2",
        f"{expected_base}-3",
    ]
    for tenant in created:
        assert client.get(f"/api/store/{tenant['slug']}").status_code == 200


def test_platform_admin_manages_tenant_lifecycle(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    slug = f"merchant-{uuid4().hex[:10]}"
    created = client.post(
        "/api/admin/tenants",
        json={
            "name": "Platform Managed Merchant",
            "slug": slug,
            "contact_email": "ops@merchant.example",
            "active": True,
        },
    )
    assert created.status_code == 201, created.text
    tenant = created.json()
    tenant_id = tenant["id"]
    assert tenant["slug"] == slug
    assert tenant["status"] == "active"
    assert tenant["active"] is True
    assert tenant["contact_email"] == "ops@merchant.example"
    assert tenant["sku_count"] == 0 and tenant["quote_count"] == 0
    assert client.get(f"/api/store/{slug}").status_code == 200

    invited_email = f"owner-{uuid4().hex}@merchant.example"
    invitation = client.post(
        f"/api/admin/tenants/{tenant_id}/member-invitations",
        json={
            "email": invited_email.upper(),
            "display_name": "Merchant Owner",
            "role": "owner",
        },
    )
    assert invitation.status_code == 201, invitation.text
    invitation_data = invitation.json()
    assert invitation_data["email"] == invited_email
    assert invitation_data["role"] == "OWNER"
    assert invitation_data["membership_status"] == "invited"
    assert invitation_data["requires_identity_provider_provisioning"] is True
    with SessionLocal() as session:
        invited_user = session.get(UserRow, UUID(invitation_data["user_id"]))
        invited_membership = session.get(
            MembershipRow, UUID(invitation_data["membership_id"])
        )
        roles = {
            role.code: role
            for role in session.scalars(
                select(RoleRow).where(RoleRow.tenant_id == UUID(tenant_id))
            ).all()
        }
        assert set(roles) == {
            "OWNER",
            "ADMIN",
            "SALES",
            "PURCHASING",
            "VIEWER",
            "CUSTOMER_SUBACCOUNT",
        }
        assert invited_user is not None and invited_membership is not None
        assert invited_user.identity_provider == "pending_oidc"
        assert invited_user.status == "invited"
        assert invited_membership.tenant_id == UUID(tenant_id)
        assert session.scalar(
            select(func.count(MembershipRoleRow.id)).where(
                MembershipRoleRow.tenant_id == UUID(tenant_id),
                MembershipRoleRow.membership_id == invited_membership.id,
                MembershipRoleRow.role_id == roles["OWNER"].id,
            )
        ) == 1
    repeated_invitation = client.post(
        f"/api/admin/tenants/{tenant_id}/member-invitations",
        json={
            "email": invited_email,
            "display_name": "Merchant Owner",
            "role": "OWNER",
        },
    )
    assert repeated_invitation.status_code == 201
    assert repeated_invitation.json()["created"] is False
    rejected_role = client.post(
        f"/api/admin/tenants/{tenant_id}/member-invitations",
        json={
            "email": f"custom-{uuid4().hex}@merchant.example",
            "display_name": "Custom Role",
            "role": "SUPERADMIN",
        },
    )
    assert rejected_role.status_code == 422

    merchant_user_id = uuid4()
    with SessionLocal() as session:
        session.add(
            UserRow(
                id=merchant_user_id,
                email_normalized=f"{merchant_user_id.hex}@suspended-tenant.test",
                display_name="Suspended Tenant Member",
                identity_provider="local-bootstrap",
                identity_subject=str(merchant_user_id),
                status="active",
            )
        )
        session.add(
            MembershipRow(
                tenant_id=UUID(tenant_id),
                user_id=merchant_user_id,
                status="active",
            )
        )
        session.commit()

    duplicate = client.post(
        "/api/admin/tenants",
        json={"name": "Duplicate", "slug": slug},
    )
    assert duplicate.status_code == 409
    assert duplicate.json()["detail"]["code"] == "TENANT_SLUG_EXISTS"

    suspended = client.patch(
        f"/api/admin/tenants/{tenant_id}",
        json={"name": "Suspended Merchant", "active": False},
    )
    assert suspended.status_code == 200, suspended.text
    assert suspended.json()["status"] == "suspended"
    assert suspended.json()["active"] is False
    assert client.get(f"/api/store/{slug}").status_code == 404
    monkeypatch.setenv("AUTH_TEST_BYPASS", "false")
    with TestClient(app) as suspended_client:
        suspended_login = suspended_client.post(
            "/api/v1/auth/login",
            json={
                "provider": "local_fake",
                "authorization_code": f"fake:{merchant_user_id}",
                "code_verifier": "S" * 43,
                "redirect_uri": "http://127.0.0.1:5173/login/callback",
            },
        )
    assert suspended_login.status_code == 403
    assert suspended_login.json()["detail"]["code"] == "AUTH_MEMBERSHIP_REQUIRED"
    monkeypatch.setenv("AUTH_TEST_BYPASS", "true")

    reactivated = client.patch(
        f"/api/admin/tenants/{tenant_id}",
        json={"contact_email": "new@merchant.example", "active": True},
    )
    assert reactivated.status_code == 200, reactivated.text
    assert reactivated.json()["status"] == "active"
    assert reactivated.json()["contact_email"] == "new@merchant.example"
    assert client.get(f"/api/store/{slug}").status_code == 200

    rows = client.get("/api/admin/tenants")
    assert rows.status_code == 200, rows.text
    assert tenant_id in {row["id"] for row in rows.json()}

    current_tenant_guard = client.patch(
        f"/api/admin/tenants/{DEFAULT_TENANT_ID}",
        json={"active": False},
    )
    assert current_tenant_guard.status_code == 409
    assert current_tenant_guard.json()["detail"]["code"] == "ACTIVE_TENANT_SUSPENSION_FORBIDDEN"


def test_platform_admin_can_open_a_password_login_owner_for_an_existing_merchant(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    suffix = uuid4().hex[:10]
    created = client.post(
        "/api/admin/tenants",
        json={
            "name": f"Password Owner Merchant {suffix}",
            "slug": f"password-owner-{suffix}",
            "active": True,
        },
    )
    assert created.status_code == 201, created.text
    tenant = created.json()
    assert tenant["owner_account"] is None

    login_identifier = f"merchant-{suffix}"
    opened = client.post(
        f"/api/admin/tenants/{tenant['id']}/owner-account",
        json={
            "display_name": "Merchant Password Owner",
            "login_identifier": login_identifier,
            "password": f"Merchant{suffix}9",
            "email": f"merchant-{suffix}@owner.test",
        },
    )
    assert opened.status_code == 201, opened.text
    owner = opened.json()
    assert owner["login_identifier"] == login_identifier
    assert owner["status"] == "active"
    assert f"Merchant{suffix}9" not in opened.text

    listed = client.get("/api/admin/tenants")
    assert listed.status_code == 200, listed.text
    listed_tenant = next(row for row in listed.json() if row["id"] == tenant["id"])
    assert listed_tenant["owner_account"]["membership_id"] == owner["membership_id"]
    assert listed_tenant["owner_account"]["login_identifier"] == login_identifier

    repeated = client.post(
        f"/api/admin/tenants/{tenant['id']}/owner-account",
        json={
            "display_name": "Another Owner",
            "login_identifier": f"another-{suffix}",
            "password": f"Another{suffix}9",
        },
    )
    assert repeated.status_code == 409
    assert repeated.json()["detail"]["code"] == "MERCHANT_OWNER_ALREADY_CONFIGURED"

    with monkeypatch.context() as auth_environment:
        auth_environment.setenv("AUTH_TEST_BYPASS", "false")
        with TestClient(app) as owner_client:
            login = owner_client.post(
                "/api/v1/auth/login",
                json={
                    "grant_type": "password",
                    "identifier": login_identifier,
                    "password": f"Merchant{suffix}9",
                },
            )
    assert login.status_code == 200, login.text
    assert login.json()["data"]["context"]["tenant_id"] == tenant["id"]
    assert login.json()["data"]["context"]["account_scope"] == "STAFF"


def test_preprovisioned_oidc_merchant_owner_can_use_an_account_without_verified_email(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    suffix = uuid4().hex[:10]
    tenant_response = client.post(
        "/api/admin/tenants",
        json={
            "name": f"OIDC Password Merchant {suffix}",
            "slug": f"oidc-password-owner-{suffix}",
            "active": True,
        },
    )
    assert tenant_response.status_code == 201, tenant_response.text
    tenant_id = tenant_response.json()["id"]
    provider_key = f"oidc:{'o' * 32}"
    subject = f"owner-{suffix}"

    def provision(
        _self: OidcIdentityProviderAdapter,
        *,
        identifier: str,
        password: str,
        display_name: str,
        email: str | None = None,
    ) -> IdentityClaim:
        assert identifier == f"oidc-owner-{suffix}"
        assert password == f"Oidc{suffix}9"
        assert display_name == "OIDC Merchant Owner"
        assert email is None
        return IdentityClaim(
            provider=provider_key,
            subject=subject,
            email_normalized=None,
            email_verified=False,
            display_name=display_name,
        )

    def authenticate(
        _self: OidcIdentityProviderAdapter,
        *,
        identifier: str,
        password: str,
    ) -> IdentityClaim:
        assert identifier == f"oidc-owner-{suffix}"
        assert password == f"Oidc{suffix}9"
        return IdentityClaim(
            provider=provider_key,
            subject=subject,
            email_normalized=None,
            email_verified=False,
            display_name="OIDC Merchant Owner",
        )

    monkeypatch.setenv("AUTH_PROFILE", "enterprise_oidc")
    monkeypatch.setattr(OidcIdentityProviderAdapter, "provision_password_user", provision)
    monkeypatch.setattr(OidcIdentityProviderAdapter, "authenticate_password", authenticate)
    with SessionLocal() as session:
        assert session.scalar(
            select(func.count(MembershipRow.id)).where(
                MembershipRow.tenant_id == UUID(tenant_id),
                MembershipRow.login_identifier == f"oidc-owner-{suffix}",
            )
        ) == 0
    opened = client.post(
        f"/api/admin/tenants/{tenant_id}/owner-account",
        json={
            "display_name": "OIDC Merchant Owner",
            "login_identifier": f"oidc-owner-{suffix}",
            "password": f"Oidc{suffix}9",
        },
    )
    assert opened.status_code == 201, opened.text

    with monkeypatch.context() as auth_environment:
        auth_environment.setenv("AUTH_TEST_BYPASS", "false")
        with TestClient(app) as owner_client:
            login = owner_client.post(
                "/api/v1/auth/login",
                json={
                    "grant_type": "password",
                    "identifier": f"oidc-owner-{suffix}",
                    "password": f"Oidc{suffix}9",
                },
            )
    assert login.status_code == 200, login.text
    assert login.json()["data"]["context"]["tenant_id"] == tenant_id


def test_platform_admin_routes_reject_regular_members(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    user_id = uuid4()
    membership_id = uuid4()
    with SessionLocal() as session:
        session.add(
            UserRow(
                id=user_id,
                email_normalized=f"{user_id.hex}@platform-admin.test",
                display_name="Regular Tenant Member",
                identity_provider="local-bootstrap",
                identity_subject=str(user_id),
                status="active",
                is_platform_admin=False,
            )
        )
        session.add(
            MembershipRow(
                id=membership_id,
                tenant_id=DEFAULT_TENANT_ID,
                user_id=user_id,
                status="active",
            )
        )
        session.commit()

    monkeypatch.setenv("AUTH_TEST_BYPASS", "false")
    with TestClient(app) as regular_client:
        login_response = regular_client.post(
            "/api/v1/auth/login",
            json={
                "provider": "local_fake",
                "authorization_code": f"fake:{user_id}",
                "code_verifier": "R" * 43,
                "redirect_uri": "http://127.0.0.1:5173/login/callback",
            },
        )
        assert login_response.status_code == 200, login_response.text
        token = login_response.json()["data"]["access_token"]
        denied = regular_client.get(
            "/api/admin/tenants",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert denied.status_code == 403
        assert denied.json()["detail"]["code"] == "PLATFORM_ADMIN_REQUIRED"
        denied_invitation = regular_client.post(
            f"/api/admin/tenants/{DEFAULT_TENANT_ID}/member-invitations",
            headers={"Authorization": f"Bearer {token}"},
            json={
                "email": f"denied-{uuid4().hex}@example.test",
                "display_name": "Denied",
                "role": "SALES",
            },
        )
        assert denied_invitation.status_code == 403
        assert denied_invitation.json()["detail"]["code"] == "PLATFORM_ADMIN_REQUIRED"
        denied_embedding_settings = regular_client.get(
            "/api/v1/ai/embedding/settings",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert denied_embedding_settings.status_code == 403
        assert (
            denied_embedding_settings.json()["detail"]["code"]
            == "PLATFORM_ADMIN_REQUIRED"
        )
        denied_system_monitoring = regular_client.get(
            "/api/v1/system/metrics",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert denied_system_monitoring.status_code == 403
        assert (
            denied_system_monitoring.json()["detail"]["code"]
            == "PLATFORM_ADMIN_REQUIRED"
        )
        denied_translation_settings = regular_client.get(
            "/api/v1/system/translation/settings",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert denied_translation_settings.status_code == 403
        assert (
            denied_translation_settings.json()["detail"]["code"]
            == "PLATFORM_ADMIN_REQUIRED"
        )


def test_system_monitoring_reports_server_resources_without_caching() -> None:
    response = client.get("/api/v1/system/metrics")
    assert response.status_code == 200, response.text
    assert response.headers["cache-control"] == "no-store"
    payload = response.json()
    assert payload["scope"] == "SERVER_HOST"
    assert payload["cpu"]["logical_cores"] >= 1
    if payload["cpu"]["utilization_percent"] is not None:
        assert 0 <= payload["cpu"]["utilization_percent"] <= 100
    if payload["memory"]["utilization_percent"] is not None:
        assert 0 <= payload["memory"]["utilization_percent"] <= 100
    assert payload["disk"]["total_bytes"] > 0
    assert payload["disk"]["used_bytes"] >= 0
    assert payload["disk"]["available_bytes"] >= 0
    assert 0 <= payload["disk"]["utilization_percent"] <= 100


def test_member_invitation_rejects_ambiguous_global_email() -> None:
    email = f"ambiguous-{uuid4().hex}@example.test"
    with SessionLocal() as session:
        for index in range(2):
            user_id = uuid4()
            session.add(
                UserRow(
                    id=user_id,
                    email_normalized=email,
                    display_name=f"Ambiguous {index}",
                    identity_provider="pending_oidc",
                    identity_subject=f"pending:{user_id}",
                    status="invited",
                )
            )
        session.commit()
    response = client.post(
        f"/api/admin/tenants/{DEFAULT_TENANT_ID}/member-invitations",
        json={"email": email, "display_name": "Ambiguous", "role": "SALES"},
    )
    assert response.status_code == 409
    assert response.json()["detail"]["code"] == "MEMBER_EMAIL_AMBIGUOUS"


def test_ai_read_projection_and_candidate_routes_enforce_rbac(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    user_id = uuid4()
    membership_id = uuid4()
    with SessionLocal() as session:
        session.add(
            UserRow(
                id=user_id,
                email_normalized=f"{user_id.hex}@ai-rbac.test",
                display_name="AI Route No-Permission User",
                identity_provider="local-bootstrap",
                identity_subject=str(user_id),
                status="active",
                is_platform_admin=False,
            )
        )
        session.add(
            MembershipRow(
                id=membership_id,
                tenant_id=DEFAULT_TENANT_ID,
                user_id=user_id,
                status="active",
            )
        )
        session.commit()

    monkeypatch.setenv("AUTH_TEST_BYPASS", "false")
    with TestClient(app) as regular_client:
        login_response = regular_client.post(
            "/api/v1/auth/login",
            json={
                "provider": "local_fake",
                "authorization_code": f"fake:{user_id}",
                "code_verifier": "R" * 43,
                "redirect_uri": "http://127.0.0.1:5173/login/callback",
            },
        )
        assert login_response.status_code == 200, login_response.text
        headers = {
            "Authorization": f"Bearer {login_response.json()['data']['access_token']}"
        }
        denied_projection = regular_client.post(
            f"/api/v1/ai/knowledge/products/{uuid4()}/project",
            headers=headers,
        )
        denied_search = regular_client.post(
            "/api/v1/ai/search/products",
            headers=headers,
            json={"query": "restricted product", "limit": 5},
        )
        denied_candidates = regular_client.get(
            f"/api/v1/ai/product-intelligence/tasks/{uuid4()}/candidates",
            headers=headers,
        )
        denied_quote_download = regular_client.get(
            f"/api/v1/public-quote-drafts/{uuid4()}/pdf",
            headers=headers,
        )

    for response in (
        denied_projection,
        denied_search,
        denied_candidates,
        denied_quote_download,
    ):
        assert response.status_code == 403, response.text
        assert response.json()["detail"]["code"] in {
            "PERMISSION_DENIED",
            "PERMISSION_REQUIRED",
        }


def test_refresh_cookie_is_secure_in_staging(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("APP_ENV", "staging")
    response = Response()
    _set_refresh_cookie(response, "opaque-refresh-token")
    cookie = response.headers["set-cookie"].lower()
    assert "httponly" in cookie
    assert "secure" in cookie
    assert "samesite=lax" in cookie
    assert "max-age=604800" in cookie


def test_multi_tenant_session_requires_server_validated_tenant_switch(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("AUTH_TEST_BYPASS", "false")
    organization_id = uuid4()
    tenant_id = uuid4()
    membership_id = uuid4()
    with SessionLocal() as session:
        session.add(
            OrganizationRow(
                id=organization_id,
                code=f"AUTH-{str(organization_id)[:8]}",
                name="Authentication Tenant Switch Test",
            )
        )
        session.add(
            TenantRow(
                id=tenant_id,
                organization_id=organization_id,
                slug=f"auth-{str(tenant_id)[:8]}",
                name="Tenant B",
            )
        )
        session.add(
            MembershipRow(
                id=membership_id,
                tenant_id=tenant_id,
                user_id=DEFAULT_OWNER_USER_ID,
                status="active",
            )
        )
        session.commit()

    login_response = client.post(
        "/api/v1/auth/login",
        json={
            "provider": "local_fake",
            "authorization_code": f"fake:{DEFAULT_OWNER_USER_ID}",
            "code_verifier": "C" * 43,
            "redirect_uri": "http://127.0.0.1:5173/login/callback",
        },
    )
    assert login_response.status_code == 200
    login_data = login_response.json()["data"]
    assert login_data["requires_tenant_selection"] is True
    assert login_data["context"]["tenant_id"] is None
    access = login_data["access_token"]
    csrf = login_data["csrf_token"]
    auth_header = {"Authorization": f"Bearer {access}"}
    membership_options = client.get(
        "/api/v1/auth/memberships", headers=auth_header
    )
    assert membership_options.status_code == 200
    assert {row["id"] for row in membership_options.json()} == {
        str(DEFAULT_MEMBERSHIP_ID),
        str(membership_id),
    }
    assert client.get("/api/v1/me", headers=auth_header).status_code == 403

    invalid_switch = client.post(
        "/api/v1/auth/tenant-context",
        json={"membership_id": str(uuid4())},
        headers={**auth_header, "X-CSRF-Token": csrf},
    )
    assert invalid_switch.status_code == 403

    switch_response = client.post(
        "/api/v1/auth/tenant-context",
        json={"membership_id": str(membership_id)},
        headers={**auth_header, "X-CSRF-Token": csrf},
    )
    assert switch_response.status_code == 200, switch_response.text
    switched = switch_response.json()["data"]
    assert switched["context"]["tenant_id"] == str(tenant_id)
    assert client.get("/api/v1/me", headers=auth_header).status_code == 401
    stale_switch = client.post(
        "/api/v1/auth/tenant-context",
        json={"membership_id": str(DEFAULT_MEMBERSHIP_ID)},
        headers={
            "Authorization": f"Bearer {access}",
            "X-CSRF-Token": switched["csrf_token"],
        },
    )
    assert stale_switch.status_code == 401
    assert stale_switch.json()["detail"]["code"] == "AUTH_SESSION_EXPIRED"
    me_response = client.get(
        "/api/v1/me",
        headers={
            "Authorization": f"Bearer {switched['access_token']}",
            "X-Tenant-ID": str(DEFAULT_TENANT_ID),
        },
    )
    assert me_response.status_code == 200, me_response.text
    assert me_response.json()["context"]["tenant_id"] == str(tenant_id)
    # A stale switch must fail before consuming the currently locked refresh
    # token; the legitimate switched session can still rotate it afterwards.
    post_stale_refresh = client.post(
        "/api/v1/auth/refresh",
        headers={"X-CSRF-Token": switched["csrf_token"]},
    )
    assert post_stale_refresh.status_code == 200, post_stale_refresh.text


def test_phase4a1a_schema_contains_only_approved_product_intelligence_tables() -> None:
    tables = set(inspect(engine).get_table_names())
    assert {
        "organizations",
        "tenants",
        "users",
        "memberships",
        "roles",
        "permissions",
        "role_permissions",
        "membership_roles",
    }.issubset(tables)
    assert {
        "products",
        "product_images",
        "product_categories",
        "product_attributes",
        "suppliers",
        "supplier_products",
        "supplier_score",
    }.issubset(tables)
    assert {"ai_tasks", "ai_provider_routes", "ai_source_evidence"}.issubset(tables)
    assert {"knowledge_documents", "knowledge_chunks", "embeddings"}.issubset(tables)
    assert {"ai_runs", "ai_task_steps", "product_field_candidates"}.issubset(tables)
    assert {
        "product_candidate_decisions",
        "product_versions",
        "outbox_events",
        "tenant_public_profiles",
        "public_catalog_offers",
        "public_quote_drafts",
        "public_quote_draft_items",
        "public_quote_download_tokens",
        "quote_excel_templates",
        "storefront_announcements",
    }.issubset(tables)
    assert {"auth_sessions", "auth_refresh_tokens", "media_objects", "worker_jobs"}.issubset(
        tables
    )
    assert {
        "inbox_events",
        "skus",
        "attribute_definitions",
        "supplier_prices",
        "product_audit_events",
        "vision_observations",
        "image_embeddings",
        "image_searches",
        "customers",
        "inquiries",
        "inquiry_items",
        "inquiry_match_results",
        "quotations",
        "quotation_versions",
        "quotation_items",
        "quotation_approvals",
    }.issubset(tables)
    assert {
        "ai_agents",
        "agents",
        "knowledge_base",
        "knowledge_bases",
        "vector_indexes",
        "orders",
    }.isdisjoint(tables)
    assert {"suppliers", "source_files", "import_jobs", "review_items"}.issubset(tables)


def test_phase4a1a_tables_are_tenant_scoped_candidate_only_and_product_detached() -> None:
    database_inspector = inspect(engine)
    for table_name in (
        "ai_runs",
        "ai_task_steps",
        "product_field_candidates",
        "product_candidate_decisions",
        "product_versions",
        "outbox_events",
        "tenant_public_profiles",
        "public_catalog_offers",
        "public_quote_drafts",
        "public_quote_draft_items",
        "public_quote_download_tokens",
        "quote_excel_templates",
    ):
        columns = {column["name"] for column in database_inspector.get_columns(table_name)}
        assert {"tenant_id", "created_at", "updated_at", "deleted_at"}.issubset(columns)

    candidate_columns = {
        column["name"]
        for column in database_inspector.get_columns("product_field_candidates")
    }
    assert {
        "ai_task_id",
        "ai_run_id",
        "source_evidence_id",
        "raw_value",
        "normalized_value",
        "review_status",
    }.issubset(candidate_columns)
    assert {"product_id", "embedding", "vector", "published_at"}.isdisjoint(
        candidate_columns
    )
    product_columns = {
        column["name"] for column in database_inspector.get_columns("products")
    }
    assert {"ai_task_id", "candidate_id", "candidate_payload"}.isdisjoint(product_columns)
    run_checks = {
        item["sqltext"] for item in database_inspector.get_check_constraints("ai_runs")
    }
    assert any("NATIVE" in expression for expression in run_checks if expression)


def test_phase3a_tables_are_tenant_scoped_and_minimize_sensitive_payloads() -> None:
    database_inspector = inspect(engine)
    for table_name in ("ai_tasks", "ai_provider_routes", "ai_source_evidence"):
        columns = {column["name"] for column in database_inspector.get_columns(table_name)}
        assert {"tenant_id", "created_at", "updated_at", "deleted_at"}.issubset(columns)

    task_columns = {column["name"] for column in database_inspector.get_columns("ai_tasks")}
    route_columns = {
        column["name"] for column in database_inspector.get_columns("ai_provider_routes")
    }
    evidence_columns = {
        column["name"] for column in database_inspector.get_columns("ai_source_evidence")
    }
    assert {"input_ref", "input_hash", "route_snapshot"}.issubset(task_columns)
    assert {"input", "output", "prompt", "embedding", "vector"}.isdisjoint(task_columns)
    assert "credential_secret_ref" in route_columns
    assert {"api_key", "credential_value", "secret_value"}.isdisjoint(route_columns)
    assert {"raw_value_ref", "raw_value_hash", "evidence_hash"}.issubset(evidence_columns)
    assert {"raw_value", "embedding", "vector"}.isdisjoint(evidence_columns)


def test_phase3b_tables_are_tenant_scoped_traceable_and_soft_deletable() -> None:
    database_inspector = inspect(engine)
    for table_name in ("knowledge_documents", "knowledge_chunks", "embeddings"):
        columns = {column["name"] for column in database_inspector.get_columns(table_name)}
        assert {"tenant_id", "created_at", "updated_at", "deleted_at"}.issubset(columns)

    document_columns = {
        column["name"] for column in database_inspector.get_columns("knowledge_documents")
    }
    chunk_columns = {
        column["name"] for column in database_inspector.get_columns("knowledge_chunks")
    }
    embedding_columns = {
        column["name"] for column in database_inspector.get_columns("embeddings")
    }
    assert {
        "source_entity_type",
        "source_entity_id",
        "source_version",
        "schema_version",
        "field_policy_version",
        "content_hash",
    }.issubset(document_columns)
    assert {"document_id", "chunk_type", "content", "content_hash", "metadata"}.issubset(
        chunk_columns
    )
    assert {
        "entity_type",
        "entity_id",
        "model_provider",
        "model_name",
        "model_version",
        "dimensions",
        "embedding",
    }.issubset(embedding_columns)
    assert {"prompt", "completion", "agent_state", "purchase_price", "profit"}.isdisjoint(
        document_columns | chunk_columns | embedding_columns
    )


def test_phase1_5_all_saas_core_tables_have_lifecycle_timestamps() -> None:
    table_names = (
        "organizations",
        "tenants",
        "users",
        "memberships",
        "roles",
        "permissions",
        "role_permissions",
        "membership_roles",
    )
    database_inspector = inspect(engine)
    for table_name in table_names:
        columns = {column["name"] for column in database_inspector.get_columns(table_name)}
        assert {"created_at", "updated_at", "deleted_at"}.issubset(columns)


def test_phase2_core_and_legacy_import_tables_have_lifecycle_and_tenant_columns() -> None:
    table_names = (
        "product_categories",
        "products",
        "product_images",
        "product_attributes",
        "suppliers",
        "supplier_products",
        "supplier_score",
        "source_files",
        "import_jobs",
        "review_items",
    )
    database_inspector = inspect(engine)
    for table_name in table_names:
        columns = {column["name"] for column in database_inspector.get_columns(table_name)}
        assert {"tenant_id", "created_at", "updated_at", "deleted_at"}.issubset(columns)

    product_columns = {column["name"] for column in database_inspector.get_columns("products")}
    image_columns = {column["name"] for column in database_inspector.get_columns("product_images")}
    assert "supplier_id" not in product_columns
    assert "embedding" not in image_columns
    assert "vector" not in image_columns
    supplier_checks = {
        item["name"] for item in database_inspector.get_check_constraints("suppliers")
    }
    image_checks = {
        item["name"] for item in database_inspector.get_check_constraints("product_images")
    }
    assert any(name and name.endswith("status_allowed") for name in supplier_checks)
    assert any(name and name.endswith("risk_level_allowed") for name in supplier_checks)
    assert any(name and name.endswith("image_role_allowed") for name in image_checks)


def test_phase1_seed_is_idempotent_and_owner_has_all_permissions() -> None:
    with SessionLocal() as session:
        before = session.scalar(select(func.count()).select_from(PermissionRow))
        seed_saas_foundation(session)
        after = session.scalar(select(func.count()).select_from(PermissionRow))
        permissions = list_permissions(
            session, tenant_id=DEFAULT_TENANT_ID, user_id=DEFAULT_OWNER_USER_ID
        )

    assert before == after == len(PERMISSION_SEEDS)
    assert permissions == frozenset(seed.code for seed in PERMISSION_SEEDS)
    assert "system.role_manage" in permissions


def test_rbac_denies_cross_tenant_permissions_and_database_rejects_cross_tenant_assignment() -> None:
    organization_id = uuid4()
    tenant_id = uuid4()
    user_id = uuid4()
    membership_id = uuid4()
    role_id = uuid4()

    with SessionLocal() as session:
        permission = session.scalar(select(PermissionRow).where(PermissionRow.code == "product.view"))
        owner_role = session.scalar(
            select(RoleRow).where(RoleRow.tenant_id == DEFAULT_TENANT_ID, RoleRow.code == "OWNER")
        )
        assert permission is not None and owner_role is not None
        session.add_all([
            OrganizationRow(id=organization_id, code=f"ORG-{organization_id.hex[:8]}", name="Tenant B Org"),
            UserRow(
                id=user_id,
                email_normalized=f"{user_id.hex}@example.test",
                display_name="Tenant B User",
                identity_provider="test",
                identity_subject=str(user_id),
                status="active",
            ),
        ])
        session.flush()
        session.add(TenantRow(
            id=tenant_id,
            organization_id=organization_id,
            slug=f"tenant-{tenant_id.hex[:8]}",
            name="Tenant B",
        ))
        session.flush()
        session.add_all([
            MembershipRow(
                id=membership_id,
                tenant_id=tenant_id,
                user_id=user_id,
                status="active",
            ),
            RoleRow(
                id=role_id,
                tenant_id=tenant_id,
                code="VIEWER",
                name="Viewer",
                status="active",
            ),
        ])
        session.flush()
        session.add_all([
            RolePermissionRow(
                tenant_id=tenant_id,
                role_id=role_id,
                permission_id=permission.id,
            ),
            MembershipRoleRow(
                tenant_id=tenant_id,
                membership_id=membership_id,
                role_id=role_id,
                assigned_by_user_id=user_id,
            ),
        ])
        session.commit()

        assert has_permission(
            session, tenant_id=tenant_id, user_id=user_id, permission_code="product.view"
        )
        assert not has_permission(
            session, tenant_id=DEFAULT_TENANT_ID, user_id=user_id, permission_code="product.view"
        )

        session.add(MembershipRoleRow(
            tenant_id=DEFAULT_TENANT_ID,
            membership_id=membership_id,
            role_id=owner_role.id,
            assigned_by_user_id=DEFAULT_OWNER_USER_ID,
        ))
        with pytest.raises(IntegrityError):
            session.commit()
        session.rollback()

def test_custom_role_and_soft_delete_contract() -> None:
    user_id = uuid4()
    membership_id = uuid4()
    role_id = uuid4()

    with SessionLocal() as session:
        permission = session.scalar(
            select(PermissionRow).where(PermissionRow.code == "quotation.approve")
        )
        assert permission is not None
        session.add(UserRow(
            id=user_id,
            email_normalized=f"{user_id.hex}@custom-role.test",
            display_name="Custom Role User",
            identity_provider="test",
            identity_subject=str(user_id),
            status="active",
        ))
        session.add(MembershipRow(
            id=membership_id,
            tenant_id=DEFAULT_TENANT_ID,
            user_id=user_id,
            status="active",
        ))
        session.add(RoleRow(
            id=role_id,
            tenant_id=DEFAULT_TENANT_ID,
            code=f"CUSTOM_{role_id.hex[:8].upper()}",
            name="Custom Approver",
            is_system=False,
            status="active",
        ))
        session.flush()
        session.add_all([
            RolePermissionRow(
                tenant_id=DEFAULT_TENANT_ID,
                role_id=role_id,
                permission_id=permission.id,
            ),
            MembershipRoleRow(
                tenant_id=DEFAULT_TENANT_ID,
                membership_id=membership_id,
                role_id=role_id,
                assigned_by_user_id=DEFAULT_OWNER_USER_ID,
            ),
        ])
        session.commit()
        assert has_permission(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            user_id=user_id,
            permission_code="quotation.approve",
        )

        role = session.get(RoleRow, role_id)
        assert role is not None
        mark_deleted(role)
        session.commit()

    with SessionLocal() as session:
        assert session.get(RoleRow, role_id) is None
        deleted_role = session.get(
            RoleRow,
            role_id,
            execution_options={"include_deleted": True},
        )
        assert deleted_role is not None and deleted_role.deleted_at is not None
        assert not has_permission(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            user_id=user_id,
            permission_code="quotation.approve",
        )
        restore_deleted(deleted_role)
        session.commit()
        assert has_permission(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            user_id=user_id,
            permission_code="quotation.approve",
        )


def test_phase2_product_supports_typed_attributes_images_and_multiple_suppliers() -> None:
    category_id = uuid4()
    product_id = uuid4()
    with SessionLocal() as session:
        session.add(ProductCategoryRow(
            id=category_id,
            tenant_id=DEFAULT_TENANT_ID,
            code=f"PET-{category_id.hex[:8]}",
            name="Pet Supplies",
        ))
        session.add(ProductRow(
            id=product_id,
            tenant_id=DEFAULT_TENANT_ID,
            product_code=f"PROD-{product_id.hex[:8]}",
            name="Test Pet Bowl",
            category_id=category_id,
            default_unit="piece",
            status="ACTIVE",
        ))
        session.flush()
        session.add_all([
            ProductAttributeRow(
                tenant_id=DEFAULT_TENANT_ID,
                product_id=product_id,
                attribute_key="weight",
                value_number=Decimal("200"),
                unit_code="g",
                confidence=Decimal("0.9500"),
                review_status="CONFIRMED",
            ),
            ProductImageRow(
                tenant_id=DEFAULT_TENANT_ID,
                product_id=product_id,
                bucket="atc-test",
                object_key=f"products/{product_id}/main.jpg",
                original_filename="main.jpg",
                content_type="image/jpeg",
                byte_size=1024,
                sha256="a" * 64,
                width=800,
                height=800,
                image_role="MAIN",
                approval_status="APPROVED",
            ),
            SupplierProductRow(
                tenant_id=DEFAULT_TENANT_ID,
                supplier_id="SUP-001",
                product_id=product_id,
                supplier_sku="BOWL-A",
                moq=Decimal("100"),
                moq_unit="piece",
                lead_time_days=15,
                status="ACTIVE",
            ),
            SupplierProductRow(
                tenant_id=DEFAULT_TENANT_ID,
                supplier_id="SUP-002",
                product_id=product_id,
                supplier_sku="BOWL-B",
                moq=Decimal("200"),
                moq_unit="piece",
                lead_time_days=20,
                status="ACTIVE",
            ),
            SupplierScoreRow(
                tenant_id=DEFAULT_TENANT_ID,
                supplier_id="SUP-001",
                quality_score=Decimal("92.50"),
                delivery_score=None,
                overall_score=None,
                sample_size=0,
                method_version="manual-v1",
                evidence_summary={"delivery": "unknown"},
            ),
        ])
        session.commit()

        sources = session.scalars(
            select(SupplierProductRow).where(SupplierProductRow.product_id == product_id)
        ).all()
        attribute = session.scalar(
            select(ProductAttributeRow).where(ProductAttributeRow.product_id == product_id)
        )
        score = session.scalar(
            select(SupplierScoreRow).where(SupplierScoreRow.supplier_id == "SUP-001")
            .order_by(SupplierScoreRow.calculated_at.desc())
        )
        assert {source.supplier_id for source in sources} == {"SUP-001", "SUP-002"}
        assert attribute is not None and attribute.value_number == Decimal("200.000000")
        assert score is not None and score.delivery_score is None and score.sample_size == 0

        product = session.get(ProductRow, product_id)
        assert product is not None
        mark_deleted(product)
        session.commit()

    with SessionLocal() as session:
        assert session.get(ProductRow, product_id) is None
        assert session.get(
            ProductRow, product_id, execution_options={"include_deleted": True}
        ) is not None


def test_phase2_typed_attribute_rejects_multiple_value_columns() -> None:
    product_id = uuid4()
    with SessionLocal() as session:
        session.add(ProductRow(
            id=product_id,
            tenant_id=DEFAULT_TENANT_ID,
            product_code=f"INVALID-ATTR-{product_id.hex[:8]}",
            name="Invalid Attribute Test",
        ))
        session.flush()
        session.add(ProductAttributeRow(
            tenant_id=DEFAULT_TENANT_ID,
            product_id=product_id,
            attribute_key="material",
            value_text="rubber",
            value_number=Decimal("1"),
        ))
        with pytest.raises(IntegrityError):
            session.commit()
        session.rollback()


def test_phase2_composite_foreign_keys_reject_cross_tenant_links() -> None:
    organization_id = uuid4()
    tenant_id = uuid4()
    supplier_id = f"SUP-{uuid4().hex[:12].upper()}"
    product_id = uuid4()

    with SessionLocal() as session:
        session.add(OrganizationRow(
            id=organization_id,
            code=f"P2-{organization_id.hex[:8]}",
            name="Phase 2 Tenant B",
        ))
        session.flush()
        session.add(TenantRow(
            id=tenant_id,
            organization_id=organization_id,
            slug=f"phase2-{tenant_id.hex[:8]}",
            name="Phase 2 Tenant B",
        ))
        session.flush()
        session.add(SupplierRow(
            id=supplier_id,
            tenant_id=tenant_id,
            supplier_code=supplier_id,
            name="Tenant B Supplier",
        ))
        session.add(ProductRow(
            id=product_id,
            tenant_id=DEFAULT_TENANT_ID,
            product_code=f"CROSS-{product_id.hex[:8]}",
            name="Tenant A Product",
        ))
        session.commit()

        session.add(SupplierProductRow(
            tenant_id=DEFAULT_TENANT_ID,
            supplier_id=supplier_id,
            product_id=product_id,
            supplier_sku="CROSS-TENANT",
        ))
        with pytest.raises(IntegrityError):
            session.commit()
        session.rollback()


def test_phase3a_task_route_and_evidence_contract_with_soft_delete() -> None:
    route_id = uuid4()
    task_id = uuid4()
    evidence_id = uuid4()

    with SessionLocal() as session:
        session.add(AIProviderRouteRow(
            id=route_id,
            tenant_id=DEFAULT_TENANT_ID,
            route_key="document-parse-default",
            route_version=1,
            capability="DOCUMENT_PARSE",
            primary_adapter="deterministic-parser",
            fallback_adapters=[],
            routing_policy={"strategy": "primary_then_fallback"},
            retention_policy={"payload": "reference_only"},
            credential_secret_ref=None,
            status="ACTIVE",
            approved_by_membership_id=DEFAULT_MEMBERSHIP_ID,
        ))
        session.flush()
        session.add(AITaskRow(
            id=task_id,
            tenant_id=DEFAULT_TENANT_ID,
            task_type="PRODUCT_DOCUMENT_PARSE",
            task_version=1,
            business_entity_type="IMPORT_JOB",
            business_entity_id="JOB-TRACE-001",
            business_entity_version=1,
            risk_level="L1_ASSISTIVE",
            status="PENDING",
            input_schema_version=1,
            input_ref="s3://controlled-inputs/job-trace-001.json",
            input_hash="1" * 64,
            policy_snapshot={"review_required": True},
            budget_snapshot={"max_attempts": 1},
            requested_by_membership_id=DEFAULT_MEMBERSHIP_ID,
            provider_route_id=route_id,
            route_snapshot={"route_key": "document-parse-default", "route_version": 1},
            idempotency_key="PRODUCT_DOCUMENT_PARSE:JOB-TRACE-001:v1",
        ))
        session.flush()
        session.add(AISourceEvidenceRow(
            id=evidence_id,
            tenant_id=DEFAULT_TENANT_ID,
            ai_task_id=task_id,
            source_entity_type="IMPORT_JOB",
            source_entity_id="JOB-TRACE-001",
            source_version=1,
            location_type="SHEET_CELL_RANGE",
            location={"sheet": "Products", "range": "A2:D2"},
            raw_value_ref="s3://controlled-evidence/job-trace-001/a2-d2",
            raw_value_hash="2" * 64,
            normalized_value_ref="product-draft://JOB-TRACE-001/row/2",
            claim_summary="Product row fields extracted for human review",
            classification="CONFIDENTIAL",
            permission_scope={"roles": ["OWNER", "PURCHASING"]},
            parser_identifier="xlsx-native-parser",
            parser_version="1.0",
            confidence=Decimal("0.9300"),
            evidence_hash="3" * 64,
        ))
        session.commit()

        task = session.get(AITaskRow, task_id)
        evidence = session.get(AISourceEvidenceRow, evidence_id)
        assert task is not None and task.provider_route_id == route_id
        assert task.route_snapshot["route_version"] == 1
        assert evidence is not None and evidence.confidence == Decimal("0.9300")
        mark_deleted(task)
        session.commit()

    with SessionLocal() as session:
        assert session.get(AITaskRow, task_id) is None
        deleted_task = session.get(
            AITaskRow,
            task_id,
            execution_options={"include_deleted": True},
        )
        assert deleted_task is not None and deleted_task.deleted_at is not None
        assert session.get(AISourceEvidenceRow, evidence_id) is not None


def test_phase3a_constraints_enforce_idempotency_and_reference_hashes() -> None:
    route_id = uuid4()
    idempotency_key = f"constraint-test:{uuid4()}"

    with SessionLocal() as session:
        session.add(AIProviderRouteRow(
            id=route_id,
            tenant_id=DEFAULT_TENANT_ID,
            route_key=f"constraint-route-{route_id.hex}",
            route_version=1,
            capability="OCR",
            primary_adapter="fake-ocr",
            fallback_adapters=[],
            routing_policy={},
            retention_policy={},
            status="DRAFT",
        ))
        session.flush()
        session.add(AITaskRow(
            tenant_id=DEFAULT_TENANT_ID,
            task_type="OCR_DOCUMENT",
            input_hash="4" * 64,
            idempotency_key=idempotency_key,
            provider_route_id=route_id,
        ))
        session.commit()

        session.add(AITaskRow(
            tenant_id=DEFAULT_TENANT_ID,
            task_type="OCR_DOCUMENT",
            input_hash="5" * 64,
            idempotency_key=idempotency_key,
            provider_route_id=route_id,
        ))
        with pytest.raises(IntegrityError):
            session.commit()
        session.rollback()

        session.add(AISourceEvidenceRow(
            tenant_id=DEFAULT_TENANT_ID,
            source_entity_type="PRODUCT",
            source_entity_id="INVALID-RAW-REF",
            location_type="ENTITY_FIELD",
            location={"field": "name"},
            raw_value_ref="s3://controlled-evidence/missing-hash",
            raw_value_hash=None,
            evidence_hash="6" * 64,
        ))
        with pytest.raises(IntegrityError):
            session.commit()
        session.rollback()


def test_phase3a_composite_foreign_keys_reject_cross_tenant_ai_links() -> None:
    organization_id = uuid4()
    tenant_id = uuid4()
    route_id = uuid4()
    tenant_b_task_id = uuid4()

    with SessionLocal() as session:
        session.add(OrganizationRow(
            id=organization_id,
            code=f"AI-{organization_id.hex[:8]}",
            name="Phase 3A Tenant B",
        ))
        session.flush()
        session.add(TenantRow(
            id=tenant_id,
            organization_id=organization_id,
            slug=f"phase3a-{tenant_id.hex[:8]}",
            name="Phase 3A Tenant B",
        ))
        session.flush()
        session.add(AIProviderRouteRow(
            id=route_id,
            tenant_id=tenant_id,
            route_key="tenant-b-route",
            route_version=1,
            capability="DOCUMENT_PARSE",
            primary_adapter="tenant-b-parser",
            fallback_adapters=[],
            routing_policy={},
            retention_policy={},
            status="ACTIVE",
        ))
        session.commit()

        session.add(AITaskRow(
            tenant_id=DEFAULT_TENANT_ID,
            task_type="CROSS_TENANT_ROUTE",
            input_hash="7" * 64,
            idempotency_key=f"cross-route:{uuid4()}",
            provider_route_id=route_id,
        ))
        with pytest.raises(IntegrityError):
            session.commit()
        session.rollback()

        session.add(AITaskRow(
            id=tenant_b_task_id,
            tenant_id=tenant_id,
            task_type="TENANT_B_TASK",
            input_hash="8" * 64,
            idempotency_key=f"tenant-b:{uuid4()}",
            provider_route_id=route_id,
        ))
        session.commit()

        session.add(AISourceEvidenceRow(
            tenant_id=DEFAULT_TENANT_ID,
            ai_task_id=tenant_b_task_id,
            source_entity_type="PRODUCT",
            source_entity_id="CROSS-TENANT-EVIDENCE",
            location_type="ENTITY_FIELD",
            location={"field": "name"},
            evidence_hash="9" * 64,
        ))
        with pytest.raises(IntegrityError):
            session.commit()
        session.rollback()


def test_phase3b_product_projection_is_filtered_idempotent_and_versioned() -> None:
    product_id = uuid4()
    supplier_id = f"KB-{uuid4().hex[:12].upper()}"
    with SessionLocal() as session:
        session.add(SupplierRow(
            id=supplier_id,
            tenant_id=DEFAULT_TENANT_ID,
            supplier_code=supplier_id,
            name="Knowledge Supplier",
            status="ACTIVE",
        ))
        session.add(ProductRow(
            id=product_id,
            tenant_id=DEFAULT_TENANT_ID,
            product_code=f"KB-{product_id.hex[:8]}",
            name="Small Waterproof Rubber Dog Toy",
            description="Durable and non-toxic toy for dogs",
            status="ACTIVE",
            current_version=1,
        ))
        session.flush()
        session.add_all([
            ProductAttributeRow(
                tenant_id=DEFAULT_TENANT_ID,
                product_id=product_id,
                attribute_key="material",
                value_text="TPR",
                review_status="CONFIRMED",
            ),
            ProductAttributeRow(
                tenant_id=DEFAULT_TENANT_ID,
                product_id=product_id,
                attribute_key="market",
                value_text="USA, Brazil",
                review_status="CONFIRMED",
            ),
            ProductAttributeRow(
                tenant_id=DEFAULT_TENANT_ID,
                product_id=product_id,
                attribute_key="internal_ai_note",
                value_text="DO-NOT-PROJECT",
                review_status="AI_SUGGESTED",
            ),
            SupplierProductRow(
                tenant_id=DEFAULT_TENANT_ID,
                supplier_id=supplier_id,
                product_id=product_id,
                supplier_sku="DOG-TPR-10",
                moq=Decimal("100"),
                moq_unit="pcs",
                lead_time_days=15,
                status="ACTIVE",
            ),
            SupplierScoreRow(
                tenant_id=DEFAULT_TENANT_ID,
                supplier_id=supplier_id,
                overall_score=Decimal("88.00"),
                method_version="supplier-score-v1",
            ),
        ])
        session.commit()

        first = project_product_knowledge(
            session, tenant_id=DEFAULT_TENANT_ID, product_id=product_id
        )
        session.commit()
        assert first.idempotent is False
        assert first.chunks == first.embeddings
        assert first.chunks >= 3
        assert first.dimensions == 384

        document = session.get(KnowledgeDocumentRow, first.document_id)
        assert document is not None
        assert document.canonical_payload["attributes"] == [
            {"key": "market", "value": "USA, Brazil"},
            {"key": "material", "value": "TPR"},
        ]
        assert document.canonical_payload["suppliers"] == [
            {"lead_time_days": 15, "moq": "100.000000", "moq_unit": "pcs"}
        ]
        assert "DO-NOT-PROJECT" not in str(document.canonical_payload)
        chunks = session.scalars(
            select(KnowledgeChunkRow).where(KnowledgeChunkRow.document_id == document.id)
        ).all()
        projected_text = "\n".join(chunk.content for chunk in chunks)
        for private_supplier_value in (
            supplier_id,
            "Knowledge Supplier",
            "DOG-TPR-10",
            "88.00",
            "supplier-score-v1",
        ):
            assert private_supplier_value not in str(document.canonical_payload)
            assert private_supplier_value not in projected_text
        assert "moq=100.000000 pcs" in projected_text
        embeddings = session.scalars(
            select(EmbeddingRow).where(EmbeddingRow.entity_id.in_([chunk.id for chunk in chunks]))
        ).all()
        assert len(embeddings) == len(chunks)
        assert all(len(embedding.embedding) == 384 for embedding in embeddings)
        assert session.get(ProductRow, product_id).search_document_version == 1

        repeated = project_product_knowledge(
            session, tenant_id=DEFAULT_TENANT_ID, product_id=product_id
        )
        session.commit()
        assert repeated.idempotent is True
        assert repeated.document_id == first.document_id

        product = session.get(ProductRow, product_id)
        product.description = "Updated durable waterproof dog toy"
        product.current_version = 2
        session.commit()
        second = project_product_knowledge(
            session, tenant_id=DEFAULT_TENANT_ID, product_id=product_id
        )
        session.commit()
        assert second.document_id != first.document_id
        assert second.source_version == 2
        assert session.get(KnowledgeDocumentRow, first.document_id).status == "STALE"
        assert session.get(KnowledgeDocumentRow, second.document_id).status == "ACTIVE"
        assert all(embedding.status == "STALE" for embedding in embeddings)


def test_chinese_rag_tokens_and_tag_signal_avoid_single_character_noise() -> None:
    query = "支持APP的6L智能宠物喂食器"
    query_tokens = _retrieval_tokens(query)

    assert "智能" in query_tokens
    assert "宠物" in query_tokens
    assert "喂食" in query_tokens
    assert "的" not in query_tokens
    assert _score_tag_relevance(query, query_tokens, ["智能喂食", "ABS"]) >= 0.80
    assert _score_tag_relevance(query, query_tokens, ["唇彩"]) < 0.50


def test_product_knowledge_promotes_sku_tags_into_rag_overview() -> None:
    chunks = build_product_chunks({
        "product": {
            "code": "TAG-RAG-001",
            "name": "便携旅行化妆套装",
            "description": "适合旅行与礼赠场景",
        },
        "category": {"name": "彩妆套装"},
        "attributes": [],
        "suppliers": [],
        "skus": [
            {"code": "TAG-RAG-001-A", "tags": ["旅行装", "防水", "礼赠"]},
            {"code": "TAG-RAG-001-B", "tags": ["旅行装", "轻量"]},
        ],
    })

    overview = next(chunk for chunk in chunks if chunk["chunk_type"] == "OVERVIEW")
    assert "Search tags / 商品标签: 旅行装, 防水, 礼赠, 轻量" in overview["content"]
    assert overview["metadata"]["search_tags"] == ["旅行装", "防水", "礼赠", "轻量"]
    assert overview["metadata"]["field_policy_version"] == 3


def test_phase3b_projection_and_hybrid_search_api_are_testable() -> None:
    waterproof_id = uuid4()
    bowl_id = uuid4()
    with SessionLocal() as session:
        session.add_all([
            ProductRow(
                id=waterproof_id,
                tenant_id=DEFAULT_TENANT_ID,
                product_code=f"WP-{waterproof_id.hex[:8]}",
                name="CometSeries Small Waterproof Rubber Dog Toy",
                description="Durable nontoxic dog toy for outdoor play",
                status="ACTIVE",
            ),
            ProductRow(
                id=bowl_id,
                tenant_id=DEFAULT_TENANT_ID,
                product_code=f"BOWL-{bowl_id.hex[:8]}",
                name="Large Ceramic Pet Bowl",
                description="Heavy ceramic feeding bowl for cats",
                status="ACTIVE",
            ),
        ])
        session.flush()
        session.add_all([
            ProductAttributeRow(
                tenant_id=DEFAULT_TENANT_ID,
                product_id=waterproof_id,
                attribute_key="material",
                value_text="TPR rubber",
                review_status="CONFIRMED",
            ),
            ProductAttributeRow(
                tenant_id=DEFAULT_TENANT_ID,
                product_id=waterproof_id,
                attribute_key="size",
                value_text="10cm small",
                review_status="CONFIRMED",
            ),
            ProductAttributeRow(
                tenant_id=DEFAULT_TENANT_ID,
                product_id=bowl_id,
                attribute_key="material",
                value_text="ceramic",
                review_status="CONFIRMED",
            ),
        ])
        session.commit()

    projection = client.post(f"/api/v1/ai/knowledge/products/{waterproof_id}/project")
    assert projection.status_code == 200
    assert projection.json()["chunks"] == projection.json()["embeddings"]
    repeated = client.post(f"/api/v1/ai/knowledge/products/{waterproof_id}/project")
    assert repeated.status_code == 200 and repeated.json()["idempotent"] is True
    assert client.post(f"/api/v1/ai/knowledge/products/{bowl_id}/project").status_code == 200

    response = client.post(
        "/api/v1/ai/search/products",
        json={"query": "CometSeries small water resistant pet toy", "limit": 10},
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["ranking_version"] == "hybrid-product-v2"
    assert payload["model"] == {
        "provider": "local",
        "name": "atc-feature-hash",
        "version": "1",
        "dimensions": 384,
    }
    assert payload["results"][0]["product_id"] == str(waterproof_id)
    assert set(payload["results"][0]["score_breakdown"]) == {
        "keyword", "semantic", "attribute", "tag", "supplier"
    }
    breakdown = payload["results"][0]["score_breakdown"]
    expected_score = (
        0.32 * breakdown["keyword"]
        + 0.45 * breakdown["semantic"]
        + 0.06 * breakdown["attribute"]
        + 0.12 * breakdown["tag"]
        + 0.05 * breakdown["supplier"]
    )
    assert payload["results"][0]["score"] == pytest.approx(expected_score, abs=0.000002)
    assert payload["results"][0]["evidence"]
    assert payload["results"][0]["supplier_signal_status"] == "UNKNOWN"

    exact = client.post(
        "/api/v1/ai/search/products",
        json={"query": f"WP-{waterproof_id.hex[:8]}", "limit": 10},
    )
    assert exact.status_code == 200
    assert exact.json()["results"][0]["product_id"] == str(waterproof_id)
    assert exact.json()["results"][0]["score_breakdown"]["keyword"] == 1.0


def test_manual_knowledge_index_update_and_full_rebuild() -> None:
    product_id = uuid4()
    with SessionLocal() as session:
        existing_document_ids = set(
            session.scalars(
                select(KnowledgeDocumentRow.id).where(
                    KnowledgeDocumentRow.tenant_id == DEFAULT_TENANT_ID
                )
            ).all()
        )
        session.add(
            ProductRow(
                id=product_id,
                tenant_id=DEFAULT_TENANT_ID,
                product_code=f"MANUAL-INDEX-{product_id.hex[:8]}",
                name="Manual index control product",
                description="Only indexed after the merchant explicitly starts the task.",
                status="ACTIVE",
                search_document_version=0,
            )
        )
        session.commit()

    status_before = client.get("/api/v1/ai/knowledge/index")
    assert status_before.status_code == 200
    assert status_before.json()["pending_products"] >= 1

    updated = client.post("/api/v1/ai/knowledge/index/update")
    assert updated.status_code == 200
    assert updated.json()["mode"] == "INCREMENTAL"
    assert updated.json()["processed_products"] >= 1
    assert updated.json()["pending_products"] == 0

    with SessionLocal() as session:
        product = session.get(ProductRow, product_id)
        assert product is not None
        assert product.search_document_version == product.current_version

    rejected = client.post(
        "/api/v1/ai/knowledge/index/rebuild",
        json={"confirm_full_rebuild": False},
    )
    assert rejected.status_code == 422

    rebuilt = client.post(
        "/api/v1/ai/knowledge/index/rebuild",
        json={"confirm_full_rebuild": True},
    )
    assert rebuilt.status_code == 200
    assert rebuilt.json()["mode"] == "FULL_REBUILD"
    assert rebuilt.json()["processed_products"] == rebuilt.json()["total_products"]
    assert rebuilt.json()["pending_products"] == 0

    with SessionLocal() as session:
        created_document_ids = set(
            session.scalars(
                select(KnowledgeDocumentRow.id).where(
                    KnowledgeDocumentRow.tenant_id == DEFAULT_TENANT_ID,
                    KnowledgeDocumentRow.id.not_in(existing_document_ids),
                )
            ).all()
        )
        created_chunk_ids = set(
            session.scalars(
                select(KnowledgeChunkRow.id).where(
                    KnowledgeChunkRow.document_id.in_(created_document_ids)
                )
            ).all()
        )
        if created_chunk_ids:
            session.execute(
                delete(EmbeddingRow).where(
                    EmbeddingRow.entity_id.in_(created_chunk_ids)
                )
            )
        if created_document_ids:
            session.execute(
                delete(KnowledgeChunkRow).where(
                    KnowledgeChunkRow.document_id.in_(created_document_ids)
                )
            )
            session.execute(
                delete(KnowledgeDocumentRow).where(
                    KnowledgeDocumentRow.id.in_(created_document_ids)
                )
            )
        session.execute(delete(ProductRow).where(ProductRow.id == product_id))
        session.commit()


def test_uncategorized_product_is_removed_from_the_smart_index() -> None:
    category_id = uuid4()
    product_id = uuid4()
    document_id: UUID | None = None
    chunk_ids: list[UUID] = []
    try:
        with SessionLocal() as session:
            session.add(
                ProductCategoryRow(
                    id=category_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    code=f"UNCATEGORIZED-{category_id.hex[:10]}",
                    name="未分类",
                    path="未分类",
                    status="ACTIVE",
                )
            )
            session.add(
                ProductRow(
                    id=product_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_code=f"UNCATEGORIZED-{product_id.hex[:8]}",
                    name="不应进入智能索引的测试商品",
                    description="This content must be retired from semantic retrieval.",
                    status="ACTIVE",
                    current_version=1,
                    search_document_version=0,
                )
            )
            session.commit()

            projected = project_product_knowledge(
                session,
                tenant_id=DEFAULT_TENANT_ID,
                product_id=product_id,
            )
            session.commit()
            document_id = projected.document_id
            chunk_ids = list(
                session.scalars(
                    select(KnowledgeChunkRow.id).where(
                        KnowledgeChunkRow.document_id == document_id
                    )
                ).all()
            )

            product = session.get(ProductRow, product_id)
            assert product is not None
            product.category_id = category_id
            product.current_version += 1
            product.search_document_version = 0
            session.commit()

            updated = update_knowledge_index(
                session,
                tenant_id=DEFAULT_TENANT_ID,
                full_rebuild=False,
            )

            assert product_id not in indexed_product_ids(
                session,
                tenant_id=DEFAULT_TENANT_ID,
            )
            assert session.get(KnowledgeDocumentRow, document_id).status == "STALE"
            assert all(
                session.get(KnowledgeChunkRow, chunk_id).status == "STALE"
                for chunk_id in chunk_ids
            )
            assert all(
                row.status == "STALE"
                for row in session.scalars(
                    select(EmbeddingRow).where(
                        EmbeddingRow.entity_id.in_(chunk_ids)
                    )
                ).all()
            )
            assert updated.pending_products == 0
            with pytest.raises(
                KnowledgeIndexExcludedError,
                match="未分类商品不会纳入智能索引",
            ):
                project_product_knowledge(
                    session,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_id=product_id,
                )
    finally:
        with SessionLocal() as session:
            if chunk_ids:
                session.execute(
                    delete(EmbeddingRow).where(
                        EmbeddingRow.entity_id.in_(chunk_ids)
                    )
                )
                session.execute(
                    delete(KnowledgeChunkRow).where(
                        KnowledgeChunkRow.id.in_(chunk_ids)
                    )
                )
            if document_id is not None:
                session.execute(
                    delete(KnowledgeDocumentRow).where(
                        KnowledgeDocumentRow.id == document_id
                    )
                )
            session.execute(delete(ProductRow).where(ProductRow.id == product_id))
            session.execute(
                delete(ProductCategoryRow).where(
                    ProductCategoryRow.id == category_id
                )
            )
            session.commit()


def test_observable_knowledge_index_job_reports_determinate_progress() -> None:
    product_id = uuid4()
    with SessionLocal() as session:
        session.add(
            ProductRow(
                id=product_id,
                tenant_id=DEFAULT_TENANT_ID,
                product_code=f"JOB-INDEX-{product_id.hex[:8]}",
                name="Observable indexing job product",
                description="A product used to verify persisted indexing progress.",
                status="ACTIVE",
                search_document_version=0,
            )
        )
        session.commit()

    started = client.post(
        "/api/v1/ai/knowledge/index/jobs",
        json={"mode": "INCREMENTAL"},
    )
    assert started.status_code == 202
    started_payload = started.json()
    assert started_payload["status"] == "QUEUED"
    assert started_payload["processed_products"] == 0
    assert started_payload["total_products"] >= 1
    assert started_payload["progress_percent"] == 0

    latest = client.get("/api/v1/ai/knowledge/index/jobs/latest")
    assert latest.status_code == 200
    assert latest.json()["id"] == started_payload["id"]

    deadline = monotonic() + 10
    final_payload = started_payload
    while monotonic() < deadline:
        progress = client.get(
            f"/api/v1/ai/knowledge/index/jobs/{started_payload['id']}"
        )
        assert progress.status_code == 200
        final_payload = progress.json()
        if final_payload["status"] in {"SUCCEEDED", "FAILED"}:
            break
        sleep(0.05)

    assert final_payload["status"] == "SUCCEEDED"
    assert final_payload["processed_products"] == final_payload["total_products"]
    assert final_payload["progress_percent"] == 100
    assert final_payload["completed_at"]

    with SessionLocal() as session:
        session.execute(delete(ProductRow).where(ProductRow.id == product_id))
        session.execute(
            delete(KnowledgeIndexJobRow).where(
                KnowledgeIndexJobRow.id == UUID(started_payload["id"])
            )
        )
        session.commit()


def test_platform_admin_manages_encrypted_embedding_configuration() -> None:
    raw_api_key = "sk-test-only-never-return-this-9876"
    with SessionLocal() as session:
        session.execute(delete(EmbeddingProviderSettingsRow))
        session.commit()

    try:
        initial = client.get("/api/v1/ai/embedding/settings")
        assert initial.status_code == 200
        assert "api_key" not in initial.json()

        saved = client.put(
            "/api/v1/ai/embedding/settings",
            json={
                "base_url": "https://embedding.example.test/v1",
                "api_key": raw_api_key,
                "model_name": "text-embedding-test",
                "dimensions": 1024,
                "timeout_seconds": 25,
            },
        )
        assert saved.status_code == 200
        payload = saved.json()
        assert payload["source"] == "database"
        assert payload["api_key_configured"] is True
        assert payload["api_key_hint"] == "••••9876"
        assert raw_api_key not in saved.text
        assert "api_key_ciphertext" not in payload

        with SessionLocal() as session:
            row = session.get(
                EmbeddingProviderSettingsRow,
                "TEXT_EMBEDDING",
            )
            assert row is not None
            assert raw_api_key not in row.api_key_ciphertext
            assert decrypt_api_key(row.api_key_ciphertext) == raw_api_key
            original_ciphertext = row.api_key_ciphertext

        retained = client.put(
            "/api/v1/ai/embedding/settings",
            json={
                "base_url": "https://embedding.example.test/v1",
                "model_name": "text-embedding-test-v2",
                "dimensions": 1024,
                "timeout_seconds": 25,
            },
        )
        assert retained.status_code == 200
        with SessionLocal() as session:
            row = session.get(
                EmbeddingProviderSettingsRow,
                "TEXT_EMBEDDING",
            )
            assert row is not None
            assert row.api_key_ciphertext == original_ciphertext
    finally:
        with SessionLocal() as session:
            session.execute(delete(EmbeddingProviderSettingsRow))
            session.commit()


def test_platform_admin_manages_encrypted_translation_configuration(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from app.use_cases import translation_management as management_use_cases

    raw_api_key = "sk-translation-test-never-return-4321"
    with SessionLocal() as session:
        session.execute(delete(TranslationProviderSettingsRow))
        session.commit()

    class TestTranslationProvider:
        identity = TranslationIdentity(
            provider="openai-compatible",
            version="test-v1",
        )

        def translate(
            self,
            text: str,
            *,
            source_locale: str,
            target_locale: str,
        ) -> str:
            assert text == "智能宠物喂食器 SF-6L20"
            assert source_locale == "zh-CN"
            assert target_locale == "en-US"
            return "Smart pet feeder SF-6L20"

    try:
        initial = client.get("/api/v1/system/translation/settings")
        assert initial.status_code == 200, initial.text
        assert initial.headers["cache-control"] == "no-store"
        assert "api_key" not in initial.json()

        saved = client.put(
            "/api/v1/system/translation/settings",
            json={
                "enabled": True,
                "base_url": "https://translation.example.test/v1",
                "api_key": raw_api_key,
                "model_name": "translation-test-model",
                "timeout_seconds": 25,
                "max_tokens": 8192,
                "requests_per_minute": 12,
                "reasoning_effort": "low",
            },
        )
        assert saved.status_code == 200, saved.text
        payload = saved.json()
        assert payload["source"] == "database"
        assert payload["enabled"] is True
        assert payload["api_key_configured"] is True
        assert payload["api_key_hint"] == "••••4321"
        assert payload["requests_per_minute"] == 12
        assert raw_api_key not in saved.text
        assert "api_key_ciphertext" not in payload

        with SessionLocal() as session:
            row = session.get(
                TranslationProviderSettingsRow,
                "CATALOG_TRANSLATION",
            )
            assert row is not None
            assert row.api_key_ciphertext is not None
            assert row.requests_per_minute == 12
            assert raw_api_key not in row.api_key_ciphertext
            assert (
                decrypt_translation_api_key(row.api_key_ciphertext)
                == raw_api_key
            )
            original_ciphertext = row.api_key_ciphertext

        retained = client.put(
            "/api/v1/system/translation/settings",
            json={
                "enabled": True,
                "base_url": "https://translation.example.test/v1",
                "model_name": "translation-test-model-v2",
                "timeout_seconds": 20,
                "max_tokens": 16384,
                "requests_per_minute": 24,
                "reasoning_effort": "minimal",
            },
        )
        assert retained.status_code == 200, retained.text
        with SessionLocal() as session:
            row = session.get(
                TranslationProviderSettingsRow,
                "CATALOG_TRANSLATION",
            )
            assert row is not None
            assert row.api_key_ciphertext == original_ciphertext
            assert row.requests_per_minute == 24
            provider = resolved_catalog_translator(
                session,
                environment_factory=lambda: (_ for _ in ()).throw(
                    AssertionError("database settings must win")
                ),
            )
            assert provider.identity.provider == "openai-compatible"
            assert "translation-test-model-v2" in provider.identity.version

        monkeypatch.setattr(
            management_use_cases,
            "candidate_translation_provider",
            lambda *_args, **_kwargs: TestTranslationProvider(),
        )
        tested = client.post(
            "/api/v1/system/translation/settings/test",
            json={
                "base_url": "https://translation.example.test/v1",
                "model_name": "translation-test-model-v2",
                "timeout_seconds": 20,
                "max_tokens": 16384,
                "requests_per_minute": 24,
                "reasoning_effort": "minimal",
            },
        )
        assert tested.status_code == 200, tested.text
        assert tested.json()["translated_text"] == "Smart pet feeder SF-6L20"
        assert tested.json()["latency_ms"] >= 0
        assert raw_api_key not in tested.text

        disabled = client.put(
            "/api/v1/system/translation/settings",
            json={
                "enabled": False,
                "base_url": "https://translation.example.test/v1",
                "model_name": "translation-test-model-v2",
                "timeout_seconds": 20,
                "max_tokens": 16384,
                "requests_per_minute": 24,
                "reasoning_effort": "minimal",
            },
        )
        assert disabled.status_code == 200, disabled.text
        assert disabled.json()["enabled"] is False
        with SessionLocal() as session:
            assert translation_provider_is_configured(
                session,
                environment_check=lambda: True,
            ) is False
    finally:
        with SessionLocal() as session:
            session.execute(delete(TranslationProviderSettingsRow))
            session.commit()


def test_platform_admin_manages_aliyun_translation_credentials() -> None:
    access_key_id = "LTAI-test-access-id-7788"
    access_key_secret = "aliyun-secret-never-return-9900"
    with SessionLocal() as session:
        session.execute(delete(TranslationProviderSettingsRow))
        session.commit()

    try:
        saved = client.put(
            "/api/v1/system/translation/settings",
            json={
                "provider": "aliyun-alimt",
                "enabled": True,
                "base_url": "mt.cn-hangzhou.aliyuncs.com",
                "access_key_id": access_key_id,
                "api_key": access_key_secret,
                "model_name": "translate_standard",
                "region_id": "cn-hangzhou",
                "timeout_seconds": 20,
                "max_tokens": 16384,
                "requests_per_minute": 90,
                "reasoning_effort": "none",
            },
        )
        assert saved.status_code == 200, saved.text
        payload = saved.json()
        assert payload["provider"] == "aliyun-alimt"
        assert payload["region_id"] == "cn-hangzhou"
        assert payload["access_key_id_configured"] is True
        assert payload["access_key_id_hint"] == "••••7788"
        assert payload["api_key_hint"] == "••••9900"
        assert payload["requests_per_minute"] == 90
        assert access_key_id not in saved.text
        assert access_key_secret not in saved.text

        with SessionLocal() as session:
            row = session.get(
                TranslationProviderSettingsRow,
                "CATALOG_TRANSLATION",
            )
            assert row is not None
            assert row.access_key_id_ciphertext is not None
            assert row.api_key_ciphertext is not None
            assert (
                decrypt_translation_api_key(row.access_key_id_ciphertext)
                == access_key_id
            )
            assert (
                decrypt_translation_api_key(row.api_key_ciphertext)
                == access_key_secret
            )
    finally:
        with SessionLocal() as session:
            session.execute(delete(TranslationProviderSettingsRow))
            session.commit()


def test_phase3b_tenant_boundaries_apply_to_links_and_retrieval() -> None:
    organization_id = uuid4()
    tenant_b = uuid4()
    product_b = uuid4()
    with SessionLocal() as session:
        session.add(OrganizationRow(
            id=organization_id,
            code=f"KB-{organization_id.hex[:8]}",
            name="Knowledge Tenant B Organization",
        ))
        session.flush()
        session.add(TenantRow(
            id=tenant_b,
            organization_id=organization_id,
            slug=f"knowledge-{tenant_b.hex[:8]}",
            name="Knowledge Tenant B",
        ))
        session.flush()
        session.add(ProductRow(
            id=product_b,
            tenant_id=tenant_b,
            product_code=f"TENANT-B-{product_b.hex[:8]}",
            name="Tenant B Exclusive Solar Lantern",
            description="Exclusive knowledge that Tenant A must never retrieve",
            status="ACTIVE",
        ))
        session.commit()

        session.add(KnowledgeDocumentRow(
            tenant_id=DEFAULT_TENANT_ID,
            source_entity_id=product_b,
            source_version=1,
            title="Illegal cross tenant projection",
            canonical_payload={},
            content_hash="a" * 64,
            permission_scope={},
        ))
        with pytest.raises(IntegrityError):
            session.commit()
        session.rollback()

        projected = project_product_knowledge(
            session, tenant_id=tenant_b, product_id=product_b
        )
        session.commit()
        assert projected.product_id == product_b
        tenant_a_results = hybrid_product_search(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            query="Tenant B Exclusive Solar Lantern",
        )
        tenant_b_results = hybrid_product_search(
            session,
            tenant_id=tenant_b,
            query="Tenant B Exclusive Solar Lantern",
        )
        assert all(result["product_id"] != product_b for result in tenant_a_results["results"])
        assert tenant_b_results["results"][0]["product_id"] == product_b


def test_phase3b_embedding_validation_rejects_wrong_dimensions_and_nonfinite_values() -> None:
    with pytest.raises(ValueError, match="dimension"):
        validate_vectors([[0.0, 1.0]], expected_count=1, dimensions=3)
    with pytest.raises(ValueError, match="non-finite"):
        validate_vectors([[0.0, float("nan")]], expected_count=1, dimensions=2)


def test_phase4a1a_single_synthetic_xlsx_contract_idempotency_recovery_and_isolation(
    tmp_path: Path,
) -> None:
    xlsx_path = tmp_path / "synthetic-product-intelligence.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(["Product Name", "Material", "Size", "Color", "MOQ", "Packing"])
    sheet.append(["Waterproof Dog Toy", "TPR", "10cm", "Red", 100, "12 pcs/carton"])
    workbook.save(xlsx_path)
    source_hash = hashlib.sha256(xlsx_path.read_bytes()).hexdigest()

    organization_id = uuid4()
    tenant_b = uuid4()
    source_a_id = f"SRC-{uuid4().hex[:12].upper()}"
    source_b_id = f"SRC-{uuid4().hex[:12].upper()}"
    recovery_source_id = f"SRC-{uuid4().hex[:12].upper()}"
    with SessionLocal() as session:
        session.add(OrganizationRow(
            id=organization_id,
            code=f"PI-{organization_id.hex[:8]}",
            name="Product Intelligence Tenant B Organization",
        ))
        session.flush()
        session.add(TenantRow(
            id=tenant_b,
            organization_id=organization_id,
            slug=f"product-intelligence-{tenant_b.hex[:8]}",
            name="Product Intelligence Tenant B",
        ))
        session.flush()
        for source_id, tenant_id in (
            (source_a_id, DEFAULT_TENANT_ID),
            (source_b_id, tenant_b),
            (recovery_source_id, DEFAULT_TENANT_ID),
        ):
            session.add(SourceFileRow(
                id=source_id,
                tenant_id=tenant_id,
                original_filename=xlsx_path.name,
                stored_filename=f"{source_id}.xlsx",
                local_path=str(xlsx_path),
                sha256=source_hash,
                byte_size=xlsx_path.stat().st_size,
                extension=".xlsx",
                detected_type="OOXML / XLSX",
                extension_matches=True,
                parser="openpyxl",
            ))
        session.commit()

        product_count_before = int(
            session.scalar(select(func.count()).select_from(ProductRow)) or 0
        )
        embedding_count_before = int(
            session.scalar(select(func.count()).select_from(EmbeddingRow)) or 0
        )

        adapter = FakeProductParserAdapter()
        first = run_product_draft_workflow(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            source_file_id=source_a_id,
            parser=adapter,
        )
        session.commit()
        assert first.status == "NEEDS_REVIEW"
        assert first.candidate_fields == 6
        assert first.idempotent is False and first.recovered is False

        candidates = session.scalars(
            select(ProductFieldCandidateRow).where(
                ProductFieldCandidateRow.tenant_id == DEFAULT_TENANT_ID,
                ProductFieldCandidateRow.ai_task_id == first.task_id,
            )
        ).all()
        assert len(candidates) == 6
        assert {candidate.field_key for candidate in candidates} == {
            "name", "material", "specification", "color", "moq", "packing"
        }
        assert all(candidate.review_status == "AI_SUGGESTED" for candidate in candidates)
        material = next(item for item in candidates if item.field_key == "material")
        evidence = session.get(AISourceEvidenceRow, material.source_evidence_id)
        assert evidence is not None
        assert evidence.location == {"sheet": "Products", "range": "B2"}
        assert material.normalized_value == {"value": "TPR"}

        repeated = run_product_draft_workflow(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            source_file_id=source_a_id,
            parser=adapter,
        )
        session.commit()
        assert repeated.idempotent is True
        assert repeated.task_id == first.task_id
        assert repeated.run_id == first.run_id
        assert session.scalar(
            select(func.count())
            .select_from(AIRunRow)
            .where(AIRunRow.ai_task_id == first.task_id)
        ) == 1

        with pytest.raises(ProductWorkflowNotFound):
            run_product_draft_workflow(
                session,
                tenant_id=DEFAULT_TENANT_ID,
                source_file_id=source_b_id,
                parser=adapter,
            )
        tenant_b_result = run_product_draft_workflow(
            session,
            tenant_id=tenant_b,
            source_file_id=source_b_id,
            parser=adapter,
        )
        session.commit()
        assert tenant_b_result.status == "NEEDS_REVIEW"
        tenant_a_candidate_ids = set(
            session.scalars(
                select(ProductFieldCandidateRow.id).where(
                    ProductFieldCandidateRow.tenant_id == DEFAULT_TENANT_ID
                )
            ).all()
        )
        tenant_b_candidate_ids = set(
            session.scalars(
                select(ProductFieldCandidateRow.id).where(
                    ProductFieldCandidateRow.tenant_id == tenant_b
                )
            ).all()
        )
        assert tenant_a_candidate_ids.isdisjoint(tenant_b_candidate_ids)

        recovering_adapter = FakeProductParserAdapter(fail_first_attempt=True)
        failed = run_product_draft_workflow(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            source_file_id=recovery_source_id,
            parser=recovering_adapter,
        )
        session.commit()
        assert failed.status == "PARTIAL"
        assert failed.error_code == "FAKE_TRANSIENT_FAILURE"
        assert failed.candidate_fields == 0
        recovered = run_product_draft_workflow(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            source_file_id=recovery_source_id,
            parser=recovering_adapter,
        )
        session.commit()
        assert recovered.status == "NEEDS_REVIEW"
        assert recovered.recovered is True and recovered.candidate_fields == 6
        recovery_runs = session.scalars(
            select(AIRunRow)
            .where(AIRunRow.ai_task_id == recovered.task_id)
            .order_by(AIRunRow.attempt_number)
        ).all()
        recovery_step = session.scalar(
            select(AITaskStepRow).where(AITaskStepRow.ai_task_id == recovered.task_id)
        )
        assert [run.status for run in recovery_runs] == ["FAILED", "SUCCEEDED"]
        assert recovery_step is not None
        assert recovery_step.status == "SUCCEEDED" and recovery_step.attempt_count == 2

        assert session.scalar(select(func.count()).select_from(ProductRow)) == product_count_before
        assert session.scalar(select(func.count()).select_from(EmbeddingRow)) == embedding_count_before

        tenant_b_candidate = session.scalar(
            select(ProductFieldCandidateRow).where(
                ProductFieldCandidateRow.tenant_id == tenant_b
            )
        )
        assert tenant_b_candidate is not None
        session.add(ProductFieldCandidateRow(
            tenant_id=DEFAULT_TENANT_ID,
            ai_task_id=tenant_b_candidate.ai_task_id,
            ai_run_id=tenant_b_candidate.ai_run_id,
            source_evidence_id=tenant_b_candidate.source_evidence_id,
            candidate_group_key="illegal-cross-tenant",
            candidate_index=0,
            field_key="name",
            raw_value="illegal",
            normalized_value={"value": "illegal"},
            extractor_key="fake-native-product-parser",
            extractor_version="1.0",
            candidate_hash="f" * 64,
        ))
        with pytest.raises(IntegrityError):
            session.commit()
        session.rollback()


def test_postgresql_offline_migration_contains_forced_rls_policies() -> None:
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option(
        "sqlalchemy.url",
        "postgresql+psycopg://migration-only:unused@localhost/ai_trade_cloud",
    )
    output = io.StringIO()
    config.output_buffer = output
    command.upgrade(config, "head", sql=True)
    sql = output.getvalue()

    for table in (
        "organizations",
        "tenants",
        "users",
        "memberships",
        "roles",
        "role_permissions",
        "membership_roles",
        "product_categories",
        "products",
        "product_images",
        "product_attributes",
        "suppliers",
        "supplier_products",
        "supplier_score",
        "source_files",
        "import_jobs",
        "review_items",
        "ai_provider_routes",
        "ai_tasks",
        "ai_source_evidence",
        "knowledge_documents",
        "knowledge_chunks",
        "embeddings",
        "knowledge_index_jobs",
        "catalog_delete_jobs",
        "storefront_chat_conversations",
        "storefront_chat_messages",
        "catalog_language_packs",
        "ai_runs",
        "ai_task_steps",
        "product_field_candidates",
        "product_candidate_decisions",
        "product_versions",
        "outbox_events",
    ):
        assert f'ALTER TABLE "{table}" ENABLE ROW LEVEL SECURITY' in sql
        assert f'ALTER TABLE "{table}" FORCE ROW LEVEL SECURITY' in sql
    assert "app.current_tenant_id" in sql
    assert "app.current_user_id" in sql
    assert "m.deleted_at IS NULL" in sql
    assert "JSONB" in sql
    assert "CREATE EXTENSION IF NOT EXISTS vector" in sql
    assert "VECTOR" in sql
    assert "vector_dims(embedding) = dimensions" in sql
    assert "USING hnsw" in sql
    assert "vector_cosine_ops" in sql
    assert "to_tsvector('simple', content)" in sql


def test_phase3a_migration_upgrades_and_downgrades_as_an_isolated_batch(tmp_path: Path) -> None:
    database_path = tmp_path / "phase3a-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)

    command.upgrade(config, "20260718_0006")
    phase2_engine = create_engine(migration_url)
    with phase2_engine.connect() as connection:
        assert {
            "ai_provider_routes", "ai_tasks", "ai_source_evidence"
        }.isdisjoint(inspect(connection).get_table_names())
    phase2_engine.dispose()

    command.upgrade(config, "20260718_0007")
    phase3a_engine = create_engine(migration_url)
    with phase3a_engine.connect() as connection:
        assert {
            "ai_provider_routes", "ai_tasks", "ai_source_evidence"
        }.issubset(inspect(connection).get_table_names())
        assert {
            "knowledge_documents", "knowledge_chunks", "embeddings"
        }.isdisjoint(inspect(connection).get_table_names())
    phase3a_engine.dispose()

    command.downgrade(config, "20260718_0006")
    downgraded_engine = create_engine(migration_url)
    with downgraded_engine.connect() as connection:
        assert {
            "ai_provider_routes", "ai_tasks", "ai_source_evidence"
        }.isdisjoint(inspect(connection).get_table_names())
        assert "suppliers" in inspect(connection).get_table_names()
    downgraded_engine.dispose()

    command.upgrade(config, "head")
    command.check(config)


def test_phase3b_migration_upgrades_and_downgrades_as_an_isolated_batch(tmp_path: Path) -> None:
    database_path = tmp_path / "phase3b-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)

    command.upgrade(config, "20260718_0007")
    phase3a_engine = create_engine(migration_url)
    with phase3a_engine.connect() as connection:
        assert {
            "knowledge_documents", "knowledge_chunks", "embeddings"
        }.isdisjoint(inspect(connection).get_table_names())
        assert "ai_tasks" in inspect(connection).get_table_names()
    phase3a_engine.dispose()

    command.upgrade(config, "head")
    phase3b_engine = create_engine(migration_url)
    with phase3b_engine.connect() as connection:
        assert {
            "knowledge_documents", "knowledge_chunks", "embeddings"
        }.issubset(inspect(connection).get_table_names())
    phase3b_engine.dispose()

    command.downgrade(config, "20260718_0007")
    downgraded_engine = create_engine(migration_url)
    with downgraded_engine.connect() as connection:
        assert {
            "knowledge_documents", "knowledge_chunks", "embeddings"
        }.isdisjoint(inspect(connection).get_table_names())
        assert "ai_tasks" in inspect(connection).get_table_names()
        assert "products" in inspect(connection).get_table_names()
    downgraded_engine.dispose()

    command.upgrade(config, "head")
    command.check(config)


def test_phase4a1a_migration_upgrades_and_downgrades_as_an_isolated_batch(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "phase4a1a-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)

    command.upgrade(config, "20260718_0008")
    phase3b_engine = create_engine(migration_url)
    with phase3b_engine.connect() as connection:
        assert {"ai_runs", "ai_task_steps", "product_field_candidates"}.isdisjoint(
            inspect(connection).get_table_names()
        )
        assert "embeddings" in inspect(connection).get_table_names()
    phase3b_engine.dispose()

    command.upgrade(config, "head")
    phase4a_engine = create_engine(migration_url)
    with phase4a_engine.connect() as connection:
        assert {"ai_runs", "ai_task_steps", "product_field_candidates"}.issubset(
            inspect(connection).get_table_names()
        )
    phase4a_engine.dispose()

    command.downgrade(config, "20260718_0008")
    downgraded_engine = create_engine(migration_url)
    with downgraded_engine.connect() as connection:
        assert {"ai_runs", "ai_task_steps", "product_field_candidates"}.isdisjoint(
            inspect(connection).get_table_names()
        )
        assert "ai_tasks" in inspect(connection).get_table_names()
        assert "embeddings" in inspect(connection).get_table_names()
    downgraded_engine.dispose()

    command.upgrade(config, "head")
    command.check(config)


def test_phase4a1b_migration_registers_native_provider_and_downgrades(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "phase4a1b-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)

    command.upgrade(config, "20260718_0009")
    phase4a1a_engine = create_engine(migration_url)
    with phase4a1a_engine.connect() as connection:
        checks = inspect(connection).get_check_constraints("ai_runs")
        assert not any("NATIVE" in (item["sqltext"] or "") for item in checks)
    phase4a1a_engine.dispose()

    command.upgrade(config, "head")
    phase4a1b_engine = create_engine(migration_url)
    with phase4a1b_engine.connect() as connection:
        checks = inspect(connection).get_check_constraints("ai_runs")
        assert any("NATIVE" in (item["sqltext"] or "") for item in checks)
    phase4a1b_engine.dispose()

    command.downgrade(config, "20260718_0009")
    downgraded_engine = create_engine(migration_url)
    with downgraded_engine.connect() as connection:
        checks = inspect(connection).get_check_constraints("ai_runs")
        assert not any("NATIVE" in (item["sqltext"] or "") for item in checks)
        assert "product_field_candidates" in inspect(connection).get_table_names()
    downgraded_engine.dispose()

    command.upgrade(config, "head")
    command.check(config)


def test_phase4a1c_migration_adds_adoption_version_and_outbox_contracts(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "phase4a1c-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)

    command.upgrade(config, "20260718_0010")
    phase4a1b_engine = create_engine(migration_url)
    with phase4a1b_engine.connect() as connection:
        tables = set(inspect(connection).get_table_names())
        assert {
            "product_candidate_decisions",
            "product_versions",
            "outbox_events",
        }.isdisjoint(tables)
        candidate_columns = {
            column["name"]
            for column in inspect(connection).get_columns("product_field_candidates")
        }
        assert "normalization_rule_version" not in candidate_columns
    phase4a1b_engine.dispose()

    command.upgrade(config, "head")
    phase4a1c_engine = create_engine(migration_url)
    with phase4a1c_engine.connect() as connection:
        tables = set(inspect(connection).get_table_names())
        assert {
            "product_candidate_decisions",
            "product_versions",
            "outbox_events",
        }.issubset(tables)
        candidate_columns = {
            column["name"]
            for column in inspect(connection).get_columns("product_field_candidates")
        }
        assert {"normalization_rule_version", "normalization_trace"}.issubset(
            candidate_columns
        )
    phase4a1c_engine.dispose()

    command.downgrade(config, "20260718_0010")
    downgraded_engine = create_engine(migration_url)
    with downgraded_engine.connect() as connection:
        tables = set(inspect(connection).get_table_names())
        assert {
            "product_candidate_decisions",
            "product_versions",
            "outbox_events",
        }.isdisjoint(tables)
        candidate_columns = {
            column["name"]
            for column in inspect(connection).get_columns("product_field_candidates")
        }
        assert {"normalization_rule_version", "normalization_trace"}.isdisjoint(
            candidate_columns
        )
    downgraded_engine.dispose()

    command.upgrade(config, "head")
    command.check(config)


def test_acg007_migration_adds_quarantine_metadata_and_durable_worker(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "acg007-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)

    command.upgrade(config, "20260718_0012")
    before_engine = create_engine(migration_url)
    with before_engine.connect() as connection:
        tables = set(inspect(connection).get_table_names())
        assert {"media_objects", "worker_jobs"}.isdisjoint(tables)
        source_columns = {
            column["name"] for column in inspect(connection).get_columns("source_files")
        }
        assert {"media_object_id", "security_status"}.isdisjoint(source_columns)
    before_engine.dispose()

    command.upgrade(config, "head")
    upgraded_engine = create_engine(migration_url)
    with upgraded_engine.connect() as connection:
        tables = set(inspect(connection).get_table_names())
        assert {"media_objects", "worker_jobs"}.issubset(tables)
        source_columns = {
            column["name"] for column in inspect(connection).get_columns("source_files")
        }
        assert {"media_object_id", "security_status"}.issubset(source_columns)
        worker_indexes = {
            index["name"] for index in inspect(connection).get_indexes("worker_jobs")
        }
        assert "ix_worker_jobs_tenant_claim" in worker_indexes
    upgraded_engine.dispose()

    command.downgrade(config, "20260718_0012")
    downgraded_engine = create_engine(migration_url)
    with downgraded_engine.connect() as connection:
        tables = set(inspect(connection).get_table_names())
        assert {"media_objects", "worker_jobs"}.isdisjoint(tables)
    downgraded_engine.dispose()

    command.upgrade(config, "head")
    command.check(config)


def test_phase1_5_migration_backfills_nonempty_identity_tables(tmp_path: Path) -> None:
    database_path = tmp_path / "phase1-nonempty.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)
    command.upgrade(config, "20260718_0001")

    populated_engine = create_engine(migration_url)
    permission_id = uuid4().hex
    user_id = uuid4().hex
    with populated_engine.begin() as connection:
        connection.exec_driver_sql(
            "INSERT INTO permissions (id, code, module, action, description, created_at) "
            "VALUES (?, 'test.permission', 'test', 'permission', NULL, CURRENT_TIMESTAMP)",
            (permission_id,),
        )
        connection.exec_driver_sql(
            "INSERT INTO users "
            "(id, email, display_name, password_hash, identity_provider, identity_subject, locale, "
            "status, is_platform_admin, last_login_at, created_at, updated_at) "
            "VALUES (?, 'same@example.test', 'Existing User', NULL, 'test', 'existing-subject', "
            "'zh-CN', 'active', 0, NULL, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)",
            (user_id,),
        )
    populated_engine.dispose()

    command.upgrade(config, "head")
    migrated_engine = create_engine(migration_url)
    with migrated_engine.connect() as connection:
        permission = connection.exec_driver_sql(
            "SELECT updated_at, deleted_at FROM permissions WHERE id = ?", (permission_id,)
        ).one()
        user_columns = {column["name"] for column in inspect(connection).get_columns("users")}
        email = connection.exec_driver_sql(
            "SELECT email_normalized FROM users WHERE id = ?", (user_id,)
        ).scalar_one()
        assert permission.updated_at is not None and permission.deleted_at is None
        assert email == "same@example.test"
        assert "password_hash" not in user_columns and "email" not in user_columns
    migrated_engine.dispose()


def test_existing_mvp_schema_is_preserved_across_phase2_upgrade_and_downgrade(tmp_path: Path) -> None:
    database_path = tmp_path / "existing-mvp.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)
    command.upgrade(config, "20260718_0000")

    legacy_engine = create_engine(migration_url)
    with legacy_engine.begin() as connection:
        connection.exec_driver_sql(
            "INSERT INTO suppliers "
            "(id, name, category, status, active_skus, health, created_at, updated_at) "
            "VALUES ('SUP-KEEP', 'Existing Supplier', 'Legacy', 'active', 12, 'good', "
            "CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)"
        )
    legacy_engine.dispose()

    command.upgrade(config, "head")

    migrated_engine = create_engine(migration_url)
    with migrated_engine.connect() as connection:
        assert connection.scalar(select(func.count()).select_from(SupplierRow.__table__)) == 1
        assert "organizations" in inspect(connection).get_table_names()
        migrated_supplier = connection.exec_driver_sql(
            "SELECT supplier_code, status, tenant_id FROM suppliers WHERE id = 'SUP-KEEP'"
        ).one()
        assert migrated_supplier.supplier_code == "SUP-KEEP"
        assert migrated_supplier.status == "ACTIVE"
        assert migrated_supplier.tenant_id is not None
    migrated_engine.dispose()

    command.downgrade(config, "20260718_0000")
    downgraded_engine = create_engine(migration_url)
    with downgraded_engine.connect() as connection:
        tables = set(inspect(connection).get_table_names())
        assert "suppliers" in tables
        assert "organizations" not in tables
        assert connection.scalar(select(func.count()).select_from(SupplierRow.__table__)) == 1
    downgraded_engine.dispose()

    command.upgrade(config, "head")
    command.check(config)


def test_detects_mislabeled_legacy_xls() -> None:
    result = detect_file_type("supplier_quote.xlsx", OLE_SIGNATURE + b"\x00" * 24)
    assert result.detected_type == "OLE / Legacy XLS"
    assert result.parser == "xlrd"
    assert result.extension_matches is False
    assert result.warning is not None


def test_detects_native_csv_and_rejects_csv_extension_masquerade() -> None:
    csv_result = detect_file_type("supplier.csv", b"Product Name,Material,MOQ\r\n")
    assert csv_result.detected_type == "TEXT / CSV"
    assert csv_result.parser == "python-csv"
    assert csv_result.extension_matches is True

    disguised = detect_file_type("supplier.csv", b"PK\x03\x04" + b"\x00" * 28)
    assert disguised.detected_type == "ZIP / OOXML"
    assert disguised.extension_matches is False
    assert disguised.warning is not None


def test_deterministic_margin_calculation() -> None:
    result = calculate_price(PriceCalculationRequest(
        purchase_price=Decimal("72"),
        quantity=500,
        target_margin_rate=Decimal("0.32"),
        currency="CNY",
    ))
    assert result.suggested_unit_price == Decimal("105.88")
    assert result.total_cost == Decimal("36000.00")
    assert result.quotation_total == Decimal("52940.00")
    assert result.gross_profit == Decimal("16940.00")


def test_product_query_and_image_filter() -> None:
    response = client.get("/api/v1/products", params={"q": "围栏", "approved_images_only": True})
    assert response.status_code == 200
    assert [row["model"] for row in response.json()] == ["PF-8G01"]


def test_sku_first_listing_is_paginated_filterable_and_tenant_scoped() -> None:
    suffix = uuid4().hex[:8].upper()
    product_id = uuid4()
    product_image_id = uuid4()
    first_sku_id = uuid4()
    second_sku_id = uuid4()
    other_tenant_id = uuid4()
    with SessionLocal() as session:
        category = ProductCategoryRow(
            tenant_id=DEFAULT_TENANT_ID,
            code=f"SKU-LIST-{suffix}",
            name=f"SKU List Category {suffix}",
            path=f"SKU-LIST-{suffix}",
            status="ACTIVE",
        )
        session.add(category)
        session.flush()
        product = ProductRow(
            id=product_id,
            tenant_id=DEFAULT_TENANT_ID,
            product_code=f"PRODUCT-{suffix}",
            name=f"SKU List Product {suffix}",
            category_id=category.id,
            status="ACTIVE",
        )
        session.add(product)
        session.flush()
        session.add_all(
            [
                SkuRow(
                    id=first_sku_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_id=product_id,
                    sku_code=f"{suffix}-ACTIVE",
                    name=f"Active SKU {suffix}",
                    option_values={},
                    default_moq=Decimal("12"),
                    moq_unit="piece",
                    status="ACTIVE",
                ),
                SkuRow(
                    id=second_sku_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_id=product_id,
                    sku_code=f"{suffix}-DRAFT",
                    name=f"Draft SKU {suffix}",
                    option_values={},
                    default_moq=Decimal("24"),
                    moq_unit="piece",
                    status="DRAFT",
                ),
            ]
        )
        session.add(
            ProductImageRow(
                id=product_image_id,
                tenant_id=DEFAULT_TENANT_ID,
                product_id=product_id,
                storage_provider="TEST",
                bucket="test",
                object_key=f"tests/{suffix}/main.jpg",
                content_type="image/jpeg",
                byte_size=128,
                sha256="c" * 64,
                image_role="MAIN",
                approval_status="APPROVED",
            )
        )
        session.flush()
        supplier = session.get(SupplierRow, "SUP-001")
        assert supplier is not None
        session.add(
            SupplierProductRow(
                tenant_id=DEFAULT_TENANT_ID,
                supplier_id=supplier.id,
                product_id=product_id,
                sku_id=first_sku_id,
                supplier_sku=f"SUP-{suffix}",
                supplier_product_name=f"Supplier SKU {suffix}",
                moq=Decimal("12"),
                moq_unit="piece",
                status="ACTIVE",
            )
        )
        session.add(
            PublicCatalogOfferRow(
                tenant_id=DEFAULT_TENANT_ID,
                sku_id=first_sku_id,
                unit_price=Decimal("123.45"),
                currency="USD",
                tags=["outdoor", "featured"],
                publication_status="DRAFT",
            )
        )

        other_category = ProductCategoryRow(
            tenant_id=other_tenant_id,
            code=f"OTHER-{suffix}",
            name=f"Other Category {suffix}",
            path=f"OTHER-{suffix}",
            status="ACTIVE",
        )
        session.add(
            TenantRow(
                id=other_tenant_id,
                organization_id=DEFAULT_ORGANIZATION_ID,
                slug=f"sku-list-{uuid4().hex[:8]}",
                name=f"Other SKU Tenant {suffix}",
            )
        )
        session.flush()
        session.add(other_category)
        session.flush()
        other_product = ProductRow(
            tenant_id=other_tenant_id,
            product_code=f"OTHER-PRODUCT-{suffix}",
            name=f"Other SKU List Product {suffix}",
            category_id=other_category.id,
            status="ACTIVE",
        )
        session.add(other_product)
        session.flush()
        session.add(
            SkuRow(
                tenant_id=other_tenant_id,
                product_id=other_product.id,
                sku_code=f"{suffix}-MUST-NOT-LEAK",
                name=f"Other Tenant SKU {suffix}",
                option_values={},
                status="ACTIVE",
            )
        )
        session.commit()

    first_page = client.get(
        "/api/v1/product-center/skus",
        params={"q": suffix, "page": 1, "page_size": 1},
    )
    assert first_page.status_code == 200, first_page.text
    assert first_page.json()["page"] == 1
    assert first_page.json()["page_size"] == 1
    assert first_page.json()["total"] == 2
    assert first_page.json()["pages"] == 2
    assert len(first_page.json()["items"]) == 1

    second_page = client.get(
        "/api/v1/product-center/skus",
        params={"q": suffix, "page": 2, "page_size": 1},
    )
    assert second_page.status_code == 200, second_page.text
    returned_codes = {
        first_page.json()["items"][0]["sku_code"],
        second_page.json()["items"][0]["sku_code"],
    }
    assert returned_codes == {f"{suffix}-ACTIVE", f"{suffix}-DRAFT"}
    assert f"{suffix}-MUST-NOT-LEAK" not in returned_codes

    complete = client.get("/api/v1/product-center/skus", params={"q": suffix})
    assert complete.status_code == 200, complete.text
    assert complete.json()["page_size"] == 50
    by_code = {item["sku_code"]: item for item in complete.json()["items"]}
    active = by_code[f"{suffix}-ACTIVE"]
    assert active["name"] == f"Active SKU {suffix}"
    assert active["product_id"] == str(product_id)
    assert active["product_code"] == f"PRODUCT-{suffix}"
    assert active["product_name"] == f"SKU List Product {suffix}"
    assert active["category"]["id"] == str(category.id)
    assert active["category"]["code"] == f"SKU-LIST-{suffix}"
    assert active["tags"] == ["outdoor", "featured"]
    assert active["supplier_summary"]["count"] == 1
    assert active["supplier_summary"]["primary_supplier_id"] == "SUP-001"
    assert active["supplier_summary"]["names"] == [supplier.name]
    assert Decimal(str(active["default_moq"])) == Decimal("12")
    assert active["moq_unit"] == "piece"
    assert Decimal(str(active["public_price"])) == Decimal("123.45")
    assert active["public_currency"] == "USD"
    assert active["public_offer_status"] == "DRAFT"
    assert active["status"] == "ACTIVE"
    assert active["version"] == 1
    assert active["updated_at"]
    assert active["source_type"] == "MANUAL"
    assert active["source_filename"] is None
    assert active["source_imported_at"] is None
    assert active["image_status"] == "APPROVED"
    assert active["thumbnail_url"] == f"/api/store/demo/media/{product_image_id}"

    no_missing_images = client.get(
        "/api/v1/product-center/skus",
        params={"q": suffix, "missing_images_only": "true"},
    )
    assert no_missing_images.status_code == 200
    assert no_missing_images.json()["total"] == 0

    without_suppliers = client.get(
        "/api/v1/product-center/skus",
        params={
            "q": suffix,
            "include_supplier_summary": "false",
        },
    )
    assert without_suppliers.status_code == 200
    without_supplier_by_code = {
        item["sku_code"]: item for item in without_suppliers.json()["items"]
    }
    assert without_supplier_by_code[f"{suffix}-ACTIVE"]["supplier_summary"] == {
        "count": 0,
        "primary_supplier_id": None,
        "primary_supplier_name": None,
        "names": [],
    }

    category_filtered = client.get(
        "/api/v1/product-center/skus",
        params={"q": suffix, "category_id": str(category.id)},
    )
    assert category_filtered.status_code == 200
    assert category_filtered.json()["total"] == 2

    status_filtered = client.get(
        "/api/v1/product-center/skus",
        params={"q": suffix, "status": "active"},
    )
    assert status_filtered.status_code == 200
    assert [item["sku_code"] for item in status_filtered.json()["items"]] == [
        f"{suffix}-ACTIVE"
    ]

    invalid_status = client.get(
        "/api/v1/product-center/skus",
        params={"status": "UNKNOWN"},
    )
    assert invalid_status.status_code == 422
    assert invalid_status.json()["detail"]["code"] == "SKU_STATUS_INVALID"

    with SessionLocal() as session:
        with pytest.raises(ApplicationError) as denied:
            list_authoritative_skus(
                session,
                tenant_id=DEFAULT_TENANT_ID,
                permissions=frozenset(),
                query=suffix,
                category_id=None,
                statuses=[],
                page=1,
                page_size=50,
            )
        assert denied.value.code == "PERMISSION_REQUIRED"

    with SessionLocal() as session:
        image = session.get(ProductImageRow, product_image_id)
        assert image is not None
        mark_deleted(image)
        session.commit()
    try:
        missing_images = client.get(
            "/api/v1/product-center/skus",
            params={"q": suffix, "missing_images_only": "true"},
        )
        assert missing_images.status_code == 200
        assert missing_images.json()["total"] == 2
        assert {
            item["sku_code"] for item in missing_images.json()["items"]
        } == {f"{suffix}-ACTIVE", f"{suffix}-DRAFT"}
        assert all(
            item["image_status"] == "NONE"
            for item in missing_images.json()["items"]
        )
        assert all(
            item["thumbnail_url"] is None
            for item in missing_images.json()["items"]
        )
    finally:
        with SessionLocal() as session:
            image = session.scalar(
                select(ProductImageRow)
                .where(ProductImageRow.id == product_image_id)
                .execution_options(include_deleted=True)
            )
            assert image is not None
            restore_deleted(image)
            session.commit()


def test_batch_delete_skus_hides_catalog_rows_and_preserves_history() -> None:
    suffix = uuid4().hex[:8].upper()
    product_id = uuid4()
    first_sku_id = uuid4()
    second_sku_id = uuid4()
    missing_sku_id = uuid4()
    with SessionLocal() as session:
        category = ProductCategoryRow(
            tenant_id=DEFAULT_TENANT_ID,
            code=f"BATCH-DELETE-{suffix}",
            name=f"Batch Delete {suffix}",
            path=f"BATCH-DELETE-{suffix}",
            status="ACTIVE",
        )
        session.add(category)
        session.flush()
        session.add(
            ProductRow(
                id=product_id,
                tenant_id=DEFAULT_TENANT_ID,
                product_code=f"BATCH-PRODUCT-{suffix}",
                name=f"Batch Product {suffix}",
                category_id=category.id,
                status="ACTIVE",
            )
        )
        session.flush()
        session.add_all(
            [
                SkuRow(
                    id=first_sku_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_id=product_id,
                    sku_code=f"BATCH-{suffix}-A",
                    name=f"Batch SKU A {suffix}",
                    option_values={},
                    status="ACTIVE",
                ),
                SkuRow(
                    id=second_sku_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_id=product_id,
                    sku_code=f"BATCH-{suffix}-B",
                    name=f"Batch SKU B {suffix}",
                    option_values={},
                    status="ACTIVE",
                ),
            ]
        )
        session.flush()
        session.add_all(
            [
                PublicCatalogOfferRow(
                    tenant_id=DEFAULT_TENANT_ID,
                    sku_id=first_sku_id,
                    unit_price=Decimal("10"),
                    currency="CNY",
                    tags=[],
                    publication_status="PUBLISHED",
                ),
                PublicCatalogOfferRow(
                    tenant_id=DEFAULT_TENANT_ID,
                    sku_id=second_sku_id,
                    unit_price=Decimal("20"),
                    currency="CNY",
                    tags=[],
                    publication_status="PUBLISHED",
                ),
            ]
        )
        session.commit()

    first_delete = client.post(
        "/api/v1/skus/batch-delete",
        json={"sku_ids": [str(first_sku_id), str(missing_sku_id)]},
    )
    assert first_delete.status_code == 200, first_delete.text
    assert first_delete.json()["success_count"] == 1
    assert first_delete.json()["failed_count"] == 1
    assert first_delete.json()["total_count"] == 2
    assert first_delete.json()["failed_items"] == [
        {
            "sku_id": str(missing_sku_id),
            "reason": "SKU 不存在或已经删除",
        }
    ]

    listing = client.get(
        "/api/v1/product-center/skus",
        params={"q": suffix},
    )
    assert listing.status_code == 200, listing.text
    assert [item["id"] for item in listing.json()["items"]] == [str(second_sku_id)]

    with SessionLocal() as session:
        deleted_sku = session.scalar(
            select(SkuRow)
            .where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.id == first_sku_id,
            )
            .execution_options(include_deleted=True)
        )
        product = session.get(ProductRow, product_id)
        offer = session.scalar(
            select(PublicCatalogOfferRow).where(
                PublicCatalogOfferRow.tenant_id == DEFAULT_TENANT_ID,
                PublicCatalogOfferRow.sku_id == first_sku_id,
            )
        )
        assert deleted_sku is not None
        assert deleted_sku.deleted_at is not None
        assert deleted_sku.status == "ARCHIVED"
        assert deleted_sku.version == 2
        assert offer is not None
        assert offer.publication_status == "SUSPENDED"
        assert product is not None
        assert product.status == "ACTIVE"
        assert product.search_document_version == 0

    second_delete = client.post(
        "/api/v1/skus/batch-delete",
        json={"sku_ids": [str(second_sku_id)]},
    )
    assert second_delete.status_code == 200, second_delete.text
    assert second_delete.json()["success_count"] == 1
    with SessionLocal() as session:
        product = session.get(ProductRow, product_id)
        assert product is not None
        assert product.status == "ARCHIVED"
        assert product.archived_at is not None

    empty_listing = client.get(
        "/api/v1/product-center/skus",
        params={"q": suffix},
    )
    assert empty_listing.status_code == 200, empty_listing.text
    assert empty_listing.json()["total"] == 0


def test_batch_merchandising_updates_category_pin_status_and_storefront_order() -> None:
    suffix = uuid4().hex[:8].upper()
    source_category_id = uuid4()
    target_category_id = uuid4()
    alpha_product_id = uuid4()
    zulu_product_id = uuid4()
    alpha_sku_id = uuid4()
    alpha_variant_sku_id = uuid4()
    zulu_sku_id = uuid4()
    product_ids = [alpha_product_id, zulu_product_id]
    sku_ids = [alpha_sku_id, alpha_variant_sku_id, zulu_sku_id]

    with SessionLocal() as session:
        session.add_all(
            [
                ProductCategoryRow(
                    id=source_category_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    code=f"BULK-SOURCE-{suffix}",
                    name=f"Bulk Source {suffix}",
                    path=f"Bulk Source {suffix}",
                    status="ACTIVE",
                    sort_order=10,
                ),
                ProductCategoryRow(
                    id=target_category_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    code=f"BULK-TARGET-{suffix}",
                    name=f"Bulk Target {suffix}",
                    path=f"Bulk Target {suffix}",
                    status="ACTIVE",
                    sort_order=20,
                ),
            ]
        )
        session.flush()
        session.add_all(
            [
                ProductRow(
                    id=alpha_product_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_code=f"BULK-ALPHA-{suffix}",
                    name=f"Alpha Product {suffix}",
                    category_id=source_category_id,
                    status="ACTIVE",
                    search_document_version=4,
                ),
                ProductRow(
                    id=zulu_product_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_code=f"BULK-ZULU-{suffix}",
                    name=f"Zulu Product {suffix}",
                    category_id=source_category_id,
                    status="ACTIVE",
                ),
            ]
        )
        session.flush()
        session.add_all(
            [
                SkuRow(
                    id=alpha_sku_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_id=alpha_product_id,
                    sku_code=f"BULK-{suffix}-ALPHA",
                    name=f"Alpha SKU {suffix}",
                    option_values={},
                    status="ACTIVE",
                ),
                SkuRow(
                    id=alpha_variant_sku_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_id=alpha_product_id,
                    sku_code=f"BULK-{suffix}-ALPHA-V2",
                    name=f"Alpha Variant {suffix}",
                    option_values={},
                    status="ACTIVE",
                ),
                SkuRow(
                    id=zulu_sku_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_id=zulu_product_id,
                    sku_code=f"BULK-{suffix}-ZULU",
                    name=f"Zulu SKU {suffix}",
                    option_values={},
                    status="ACTIVE",
                ),
            ]
        )
        session.flush()
        session.add_all(
            [
                PublicCatalogOfferRow(
                    tenant_id=DEFAULT_TENANT_ID,
                    sku_id=sku_id,
                    unit_price=Decimal("10"),
                    currency="CNY",
                    tags=[f"Bulk Pin {suffix}"],
                    publication_status="PUBLISHED",
                )
                for sku_id in sku_ids
            ]
        )
        session.commit()

    move = client.post(
        "/api/v1/skus/batch-update-category",
        json={
            "sku_ids": [str(alpha_sku_id), str(alpha_variant_sku_id)],
            "category_id": str(target_category_id),
        },
    )
    assert move.status_code == 200, move.text
    assert move.json()["success_count"] == 2
    assert move.json()["affected_product_count"] == 1
    with SessionLocal() as session:
        alpha = session.get(ProductRow, alpha_product_id)
        assert alpha is not None
        assert alpha.category_id == target_category_id
        assert alpha.search_document_version == 0

    pin_later_category = client.post(
        "/api/v1/skus/batch-update-pinned",
        json={"sku_ids": [str(alpha_sku_id)], "pinned": True},
    )
    assert pin_later_category.status_code == 200, pin_later_category.text
    global_products = client.get(
        "/api/store/demo/products",
        params={
            "tags": f"Bulk Pin {suffix}",
            "include_facets": "false",
        },
    )
    assert global_products.status_code == 200, global_products.text
    assert [item["name"] for item in global_products.json()["items"]] == [
        f"Alpha Product {suffix}",
        f"Zulu Product {suffix}",
    ]
    global_skus = client.get(
        "/api/store/demo/skus",
        params={
            "tags": f"Bulk Pin {suffix}",
            "include_facets": "false",
        },
    )
    assert global_skus.status_code == 200, global_skus.text
    assert [item["product_id"] for item in global_skus.json()["items"][:2]] == [
        str(alpha_product_id),
        str(alpha_product_id),
    ]
    unpin_later_category = client.post(
        "/api/v1/skus/batch-update-pinned",
        json={"sku_ids": [str(alpha_sku_id)], "pinned": False},
    )
    assert unpin_later_category.status_code == 200, unpin_later_category.text

    listing = client.get(
        "/api/v1/product-center/skus",
        params={"q": suffix},
    )
    assert listing.status_code == 200, listing.text
    assert all(item["is_pinned"] is False for item in listing.json()["items"])

    move_back = client.post(
        "/api/v1/skus/batch-update-category",
        json={
            "sku_ids": [str(alpha_sku_id)],
            "category_id": str(source_category_id),
        },
    )
    assert move_back.status_code == 200, move_back.text

    pin = client.post(
        "/api/v1/skus/batch-update-pinned",
        json={"sku_ids": [str(zulu_sku_id)], "pinned": True},
    )
    assert pin.status_code == 200, pin.text
    assert pin.json()["affected_product_count"] == 1
    public_pinned = client.get(
        "/api/store/demo/products",
        params={
            "category": f"Bulk Source {suffix}",
            "include_facets": "false",
        },
    )
    assert public_pinned.status_code == 200, public_pinned.text
    assert [item["name"] for item in public_pinned.json()["items"]] == [
        f"Zulu Product {suffix}",
        f"Alpha Product {suffix}",
    ]

    deactivate = client.post(
        "/api/v1/skus/batch-update-status",
        json={"sku_ids": [str(zulu_sku_id)], "status": "INACTIVE"},
    )
    assert deactivate.status_code == 200, deactivate.text
    public_inactive = client.get(
        "/api/store/demo/products",
        params={
            "category": f"Bulk Source {suffix}",
            "include_facets": "false",
        },
    )
    assert public_inactive.status_code == 200, public_inactive.text
    assert [item["name"] for item in public_inactive.json()["items"]] == [
        f"Alpha Product {suffix}"
    ]

    activate = client.post(
        "/api/v1/skus/batch-update-status",
        json={"sku_ids": [str(zulu_sku_id)], "status": "ACTIVE"},
    )
    assert activate.status_code == 200, activate.text
    unpin = client.post(
        "/api/v1/skus/batch-update-pinned",
        json={"sku_ids": [str(zulu_sku_id)], "pinned": False},
    )
    assert unpin.status_code == 200, unpin.text
    public_unpinned = client.get(
        "/api/store/demo/products",
        params={
            "category": f"Bulk Source {suffix}",
            "include_facets": "false",
        },
    )
    assert public_unpinned.status_code == 200, public_unpinned.text
    assert [item["name"] for item in public_unpinned.json()["items"]] == [
        f"Alpha Product {suffix}",
        f"Zulu Product {suffix}",
    ]

    missing_id = uuid4()
    missing = client.post(
        "/api/v1/skus/batch-update-pinned",
        json={"sku_ids": [str(missing_id)], "pinned": True},
    )
    assert missing.status_code == 200, missing.text
    assert missing.json()["failed_items"] == [
        {
            "sku_id": str(missing_id),
            "reason": "SKU 不存在、已删除或已经归档",
        }
    ]

    with SessionLocal() as session:
        session.execute(
            delete(ProductAuditEventRow).where(
                ProductAuditEventRow.product_id.in_(product_ids)
            )
        )
        session.execute(
            delete(PublicCatalogOfferRow).where(
                PublicCatalogOfferRow.sku_id.in_(sku_ids)
            )
        )
        session.execute(delete(SkuRow).where(SkuRow.id.in_(sku_ids)))
        session.execute(delete(ProductRow).where(ProductRow.id.in_(product_ids)))
        session.execute(
            delete(ProductCategoryRow).where(
                ProductCategoryRow.id.in_([source_category_id, target_category_id])
            )
        )
        session.commit()


def test_hot_product_merchandising_uses_recent_views_and_submitted_quotes() -> None:
    suffix = uuid4().hex[:8].upper()
    category_id = uuid4()
    alpha_product_id = uuid4()
    viewed_product_id = uuid4()
    ordered_product_id = uuid4()
    alpha_sku_id = uuid4()
    viewed_sku_id = uuid4()
    ordered_sku_id = uuid4()
    quote_id = uuid4()
    product_ids = [alpha_product_id, viewed_product_id, ordered_product_id]
    sku_ids = [alpha_sku_id, viewed_sku_id, ordered_sku_id]
    category_name = f"Hot Ranking {suffix}"
    alpha_name = f"A Quiet {suffix}"
    viewed_name = f"B Viewed {suffix}"
    ordered_name = f"C Ordered {suffix}"
    now = datetime.now(UTC)

    with SessionLocal() as session:
        profile = session.get(TenantPublicProfileRow, DEFAULT_TENANT_ID)
        assert profile is not None
        original_hot_products_enabled = bool(profile.hot_products_enabled)
        profile.hot_products_enabled = True
        session.add(
            ProductCategoryRow(
                id=category_id,
                tenant_id=DEFAULT_TENANT_ID,
                code=f"HOT-{suffix}",
                name=category_name,
                path=category_name,
                status="ACTIVE",
                sort_order=50_000,
            )
        )
        session.flush()
        session.add_all(
            [
                ProductRow(
                    id=product_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_code=f"HOT-{suffix}-{index}",
                    name=name,
                    category_id=category_id,
                    status="ACTIVE",
                )
                for index, (product_id, name) in enumerate(
                    [
                        (alpha_product_id, alpha_name),
                        (viewed_product_id, viewed_name),
                        (ordered_product_id, ordered_name),
                    ],
                    start=1,
                )
            ]
        )
        session.flush()
        session.add_all(
            [
                SkuRow(
                    id=sku_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_id=product_id,
                    sku_code=f"HOT-{suffix}-{index}",
                    name=name,
                    default_moq=Decimal("1"),
                    moq_unit="piece",
                    option_values={},
                    status="ACTIVE",
                )
                for index, (sku_id, product_id, name) in enumerate(
                    [
                        (alpha_sku_id, alpha_product_id, alpha_name),
                        (viewed_sku_id, viewed_product_id, viewed_name),
                        (ordered_sku_id, ordered_product_id, ordered_name),
                    ],
                    start=1,
                )
            ]
        )
        session.flush()
        session.add_all(
            [
                PublicCatalogOfferRow(
                    tenant_id=DEFAULT_TENANT_ID,
                    sku_id=sku_id,
                    unit_price=Decimal("10"),
                    currency="CNY",
                    tags=[],
                    publication_status="PUBLISHED",
                    published_at=now,
                )
                for sku_id in sku_ids
            ]
        )
        session.add(
            StorefrontProductViewDailyRow(
                tenant_id=DEFAULT_TENANT_ID,
                viewed_on=now.date(),
                product_id=viewed_product_id,
                sku_id=viewed_sku_id,
                sku_code_snapshot=f"HOT-{suffix}-2",
                product_name_snapshot=viewed_name,
                country_code="ZZ",
                view_count=8,
            )
        )
        session.add(
            PublicQuoteDraftRow(
                id=quote_id,
                tenant_id=DEFAULT_TENANT_ID,
                request_number=f"HOT-Q-{suffix}",
                status="PENDING_CONFIRMATION",
                customer_name="Hot Ranking Customer",
                currency="CNY",
                subtotal_amount=Decimal("10"),
                estimated_total=Decimal("10"),
                expires_at=now + timedelta(days=7),
                snapshot={"status": "PENDING_CONFIRMATION", "items": []},
                content_hash="c" * 64,
                disclaimer_version="public-draft-v1",
            )
        )
        session.flush()
        session.add(
            PublicQuoteDraftItemRow(
                tenant_id=DEFAULT_TENANT_ID,
                quote_draft_id=quote_id,
                sku_id=ordered_sku_id,
                position=1,
                quantity=Decimal("1"),
                product_id_snapshot=ordered_product_id,
                product_version=1,
                sku_version=1,
                sku_code_snapshot=f"HOT-{suffix}-3",
                name_snapshot=ordered_name,
                description_snapshot=None,
                specification_snapshot=None,
                option_values_snapshot={},
                category_snapshot=category_name,
                tags_snapshot=[],
                image_url_snapshot=None,
                minimum_order_quantity=Decimal("1"),
                unit_code_snapshot="piece",
                currency_snapshot="CNY",
                unit_price_snapshot=Decimal("10"),
                line_total=Decimal("10"),
            )
        )
        session.commit()

    try:
        hot_listing = client.get(
            "/api/store/demo/products",
            params={"page_size": 100, "include_facets": "false"},
        )
        assert hot_listing.status_code == 200, hot_listing.text
        hot_payload = hot_listing.json()
        assert hot_payload["hot_products_enabled"] is True
        assert hot_payload["hot_sort_applied"] is True
        hot_names = [item["name"] for item in hot_payload["items"]]
        assert hot_names.index(ordered_name) < hot_names.index(viewed_name)
        assert hot_names.index(viewed_name) < hot_names.index(alpha_name)

        category_listing = client.get(
            "/api/store/demo/products",
            params={
                "category": category_name,
                "include_facets": "false",
            },
        )
        assert category_listing.status_code == 200, category_listing.text
        assert category_listing.json()["hot_sort_applied"] is False
        assert [item["name"] for item in category_listing.json()["items"]] == [
            alpha_name,
            viewed_name,
            ordered_name,
        ]

        with SessionLocal() as session:
            alpha = session.get(ProductRow, alpha_product_id)
            assert alpha is not None
            alpha.storefront_pinned_at = now + timedelta(seconds=1)
            session.commit()
        pinned_listing = client.get(
            "/api/store/demo/products",
            params={"page_size": 100, "include_facets": "false"},
        )
        assert pinned_listing.status_code == 200, pinned_listing.text
        pinned_names = [item["name"] for item in pinned_listing.json()["items"]]
        assert pinned_names.index(alpha_name) < pinned_names.index(ordered_name)
        assert pinned_names.index(ordered_name) < pinned_names.index(viewed_name)
    finally:
        with SessionLocal() as session:
            profile = session.get(TenantPublicProfileRow, DEFAULT_TENANT_ID)
            assert profile is not None
            profile.hot_products_enabled = original_hot_products_enabled
            session.execute(
                delete(PublicQuoteDraftItemRow).where(
                    PublicQuoteDraftItemRow.quote_draft_id == quote_id
                )
            )
            session.execute(
                delete(PublicQuoteDraftRow).where(PublicQuoteDraftRow.id == quote_id)
            )
            session.execute(
                delete(StorefrontProductViewDailyRow).where(
                    StorefrontProductViewDailyRow.product_id.in_(product_ids)
                )
            )
            session.execute(
                delete(PublicCatalogOfferRow).where(
                    PublicCatalogOfferRow.sku_id.in_(sku_ids)
                )
            )
            session.execute(delete(SkuRow).where(SkuRow.id.in_(sku_ids)))
            session.execute(delete(ProductRow).where(ProductRow.id.in_(product_ids)))
            session.execute(
                delete(ProductCategoryRow).where(ProductCategoryRow.id == category_id)
            )
            session.commit()


def test_demo_seed_does_not_reinsert_a_soft_deleted_sku() -> None:
    sku_code = "AQ-320S"
    with SessionLocal() as session:
        original = session.scalar(
            select(SkuRow)
            .where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.sku_code == sku_code,
            )
            .execution_options(include_deleted=True)
        )
        assert original is not None
        original_status = original.status
        original_deleted_at = original.deleted_at
        original.status = "ARCHIVED"
        original.deleted_at = datetime.now(UTC)
        session.commit()

    try:
        with SessionLocal() as session:
            seed_product_center_demo(session)

        with SessionLocal() as session:
            rows = session.scalars(
                select(SkuRow)
                .where(
                    SkuRow.tenant_id == DEFAULT_TENANT_ID,
                    SkuRow.sku_code == sku_code,
                )
                .execution_options(include_deleted=True)
            ).all()
            assert len(rows) == 1
            assert rows[0].status == "ARCHIVED"
            assert rows[0].deleted_at is not None
    finally:
        with SessionLocal() as session:
            row = session.scalar(
                select(SkuRow)
                .where(
                    SkuRow.tenant_id == DEFAULT_TENANT_ID,
                    SkuRow.sku_code == sku_code,
                )
                .execution_options(include_deleted=True)
            )
            assert row is not None
            row.status = original_status
            row.deleted_at = original_deleted_at
            session.commit()


def test_delete_all_products_route_requires_the_current_password(monkeypatch) -> None:
    calls: list[dict[str, object]] = []
    job_id = uuid4()
    created_at = datetime.now(UTC)

    def fake_start_catalog_delete_job(_session, **kwargs):
        calls.append(kwargs)
        return {
            "id": job_id,
            "status": "QUEUED",
            "stage": "QUEUED",
            "progress": 0,
            "total_products": 0,
            "total_skus": 0,
            "deleted_product_count": 0,
            "deleted_sku_count": 0,
            "error_message": None,
            "created_at": created_at,
            "started_at": None,
            "completed_at": None,
        }

    monkeypatch.setattr(
        "app.routers.product_center.catalog_deletion.start_catalog_delete_job",
        fake_start_catalog_delete_job,
    )

    rejected = client.post(
        "/api/v1/product-center/products/delete-all",
        json={"password": "wrong-password"},
    )
    assert rejected.status_code == 422, rejected.text
    assert rejected.json()["detail"] == {
        "code": "CURRENT_PASSWORD_INVALID",
        "message": "密码错误，请重新输入。",
    }
    assert "wrong-password" not in rejected.text
    assert calls == []

    accepted = client.post(
        "/api/v1/product-center/products/delete-all",
        json={"password": "zhimaoyun123"},
    )
    assert accepted.status_code == 202, accepted.text
    assert accepted.json()["id"] == str(job_id)
    assert accepted.json()["status"] == "QUEUED"
    assert accepted.json()["progress"] == 0
    assert calls[0]["tenant_id"] == DEFAULT_TENANT_ID
    assert calls[0]["user_id"] == DEFAULT_OWNER_USER_ID
    assert calls[0]["organization_id"] == DEFAULT_ORGANIZATION_ID

    def fake_get_catalog_delete_job(_session, **kwargs):
        calls.append(kwargs)
        return {
            "id": job_id,
            "status": "SUCCEEDED",
            "stage": "COMPLETED",
            "progress": 100,
            "total_products": 7,
            "total_skus": 11,
            "deleted_product_count": 7,
            "deleted_sku_count": 11,
            "error_message": None,
            "created_at": created_at,
            "started_at": created_at,
            "completed_at": created_at,
        }

    monkeypatch.setattr(
        "app.routers.product_center.catalog_deletion.get_catalog_delete_job",
        fake_get_catalog_delete_job,
    )
    completed = client.get(
        f"/api/v1/product-center/products/delete-all/{job_id}"
    )
    assert completed.status_code == 200, completed.text
    assert completed.json()["deleted_product_count"] == 7
    assert completed.json()["deleted_sku_count"] == 11
    assert calls[-1]["job_id"] == job_id


def test_current_password_verification_supports_local_merchant_accounts() -> None:
    user_id = uuid4()
    identifier = f"delete-all-{uuid4().hex}@example.test"
    password = "MerchantPass42"
    salt, password_hash = new_local_password_material(password)
    try:
        with SessionLocal() as session:
            session.add(
                UserRow(
                    id=user_id,
                    email_normalized=identifier,
                    display_name="Delete All Merchant",
                    identity_provider="local-password",
                    identity_subject=f"local-password:{user_id}",
                    status="active",
                )
            )
            session.flush()
            session.add(
                LocalAccountCredentialRow(
                    user_id=user_id,
                    identifier_normalized=identifier,
                    password_salt=salt,
                    password_hash=password_hash,
                )
            )
            session.commit()

        with SessionLocal() as session:
            verify_current_user_password(
                session,
                user_id=user_id,
                password=password,
            )
            with pytest.raises(AuthError) as rejected:
                verify_current_user_password(
                    session,
                    user_id=user_id,
                    password="NotThePassword42",
                )
            assert rejected.value.code == "CURRENT_PASSWORD_INVALID"
    finally:
        with SessionLocal() as session:
            session.execute(
                delete(LocalAccountCredentialRow).where(
                    LocalAccountCredentialRow.user_id == user_id
                )
            )
            session.execute(delete(UserRow).where(UserRow.id == user_id))
            session.commit()


def test_delete_all_products_is_tenant_scoped_and_uses_bulk_soft_delete() -> None:
    suffix = uuid4().hex[:10]
    organization_id = uuid4()
    tenant_id = uuid4()
    user_id = uuid4()
    membership_id = uuid4()
    product_id = uuid4()
    sku_id = uuid4()
    try:
        with SessionLocal() as session:
            session.add(
                OrganizationRow(
                    id=organization_id,
                    code=f"delete-all-{suffix}",
                    name=f"Delete All {suffix}",
                    status="active",
                )
            )
            session.add(
                TenantRow(
                    id=tenant_id,
                    organization_id=organization_id,
                    slug=f"delete-all-{suffix}",
                    name=f"Delete All {suffix}",
                    status="active",
                )
            )
            session.add(
                UserRow(
                    id=user_id,
                    email_normalized=f"delete-all-{suffix}@example.test",
                    display_name="Catalog Owner",
                    identity_provider="local-password",
                    identity_subject=f"delete-all:{suffix}",
                    status="active",
                )
            )
            session.add(
                MembershipRow(
                    id=membership_id,
                    tenant_id=tenant_id,
                    user_id=user_id,
                    status="active",
                    joined_at=datetime.now(UTC),
                )
            )
            session.flush()
            session.add(
                ProductRow(
                    id=product_id,
                    tenant_id=tenant_id,
                    product_code=f"DELETE-ALL-{suffix}",
                    name="Bulk deletion test product",
                    status="ACTIVE",
                    current_version=3,
                    search_document_version=3,
                )
            )
            session.flush()
            session.add(
                SkuRow(
                    id=sku_id,
                    tenant_id=tenant_id,
                    product_id=product_id,
                    sku_code=f"DELETE-ALL-SKU-{suffix}",
                    name="Bulk deletion test SKU",
                    option_values={},
                    status="ACTIVE",
                    version=4,
                )
            )
            session.flush()
            session.add(
                PublicCatalogOfferRow(
                    tenant_id=tenant_id,
                    sku_id=sku_id,
                    unit_price=Decimal("12.00"),
                    currency="CNY",
                    tags=[],
                    publication_status="PUBLISHED",
                )
            )
            session.commit()

        with SessionLocal() as session:
            result = delete_all_products_use_case(
                session,
                tenant_id=tenant_id,
                user_id=user_id,
                membership_id=membership_id,
                permissions=frozenset({"product.edit"}),
            )
            assert result == {
                "deleted_product_count": 1,
                "deleted_sku_count": 1,
            }

        with SessionLocal() as session:
            sku = session.scalar(
                select(SkuRow)
                .where(SkuRow.tenant_id == tenant_id, SkuRow.id == sku_id)
                .execution_options(include_deleted=True)
            )
            product = session.get(ProductRow, product_id)
            offer = session.scalar(
                select(PublicCatalogOfferRow).where(
                    PublicCatalogOfferRow.tenant_id == tenant_id,
                    PublicCatalogOfferRow.sku_id == sku_id,
                )
            )
            event = session.scalar(
                select(ProductAuditEventRow).where(
                    ProductAuditEventRow.tenant_id == tenant_id,
                    ProductAuditEventRow.action == "catalog.deleted_all",
                )
            )
            assert sku is not None
            assert sku.status == "ARCHIVED"
            assert sku.deleted_at is not None
            assert sku.version == 5
            assert product is not None
            assert product.status == "ARCHIVED"
            assert product.archived_at is not None
            assert product.current_version == 4
            assert product.search_document_version == 0
            assert offer is not None
            assert offer.publication_status == "SUSPENDED"
            assert event is not None
            assert event.before == {"product_count": 1, "sku_count": 1}

            repeated = delete_all_products_use_case(
                session,
                tenant_id=tenant_id,
                user_id=user_id,
                membership_id=membership_id,
                permissions=frozenset({"product.edit"}),
            )
            assert repeated == {
                "deleted_product_count": 0,
                "deleted_sku_count": 0,
            }
    finally:
        with SessionLocal() as session:
            session.execute(
                delete(ProductAuditEventRow).where(
                    ProductAuditEventRow.tenant_id == tenant_id
                )
            )
            session.execute(
                delete(PublicCatalogOfferRow).where(
                    PublicCatalogOfferRow.tenant_id == tenant_id
                )
            )
            session.execute(delete(SkuRow).where(SkuRow.tenant_id == tenant_id))
            session.execute(
                delete(ProductRow).where(ProductRow.tenant_id == tenant_id)
            )
            session.execute(
                delete(MembershipRow).where(MembershipRow.tenant_id == tenant_id)
            )
            session.execute(delete(TenantRow).where(TenantRow.id == tenant_id))
            session.execute(delete(UserRow).where(UserRow.id == user_id))
            session.execute(
                delete(OrganizationRow).where(OrganizationRow.id == organization_id)
            )
            session.commit()


def test_upload_parse_review_and_approve_xlsx() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "产品表"
    sheet.append(["产品名称", "型号", "单价", "MOQ"])
    sheet.append(["测试宠物碗", "BOWL-001", 12.5, 100])
    content = BytesIO()
    workbook.save(content)

    response = client.post(
        "/api/v1/imports",
        files={"file": ("supplier.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"supplier_id": "SUP-001", "source_type": "SUPPLIER_QUOTE"},
    )
    assert response.status_code == 201
    job = response.json()
    assert job["status"] == "needs_review"
    assert job["products"] == 1
    assert job["detected_type"] == "OOXML / XLSX"
    assert job["candidate_status"] == "NEEDS_REVIEW"
    assert job["candidate_fields"] == 4
    assert job["candidate_idempotent"] is False
    assert job["ai_task_id"]

    review_response = client.get("/api/v1/review-items", params={"job_id": job["id"]})
    assert review_response.status_code == 200
    items = review_response.json()
    assert len(items) == 1
    assert items[0]["name"] == "测试宠物碗"
    assert items[0]["status"] == "pending"

    item_id = items[0]["id"]
    update_response = client.patch(
        f"/api/v1/review-items/{item_id}",
        json={"normalized_values": {"name": "测试宠物食盆"}},
    )
    assert update_response.status_code == 200
    name_field = next(field for field in update_response.json()["fields"] if field["key"] == "name")
    assert name_field["normalized"] == "测试宠物食盆"

    approval = client.post(f"/api/v1/review-items/{item_id}/approve")
    assert approval.status_code == 200
    assert approval.json() == {"id": item_id, "status": "approved", "image_status": "SOURCE"}


def test_product_template_download_matches_the_strict_import_contract() -> None:
    response = client.get("/api/v1/product-template.xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename*=UTF-8''" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content), read_only=False, data_only=True)
    try:
        assert workbook.sheetnames == [
            PRODUCT_MASTER_TEMPLATE_SHEET,
            SKU_DETAIL_TEMPLATE_SHEET,
        ]
        product_sheet = workbook[PRODUCT_MASTER_TEMPLATE_SHEET]
        sku_sheet = workbook[SKU_DETAIL_TEMPLATE_SHEET]
        assert product_sheet.max_row == 1
        assert sku_sheet.max_row == 1
        assert tuple(
            product_sheet.cell(row=1, column=index).value
            for index in range(1, len(PRODUCT_MASTER_TEMPLATE_HEADERS) + 1)
        ) == PRODUCT_MASTER_TEMPLATE_HEADERS
        assert tuple(
            sku_sheet.cell(row=1, column=index).value
            for index in range(1, len(SKU_DETAIL_TEMPLATE_HEADERS) + 1)
        ) == SKU_DETAIL_TEMPLATE_HEADERS
        assert product_sheet.freeze_panes == "A2"
        assert sku_sheet.freeze_panes == "D2"
        assert product_sheet.auto_filter.ref == "A1:R1"
        assert sku_sheet.auto_filter.ref == "A1:Z1"
        assert product_sheet["A1"].fill.fgColor.rgb == "002D1B69"
        assert sku_sheet["A1"].fill.fgColor.rgb == "0023453B"
        assert product_sheet["A1"].font.bold is True
        assert sku_sheet["A1"].font.bold is True
        assert product_sheet["A1"].comment is not None
        assert "连接 Product 与 SKU" in product_sheet["A1"].comment.text
        assert product_sheet["I1"].comment is not None
        assert "直接插入" in product_sheet["I1"].comment.text
        assert "HTTP(S)" in product_sheet["I1"].comment.text
        assert sku_sheet["A1"].comment is not None
        assert "Product 表" in sku_sheet["A1"].comment.text
        assert "候选值" in sku_sheet["D1"].comment.text
        assert "自动组合" in sku_sheet["E1"].comment.text
        assert sku_sheet["V1"].comment is not None
        assert "进销存" in sku_sheet["V1"].comment.text
        assert sku_sheet["W1"].comment is not None
        assert "继承商品价格" in sku_sheet["W1"].comment.text
        assert "kg" in sku_sheet["X1"].comment.text
        assert "最小起订" in sku_sheet["Y1"].comment.text
        assert "每箱" in sku_sheet["Z1"].comment.text
        assert response.headers["cache-control"] == "no-store"
    finally:
        workbook.close()


def test_product_template_imports_embedded_images_into_managed_storage(
    tmp_path: Path,
    request: pytest.FixtureRequest,
) -> None:
    import_job_ids: list[str] = []
    sku_code = "EMBEDDED-API-001"
    category_name = "内嵌图片测试分类"
    _cleanup_template_test_records(
        import_job_ids=[],
        sku_codes=[sku_code],
        category_names=[category_name],
    )
    request.addfinalizer(
        lambda: _cleanup_template_test_records(
            import_job_ids=import_job_ids,
            sku_codes=[sku_code],
            category_names=[category_name],
        )
    )
    image_bytes = (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
        b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00"
        b"\x1f\x15\xc4\x89\x00\x00\x00\rIDAT\x08\xd7c\xf8\xcf\xc0"
        b"\xf0\x1f\x00\x05\x00\x01\xff\x89\x99=\x1d\x00\x00\x00"
        b"\x00IEND\xaeB`\x82"
    )
    image_path = tmp_path / "embedded-product.png"
    image_path.write_bytes(image_bytes)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = PRODUCT_TEMPLATE_SHEET
    sheet.append(list(PRODUCT_TEMPLATE_HEADERS))
    sheet.append([
        "内嵌图片 API 商品",
        category_name,
        sku_code,
        None,
        None,
        "两张图片直接放在 Excel 单元格位置",
        None,
        None,
        *([None] * 10),
    ])
    first = OpenpyxlImage(image_path)
    first.anchor = "I2"
    sheet.add_image(first)
    second = OpenpyxlImage(image_path)
    second.anchor = "J2"
    sheet.add_image(second)
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    response = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "内嵌图片商品.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"source_type": "PRODUCT_TEMPLATE"},
    )

    assert response.status_code == 201, response.text
    import_job_ids.append(response.json()["id"])
    assert response.json()["status"] == "published"
    with SessionLocal() as session:
        sku = session.scalar(
            select(SkuRow).where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.sku_code == sku_code,
            )
        )
        assert sku is not None
        images = session.scalars(
            select(ProductImageRow)
            .where(
                ProductImageRow.tenant_id == DEFAULT_TENANT_ID,
                ProductImageRow.product_id == sku.product_id,
                ProductImageRow.deleted_at.is_(None),
            )
            .order_by(ProductImageRow.sort_order)
        ).all()
        assert len(images) == 2
        assert [image.image_role for image in images] == ["MAIN", "GALLERY"]
        assert all(image.storage_provider == "LOCAL" for image in images)
        assert all(image.bucket == "product-template" for image in images)
        assert all(image.content_type == "image/png" for image in images)
        assert all(image.byte_size == len(image_bytes) for image in images)
        assert all(
            image.sha256 == hashlib.sha256(image_bytes).hexdigest()
            for image in images
        )
        assert all(
            get_object_storage().exists(image.object_key)
            for image in images
        )

    listing = client.get("/api/store/demo/skus", params={"q": sku_code})
    assert listing.status_code == 200, listing.text
    assert listing.json()["total"] == 1
    public_image_url = listing.json()["items"][0]["image_url"]
    assert public_image_url.startswith("/api/store/demo/media/")
    public_image = client.get(public_image_url)
    assert public_image.status_code == 200
    assert public_image.content == image_bytes
    assert public_image.headers["content-type"].startswith("image/png")


def test_product_variant_template_imports_one_product_with_multiple_public_skus(
    tmp_path: Path,
    request: pytest.FixtureRequest,
) -> None:
    import_job_ids: list[str] = []
    category_name = "商品规格模板测试"
    workbook_path = tmp_path / "商品图册模板（更新了商品规格分类）.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(list(PRODUCT_VARIANT_TEMPLATE_HEADERS))
    sheet.append([
        "规格模板 API 摄像头",
        category_name,
        "VARIANT-API-CAMERA",
        None,
        "同一个商品下包含两个可报价规格",
        "接口回归测试",
        "是",
        12,
        "普通款",
        10,
        None,
    ])
    sheet.append([
        "规格模板 API 摄像头",
        category_name,
        "VARIANT-API-CAMERA",
        None,
        "同一个商品下包含两个可报价规格",
        None,
        None,
        12,
        "蓝牙款",
        20,
        None,
    ])
    workbook.save(workbook_path)
    workbook.close()

    parsed = parse_product_template(workbook_path)
    sku_codes = [row.sku_code for row in parsed.rows]
    _cleanup_template_test_records(
        import_job_ids=[],
        sku_codes=sku_codes,
        category_names=[category_name],
    )
    request.addfinalizer(
        lambda: _cleanup_template_test_records(
            import_job_ids=import_job_ids,
            sku_codes=sku_codes,
            category_names=[category_name],
        )
    )

    response = client.post(
        "/api/v1/imports",
        files={
            "file": (
                workbook_path.name,
                workbook_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"source_type": "PRODUCT_TEMPLATE"},
    )
    assert response.status_code == 201, response.text
    job = response.json()
    import_job_ids.append(job["id"])
    assert job["status"] == "published"
    assert any(
        "商品+规格模板" in warning
        for warning in job["warning_messages"]
    )

    sku_response = client.get(
        "/api/v1/product-center/skus",
        params={"q": "VARIANT-API-CAMERA", "page_size": 10},
    )
    assert sku_response.status_code == 200, sku_response.text
    sku_items = sku_response.json()["items"]
    assert len(sku_items) == 2
    assert len({item["product_id"] for item in sku_items}) == 1
    assert {item["name"] for item in sku_items} == {
        "规格模板 API 摄像头 · 普通款",
        "规格模板 API 摄像头 · 蓝牙款",
    }

    product_listing = client.get(
        "/api/store/demo/products",
        params={"q": "规格模板 API 摄像头"},
    )
    assert product_listing.status_code == 200, product_listing.text
    payload = product_listing.json()
    assert payload["total"] == 1
    assert payload["items"][0]["product_code"] == "VARIANT-API-CAMERA"
    assert payload["items"][0]["sku_count"] == 2
    assert payload["items"][0]["price_from"] == "10.00"
    assert payload["items"][0]["price_to"] == "20.00"

    product_detail = client.get(
        f"/api/store/demo/products/{payload['items'][0]['id']}"
    )
    assert product_detail.status_code == 200, product_detail.text
    detail = product_detail.json()
    assert detail["name"] == "规格模板 API 摄像头"
    assert {
        item["specification"] for item in detail["skus"]
    } == {"普通款", "蓝牙款"}


def test_product_template_failure_returns_complete_structured_issue_details(
    request: pytest.FixtureRequest,
) -> None:
    import_job_ids: list[str] = []
    request.addfinalizer(
        lambda: _cleanup_template_test_records(
            import_job_ids=import_job_ids,
            sku_codes=[],
            category_names=[],
        )
    )
    content = _product_template_bytes([
        [
            "",
            "测试分类",
            "INVALID-DETAIL-A",
            None,
            "not-a-price",
            None,
            None,
            None,
            "javascript:alert(1)",
            *([None] * 9),
        ],
        [
            "错误商品 B",
            "一级/二级/三级",
            "INVALID-DETAIL-B",
            None,
            None,
            None,
            None,
            "x" * 81,
            *([None] * 10),
        ],
    ])

    response = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "待修正商品.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "source_type": "PRODUCT_TEMPLATE",
            "defer_processing": "true",
        },
    )

    assert response.status_code == 201, response.text
    queued = response.json()
    import_job_ids.append(queued["id"])
    assert queued["status"] == "scanning"
    assert queued["progress"] == 5

    detailed = client.get(f"/api/v1/imports/{queued['id']}")
    assert detailed.status_code == 200
    payload = detailed.json()
    assert payload["status"] == "failed"
    assert payload["progress"] == 100
    assert payload["products"] == 0
    assert payload["warnings"] == 5
    assert payload["result_details"]["issue_total"] == 5
    assert payload["result_details"]["issues_truncated"] == 0
    assert payload["result_details"]["import_stage"] == "VALIDATION_FAILED"
    assert {
        (issue["row_number"], issue["column"], issue["code"])
        for issue in payload["result_details"]["issues"]
    } == {
        (2, "商品名称", "REQUIRED_VALUE_MISSING"),
        (2, "商品价格", "PRICE_INVALID"),
        (2, "商品图片1", "IMAGE_URL_INVALID"),
        (3, "商品分类", "CATEGORY_INVALID"),
        (3, "标签", "TAGS_INVALID"),
    }
    assert all(
        issue["message"] and issue["suggestion"]
        for issue in payload["result_details"]["issues"]
    )

    assert len(payload["result_details"]["issues"]) == 5


def test_fixed_product_template_imports_optional_supplier_and_publishes_blank_prices_as_zero(
    request: pytest.FixtureRequest,
) -> None:
    created_import_job_ids: list[str] = []

    def cleanup_template_products() -> None:
        with SessionLocal() as session:
            sku_rows = session.scalars(
                select(SkuRow)
                .where(
                    SkuRow.tenant_id == DEFAULT_TENANT_ID,
                    SkuRow.sku_code.in_(["TPL-API-001", "TPL-API-002"]),
                )
                .execution_options(include_deleted=True)
            ).all()
            sku_ids = [row.id for row in sku_rows]
            category_ids = session.scalars(
                select(ProductCategoryRow.id).where(
                    ProductCategoryRow.tenant_id == DEFAULT_TENANT_ID,
                    ProductCategoryRow.name == "模版测试分类",
                )
            ).all()
            category_product_ids = (
                session.scalars(
                    select(ProductRow.id)
                    .where(
                        ProductRow.tenant_id == DEFAULT_TENANT_ID,
                        ProductRow.category_id.in_(category_ids),
                    )
                    .execution_options(include_deleted=True)
                ).all()
                if category_ids
                else []
            )
            product_ids = list(
                dict.fromkeys(
                    [*(row.product_id for row in sku_rows), *category_product_ids]
                )
            )
            if sku_ids:
                session.execute(
                    delete(PublicCatalogOfferRow).where(
                        PublicCatalogOfferRow.tenant_id == DEFAULT_TENANT_ID,
                        PublicCatalogOfferRow.sku_id.in_(sku_ids),
                    )
                )
            if product_ids:
                session.execute(
                    delete(ProductImageRow).where(
                        ProductImageRow.tenant_id == DEFAULT_TENANT_ID,
                        ProductImageRow.product_id.in_(product_ids),
                    )
                )
            if sku_ids:
                session.execute(
                    delete(SkuRow).where(
                        SkuRow.tenant_id == DEFAULT_TENANT_ID,
                        SkuRow.id.in_(sku_ids),
                    )
                )
            if product_ids:
                session.execute(
                    delete(ProductAuditEventRow).where(
                        ProductAuditEventRow.tenant_id == DEFAULT_TENANT_ID,
                        ProductAuditEventRow.product_id.in_(product_ids),
                    )
                )
                session.execute(
                    delete(ProductRow).where(
                        ProductRow.tenant_id == DEFAULT_TENANT_ID,
                        ProductRow.id.in_(product_ids),
                    )
                )
            if created_import_job_ids:
                worker_rows = session.scalars(
                    select(WorkerJobRow).where(
                        WorkerJobRow.tenant_id == DEFAULT_TENANT_ID,
                        WorkerJobRow.import_job_id.in_(created_import_job_ids),
                    )
                ).all()
                source_file_ids = [row.source_file_id for row in worker_rows]
                media_object_ids = [row.media_object_id for row in worker_rows]
                session.execute(
                    delete(WorkerJobRow).where(
                        WorkerJobRow.tenant_id == DEFAULT_TENANT_ID,
                        WorkerJobRow.import_job_id.in_(created_import_job_ids),
                    )
                )
                session.execute(
                    delete(ImportJobRow).where(
                        ImportJobRow.tenant_id == DEFAULT_TENANT_ID,
                        ImportJobRow.id.in_(created_import_job_ids),
                    )
                )
                if source_file_ids:
                    session.execute(
                        delete(SourceFileRow).where(
                            SourceFileRow.tenant_id == DEFAULT_TENANT_ID,
                            SourceFileRow.id.in_(source_file_ids),
                        )
                    )
                if media_object_ids:
                    session.execute(
                        delete(MediaObjectRow).where(
                            MediaObjectRow.tenant_id == DEFAULT_TENANT_ID,
                            MediaObjectRow.id.in_(media_object_ids),
                        )
                    )
            session.execute(
                delete(ProductCategoryRow).where(
                    ProductCategoryRow.tenant_id == DEFAULT_TENANT_ID,
                    ProductCategoryRow.name == "模版测试分类",
                )
            )
            session.execute(
                delete(SupplierRow).where(
                    SupplierRow.tenant_id == DEFAULT_TENANT_ID,
                    SupplierRow.name == "模版供应商 A",
                )
            )
            session.commit()

    cleanup_template_products()
    request.addfinalizer(cleanup_template_products)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品列表"
    sheet.append(list(PRODUCT_TEMPLATE_HEADERS))
    sheet.append([
        "模版商品 A",
        "模版测试分类",
        "TPL-API-001",
        "模版供应商 A",
        "12.5",
        "固定模版商品描述",
        "10",
        "新品，热卖,新品",
        "https://img.example.com/tpl-api-001.jpg",
        *([None] * 9),
    ])
    sheet.append([
        "重复商品 A",
        "模版测试分类",
        "TPL-API-001",
        "模版供应商 A",
        99,
        None,
        None,
        None,
        *([None] * 10),
    ])
    sheet.append([
        "模版商品 B",
        "模版测试分类",
        "TPL-API-002",
        None,
        "",
        None,
        None,
        None,
        *([None] * 10),
    ])
    for duplicate_index in range(1003):
        sheet.append([
            f"额外重复商品 {duplicate_index + 1}",
            "模版测试分类",
            "TPL-API-001",
            "模版供应商 A",
            88 + duplicate_index,
            None,
            None,
            None,
            *([None] * 10),
        ])
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    response = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "商品模版.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"source_type": "PRODUCT_TEMPLATE"},
    )

    assert response.status_code == 201, response.text
    job = response.json()
    created_import_job_ids.append(job["id"])
    assert job["source_type"] == "PRODUCT_TEMPLATE"
    assert job["supplier"] == "商品模版"
    assert job["status"] == "published"
    assert job["products"] == 2
    assert job["warnings"] == 1004
    assert len(job["warning_messages"]) == 1000
    assert sum("重复" in warning for warning in job["warning_messages"]) == 1000
    assert any("第 1002 行" in warning for warning in job["warning_messages"])
    assert "另有 1001 条提醒" in job["error_message"]
    assert job["result_details"]["warnings_truncated"] == 4
    assert job["result_details"]["outcome"] == "TEMPLATE_IMPORTED"
    assert job["result_details"]["imported"] == 2
    assert job["result_details"]["processed_rows"] == 2
    assert job["result_details"]["total_rows"] == 2
    assert job["candidate_fields"] == 0

    listed_job = next(
        row
        for row in client.get("/api/v1/imports", params={"limit": 500}).json()
        if row["id"] == job["id"]
    )
    assert len(listed_job["warning_messages"]) == 20
    assert len(listed_job["result_details"]["warnings"]) == 20
    assert listed_job["result_details"]["warnings_truncated"] == 984
    detailed_job = client.get(f"/api/v1/imports/{job['id']}").json()
    assert len(detailed_job["warning_messages"]) == 1000
    assert len(detailed_job["result_details"]["warnings"]) == 1000
    assert detailed_job["result_details"]["warnings_truncated"] == 4

    sku_response = client.get(
        "/api/v1/product-center/skus",
        params={"q": "TPL-API", "page_size": 10},
    )
    assert sku_response.status_code == 200
    assert sku_response.json()["total"] == 2
    sku_by_code = {
        row["sku_code"]: row for row in sku_response.json()["items"]
    }
    assert sku_by_code["TPL-API-001"]["default_moq"] is None
    assert sku_by_code["TPL-API-001"]["public_price"] == "12.50"
    assert sku_by_code["TPL-API-001"]["public_offer_status"] == "PUBLISHED"
    assert sku_by_code["TPL-API-001"]["image_status"] == "APPROVED"
    assert sku_by_code["TPL-API-001"]["tags"] == ["新品", "热卖"]
    assert sku_by_code["TPL-API-001"]["source_type"] == "PRODUCT_TEMPLATE"
    assert sku_by_code["TPL-API-001"]["source_filename"] == "商品模版.xlsx"
    assert sku_by_code["TPL-API-001"]["source_imported_at"]
    assert sku_by_code["TPL-API-002"]["public_price"] == "0.00"
    assert sku_by_code["TPL-API-002"]["public_offer_status"] == "PUBLISHED"

    inventory_response = client.get(
        "/api/v1/inventory/stocks",
        params={"q": "TPL-API-001", "page_size": 10},
    )
    assert inventory_response.status_code == 200, inventory_response.text
    inventory_item = inventory_response.json()["items"][0]
    assert inventory_item["supplier_name"] == "模版供应商 A"
    assert inventory_item["supplier_id"]

    # Re-importing the fixed template may refresh its note/marker, but must
    # preserve product-center variant attributes maintained outside Excel.
    with SessionLocal() as session:
        sku_a = session.scalar(
            select(SkuRow).where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.sku_code == "TPL-API-001",
            )
        )
        assert sku_a is not None
        supplier_a = session.scalar(
            select(SupplierRow).where(
                SupplierRow.tenant_id == DEFAULT_TENANT_ID,
                SupplierRow.name == "模版供应商 A",
            )
        )
        assert supplier_a is not None
        assert sku_a.supplier_id == supplier_a.id
        assert supplier_a.active_skus == 1
        sku_b = session.scalar(
            select(SkuRow).where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.sku_code == "TPL-API-002",
            )
        )
        assert sku_b is not None
        assert sku_b.supplier_id is None
        sku_a_id = sku_a.id
        sku_a_version = sku_a.version
        offer_a = session.scalar(
            select(PublicCatalogOfferRow).where(
                PublicCatalogOfferRow.tenant_id == DEFAULT_TENANT_ID,
                PublicCatalogOfferRow.sku_id == sku_a.id,
            )
        )
        assert offer_a is not None
        assert offer_a.display_tag == "新品"
        offer_a.display_tag = "热卖"
        offer_a.tag_color = "#725B9B"
        session.commit()
    sku_update = client.patch(
        f"/api/v1/skus/{sku_a_id}",
        json={
            "expected_version": sku_a_version,
            "option_values": {
                "颜色": "红色",
                "_sku2quotation": "客户端不能覆盖内部标记",
            },
        },
    )
    assert sku_update.status_code == 200, sku_update.text
    assert sku_update.json()["option_values"]["颜色"] == "红色"
    assert sku_update.json()["option_values"]["_sku2quotation"] == {
        "source": "PRODUCT_TEMPLATE",
        "schema": 1,
    }

    review_response = client.get(
        "/api/v1/review-items",
        params={"job_id": job["id"]},
    )
    assert review_response.status_code == 200
    assert review_response.json() == []

    repeated = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "商品更新.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"source_type": "PRODUCT_TEMPLATE"},
    )
    assert repeated.status_code == 201, repeated.text
    created_import_job_ids.append(repeated.json()["id"])
    assert repeated.json()["status"] == "published"
    assert repeated.json()["products"] == 2
    assert "未变化 2" in repeated.json()["error_message"]
    repeated_skus = client.get(
        "/api/v1/product-center/skus",
        params={"q": "TPL-API", "page_size": 10},
    ).json()
    assert repeated_skus["total"] == 2
    repeated_by_code = {
        row["sku_code"]: row for row in repeated_skus["items"]
    }
    assert repeated_by_code["TPL-API-001"]["version"] == 2
    assert repeated_by_code["TPL-API-002"]["version"] == 1
    assert repeated_by_code["TPL-API-001"]["source_filename"] == "商品更新.xlsx"
    assert repeated_by_code["TPL-API-002"]["source_filename"] == "商品更新.xlsx"
    with SessionLocal() as session:
        preserved_sku = session.scalar(
            select(SkuRow).where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.sku_code == "TPL-API-001",
            )
        )
        assert preserved_sku is not None
        assert preserved_sku.option_values["颜色"] == "红色"
        preserved_offer = session.scalar(
            select(PublicCatalogOfferRow).where(
                PublicCatalogOfferRow.tenant_id == DEFAULT_TENANT_ID,
                PublicCatalogOfferRow.sku_id == preserved_sku.id,
            )
        )
        assert preserved_offer is not None
        assert preserved_offer.display_tag == "热卖"
        assert preserved_offer.tag_color == "#725B9B"

    with SessionLocal() as session:
        sku_a = session.scalar(
            select(SkuRow).where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.sku_code == "TPL-API-001",
            )
        )
        assert sku_a is not None
        product_a_id = sku_a.product_id
        offer_a = session.scalar(
            select(PublicCatalogOfferRow).where(
                PublicCatalogOfferRow.tenant_id == DEFAULT_TENANT_ID,
                PublicCatalogOfferRow.sku_id == sku_a.id,
            )
        )
        image_a = session.scalar(
            select(ProductImageRow).where(
                ProductImageRow.tenant_id == DEFAULT_TENANT_ID,
                ProductImageRow.object_key
                == "https://img.example.com/tpl-api-001.jpg",
            )
        )
        assert offer_a is not None
        assert image_a is not None
        offer_a_id = offer_a.id
        image_a_id = image_a.id

    reduced_workbook = load_workbook(
        BytesIO(client.get("/api/v1/product-template.xlsx").content)
    )
    reduced_product_sheet = reduced_workbook[PRODUCT_MASTER_TEMPLATE_SHEET]
    reduced_sku_sheet = reduced_workbook[SKU_DETAIL_TEMPLATE_SHEET]
    reduced_product_sheet.append([
        "TPL-PRODUCT-B",
        "模版商品 B",
        "模版测试分类",
        None,
        None,
        None,
        None,
        None,
        *([None] * 10),
    ])
    reduced_sku_sheet.append([
        "TPL-PRODUCT-B",
        "TPL-API-002",
        None,
        *([None] * 18),
        None,
        None,
        "2.5",
        10,
        24,
    ])
    reduced_content = BytesIO()
    reduced_workbook.save(reduced_content)
    reduced_workbook.close()
    reduced = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "商品模版.xlsx",
                reduced_content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"source_type": "PRODUCT_TEMPLATE"},
    )
    assert reduced.status_code == 201, reduced.text
    created_import_job_ids.append(reduced.json()["id"])
    assert "保留未包含商品" in reduced.json()["error_message"]
    assert "归档" not in reduced.json()["error_message"]

    with SessionLocal() as session:
        sku_a = session.scalar(
            select(SkuRow)
            .where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.sku_code == "TPL-API-001",
            )
            .execution_options(include_deleted=True)
        )
        product_a = session.get(
            ProductRow,
            product_a_id,
            execution_options={"include_deleted": True},
        )
        offer_a = session.get(
            PublicCatalogOfferRow,
            offer_a_id,
            execution_options={"include_deleted": True},
        )
        image_a = session.get(
            ProductImageRow,
            image_a_id,
            execution_options={"include_deleted": True},
        )
        assert sku_a is not None and sku_a.status == "ACTIVE"
        sku_b = session.scalar(
            select(SkuRow).where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.sku_code == "TPL-API-002",
            )
        )
        assert sku_b is not None
        assert sku_b.weight == Decimal("2.500000")
        assert sku_b.weight_unit == "kg"
        assert sku_b.default_moq == Decimal("10.000000")
        assert sku_b.option_values["装箱数"] == "24"
        assert sku_b.option_values["毛重"] == "2.5"
        assert sku_b.option_values["起定数"] == "10"
        assert "一箱个数" not in sku_b.option_values
        assert "是否是新品" not in sku_b.option_values
        assert product_a is not None and product_a.status == "ACTIVE"
        assert offer_a is not None and offer_a.publication_status == "PUBLISHED"
        assert image_a is not None and image_a.deleted_at is None

    restored = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "商品模版.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"source_type": "PRODUCT_TEMPLATE"},
    )
    assert restored.status_code == 201, restored.text
    created_import_job_ids.append(restored.json()["id"])
    with SessionLocal() as session:
        sku_a = session.scalar(
            select(SkuRow).where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.sku_code == "TPL-API-001",
            )
        )
        product_a = session.get(ProductRow, product_a_id)
        offer_a = session.get(PublicCatalogOfferRow, offer_a_id)
        image_a = session.get(ProductImageRow, image_a_id)
        assert sku_a is not None and sku_a.status == "ACTIVE"
        assert product_a is not None and product_a.status == "ACTIVE"
        assert offer_a is not None and offer_a.publication_status == "PUBLISHED"
        assert offer_a.unit_price == Decimal("12.50")
        assert image_a is not None and image_a.deleted_at is None


def test_file_security_clean_upload_promotes_object_before_parsing() -> None:
    response = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "clean-products.csv",
                b"name,model,material\nClean Toy,CLEAN-1,TPR\n",
                "text/csv",
            )
        },
        data={"supplier_id": "SUP-001", "source_type": "SUPPLIER_QUOTE"},
    )
    assert response.status_code == 201, response.text
    job_id = response.json()["id"]
    with SessionLocal() as session:
        import_job = session.get(ImportJobRow, job_id)
        assert import_job is not None
        source = import_job.source_file
        assert source.security_status == "ACCEPTED"
        assert source.media_object_id is not None
        media = session.get(MediaObjectRow, source.media_object_id)
        assert media is not None
        assert (media.zone, media.status, media.scan_status) == (
            "SOURCE",
            "AVAILABLE",
            "CLEAN",
        )
        worker_job = session.scalar(
            select(WorkerJobRow).where(WorkerJobRow.import_job_id == job_id)
        )
        assert worker_job is not None
        assert worker_job.status == "SUCCEEDED"
        assert worker_job.checkpoint["outcome"] == "PARSED"
        source_key = media.object_key
    storage = get_object_storage()
    source_path = storage.local_path(source_key)
    assert source_path is not None and source_path.is_file()
    quarantine_path = storage.local_path(
        source_key.replace("/source/", "/quarantine/", 1)
    )
    assert quarantine_path is not None and not quarantine_path.exists()


def test_malware_marker_stays_quarantined_and_never_reaches_parser() -> None:
    response = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "malicious.csv",
                b"name,model\nATC-MALWARE-TEST,BLOCKED-1\n",
                "text/csv",
            )
        },
        data={"supplier_id": "SUP-001", "source_type": "SUPPLIER_QUOTE"},
    )
    assert response.status_code == 201, response.text
    payload = response.json()
    assert payload["status"] == "failed"
    assert payload["candidate_fields"] == 0
    with SessionLocal() as session:
        import_job = session.get(ImportJobRow, payload["id"])
        assert import_job is not None
        source = import_job.source_file
        media = session.get(MediaObjectRow, source.media_object_id)
        worker_job = session.scalar(
            select(WorkerJobRow).where(WorkerJobRow.import_job_id == import_job.id)
        )
        assert media is not None and worker_job is not None
        assert source.security_status == "QUARANTINED"
        assert (media.zone, media.status, media.scan_status) == (
            "QUARANTINE",
            "REJECTED",
            "INFECTED",
        )
        assert worker_job.status == "SUCCEEDED"
        assert worker_job.checkpoint["outcome"] == "QUARANTINED"
        assert session.scalar(
            select(func.count()).select_from(ReviewItemRow).where(
                ReviewItemRow.job_id == import_job.id
            )
        ) == 0
        assert session.scalar(
            select(func.count()).select_from(AITaskRow).where(
                AITaskRow.business_entity_id == source.id
            )
        ) == 0


def test_inline_startup_immediately_resumes_worker_from_stopped_api_process(
    monkeypatch: pytest.MonkeyPatch,
    request: pytest.FixtureRequest,
) -> None:
    monkeypatch.setenv("FILE_WORKER_INLINE", "false")
    response = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "restart-recovery.csv",
                b"name,model\nRestart Recovery,RESTART-RECOVERY-1\n",
                "text/csv",
            )
        },
        data={"supplier_id": "SUP-001", "source_type": "SUPPLIER_QUOTE"},
    )
    assert response.status_code == 201, response.text
    import_job_id = response.json()["id"]
    request.addfinalizer(
        lambda: _cleanup_template_test_records(
            import_job_ids=[import_job_id],
            sku_codes=[],
            category_names=[],
        )
    )

    future_lease = datetime.now(UTC) + timedelta(minutes=10)
    with SessionLocal() as session:
        worker_job = session.scalar(
            select(WorkerJobRow).where(
                WorkerJobRow.tenant_id == DEFAULT_TENANT_ID,
                WorkerJobRow.import_job_id == import_job_id,
            )
        )
        assert worker_job is not None
        worker_job.status = "RUNNING"
        worker_job.attempt_count = 1
        worker_job.lease_owner = "stopped-api-process"
        worker_job.lease_expires_at = future_lease
        session.commit()

    submitted: list[tuple[object, UUID, str]] = []

    def capture_submit(
        function: object,
        *,
        tenant_id: UUID,
        import_job_id: str,
    ) -> SimpleNamespace:
        submitted.append((function, tenant_id, import_job_id))
        return SimpleNamespace()

    monkeypatch.setenv("FILE_WORKER_INLINE", "true")
    monkeypatch.setattr(
        legacy_operations_use_cases._deferred_import_executor,
        "submit",
        capture_submit,
    )

    resumed = legacy_operations_use_cases.resume_deferred_imports()

    assert resumed >= 1
    assert (
        legacy_operations_use_cases.process_deferred_import,
        DEFAULT_TENANT_ID,
        import_job_id,
    ) in submitted
    with SessionLocal() as session:
        worker_job = session.scalar(
            select(WorkerJobRow).where(
                WorkerJobRow.tenant_id == DEFAULT_TENANT_ID,
                WorkerJobRow.import_job_id == import_job_id,
            )
        )
        assert worker_job is not None
        assert worker_job.status == "RUNNING"
        assert worker_job.lease_expires_at is not None
        recovered_lease = worker_job.lease_expires_at
        if recovered_lease.tzinfo is None:
            recovered_lease = recovered_lease.replace(tzinfo=UTC)
        assert recovered_lease < future_lease


def test_persistent_file_worker_recovers_scanner_failure_and_expired_lease(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("FILE_WORKER_INLINE", "false")
    response = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "recoverable.csv",
                b"name,model\nRecovered Toy,RECOVER-1\n",
                "text/csv",
            )
        },
        data={"supplier_id": "SUP-001", "source_type": "SUPPLIER_QUOTE"},
    )
    assert response.status_code == 201, response.text
    assert response.json()["status"] == "scanning"
    import_job_id = response.json()["id"]
    with SessionLocal() as session:
        worker_job = session.scalar(
            select(WorkerJobRow).where(WorkerJobRow.import_job_id == import_job_id)
        )
        assert worker_job is not None and worker_job.status == "PENDING"
        worker_job_id = worker_job.id

    class UnavailableScanner:
        engine_name = "unavailable-test-scanner"

        def scan(self, _path: Path) -> object:
            raise ConnectionError("scanner unavailable")

    first_now = datetime.now(UTC)
    with SessionLocal() as session:
        failed = process_file_worker_job(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            job_id=worker_job_id,
            worker_id="worker-before-restart",
            scanner=UnavailableScanner(),  # type: ignore[arg-type]
            now=first_now,
        )
    assert failed.status == "RETRY"

    with SessionLocal() as session:
        worker_job = session.get(WorkerJobRow, worker_job_id)
        assert worker_job is not None
        worker_job.status = "RUNNING"
        worker_job.lease_owner = "crashed-worker"
        worker_job.lease_expires_at = first_now - timedelta(seconds=1)
        session.commit()

    with SessionLocal() as session:
        recovered = process_file_worker_job(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            job_id=worker_job_id,
            worker_id="worker-after-restart",
            scanner=DeterministicDevelopmentScanner(),
            now=first_now + timedelta(seconds=5),
        )
    assert recovered.status == "SUCCEEDED"
    assert recovered.outcome == "PARSED"
    with SessionLocal() as session:
        worker_job = session.get(WorkerJobRow, worker_job_id)
        import_job = session.get(ImportJobRow, import_job_id)
        assert worker_job is not None and import_job is not None
        assert worker_job.attempt_count == 2
        assert worker_job.lease_owner is None
        assert import_job.status == "needs_review"


def test_product_template_worker_retry_resumes_from_promoted_source(
    monkeypatch: pytest.MonkeyPatch,
    request: pytest.FixtureRequest,
) -> None:
    monkeypatch.setenv("FILE_WORKER_INLINE", "false")
    template_bytes = client.get("/api/v1/product-template.xlsx").content
    response = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "商品模版.xlsx",
                template_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"source_type": "PRODUCT_TEMPLATE"},
    )
    assert response.status_code == 201, response.text
    import_job_id = response.json()["id"]

    def cleanup_import_graph() -> None:
        with SessionLocal() as session:
            worker_rows = session.scalars(
                select(WorkerJobRow).where(
                    WorkerJobRow.tenant_id == DEFAULT_TENANT_ID,
                    WorkerJobRow.import_job_id == import_job_id,
                )
            ).all()
            source_file_ids = [row.source_file_id for row in worker_rows]
            media_object_ids = [row.media_object_id for row in worker_rows]
            session.execute(
                delete(WorkerJobRow).where(
                    WorkerJobRow.tenant_id == DEFAULT_TENANT_ID,
                    WorkerJobRow.import_job_id == import_job_id,
                )
            )
            session.execute(
                delete(ImportJobRow).where(
                    ImportJobRow.tenant_id == DEFAULT_TENANT_ID,
                    ImportJobRow.id == import_job_id,
                )
            )
            if source_file_ids:
                session.execute(
                    delete(SourceFileRow).where(
                        SourceFileRow.tenant_id == DEFAULT_TENANT_ID,
                        SourceFileRow.id.in_(source_file_ids),
                    )
                )
            if media_object_ids:
                session.execute(
                    delete(MediaObjectRow).where(
                        MediaObjectRow.tenant_id == DEFAULT_TENANT_ID,
                        MediaObjectRow.id.in_(media_object_ids),
                    )
                )
            session.commit()

    request.addfinalizer(cleanup_import_graph)
    with SessionLocal() as session:
        worker_job = session.scalar(
            select(WorkerJobRow).where(WorkerJobRow.import_job_id == import_job_id)
        )
        assert worker_job is not None
        worker_job_id = worker_job.id

    original_processor = file_processing_worker.process_product_template_import

    def transient_import_failure(*_args: object, **_kwargs: object) -> object:
        raise ConnectionError("temporary importer dependency failure")

    monkeypatch.setattr(
        file_processing_worker,
        "process_product_template_import",
        transient_import_failure,
    )
    first_now = datetime.now(UTC)
    with SessionLocal() as session:
        first = process_file_worker_job(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            job_id=worker_job_id,
            worker_id="template-import-first-attempt",
            scanner=DeterministicDevelopmentScanner(),
            now=first_now,
        )
    assert first.status == "RETRY"

    with SessionLocal() as session:
        worker_job = session.get(WorkerJobRow, worker_job_id)
        import_job = session.get(ImportJobRow, import_job_id)
        assert worker_job is not None and import_job is not None
        media = session.get(MediaObjectRow, worker_job.media_object_id)
        source = session.get(SourceFileRow, worker_job.source_file_id)
        assert media is not None and source is not None
        assert worker_job.checkpoint["promoted"] is True
        assert worker_job.checkpoint["last_error_stage"] == "IMPORT"
        assert (media.zone, media.status, media.scan_status) == (
            "SOURCE",
            "AVAILABLE",
            "CLEAN",
        )
        assert source.security_status == "ACCEPTED"
        source_key = media.object_key

    class MustNotRescan:
        engine_name = "must-not-rescan"

        def scan(self, _path: Path) -> object:
            raise AssertionError("promoted source must not be scanned or promoted again")

    monkeypatch.setattr(
        file_processing_worker,
        "process_product_template_import",
        original_processor,
    )
    with SessionLocal() as session:
        recovered = process_file_worker_job(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            job_id=worker_job_id,
            worker_id="template-import-retry",
            scanner=MustNotRescan(),  # type: ignore[arg-type]
            now=first_now + timedelta(seconds=5),
        )
    assert recovered.status == "SUCCEEDED"
    assert recovered.outcome == "TEMPLATE_REJECTED"

    storage = get_object_storage()
    source_path = storage.local_path(source_key)
    quarantine_path = storage.local_path(
        source_key.replace("/source/", "/quarantine/", 1)
    )
    assert source_path is not None and source_path.is_file()
    assert quarantine_path is not None and not quarantine_path.exists()
    import_response = client.get(f"/api/v1/imports/{import_job_id}")
    assert import_response.status_code == 200
    assert import_response.json()["warning_messages"] == [
        "Product 表中没有可导入的有效商品。"
    ]


def test_older_product_template_retry_cannot_override_newer_snapshot(
    monkeypatch: pytest.MonkeyPatch,
    request: pytest.FixtureRequest,
) -> None:
    import_job_ids: list[str] = []
    sku_codes = ["SNAP-COMMON", "SNAP-OLD-ONLY", "SNAP-NEW-ONLY"]
    request.addfinalizer(
        lambda: _cleanup_template_test_records(
            import_job_ids=import_job_ids,
            sku_codes=sku_codes,
            category_names=["快照顺序测试"],
        )
    )
    _cleanup_template_test_records(
        import_job_ids=import_job_ids,
        sku_codes=sku_codes,
        category_names=["快照顺序测试"],
    )

    old_template = _product_template_bytes([
        [
            "旧版共同商品",
            "快照顺序测试",
            "SNAP-COMMON",
            "10",
            None,
            None,
            *([None] * 10),
        ],
        [
            "旧版独有商品",
            "快照顺序测试",
            "SNAP-OLD-ONLY",
            "11",
            None,
            None,
            *([None] * 10),
        ],
    ])
    newer_template = _product_template_bytes([
        [
            "新版共同商品",
            "快照顺序测试",
            "SNAP-COMMON",
            "20",
            None,
            None,
            *([None] * 10),
        ],
        [
            "新版独有商品",
            "快照顺序测试",
            "SNAP-NEW-ONLY",
            "21",
            None,
            None,
            *([None] * 10),
        ],
    ])

    monkeypatch.setenv("FILE_WORKER_INLINE", "false")
    old_response = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "旧商品模版.xlsx",
                old_template,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"source_type": "PRODUCT_TEMPLATE"},
    )
    assert old_response.status_code == 201, old_response.text
    old_job_id = old_response.json()["id"]
    import_job_ids.append(old_job_id)
    with SessionLocal() as session:
        old_worker = session.scalar(
            select(WorkerJobRow).where(WorkerJobRow.import_job_id == old_job_id)
        )
        assert old_worker is not None
        old_worker_id = old_worker.id

    monkeypatch.setenv("FILE_WORKER_INLINE", "true")
    newer_response = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "新商品模版.xlsx",
                newer_template,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"source_type": "PRODUCT_TEMPLATE"},
    )
    assert newer_response.status_code == 201, newer_response.text
    assert newer_response.json()["status"] == "published"
    newer_job_id = newer_response.json()["id"]
    import_job_ids.append(newer_job_id)

    # Remove clock-resolution and scheduling variance from the ordering
    # assertion: the newer snapshot is explicitly one second later.
    ordering_base = datetime.now(UTC) - timedelta(minutes=1)
    with SessionLocal() as session:
        old_job = session.get(ImportJobRow, old_job_id)
        newer_job = session.get(ImportJobRow, newer_job_id)
        common_before = session.scalar(
            select(SkuRow).where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.sku_code == "SNAP-COMMON",
            )
        )
        assert old_job is not None and newer_job is not None and common_before is not None
        old_job.created_at = ordering_base
        newer_job.created_at = ordering_base + timedelta(seconds=1)
        common_id = common_before.id
        common_version = common_before.version
        common_product_id = common_before.product_id
        common_offer = session.scalar(
            select(PublicCatalogOfferRow).where(
                PublicCatalogOfferRow.tenant_id == DEFAULT_TENANT_ID,
                PublicCatalogOfferRow.sku_id == common_before.id,
            )
        )
        assert common_offer is not None
        common_offer_id = common_offer.id
        newer_worker = session.scalar(
            select(WorkerJobRow).where(WorkerJobRow.import_job_id == newer_job_id)
        )
        assert newer_worker is not None
        newer_worker_id = newer_worker.id
        session.commit()

    # Simulate a failure after the template transaction has committed but
    # before the outer worker checkpoint is durable. The applied snapshot must
    # remain published so an older retry can still see it.
    with SessionLocal() as session:
        checkpoint_retry = file_processing_worker._record_retry(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            job_id=newer_worker_id,
            error=ConnectionError("checkpoint commit interrupted"),
            now=datetime.now(UTC),
        )
    assert checkpoint_retry.status == "RETRY"
    with SessionLocal() as session:
        newer_job = session.get(ImportJobRow, newer_job_id)
        newer_worker = session.get(WorkerJobRow, newer_worker_id)
        assert newer_job is not None and newer_job.status == "published"
        assert newer_worker is not None
        assert newer_worker.checkpoint["template_snapshot_committed"] is True

    with SessionLocal() as session:
        stale_result = process_file_worker_job(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            job_id=old_worker_id,
            worker_id="stale-template-retry",
            scanner=DeterministicDevelopmentScanner(),
            now=datetime.now(UTC),
        )
    assert stale_result.status == "SUCCEEDED"
    assert stale_result.outcome == "TEMPLATE_SUPERSEDED"

    stale_detail = client.get(f"/api/v1/imports/{old_job_id}")
    assert stale_detail.status_code == 200
    assert stale_detail.json()["status"] == "failed"
    assert stale_detail.json()["warnings"] == 1
    assert len(stale_detail.json()["warning_messages"]) == 1
    assert "早于已经生效的新版本" in stale_detail.json()["error_message"]

    with SessionLocal() as session:
        common_after = session.get(SkuRow, common_id)
        common_product = session.get(ProductRow, common_product_id)
        common_offer = session.get(PublicCatalogOfferRow, common_offer_id)
        new_only = session.scalar(
            select(SkuRow).where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.sku_code == "SNAP-NEW-ONLY",
            )
        )
        old_only = session.scalar(
            select(SkuRow)
            .where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.sku_code == "SNAP-OLD-ONLY",
            )
            .execution_options(include_deleted=True)
        )
        assert common_after is not None and common_after.version == common_version
        assert common_product is not None and common_product.name == "新版共同商品"
        assert common_offer is not None and common_offer.unit_price == Decimal("20.00")
        assert new_only is not None and new_only.status == "ACTIVE"
        assert old_only is None


@pytest.mark.parametrize("infected_after_recovery", [False, True])
def test_file_worker_recovers_promotion_completed_before_checkpoint_commit(
    monkeypatch: pytest.MonkeyPatch,
    request: pytest.FixtureRequest,
    infected_after_recovery: bool,
) -> None:
    import_job_ids: list[str] = []
    request.addfinalizer(
        lambda: _cleanup_template_test_records(
            import_job_ids=import_job_ids,
            sku_codes=[],
            category_names=[],
        )
    )
    monkeypatch.setenv("FILE_WORKER_INLINE", "false")
    response = client.post(
        "/api/v1/imports",
        files={
            "file": (
                "商品模版.xlsx",
                _product_template_bytes([]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"source_type": "PRODUCT_TEMPLATE"},
    )
    assert response.status_code == 201, response.text
    import_job_id = response.json()["id"]
    import_job_ids.append(import_job_id)

    crash_time = datetime.now(UTC)
    with SessionLocal() as session:
        worker = session.scalar(
            select(WorkerJobRow).where(WorkerJobRow.import_job_id == import_job_id)
        )
        assert worker is not None
        media = session.get(MediaObjectRow, worker.media_object_id)
        source = session.get(SourceFileRow, worker.source_file_id)
        import_job = session.get(ImportJobRow, import_job_id)
        assert media is not None and source is not None and import_job is not None
        worker_id = worker.id
        quarantine_key = media.object_key
        source_key = quarantine_key.replace("/quarantine/", "/source/", 1)
        worker.status = "RUNNING"
        worker.attempt_count = 1
        worker.lease_owner = "worker-crashed-after-promotion"
        worker.lease_expires_at = crash_time - timedelta(seconds=1)
        worker.checkpoint = {}
        media.status = "SCANNING"
        media.scan_status = "RUNNING"
        source.security_status = "SCANNING"
        import_job.status = "scanning"
        import_job.progress = 10
        session.commit()

    storage = get_object_storage()
    storage.promote(quarantine_key=quarantine_key, source_key=source_key)
    assert not storage.exists(quarantine_key)
    assert storage.exists(source_key)
    with SessionLocal() as session:
        worker = session.get(WorkerJobRow, worker_id)
        media = session.get(MediaObjectRow, worker.media_object_id) if worker else None
        assert worker is not None and media is not None
        assert worker.checkpoint == {}
        assert media.object_key == quarantine_key
        assert media.zone == "QUARANTINE"

    recovery_promotions: list[tuple[str, str]] = []
    expected_recovered_source_key = source_key
    expected_quarantine_key = quarantine_key

    class RecoveryStorage:
        backend_name = storage.backend_name

        def put_file(self, *args: object, **kwargs: object) -> None:
            storage.put_file(*args, **kwargs)  # type: ignore[arg-type]

        def promote(self, *, quarantine_key: str, source_key: str) -> None:
            recovery_promotions.append((quarantine_key, source_key))
            if not infected_after_recovery:
                raise AssertionError("clean recovered source must not be promoted a second time")
            assert (quarantine_key, source_key) == (
                expected_recovered_source_key,
                expected_quarantine_key,
            )
            storage.promote(quarantine_key=quarantine_key, source_key=source_key)

        def exists(self, object_key: str) -> bool:
            return storage.exists(object_key)

        def delete(self, object_key: str) -> None:
            storage.delete(object_key)

        def materialize(self, object_key: str):
            return storage.materialize(object_key)

        def local_path(self, object_key: str):
            return storage.local_path(object_key)

    scanned_paths: list[Path] = []

    class CountingScanner:
        engine_name = "counting-development-scanner"

        def scan(self, path: Path):
            scanned_paths.append(path)
            if infected_after_recovery:
                return FileScanResult(
                    clean=False,
                    engine=self.engine_name,
                    signature="RECOVERED-INFECTED-TEST",
                    detail_code="RECOVERED_INFECTED",
                )
            return DeterministicDevelopmentScanner().scan(path)

    with SessionLocal() as session:
        recovered = process_file_worker_job(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            job_id=worker_id,
            worker_id="worker-after-promotion-crash",
            storage=RecoveryStorage(),  # type: ignore[arg-type]
            scanner=CountingScanner(),  # type: ignore[arg-type]
            now=crash_time,
        )
    assert recovered.status == "SUCCEEDED"
    assert recovered.outcome == (
        "QUARANTINED" if infected_after_recovery else "TEMPLATE_REJECTED"
    )
    assert scanned_paths == [storage.local_path(source_key)]

    with SessionLocal() as session:
        worker = session.get(WorkerJobRow, worker_id)
        media = session.get(MediaObjectRow, worker.media_object_id) if worker else None
        source = session.get(SourceFileRow, worker.source_file_id) if worker else None
        assert worker is not None and media is not None and source is not None
        assert worker.attempt_count == 2
        assert worker.checkpoint["promotion_recovered"] is True
        if infected_after_recovery:
            assert not worker.checkpoint.get("promoted", False)
            assert worker.checkpoint["outcome"] == "QUARANTINED"
            assert (media.object_key, media.zone, media.status, media.scan_status) == (
                quarantine_key,
                "QUARANTINE",
                "REJECTED",
                "INFECTED",
            )
            assert source.security_status == "QUARANTINED"
        else:
            assert worker.checkpoint["promoted"] is True
            assert worker.checkpoint["outcome"] == "TEMPLATE_REJECTED"
            assert (media.object_key, media.zone, media.status, media.scan_status) == (
                source_key,
                "SOURCE",
                "AVAILABLE",
                "CLEAN",
            )
            assert source.security_status == "ACCEPTED"
    if infected_after_recovery:
        assert recovery_promotions == [(source_key, quarantine_key)]
        assert storage.exists(quarantine_key)
        assert not storage.exists(source_key)
    else:
        assert recovery_promotions == []
        assert storage.exists(source_key)
        assert not storage.exists(quarantine_key)


def test_development_scanner_is_fail_closed_in_production(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("APP_ENV", "production")
    monkeypatch.setenv("FILE_SCANNER_PROFILE", "development")
    with pytest.raises(RuntimeError, match="forbidden in production"):
        get_file_scanner()


def test_phase4a1b_real_csv_upload_candidate_evidence_and_content_idempotency() -> None:
    content = (
        "Product Name,SKU,Material,Size,MOQ,Packing\r\n"
        "Waterproof Dog Toy,DOG-001,TPR,10cm,100,12 pcs/carton\r\n"
    ).encode("utf-8-sig")

    with SessionLocal() as session:
        product_count_before = int(
            session.scalar(select(func.count()).select_from(ProductRow)) or 0
        )
        embedding_count_before = int(
            session.scalar(select(func.count()).select_from(EmbeddingRow)) or 0
        )

    first_response = client.post(
        "/api/v1/imports",
        files={"file": ("supplier_products.csv", content, "text/csv")},
        data={"supplier_id": "SUP-001", "source_type": "SUPPLIER_CATALOG"},
    )
    assert first_response.status_code == 201
    first = first_response.json()
    assert first["detected_type"] == "TEXT / CSV"
    assert first["parser"] == "python-csv"
    assert first["products"] == 1
    assert first["candidate_status"] == "NEEDS_REVIEW"
    assert first["candidate_fields"] == 6
    assert first["candidate_idempotent"] is False

    candidates_response = client.get(
        f"/api/v1/ai/product-intelligence/tasks/{first['ai_task_id']}/candidates"
    )
    assert candidates_response.status_code == 200
    candidates = candidates_response.json()
    assert len(candidates) == 6
    assert all(item["review_status"] == "AI_SUGGESTED" for item in candidates)
    material = next(item for item in candidates if item["field_key"] == "material")
    assert material["raw_value"] == "TPR"
    assert material["evidence"]["location"] == {
        "sheet": "supplier_products",
        "range": "C2",
    }
    assert len(material["evidence"]["raw_value_hash"]) == 64

    repeated_response = client.post(
        "/api/v1/imports",
        files={"file": ("renamed_supplier_file.csv", content, "text/csv")},
        data={"supplier_id": "SUP-001", "source_type": "SUPPLIER_CATALOG"},
    )
    assert repeated_response.status_code == 201
    repeated = repeated_response.json()
    assert repeated["ai_task_id"] == first["ai_task_id"]
    assert repeated["candidate_fields"] == 6
    assert repeated["candidate_idempotent"] is True

    with SessionLocal() as session:
        task_id = UUID(first["ai_task_id"])
        runs = session.scalars(
            select(AIRunRow).where(AIRunRow.ai_task_id == task_id)
        ).all()
        assert len(runs) == 1
        assert runs[0].provider_type == "NATIVE"
        assert runs[0].usage == {}
        assert session.scalar(select(func.count()).select_from(ProductRow)) == product_count_before
        assert session.scalar(select(func.count()).select_from(EmbeddingRow)) == embedding_count_before


def test_phase4a1b_native_pipeline_recovers_when_source_file_becomes_available(
    tmp_path: Path,
) -> None:
    content = (
        "Product Name,Material,MOQ\r\n"
        "Recoverable Dog Toy,TPR,200\r\n"
    ).encode("utf-8")
    missing_path = tmp_path / "temporarily-unavailable.csv"
    source_id = f"SRC-{uuid4().hex[:12].upper()}"
    source_hash = hashlib.sha256(content).hexdigest()

    with SessionLocal() as session:
        session.add(SourceFileRow(
            id=source_id,
            tenant_id=DEFAULT_TENANT_ID,
            original_filename="recoverable.csv",
            stored_filename=missing_path.name,
            local_path=str(missing_path),
            sha256=source_hash,
            byte_size=len(content),
            extension=".csv",
            detected_type="TEXT / CSV",
            extension_matches=True,
            parser="python-csv",
        ))
        session.commit()

        first = run_product_draft_workflow(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            source_file_id=source_id,
            parser=NativeSupplierFileParserAdapter(),
        )
        session.commit()
        assert first.status == "PARTIAL"
        assert first.error_code == "PRODUCT_DRAFT_PIPELINE_FAILURE"
        assert first.candidate_fields == 0

        missing_path.write_bytes(content)
        second = run_product_draft_workflow(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            source_file_id=source_id,
            parser=NativeSupplierFileParserAdapter(),
        )
        session.commit()
        assert second.status == "NEEDS_REVIEW"
        assert second.recovered is True
        assert second.candidate_fields == 3
        runs = session.scalars(
            select(AIRunRow)
            .where(AIRunRow.ai_task_id == second.task_id)
            .order_by(AIRunRow.attempt_number)
        ).all()
        assert [(run.attempt_number, run.status) for run in runs] == [
            (1, "FAILED"),
            (2, "SUCCEEDED"),
        ]


def test_phase4a1c_deterministic_normalization_preserves_rules_and_review_warnings() -> None:
    capacity = normalize_product_field("capacity", "3000 ml")
    assert capacity.value == {"value": "3", "unit": "L"}
    assert capacity.unit == "L"
    assert capacity.rule_version == "product-normalization-v1"
    assert any(step["rule"] == "millilitre-to-litre" for step in capacity.trace)

    dimensions = normalize_product_field("specification", "10 x 8 x 5 CM")
    assert dimensions.value == {
        "dimensions": ["10", "8", "5"],
        "axis_order": "UNCONFIRMED",
        "unit": "cm",
    }
    assert dimensions.validation_status == "WARNING"
    assert "DIMENSION_AXIS_ORDER_REVIEW_REQUIRED" in dimensions.warnings

    unknown_material = normalize_product_field("material", "Supplier Secret Blend")
    assert unknown_material.value == {"value": "Supplier Secret Blend"}
    assert "MATERIAL_VOCABULARY_REVIEW_REQUIRED" in unknown_material.warnings

    colors = normalize_product_field("color", "Red / Blue")
    assert colors.value["variant_relation"] == "UNCONFIRMED"
    assert "COLOR_VARIANT_RELATION_REVIEW_REQUIRED" in colors.warnings

    invalid_moq = normalize_product_field("moq", "-10 pcs")
    assert invalid_moq.validation_status == "FAILED"
    assert "MOQ_MUST_BE_NONNEGATIVE" in invalid_moq.warnings


def test_phase4a1c_human_adoption_versions_outbox_projection_and_failure_isolation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook_path = tmp_path / "phase4a1c-products.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(["Product Name", "Item No", "Material", "Size", "Color", "MOQ", "Packing"])
    sheet.append(["Waterproof Dog Toy", "DOG-100", "TPR", "10 x 8 x 5 cm", "Red / Blue", 100, "12 pcs/carton"])
    sheet.append(["Durable Dog Ball", "DOG-200", "Rubber", "8 cm", "Green", 200, "24 pcs/carton"])
    workbook.save(workbook_path)
    source_hash = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    source_id = f"SRC-{uuid4().hex[:12].upper()}"
    job_id = f"JOB-{uuid4().hex[:12].upper()}"

    with SessionLocal() as session:
        supplier = session.scalar(
            select(SupplierRow).where(
                SupplierRow.tenant_id == DEFAULT_TENANT_ID,
                SupplierRow.status == "ACTIVE",
            )
        )
        assert supplier is not None
        session.add(SourceFileRow(
            id=source_id,
            tenant_id=DEFAULT_TENANT_ID,
            original_filename=workbook_path.name,
            stored_filename=workbook_path.name,
            local_path=str(workbook_path),
            sha256=source_hash,
            byte_size=workbook_path.stat().st_size,
            extension=".xlsx",
            detected_type="OOXML / XLSX",
            extension_matches=True,
            parser="openpyxl",
        ))
        session.add(ImportJobRow(
            id=job_id,
            tenant_id=DEFAULT_TENANT_ID,
            source_file_id=source_id,
            supplier_id=supplier.id,
            supplier_name=supplier.name,
            source_type="SUPPLIER_CATALOG",
            status="needs_review",
            progress=100,
            products_count=2,
        ))
        session.commit()
        workflow = run_product_draft_workflow(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            source_file_id=source_id,
            parser=NativeSupplierFileParserAdapter(),
            idempotency_context=f"supplier:{supplier.id}",
        )
        session.commit()
        candidates = session.scalars(
            select(ProductFieldCandidateRow)
            .where(
                ProductFieldCandidateRow.tenant_id == DEFAULT_TENANT_ID,
                ProductFieldCandidateRow.ai_task_id == workflow.task_id,
            )
            .order_by(ProductFieldCandidateRow.candidate_index, ProductFieldCandidateRow.field_key)
        ).all()
        groups: dict[str, dict[str, str]] = {}
        for candidate in candidates:
            groups.setdefault(candidate.candidate_group_key, {})[candidate.field_key] = candidate.raw_value
        ordered_groups = list(groups.items())
        assert len(ordered_groups) == 2
        assert all(candidate.normalization_rule_version == "product-normalization-v1" for candidate in candidates)
        assert all(candidate.normalization_trace for candidate in candidates)
        product_count_before = int(
            session.scalar(select(func.count()).select_from(ProductRow)) or 0
        )
        knowledge_count_before = int(
            session.scalar(select(func.count()).select_from(KnowledgeDocumentRow)) or 0
        )

    first_group_key, first_values = ordered_groups[0]
    approval_payload = {
        "idempotency_key": f"approve-{uuid4()}",
        "confirmed_values": first_values,
        "activate": True,
        "product_code": f"ATC-{uuid4().hex[:10].upper()}",
        "change_reason": "Phase 4A-1C human acceptance test",
    }
    approve_response = client.post(
        f"/api/v1/ai/product-intelligence/tasks/{workflow.task_id}/groups/{first_group_key}/approve",
        json=approval_payload,
    )
    assert approve_response.status_code == 202
    approved = approve_response.json()
    assert approved["product_version"] == 1
    assert approved["outbox_status"] == "PENDING"
    assert approved["idempotent"] is False

    repeated_response = client.post(
        f"/api/v1/ai/product-intelligence/tasks/{workflow.task_id}/groups/{first_group_key}/approve",
        json=approval_payload,
    )
    assert repeated_response.status_code == 202
    repeated = repeated_response.json()
    assert repeated["decision_id"] == approved["decision_id"]
    assert repeated["product_id"] == approved["product_id"]
    assert repeated["idempotent"] is True

    reviewed_candidates = client.get(
        f"/api/v1/ai/product-intelligence/tasks/{workflow.task_id}/candidates"
    )
    assert reviewed_candidates.status_code == 200
    reviewed_first_group = [
        item
        for item in reviewed_candidates.json()
        if item["candidate_group_key"] == first_group_key
    ]
    assert reviewed_first_group
    assert all(item["review_status"] == "AI_SUGGESTED" for item in reviewed_first_group)
    assert all(item["latest_decision"]["action"] == "APPROVE" for item in reviewed_first_group)
    assert all(item["latest_decision"]["status"] == "APPLIED" for item in reviewed_first_group)
    assert all(item["latest_decision"]["product_id"] == approved["product_id"] for item in reviewed_first_group)

    with SessionLocal() as session:
        product_id = UUID(approved["product_id"])
        outbox_event_id = UUID(approved["outbox_event_id"])
        product = session.scalar(
            select(ProductRow).where(
                ProductRow.tenant_id == DEFAULT_TENANT_ID,
                ProductRow.id == product_id,
            )
        )
        assert product is not None
        assert product.name == "Waterproof Dog Toy"
        assert product.status == "ACTIVE"
        assert product.current_version == 1
        assert product.search_document_version == 0
        assert session.scalar(
            select(func.count())
            .select_from(ProductVersionRow)
            .where(ProductVersionRow.product_id == product_id)
        ) == 1
        decision = session.get(ProductCandidateDecisionRow, UUID(approved["decision_id"]))
        assert decision is not None
        assert decision.normalization_snapshot["material"]["raw_value"] == "TPR"
        assert decision.normalization_snapshot["material"]["human_value"] == "TPR"
        assert decision.reviewed_by_membership_id == DEFAULT_MEMBERSHIP_ID
        task = session.get(AITaskRow, workflow.task_id)
        assert task is not None and task.status == "NEEDS_REVIEW"
        attributes = session.scalars(
            select(ProductAttributeRow).where(
                ProductAttributeRow.tenant_id == DEFAULT_TENANT_ID,
                ProductAttributeRow.product_id == product_id,
            )
        ).all()
        assert attributes
        assert all(attribute.review_status == "CONFIRMED" for attribute in attributes)
        assert any(attribute.attribute_key == "material" and attribute.value_text == "TPR" for attribute in attributes)
        supplier_product = session.scalar(
            select(SupplierProductRow).where(
                SupplierProductRow.tenant_id == DEFAULT_TENANT_ID,
                SupplierProductRow.product_id == product_id,
            )
        )
        assert supplier_product is not None
        assert supplier_product.supplier_sku == "DOG-100"
        assert supplier_product.moq == Decimal("100")
        assert supplier_product.moq_unit == "piece"
        assert session.scalar(
            select(func.count())
            .select_from(KnowledgeDocumentRow)
            .where(KnowledgeDocumentRow.source_entity_id == product_id)
        ) == 0

        dispatched = dispatch_product_committed_event(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            event_id=outbox_event_id,
        )
        session.commit()
        assert dispatched.status == "PUBLISHED"
        assert dispatched.document_id is not None
        session.refresh(product)
        assert product.search_document_version == 1
        assert session.scalar(
            select(func.count())
            .select_from(KnowledgeDocumentRow)
            .where(
                KnowledgeDocumentRow.source_entity_id == product_id,
                KnowledgeDocumentRow.source_version == 1,
                KnowledgeDocumentRow.status == "ACTIVE",
            )
        ) == 1
        repeated_dispatch = dispatch_product_committed_event(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            event_id=outbox_event_id,
        )
        assert repeated_dispatch.idempotent is True

        with pytest.raises(ProductAdoptionError) as duplicate_group:
            approve_candidate_group(
                session,
                tenant_id=DEFAULT_TENANT_ID,
                task_id=workflow.task_id,
                candidate_group_key=first_group_key,
                reviewer_membership_id=DEFAULT_MEMBERSHIP_ID,
                idempotency_key=f"approve-{uuid4()}",
                confirmed_values=first_values,
                activate=True,
            )
        assert duplicate_group.value.code == "CANDIDATE_GROUP_ALREADY_APPLIED"

    second_group_key, second_values = ordered_groups[1]
    reject_response = client.post(
        f"/api/v1/ai/product-intelligence/tasks/{workflow.task_id}/groups/{second_group_key}/reject",
        json={
            "idempotency_key": f"reject-{uuid4()}",
            "reason": "Human reviewer requests a corrected source row",
        },
    )
    assert reject_response.status_code == 200
    assert reject_response.json()["status"] == "RECORDED"
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ProductRow)) == product_count_before + 1
        task = session.get(AITaskRow, workflow.task_id)
        assert task is not None and task.status == "SUCCEEDED"

    second_approval = client.post(
        f"/api/v1/ai/product-intelligence/tasks/{workflow.task_id}/groups/{second_group_key}/approve",
        json={
            "idempotency_key": f"approve-{uuid4()}",
            "confirmed_values": second_values,
            "activate": True,
            "product_code": f"ATC-{uuid4().hex[:10].upper()}",
        },
    )
    assert second_approval.status_code == 202
    second = second_approval.json()

    import app.services.product_intelligence.adoption as adoption_module

    def fail_projection(*args: object, **kwargs: object) -> None:
        raise RuntimeError("synthetic projector outage")

    monkeypatch.setattr(adoption_module, "project_product_knowledge", fail_projection)
    with SessionLocal() as session:
        failed_dispatch = dispatch_product_committed_event(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            event_id=UUID(second["outbox_event_id"]),
        )
        session.commit()
        assert failed_dispatch.status == "FAILED"
        assert failed_dispatch.error_code == "KNOWLEDGE_PROJECTION_FAILED"
        persisted_product = session.scalar(
            select(ProductRow).where(ProductRow.id == UUID(second["product_id"]))
        )
        assert persisted_product is not None
        assert persisted_product.status == "ACTIVE"
        assert persisted_product.search_document_version == 0
        failed_event = session.get(OutboxEventRow, UUID(second["outbox_event_id"]))
        assert failed_event is not None and failed_event.attempt_count == 1
        assert session.scalar(select(func.count()).select_from(ProductRow)) == product_count_before + 2
        assert session.scalar(select(func.count()).select_from(KnowledgeDocumentRow)) == knowledge_count_before + 1

        organization_id = uuid4()
        tenant_b = uuid4()
        membership_b = uuid4()
        session.add(OrganizationRow(
            id=organization_id,
            code=f"P1C-{organization_id.hex[:8]}",
            name="Phase 4A-1C Isolation Organization",
        ))
        session.flush()
        session.add(TenantRow(
            id=tenant_b,
            organization_id=organization_id,
            slug=f"phase4a1c-{tenant_b.hex[:8]}",
            name="Phase 4A-1C Tenant B",
        ))
        session.flush()
        session.add(MembershipRow(
            id=membership_b,
            tenant_id=tenant_b,
            user_id=DEFAULT_OWNER_USER_ID,
            status="active",
        ))
        session.commit()
        with pytest.raises(ProductAdoptionError) as cross_tenant:
            approve_candidate_group(
                session,
                tenant_id=tenant_b,
                task_id=workflow.task_id,
                candidate_group_key=first_group_key,
                reviewer_membership_id=membership_b,
                idempotency_key=f"approve-{uuid4()}",
                confirmed_values=first_values,
                activate=True,
            )
        assert cross_tenant.value.code == "CANDIDATE_GROUP_NOT_FOUND"


def test_phase4a1c_update_requires_expected_version_and_creates_new_snapshot(
    tmp_path: Path,
) -> None:
    source_path = tmp_path / "phase4a1c-update.csv"
    source_path.write_text(
        "Product Name,Material,MOQ\nVersioned Product Updated,Silicone,300\n",
        encoding="utf-8",
    )
    source_id = f"SRC-{uuid4().hex[:12].upper()}"
    job_id = f"JOB-{uuid4().hex[:12].upper()}"
    product_id = uuid4()
    with SessionLocal() as session:
        supplier = session.scalar(
            select(SupplierRow).where(SupplierRow.tenant_id == DEFAULT_TENANT_ID)
        )
        assert supplier is not None
        session.add(ProductRow(
            id=product_id,
            tenant_id=DEFAULT_TENANT_ID,
            product_code=f"LEGACY-{uuid4().hex[:8]}",
            name="Versioned Product Original",
            status="ACTIVE",
            current_version=1,
            search_document_version=0,
            created_by=DEFAULT_OWNER_USER_ID,
            updated_by=DEFAULT_OWNER_USER_ID,
        ))
        session.add(SourceFileRow(
            id=source_id,
            tenant_id=DEFAULT_TENANT_ID,
            original_filename=source_path.name,
            stored_filename=source_path.name,
            local_path=str(source_path),
            sha256=hashlib.sha256(source_path.read_bytes()).hexdigest(),
            byte_size=source_path.stat().st_size,
            extension=".csv",
            detected_type="TEXT / CSV",
            extension_matches=True,
            parser="python-csv",
        ))
        session.add(ImportJobRow(
            id=job_id,
            tenant_id=DEFAULT_TENANT_ID,
            source_file_id=source_id,
            supplier_id=supplier.id,
            supplier_name=supplier.name,
            status="needs_review",
            progress=100,
        ))
        session.commit()
        workflow = run_product_draft_workflow(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            source_file_id=source_id,
            parser=NativeSupplierFileParserAdapter(),
            idempotency_context=f"supplier:{supplier.id}",
        )
        session.commit()
        candidates = session.scalars(
            select(ProductFieldCandidateRow).where(
                ProductFieldCandidateRow.ai_task_id == workflow.task_id
            )
        ).all()
        group_key = candidates[0].candidate_group_key
        values = {candidate.field_key: candidate.raw_value for candidate in candidates}

        with pytest.raises(ProductAdoptionError) as missing_version:
            approve_candidate_group(
                session,
                tenant_id=DEFAULT_TENANT_ID,
                task_id=workflow.task_id,
                candidate_group_key=group_key,
                reviewer_membership_id=DEFAULT_MEMBERSHIP_ID,
                idempotency_key=f"approve-{uuid4()}",
                confirmed_values=values,
                activate=True,
                target_product_id=product_id,
            )
        assert missing_version.value.code == "EXPECTED_PRODUCT_VERSION_REQUIRED"
        session.rollback()

        result = approve_candidate_group(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            task_id=workflow.task_id,
            candidate_group_key=group_key,
            reviewer_membership_id=DEFAULT_MEMBERSHIP_ID,
            idempotency_key=f"approve-{uuid4()}",
            confirmed_values=values,
            activate=True,
            target_product_id=product_id,
            expected_product_version=1,
        )
        session.commit()
        assert result.product_version == 2
        product = session.get(ProductRow, product_id)
        assert product is not None and product.current_version == 2
        assert product.name == "Versioned Product Updated"
        versions = session.scalars(
            select(ProductVersionRow)
            .where(ProductVersionRow.product_id == product_id)
            .order_by(ProductVersionRow.version_number)
        ).all()
        assert [version.version_number for version in versions] == [1, 2]
        assert versions[0].snapshot["product"]["name"] == "Versioned Product Original"
        assert versions[1].snapshot["product"]["name"] == "Versioned Product Updated"
        assert versions[0].content_hash != versions[1].content_hash

        versions[1].change_reason = "attempted history rewrite"
        with pytest.raises(IntegrityError):
            session.commit()
        session.rollback()

        with pytest.raises(ProductAdoptionError) as conflict:
            approve_candidate_group(
                session,
                tenant_id=DEFAULT_TENANT_ID,
                task_id=workflow.task_id,
                candidate_group_key=group_key,
                reviewer_membership_id=DEFAULT_MEMBERSHIP_ID,
                idempotency_key=f"approve-{uuid4()}",
                confirmed_values=values,
                activate=True,
                target_product_id=product_id,
                expected_product_version=1,
            )
        assert conflict.value.code == "CANDIDATE_GROUP_ALREADY_APPLIED"


def test_multiline_catalog_header_does_not_pollute_field_mapping(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["某供应商产品规格与联系方式"])
    sheet.append(["大类", "产品名称", "产品尺寸", "叠装箱规"])
    sheet.append(["饮水机", "宠物饮水机", "20*20*16", "60*60*60/24套"])
    path = tmp_path / "catalog.xlsx"
    workbook.save(path)

    result = parse_document(path, detect_file_path(path, path.name))
    assert len(result.records) == 1
    fields = {field["key"]: field["source"] for field in result.records[0].fields}
    assert fields == {
        "name": "宠物饮水机",
        "category": "饮水机",
        "specification": "20*20*16",
        "packing": "60*60*60/24套",
    }


def test_compact_inline_database_approval_projects_and_publishes_outbox(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    event_id = _create_pending_product_event(tmp_path, suffix=uuid4().hex[:6])
    with SessionLocal() as session:
        event = session.get(OutboxEventRow, event_id)
        assert event is not None and event.decision_id is not None
        decision = session.get(ProductCandidateDecisionRow, event.decision_id)
        assert decision is not None
        task_id = decision.ai_task_id
        group_key = decision.candidate_group_key
        idempotency_key = decision.idempotency_key
        confirmed_values = dict(decision.human_values)
        product_id = decision.product_id
        assert product_id is not None
        product = session.get(ProductRow, product_id)
        assert product is not None
        product_name = product.name
        assert event.status == "PENDING"
        assert product.search_document_version == 0

    monkeypatch.setenv("ATC_RUNTIME_PROFILE", "compact")
    monkeypatch.setenv("OUTBOX_PUBLISHER_PROFILE", "inline_database")
    response = client.post(
        f"/api/v1/ai/product-intelligence/tasks/{task_id}/groups/{group_key}/approve",
        json={
            "idempotency_key": idempotency_key,
            "confirmed_values": confirmed_values,
            "activate": True,
        },
    )
    assert response.status_code == 202, response.text
    payload = response.json()
    assert payload["idempotent"] is True
    assert payload["outbox_event_id"] == str(event_id)
    assert payload["outbox_status"] == "PUBLISHED"

    with SessionLocal() as session:
        event = session.get(OutboxEventRow, event_id)
        product = session.get(ProductRow, product_id)
        assert event is not None and event.status == "PUBLISHED"
        assert event.published_at is not None
        assert product is not None and product.search_document_version == 1
        assert (
            session.scalar(
                select(func.count())
                .select_from(KnowledgeDocumentRow)
                .where(
                    KnowledgeDocumentRow.tenant_id == DEFAULT_TENANT_ID,
                    KnowledgeDocumentRow.source_entity_id == product_id,
                    KnowledgeDocumentRow.source_version == 1,
                    KnowledgeDocumentRow.status == "ACTIVE",
                )
            )
            == 1
        )

    search = client.post(
        "/api/v1/ai/search/products",
        json={"query": product_name, "limit": 10},
    )
    assert search.status_code == 200, search.text
    assert any(
        row["product_id"] == str(product_id)
        for row in search.json()["results"]
    )


def test_outbox_relay_retry_dead_letter_expired_lease_and_metrics(
    tmp_path: Path,
) -> None:
    event_id = _create_pending_product_event(tmp_path, suffix=uuid4().hex[:6])
    start = datetime.now(UTC)
    failing = InMemoryOutboxPublisher(fail_with=RuntimeError("broker unavailable"))
    with SessionLocal() as session:
        event = session.get(OutboxEventRow, event_id)
        assert event is not None
        event.max_attempts = 2
        session.commit()
        first = relay_one_outbox_event(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            relay_id="relay-a",
            publisher=failing,
            event_id=event_id,
            now=start,
        )
        assert first.status == "FAILED"
        assert first.outcome == "RETRY_SCHEDULED"
        assert first.attempt_count == 1
        assert first.next_attempt_at is not None
        not_due = relay_one_outbox_event(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            relay_id="relay-a",
            publisher=failing,
            event_id=event_id,
            now=start,
        )
        assert not_due.status == "IDLE"
        second = relay_one_outbox_event(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            relay_id="relay-b",
            publisher=failing,
            event_id=event_id,
            now=first.next_attempt_at + timedelta(seconds=1),
        )
        assert second.status == "DEAD"
        assert second.outcome == "DEAD_LETTERED"
        assert second.attempt_count == 2

    lease_event_id = _create_pending_product_event(tmp_path, suffix=uuid4().hex[:6])
    publisher = InMemoryOutboxPublisher()
    with SessionLocal() as session:
        lease_event = session.get(OutboxEventRow, lease_event_id)
        assert lease_event is not None
        lease_event.status = "PROCESSING"
        lease_event.lease_owner = "crashed-relay"
        lease_event.lease_expires_at = start - timedelta(seconds=1)
        session.commit()
        recovered = relay_one_outbox_event(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            relay_id="replacement-relay",
            publisher=publisher,
            event_id=lease_event_id,
            now=start,
        )
        assert recovered.status == "PUBLISHED"
        assert recovered.attempt_count == 1
        assert [message.event_id for message in publisher.messages] == [lease_event_id]

    metrics = client.get("/api/v1/system/outbox/metrics")
    assert metrics.status_code == 200
    assert metrics.json()["dead_count"] >= 1
    assert metrics.json()["lag_seconds"] >= 0


def test_outbox_relay_at_least_once_inbox_idempotency_and_tenant_boundary(
    tmp_path: Path,
) -> None:
    event_id = _create_pending_product_event(tmp_path, suffix=uuid4().hex[:6])
    publisher = InMemoryOutboxPublisher()
    with SessionLocal() as session:
        wrong_tenant = relay_one_outbox_event(
            session,
            tenant_id=uuid4(),
            relay_id="relay-wrong-tenant",
            publisher=publisher,
            event_id=event_id,
        )
        assert wrong_tenant.status == "IDLE"
        delivered = relay_one_outbox_event(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            relay_id="relay-primary",
            publisher=publisher,
            event_id=event_id,
        )
        assert delivered.status == "PUBLISHED"
        duplicate_relay = relay_one_outbox_event(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            relay_id="relay-primary",
            publisher=publisher,
            event_id=event_id,
        )
        assert duplicate_relay.outcome == "ALREADY_PUBLISHED"
        assert len(publisher.messages) == 1

    message = publisher.messages[0]
    with SessionLocal() as session:
        first = consume_product_committed_message(session, message=message)
        session.commit()
        assert first.status == "COMPLETED"
        assert first.idempotent is False
        repeated = consume_product_committed_message(session, message=message)
        session.commit()
        assert repeated.status == "COMPLETED"
        assert repeated.idempotent is True
        assert repeated.inbox_id == first.inbox_id
        forged = replace(message, payload={**message.payload, "product_version": 999})
        with pytest.raises(ProductAdoptionError) as integrity_error:
            consume_product_committed_message(session, message=forged)
        assert integrity_error.value.code == "OUTBOX_MESSAGE_INTEGRITY_FAILED"
        assert session.scalar(
            select(func.count()).select_from(InboxEventRow).where(
                InboxEventRow.tenant_id == DEFAULT_TENANT_ID,
                InboxEventRow.event_id == event_id,
            )
        ) == 1


def test_memory_outbox_publisher_is_fail_closed_in_production(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("APP_ENV", "production")
    monkeypatch.setenv("OUTBOX_PUBLISHER_PROFILE", "memory")
    with pytest.raises(RuntimeError, match="forbidden"):
        get_outbox_publisher()


def test_acg008_through_acg010_migrations_are_reversible(tmp_path: Path) -> None:
    database_path = tmp_path / "acg008-009-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)

    command.upgrade(config, "20260718_0013")
    before_relay = create_engine(migration_url)
    with before_relay.connect() as connection:
        assert "inbox_events" not in inspect(connection).get_table_names()
        outbox_columns = {
            column["name"] for column in inspect(connection).get_columns("outbox_events")
        }
        assert {"available_at", "lease_owner", "dead_lettered_at"}.isdisjoint(
            outbox_columns
        )
    before_relay.dispose()

    command.upgrade(config, "20260718_0014")
    relay_engine = create_engine(migration_url)
    with relay_engine.connect() as connection:
        assert "inbox_events" in inspect(connection).get_table_names()
        outbox_columns = {
            column["name"] for column in inspect(connection).get_columns("outbox_events")
        }
        assert {"available_at", "lease_owner", "dead_lettered_at"}.issubset(
            outbox_columns
        )
        assert {
            "skus",
            "attribute_definitions",
            "supplier_prices",
            "product_audit_events",
        }.isdisjoint(inspect(connection).get_table_names())
    relay_engine.dispose()

    command.upgrade(config, "head")
    product_engine = create_engine(migration_url)
    with product_engine.connect() as connection:
        tables = set(inspect(connection).get_table_names())
        assert {
            "skus",
            "attribute_definitions",
            "supplier_prices",
            "product_audit_events",
            "vision_observations",
            "image_embeddings",
            "image_searches",
            "customers",
            "inquiries",
            "inquiry_items",
            "inquiry_match_results",
            "quotations",
            "quotation_versions",
            "quotation_items",
            "quotation_approvals",
            "tenant_public_profiles",
            "public_catalog_offers",
            "public_quote_drafts",
            "public_quote_draft_items",
            "public_quote_download_tokens",
        }.issubset(tables)
        assert "sku_id" in {
            column["name"]
            for column in inspect(connection).get_columns("supplier_products")
        }
        assert "attribute_definition_id" in {
            column["name"]
            for column in inspect(connection).get_columns("product_attributes")
        }
    product_engine.dispose()

    command.downgrade(config, "20260718_0014")
    downgraded_engine = create_engine(migration_url)
    with downgraded_engine.connect() as connection:
        tables = set(inspect(connection).get_table_names())
        assert {
            "skus",
            "attribute_definitions",
            "supplier_prices",
            "product_audit_events",
        }.isdisjoint(tables)
    downgraded_engine.dispose()
    command.upgrade(config, "head")
    command.check(config)


def test_manual_product_creation_builds_product_sku_offer_and_audit(
    request: pytest.FixtureRequest,
) -> None:
    created_product_ids: list[UUID] = []

    def cleanup() -> None:
        if not created_product_ids:
            return
        with SessionLocal() as session:
            sku_ids = select(SkuRow.id).where(
                SkuRow.tenant_id == DEFAULT_TENANT_ID,
                SkuRow.product_id.in_(created_product_ids),
            )
            session.execute(
                delete(PublicCatalogOfferRow).where(
                    PublicCatalogOfferRow.tenant_id == DEFAULT_TENANT_ID,
                    PublicCatalogOfferRow.sku_id.in_(sku_ids),
                )
            )
            session.execute(
                delete(ProductAuditEventRow).where(
                    ProductAuditEventRow.tenant_id == DEFAULT_TENANT_ID,
                    ProductAuditEventRow.product_id.in_(created_product_ids),
                )
            )
            session.execute(
                delete(ProductImageRow).where(
                    ProductImageRow.tenant_id == DEFAULT_TENANT_ID,
                    ProductImageRow.product_id.in_(created_product_ids),
                )
            )
            session.execute(
                delete(SkuRow).where(
                    SkuRow.tenant_id == DEFAULT_TENANT_ID,
                    SkuRow.product_id.in_(created_product_ids),
                )
            )
            session.execute(
                delete(ProductRow).where(
                    ProductRow.tenant_id == DEFAULT_TENANT_ID,
                    ProductRow.id.in_(created_product_ids),
                )
            )
            session.commit()

    request.addfinalizer(cleanup)
    category_id = client.get(
        "/api/v1/products/71000000-0000-0000-0000-000000000002"
    ).json()["category"]["id"]
    prefix = uuid4().hex[:10].upper()
    payload = {
        "name": f"Manual product {prefix}",
        "product_code": f"MP-{prefix}",
        "description": "Created from the single-product form.",
        "category_id": category_id,
        "default_unit": "piece",
        "image_url": f"https://cdn.example.com/catalog/{prefix}.jpg",
        "sku_code": f"MS-{prefix}",
        "sku_name": f"Manual SKU {prefix}",
        "barcode": f"BC-{prefix}",
        "default_moq": "12",
        "moq_unit": "piece",
        "weight": "0.75",
        "weight_unit": "kg",
        "unit_price": "19.90",
        "currency": "usd",
        "tags": ["manual", "new", "manual"],
        "display_tag": "manual",
        "tag_color": "#336699",
        "publish_to_storefront": True,
    }

    created = client.post("/api/v1/products", json=payload)
    assert created.status_code == 201, created.text
    detail = created.json()
    created_product_ids.append(UUID(detail["id"]))
    assert detail["product_code"] == payload["product_code"]
    assert detail["name"] == payload["name"]
    assert detail["description"] == payload["description"]
    assert detail["status"] == "ACTIVE"
    assert detail["image_status"] == "APPROVED"
    assert len(detail["skus"]) == 1
    assert detail["skus"][0]["sku_code"] == payload["sku_code"]
    assert detail["skus"][0]["status"] == "ACTIVE"
    assert Decimal(detail["skus"][0]["default_moq"]) == Decimal("12")

    offers = client.get(f"/api/v1/products/{detail['id']}/public-offers")
    assert offers.status_code == 200, offers.text
    assert Decimal(offers.json()[0]["unit_price"]) == Decimal("19.90")
    assert offers.json()[0]["currency"] == "USD"
    assert offers.json()[0]["tags"] == ["manual", "new"]
    assert offers.json()[0]["publication_status"] == "PUBLISHED"

    listed = client.get(
        "/api/v1/product-center/skus",
        params={"q": payload["sku_code"], "page": 1, "page_size": 20},
    )
    assert listed.status_code == 200, listed.text
    assert listed.json()["total"] == 1
    assert listed.json()["items"][0]["source_type"] == "MANUAL"
    assert Decimal(listed.json()["items"][0]["public_price"]) == Decimal("19.90")
    assert listed.json()["items"][0]["public_currency"] == "USD"
    assert {
        "product.created",
        "sku.created",
        "public_offer.published",
    }.issubset({event["action"] for event in detail["activity"]})

    duplicate = client.post(
        "/api/v1/products",
        json={**payload, "sku_code": f"MS-{uuid4().hex[:10].upper()}"},
    )
    assert duplicate.status_code == 409
    assert duplicate.json()["detail"]["code"] == "PRODUCT_CODE_CONFLICT"


def test_product_main_image_upload_is_indexed_and_included_in_sku_export(
    request: pytest.FixtureRequest,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    created_product_ids: list[UUID] = []
    stored_object_keys: list[str] = []

    def cleanup() -> None:
        with SessionLocal() as session:
            image_rows = (
                session.scalars(
                    select(ProductImageRow)
                    .where(
                        ProductImageRow.tenant_id == DEFAULT_TENANT_ID,
                        ProductImageRow.product_id.in_(created_product_ids),
                    )
                    .execution_options(include_deleted=True)
                ).all()
                if created_product_ids
                else []
            )
            stored_object_keys.extend(
                row.object_key
                for row in image_rows
                if not row.object_key.startswith(("http://", "https://"))
            )
            if created_product_ids:
                sku_ids = select(SkuRow.id).where(
                    SkuRow.tenant_id == DEFAULT_TENANT_ID,
                    SkuRow.product_id.in_(created_product_ids),
                )
                session.execute(
                    delete(PublicCatalogOfferRow).where(
                        PublicCatalogOfferRow.tenant_id == DEFAULT_TENANT_ID,
                        PublicCatalogOfferRow.sku_id.in_(sku_ids),
                    )
                )
                session.execute(
                    delete(ProductAuditEventRow).where(
                        ProductAuditEventRow.tenant_id == DEFAULT_TENANT_ID,
                        ProductAuditEventRow.product_id.in_(created_product_ids),
                    )
                )
                session.execute(
                    delete(ProductImageRow).where(
                        ProductImageRow.tenant_id == DEFAULT_TENANT_ID,
                        ProductImageRow.product_id.in_(created_product_ids),
                    )
                )
                session.execute(
                    delete(SkuRow).where(
                        SkuRow.tenant_id == DEFAULT_TENANT_ID,
                        SkuRow.product_id.in_(created_product_ids),
                    )
                )
                session.execute(
                    delete(ProductRow).where(
                        ProductRow.tenant_id == DEFAULT_TENANT_ID,
                        ProductRow.id.in_(created_product_ids),
                    )
                )
                session.commit()
        storage = get_object_storage()
        for object_key in set(stored_object_keys):
            storage.delete(object_key)

    request.addfinalizer(cleanup)
    monkeypatch.setenv(
        "PUBLIC_MEDIA_BASE_URL",
        "https://resources.example.test",
    )
    category_id = client.get(
        "/api/v1/products/71000000-0000-0000-0000-000000000002"
    ).json()["category"]["id"]
    suffix = uuid4().hex[:10].upper()
    created = client.post(
        "/api/v1/products",
        json={
            "name": f"Image export product {suffix}",
            "product_code": f"IMG-{suffix}",
            "description": "Image upload and workbook export test.",
            "category_id": category_id,
            "sku_code": f"IMG-SKU-{suffix}",
            "unit_price": "0",
            "currency": "USD",
            "publish_to_storefront": True,
        },
    )
    assert created.status_code == 201, created.text
    detail = created.json()
    product_id = UUID(detail["id"])
    sku_id = detail["skus"][0]["id"]
    created_product_ids.append(product_id)

    image_buffer = BytesIO()
    Image.new("RGB", (320, 240), color=(45, 27, 105)).save(
        image_buffer,
        format="PNG",
    )
    uploaded = client.post(
        f"/api/v1/products/{product_id}/images/main",
        files={"image": ("catalog-main.png", image_buffer.getvalue(), "image/png")},
    )
    assert uploaded.status_code == 201, uploaded.text
    uploaded_payload = uploaded.json()
    assert uploaded_payload["content_type"] == "image/webp"
    assert uploaded_payload["width"] == 320
    assert uploaded_payload["height"] == 240
    assert uploaded_payload["image_role"] == "MAIN"
    assert uploaded_payload["approval_status"] == "APPROVED"
    assert uploaded_payload["url"].startswith(
        f"https://resources.example.test/tenants/{DEFAULT_TENANT_ID}/products/{product_id}/images/"
    )

    with SessionLocal() as session:
        image_row = session.get(ProductImageRow, UUID(uploaded_payload["id"]))
        assert image_row is not None
        assert image_row.object_key.startswith(
            f"tenants/{DEFAULT_TENANT_ID}/products/{product_id}/images/"
        )
        assert image_row.sha256
        assert session.scalar(
            select(ProductImageRow.id).where(
                ProductImageRow.tenant_id == DEFAULT_TENANT_ID,
                ProductImageRow.sha256 == image_row.sha256,
            )
        ) == image_row.id
        assert get_object_storage().exists(image_row.object_key)

    listed = client.get(
        "/api/v1/product-center/skus",
        params={"q": f"IMG-SKU-{suffix}", "page": 1, "page_size": 20},
    )
    assert listed.status_code == 200, listed.text
    assert listed.json()["items"][0]["thumbnail_url"] == uploaded_payload["url"]

    exported = client.post(
        "/api/v1/product-center/skus/export",
        json={"sku_ids": [sku_id]},
    )
    assert exported.status_code == 200, exported.text
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in exported.headers["content-type"]
    workbook = load_workbook(BytesIO(exported.content))
    try:
        assert workbook.sheetnames == ["商品", "SKU"]
        product_sheet = workbook["商品"]
        sku_sheet = workbook["SKU"]
        product_headers = [cell.value for cell in product_sheet[1]]
        sku_headers = [cell.value for cell in sku_sheet[1]]
        product_values = {
            product_headers[column - 1]: product_sheet.cell(2, column).value
            for column in range(1, len(product_headers) + 1)
        }
        sku_values = {
            sku_headers[column - 1]: sku_sheet.cell(2, column).value
            for column in range(1, len(sku_headers) + 1)
        }
        assert product_values["商品ID"] == str(product_id)
        assert product_values["商品名称"] == detail["name"]
        assert product_values["图片地址1"] == uploaded_payload["url"]
        image_url_cell = product_sheet.cell(
            2,
            product_headers.index("图片地址1") + 1,
        )
        assert image_url_cell.hyperlink is not None
        assert image_url_cell.hyperlink.target == uploaded_payload["url"]
        assert len(product_sheet._images) == 0
        assert sku_values["SKU ID"] == sku_id
        assert sku_values["商品ID"] == str(product_id)
        assert sku_values["SKU编号"] == f"IMG-SKU-{suffix}"
    finally:
        workbook.close()


def test_product_center_sku_matrix_price_history_attributes_and_audit() -> None:
    product_id = UUID("71000000-0000-0000-0000-000000000002")
    detail_response = client.get(f"/api/v1/products/{product_id}")
    assert detail_response.status_code == 200, detail_response.text
    detail = detail_response.json()
    assert detail["product_code"] == "SKU-18211"
    assert detail["sources"]
    category_id = detail["category"]["id"]

    definition_response = client.post(
        "/api/v1/attribute-definitions",
        json={
            "category_id": category_id,
            "attribute_key": f"finish_{uuid4().hex[:8]}",
            "display_name": "表面工艺",
            "data_type": "ENUM",
            "enum_values": ["powder-coated", "galvanized"],
            "is_filterable": True,
            "is_matchable": True,
        },
    )
    assert definition_response.status_code == 201, definition_response.text

    prefix = uuid4().hex[:7].upper()
    sku_items = [
        {
            "sku_code": f"{prefix}-{color}-{size}",
            "name": f"Fence {color} {size}",
            "option_values": {"color": color, "size": size},
            "default_moq": "10",
            "moq_unit": "piece",
            "status": "ACTIVE",
        }
        for color in ("BLACK", "WHITE")
        for size in ("S", "M", "L")
    ]
    sku_response = client.post(
        f"/api/v1/products/{product_id}/skus", json={"items": sku_items}
    )
    assert sku_response.status_code == 201, sku_response.text
    created_skus = sku_response.json()
    assert len(created_skus) == 6

    duplicate = client.post(
        f"/api/v1/products/{product_id}/skus", json={"items": [sku_items[0]]}
    )
    assert duplicate.status_code == 409
    assert duplicate.json()["detail"]["code"] == "SKU_CODE_CONFLICT"

    first_sku = created_skus[0]
    updated = client.patch(
        f"/api/v1/skus/{first_sku['id']}",
        json={"expected_version": 1, "barcode": f"BC-{prefix}"},
    )
    assert updated.status_code == 200
    assert updated.json()["version"] == 2
    stale = client.patch(
        f"/api/v1/skus/{first_sku['id']}",
        json={"expected_version": 1, "barcode": "STALE"},
    )
    assert stale.status_code == 409
    assert stale.json()["detail"]["code"] == "SKU_VERSION_CONFLICT"

    exact = client.get("/api/v1/products", params={"q": sku_items[0]["sku_code"]})
    assert exact.status_code == 200
    assert exact.json()[0]["id"] == str(product_id)

    supplier_product_id = detail["sources"][0]["supplier_product_id"]
    now = datetime.now(UTC)
    valid_price = client.post(
        "/api/v1/product-prices",
        json={
            "supplier_product_id": supplier_product_id,
            "sku_id": first_sku["id"],
            "min_quantity": "100",
            "unit_price": "139.50",
            "currency": "cny",
            "unit_code": "piece",
            "valid_from": now.isoformat(),
            "valid_to": (now + timedelta(days=90)).isoformat(),
        },
    )
    assert valid_price.status_code == 201, valid_price.text
    assert valid_price.json()["price_validity"] == "VALID"
    expired_price = client.post(
        "/api/v1/product-prices",
        json={
            "supplier_product_id": supplier_product_id,
            "min_quantity": "500",
            "unit_price": "128",
            "currency": "CNY",
            "unit_code": "piece",
            "valid_from": (now - timedelta(days=120)).isoformat(),
            "valid_to": (now - timedelta(days=30)).isoformat(),
        },
    )
    assert expired_price.status_code == 201, expired_price.text
    assert expired_price.json()["price_validity"] == "EXPIRED"
    history = client.get(f"/api/v1/products/{product_id}/prices")
    assert history.status_code == 200
    assert {valid_price.json()["id"], expired_price.json()["id"]}.issubset(
        {item["id"] for item in history.json()}
    )

    refreshed = client.get(f"/api/v1/products/{product_id}").json()
    actions = {event["action"] for event in refreshed["activity"]}
    assert {"sku.created", "sku.updated", "price.confirmed"}.issubset(actions)

    with SessionLocal() as session:
        hidden = list_authoritative_products(
            session,
            tenant_id=DEFAULT_TENANT_ID,
            permissions=frozenset({"product.view"}),
            query=sku_items[0]["sku_code"],
            category_id=None,
            supplier_id=None,
            statuses=[],
            approved_images_only=False,
            limit=10,
        )
        assert hidden and hidden[0].price is None
        assert hidden[0].current_offer is not None
        assert hidden[0].current_offer.unit_price is None


def test_product_review_queue_uses_candidate_evidence_and_human_adoption() -> None:
    csv_content = (
        "Product Name,Item No,Material,Color,Size,MOQ\n"
        f"Review Queue Product {uuid4().hex[:6]},RQ-{uuid4().hex[:6]},TPR,Blue,10cm,120\n"
    ).encode("utf-8")
    upload = client.post(
        "/api/v1/imports",
        files={"file": ("review-queue.csv", csv_content, "text/csv")},
        data={"supplier_id": "SUP-001", "source_type": "SUPPLIER_CATALOG"},
    )
    assert upload.status_code == 201, upload.text
    assert upload.json()["candidate_status"] == "NEEDS_REVIEW"
    task_id = upload.json()["ai_task_id"]

    queue = client.get("/api/v1/product-review-items", params={"limit": 500})
    assert queue.status_code == 200, queue.text
    item = next(row for row in queue.json() if row["task_id"] == task_id)
    assert item["status"] == "pending"
    assert item["source"].startswith("SRC-")
    assert item["fields"] and all(field["source"] for field in item["fields"])
    confirmed = {field["key"]: field["normalized"] for field in item["fields"]}
    approval = client.post(
        f"/api/v1/ai/product-intelligence/tasks/{task_id}/groups/{item['candidate_group_key']}/approve",
        json={
            "idempotency_key": f"review-ui-{uuid4()}",
            "confirmed_values": confirmed,
            "activate": True,
            "product_code": f"ATC-RQ-{uuid4().hex[:8].upper()}",
            "change_reason": "ACG-009 Product review page acceptance",
        },
    )
    assert approval.status_code == 202, approval.text
    refreshed = client.get("/api/v1/product-review-items", params={"limit": 500})
    reviewed = next(row for row in refreshed.json() if row["task_id"] == task_id)
    assert reviewed["status"] == "approved"
    assert reviewed["applied_product_id"] == approval.json()["product_id"]


def test_image_intelligence_projection_search_and_media_gate(tmp_path: Path) -> None:
    product_id = UUID("71000000-0000-0000-0000-000000000002")
    image_bytes = (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
        b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00"
        b"\x1f\x15\xc4\x89\x00\x00\x00\rIDAT\x08\xd7c\xf8\xcf\xc0\xf0\x1f\x00\x05\x00\x01\xff\x89\x99=\x1d\x00\x00\x00\x00IEND\xaeB`\x82"
    )
    image_hash = hashlib.sha256(image_bytes).hexdigest()
    image_id = uuid4()
    object_key = f"tenants/{DEFAULT_TENANT_ID}/source/product-images/{image_id}.png"
    path = tmp_path / "approved-product.png"
    path.write_bytes(image_bytes)
    get_object_storage().put_file(path, object_key=object_key, content_type="image/png")
    with SessionLocal() as session:
        session.add(ProductImageRow(id=image_id, tenant_id=DEFAULT_TENANT_ID, product_id=product_id, storage_provider="local", bucket="local", object_key=object_key, original_filename=path.name, content_type="image/png", byte_size=len(image_bytes), sha256=image_hash, width=1, height=1, image_role="MAIN", sort_order=99, approval_status="APPROVED"))
        source_id = uuid4()
        session.add(ProductImageRow(id=source_id, tenant_id=DEFAULT_TENANT_ID, product_id=product_id, storage_provider="local", bucket="local", object_key=f"source/{source_id}.png", original_filename="source.png", content_type="image/png", byte_size=len(image_bytes), sha256=image_hash, width=1, height=1, image_role="GALLERY", sort_order=100, approval_status="SOURCE"))
        session.commit()

    projection = client.post(f"/api/v1/product-images/{image_id}/intelligence")
    assert projection.status_code == 200, projection.text
    assert projection.json()["idempotent"] is False
    repeated = client.post(f"/api/v1/product-images/{image_id}/intelligence")
    assert repeated.status_code == 200
    assert repeated.json()["idempotent"] is True
    assert repeated.json()["embedding_id"] == projection.json()["embedding_id"]
    rejected = client.post(f"/api/v1/product-images/{source_id}/intelligence")
    assert rejected.status_code == 409
    assert rejected.json()["detail"]["code"] == "IMAGE_NOT_APPROVED"

    search = client.post("/api/v1/image-searches?limit=5", files={"file": ("query.png", image_bytes, "image/png")})
    assert search.status_code == 200, search.text
    payload = search.json()
    assert payload["status"] == "COMPLETED"
    assert payload["results"][0]["product_id"] == str(product_id)
    assert payload["results"][0]["classification"] == "POSSIBLE_SAME_ITEM"
    assert "never proves" in payload["warnings"][0]
    with SessionLocal() as session:
        expired_search = session.get(ImageSearchRow, UUID(payload["id"]))
        assert expired_search is not None
        expired_search.expires_at = datetime.now(UTC) - timedelta(seconds=1)
        expired_key = expired_search.query_object_key
        session.commit()
    assert get_object_storage().local_path(expired_key).is_file()
    cleanup_trigger = client.post("/api/v1/image-searches", files={"file": ("query-2.png", image_bytes, "image/png")})
    assert cleanup_trigger.status_code == 200
    with SessionLocal() as session:
        expired_search = session.get(ImageSearchRow, UUID(payload["id"]))
        assert expired_search.status == "EXPIRED"
        assert expired_search.query_embedding is None
    assert not get_object_storage().local_path(expired_key).exists()
    malware = client.post("/api/v1/image-searches", files={"file": ("evil.png", image_bytes + b"ATC-MALWARE-TEST", "image/png")})
    assert malware.status_code == 403
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ImageEmbeddingRow).where(ImageEmbeddingRow.product_image_id == image_id)) == 1
        assert session.scalar(select(func.count()).select_from(VisionObservationRow).where(VisionObservationRow.product_image_id == image_id)) == 1
        assert session.scalar(select(func.count()).select_from(ImageSearchRow).where(ImageSearchRow.tenant_id == DEFAULT_TENANT_ID)) >= 1
        assert session.get(ProductImageRow, source_id).approval_status == "SOURCE"


def test_development_image_provider_is_fail_closed_in_production(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("APP_ENV", "production")
    monkeypatch.setenv("IMAGE_INTELLIGENCE_PROFILE", "deterministic")
    with pytest.raises(RuntimeError, match="forbidden"):
        get_image_intelligence_provider()


def test_disabled_image_intelligence_returns_service_unavailable(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("IMAGE_INTELLIGENCE_PROFILE", "disabled")
    with SessionLocal() as session:
        product = session.scalar(
            select(ProductRow).where(ProductRow.tenant_id == DEFAULT_TENANT_ID)
        )
        assert product is not None
        image_id = uuid4()
        session.add(
            ProductImageRow(
                id=image_id,
                tenant_id=DEFAULT_TENANT_ID,
                product_id=product.id,
                storage_provider="local",
                bucket="local",
                object_key=f"disabled-provider/{image_id}.png",
                original_filename="disabled-provider.png",
                content_type="image/png",
                byte_size=16,
                sha256="0" * 64,
                image_role="GALLERY",
                sort_order=0,
                approval_status="APPROVED",
            )
        )
        session.commit()

    projection = client.post(f"/api/v1/product-images/{image_id}/intelligence")
    assert projection.status_code == 503
    assert projection.json()["detail"]["code"] == "IMAGE_INTELLIGENCE_UNAVAILABLE"

    search = client.post(
        "/api/v1/image-searches",
        files={
            "file": (
                "query.png",
                b"\x89PNG\r\n\x1a\nquery",
                "image/png",
            )
        },
    )
    assert search.status_code == 503
    assert search.json()["detail"]["code"] == "IMAGE_INTELLIGENCE_UNAVAILABLE"


def test_dashboard_and_supplier_profiles_use_tenant_scoped_authoritative_data() -> None:
    dashboard = client.get("/api/v1/dashboard")
    assert dashboard.status_code == 200, dashboard.text
    payload = dashboard.json()
    assert payload["data_scope"] == "TENANT"
    metric_keys = {metric["key"] for metric in payload["metrics"]}
    assert {"active_skus", "today_inquiries", "pending_quotations", "active_suppliers"}.issubset(metric_keys)
    assert payload["data_health"]["active_products"] >= 1
    assert 0 <= payload["data_health"]["score"] <= 100
    for coverage_field in (
        "approved_image_coverage",
        "supplier_source_coverage",
        "valid_price_coverage",
    ):
        assert 0 <= payload["data_health"][coverage_field] <= 1

    directory = client.get("/api/v1/supplier-profiles")
    assert directory.status_code == 200, directory.text
    suppliers = directory.json()
    assert suppliers
    seeded = next(row for row in suppliers if row["id"] == "SUP-001")
    assert seeded["active_products"] >= 1
    assert seeded["active_skus"] >= 1
    assert seeded["valid_prices"] >= 1
    detail = client.get("/api/v1/supplier-profiles/SUP-001")
    assert detail.status_code == 200, detail.text
    assert detail.json()["sources"]
    assert detail.json()["sources"][0]["product_code"]

    other_tenant_id = uuid4()
    other_supplier_id = f"SUP-X-{uuid4().hex[:8]}"
    with SessionLocal() as session:
        session.add(TenantRow(id=other_tenant_id, organization_id=DEFAULT_ORGANIZATION_ID, slug=f"other-{uuid4().hex[:8]}", name="Other isolated tenant"))
        session.flush()
        session.add(SupplierRow(id=other_supplier_id, tenant_id=other_tenant_id, supplier_code=other_supplier_id, name="Must not leak", category="isolated"))
        session.commit()
    isolated_directory = client.get("/api/v1/supplier-profiles")
    assert all(row["id"] != other_supplier_id for row in isolated_directory.json())
    assert client.get(f"/api/v1/supplier-profiles/{other_supplier_id}").status_code == 404


def test_inquiry_matching_selection_and_human_gated_quotation() -> None:
    customer = client.post("/api/v1/customers", json={"company_name": f"ACG Customer {uuid4().hex[:6]}", "country_code": "US", "language": "en", "default_currency": "CNY"})
    assert customer.status_code == 201, customer.text
    inquiry = client.post("/api/v1/inquiries", json={"customer_id": customer.json()["id"], "currency": "CNY", "language": "en", "items": [{"requirement": "SKU-18211", "quantity": 10, "unit_code": "PCS"}]})
    assert inquiry.status_code == 201, inquiry.text
    inquiry_payload = inquiry.json()
    assert inquiry_payload["status"] == "MATCHING"
    item_id = inquiry_payload["items"][0]["id"]

    matched = client.post(f"/api/v1/inquiries/{inquiry_payload['id']}/match")
    assert matched.status_code == 200, matched.text
    candidates = matched.json()["candidates"][item_id]
    assert candidates
    assert candidates[0]["total_score"] == 1.0
    assert "精确产品编码命中" in candidates[0]["reasons"]

    rematched = client.post(f"/api/v1/inquiries/{inquiry_payload['id']}/match")
    assert rematched.status_code == 200, rematched.text
    stale_selection = client.post(f"/api/v1/inquiry-items/{item_id}/selection", json={"match_result_id": candidates[0]["id"]})
    assert stale_selection.status_code == 409
    assert stale_selection.json()["detail"]["code"] == "MATCH_RESULT_STALE"
    candidates = rematched.json()["candidates"][item_id]

    selected = client.post(f"/api/v1/inquiry-items/{item_id}/selection", json={"match_result_id": candidates[0]["id"]})
    assert selected.status_code == 200, selected.text
    assert selected.json()["status"] == "SELECTED"
    ready = client.get(f"/api/v1/inquiries/{inquiry_payload['id']}").json()
    assert ready["status"] == "READY_FOR_QUOTE"

    with SessionLocal() as session:
        selected_product = session.get(ProductRow, UUID(candidates[0]["product_id"]))
        original_version = selected_product.current_version
        selected_product.current_version += 1
        session.commit()
    stale_quote = client.post(f"/api/v1/inquiries/{inquiry_payload['id']}/quotation", json={"target_margin_rate": "0.20", "expires_in_days": 30})
    assert stale_quote.status_code == 409
    assert stale_quote.json()["detail"]["code"] == "MATCH_STALE"
    with SessionLocal() as session:
        selected_product = session.get(ProductRow, UUID(candidates[0]["product_id"]))
        selected_product.current_version = original_version
        session.commit()

    with SessionLocal() as session:
        selected_source = session.get(SupplierProductRow, UUID(candidates[0]["supplier_product_id"]))
        original_source_status = selected_source.status
        selected_source.status = "INACTIVE"
        session.commit()
    inactive_source_quote = client.post(f"/api/v1/inquiries/{inquiry_payload['id']}/quotation", json={"target_margin_rate": "0.20", "expires_in_days": 30})
    assert inactive_source_quote.status_code == 409
    assert inactive_source_quote.json()["detail"]["code"] == "SUPPLIER_SOURCE_MISSING"
    with SessionLocal() as session:
        selected_source = session.get(SupplierProductRow, UUID(candidates[0]["supplier_product_id"]))
        selected_source.status = original_source_status
        session.commit()

    quote = client.post(f"/api/v1/inquiries/{inquiry_payload['id']}/quotation", json={"target_margin_rate": "0.20", "expires_in_days": 30})
    assert quote.status_code == 201, quote.text
    quote_payload = quote.json()
    assert quote_payload["status"] == "CALCULATED"
    assert quote_payload["approval_status"] == "PENDING"
    assert quote_payload["items"][0]["unit_cost"] is not None
    assert Decimal(quote_payload["items"][0]["line_total"]) == Decimal(quote_payload["items"][0]["unit_price"]) * Decimal("10")

    approved = client.post(f"/api/v1/quotations/{quote_payload['id']}/decision", json={"decision": "APPROVED", "reason": "Owner verified customer, source and margin"})
    assert approved.status_code == 200, approved.text
    assert approved.json()["status"] == "APPROVED"
    assert approved.json()["version_hash"] == quote_payload["version_hash"]
    quote_list = client.get("/api/v1/quotations")
    assert quote_list.status_code == 200
    assert any(row["id"] == quote_payload["id"] for row in quote_list.json())
    repeated = client.post(f"/api/v1/quotations/{quote_payload['id']}/decision", json={"decision": "APPROVED", "reason": "duplicate"})
    assert repeated.status_code == 409
    revised = client.post(
        f"/api/v1/quotations/{quote_payload['id']}/revisions",
        json={
            "expected_version": 1,
            "change_reason": "Customer increased the confirmed quantity",
            "items": [{"item_id": quote_payload["items"][0]["id"], "quantity": "20", "target_margin_rate": "0.25"}],
        },
    )
    assert revised.status_code == 201, revised.text
    revised_payload = revised.json()
    assert revised_payload["current_version"] == 2
    assert revised_payload["status"] == "CALCULATED"
    assert revised_payload["approval_status"] == "PENDING"
    assert revised_payload["version_hash"] != quote_payload["version_hash"]
    assert Decimal(revised_payload["items"][0]["quantity"]) == Decimal("20")
    assert len(revised_payload["versions"]) == 2
    assert revised_payload["versions"][0]["version_number"] == 2
    stale_revision = client.post(
        f"/api/v1/quotations/{quote_payload['id']}/revisions",
        json={
            "expected_version": 1,
            "change_reason": "Stale browser must not overwrite",
            "items": [{"item_id": revised_payload["items"][0]["id"], "quantity": "30", "target_margin_rate": "0.25"}],
        },
    )
    assert stale_revision.status_code == 409
    assert stale_revision.json()["detail"]["code"] == "QUOTATION_VERSION_CONFLICT"
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(InquiryRow).where(InquiryRow.id == UUID(inquiry_payload["id"]))) == 1
        assert session.scalar(select(func.count()).select_from(InquiryMatchResultRow).where(InquiryMatchResultRow.inquiry_item_id == UUID(item_id), InquiryMatchResultRow.status == "SELECTED")) == 1
        assert session.scalar(select(func.count()).select_from(QuotationRow).where(QuotationRow.id == UUID(quote_payload["id"]))) == 1
        assert session.scalar(select(func.count()).select_from(QuotationVersionRow).where(QuotationVersionRow.quotation_id == UUID(quote_payload["id"]))) == 2
        assert session.scalar(select(func.count()).select_from(QuotationItemRow).join(QuotationVersionRow, QuotationVersionRow.id == QuotationItemRow.quotation_version_id).where(QuotationVersionRow.quotation_id == UUID(quote_payload["id"]))) == 2
        assert session.scalar(select(func.count()).select_from(QuotationApprovalRow).where(QuotationApprovalRow.quotation_id == UUID(quote_payload["id"]), QuotationApprovalRow.status == "APPROVED")) == 1
        assert session.scalar(select(func.count()).select_from(QuotationApprovalRow).where(QuotationApprovalRow.quotation_id == UUID(quote_payload["id"]), QuotationApprovalRow.status == "PENDING")) == 1
        assert session.scalar(select(func.count()).select_from(OutboxEventRow).where(OutboxEventRow.aggregate_type == "QUOTATION", OutboxEventRow.aggregate_id == quote_payload["id"])) == 3


def test_public_catalog_lists_only_published_active_facts_and_approved_images(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("PUBLIC_MEDIA_BASE_URL", "https://cdn.example.test/catalog")

    store_response = client.get("/api/store/demo")
    assert store_response.status_code == 200, store_response.text
    store = store_response.json()
    assert store["slug"] == "demo"
    assert store["name"] == "Local Demo Company"
    assert "待人工确认" in store["quote_notice"]

    listing_response = client.get("/api/store/demo/skus")
    assert listing_response.status_code == 200, listing_response.text
    listing = listing_response.json()
    assert listing["total"] == 3
    by_code = {item["sku_code"]: item for item in listing["items"]}
    assert Decimal(str(by_code["AQ-320S"]["price"])) == Decimal("99.00")
    assert Decimal(str(by_code["PF-8G01"]["price"])) == Decimal("229.00")
    assert Decimal(str(by_code["SF-6L20"]["price"])) == Decimal("299.00")
    assert by_code["AQ-320S"]["image_url"] is None
    assert by_code["SF-6L20"]["image_url"] is None
    assert by_code["PF-8G01"]["image_url"].startswith(
        "https://cdn.example.test/catalog/tenants/"
    )
    for item in listing["items"]:
        assert "unit_cost" not in item
        assert "supplier_product_id" not in item
        assert "supplier_price" not in item
        assert "moq" not in item

    product_listing_response = client.get("/api/store/demo/products")
    assert product_listing_response.status_code == 200, product_listing_response.text
    product_listing = product_listing_response.json()
    assert product_listing["total"] == 3
    product_by_name = {
        item["name"]: item for item in product_listing["items"]
    }
    product_summary = product_by_name["八片带门宠物围栏"]
    assert product_summary["sku_count"] == 1
    assert Decimal(str(product_summary["price_from"])) == Decimal("229.00")
    assert Decimal(str(product_summary["price_to"])) == Decimal("229.00")
    product_detail_response = client.get(
        f"/api/store/demo/products/{product_summary['id']}"
    )
    assert product_detail_response.status_code == 200, product_detail_response.text
    assert "max-age=30" in product_detail_response.headers["cache-control"]
    assert (
        "stale-while-revalidate=120"
        in product_detail_response.headers["cache-control"]
    )
    product_detail = product_detail_response.json()
    assert product_detail["name"] == product_summary["name"]
    assert [item["sku_code"] for item in product_detail["skus"]] == ["PF-8G01"]
    assert client.get(f"/api/store/demo/products/{uuid4()}").status_code == 404

    detail_response = client.get(
        f"/api/store/demo/skus/{by_code['PF-8G01']['id']}"
    )
    assert detail_response.status_code == 200, detail_response.text
    detail = detail_response.json()
    assert detail["sku_code"] == "PF-8G01"
    assert detail["description"] == by_code["PF-8G01"]["description"]
    assert detail["display_tag"] == by_code["PF-8G01"]["display_tag"]
    assert "supplier_id" not in detail
    assert "supplier_name" not in detail
    assert client.get(f"/api/store/demo/skus/{uuid4()}").status_code == 404

    filtered = client.get(
        "/api/store/demo/skus",
        params={"category": "围栏与玩具", "tags": "宠物围栏"},
    )
    assert filtered.status_code == 200
    assert [item["sku_code"] for item in filtered.json()["items"]] == ["PF-8G01"]

    tag_query = client.get("/api/store/demo/skus", params={"q": "宠物围栏"})
    assert tag_query.status_code == 200
    assert [item["sku_code"] for item in tag_query.json()["items"]] == ["PF-8G01"]

    expanded = client.get(
        "/api/store/demo/skus", params={"q": "围栏 PF", "semantic": "true"}
    )
    assert expanded.status_code == 200
    assert [item["sku_code"] for item in expanded.json()["items"]] == ["PF-8G01"]

    for natural_query, expected_sku in (
        ("不锈钢无线宠物饮水机", "AQ-320S"),
        ("支持APP的6L智能宠物喂食器", "SF-6L20"),
        ("可折叠带门宠物围栏", "PF-8G01"),
    ):
        natural = client.get(
            "/api/store/demo/skus",
            params={"q": natural_query, "semantic": "true"},
        )
        assert natural.status_code == 200
        assert [item["sku_code"] for item in natural.json()["items"]] == [expected_sku]

    with SessionLocal() as session:
        inactive_sku = session.scalar(select(SkuRow).where(SkuRow.sku_code == "SF-6L20"))
        hidden_product = session.scalar(
            select(ProductRow).where(ProductRow.product_code == "SKU-24018")
        )
        assert inactive_sku is not None and hidden_product is not None
        inactive_sku.status = "INACTIVE"
        hidden_product.status = "DRAFT"
        session.commit()
    try:
        active_only = client.get("/api/store/demo/skus")
        assert active_only.status_code == 200
        assert [item["sku_code"] for item in active_only.json()["items"]] == ["PF-8G01"]
        assert (
            client.get(
                f"/api/store/demo/skus/{by_code['SF-6L20']['id']}"
            ).status_code
            == 404
        )
    finally:
        with SessionLocal() as session:
            inactive_sku = session.scalar(select(SkuRow).where(SkuRow.sku_code == "SF-6L20"))
            hidden_product = session.scalar(
                select(ProductRow).where(ProductRow.product_code == "SKU-24018")
            )
            assert inactive_sku is not None and hidden_product is not None
            inactive_sku.status = "ACTIVE"
            hidden_product.status = "ACTIVE"
            session.commit()


def test_storefront_detail_views_are_idempotent_and_tenant_analytics_are_aggregated(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("TRUST_CLOUDFLARE_VISITOR_HEADERS", "true")
    listing = client.get("/api/store/demo/skus", params={"page_size": 1})
    assert listing.status_code == 200, listing.text
    sku = listing.json()["items"][0]
    event_id = f"view-{uuid4()}"

    with SessionLocal() as session:
        before_total = int(
            session.scalar(
                select(
                    func.coalesce(
                        func.sum(StorefrontProductViewDailyRow.view_count),
                        0,
                    )
                ).where(
                    StorefrontProductViewDailyRow.tenant_id == DEFAULT_TENANT_ID
                )
            )
            or 0
        )

    headers = {
        "CF-Connecting-IP": "203.0.113.17",
        "CF-IPCountry": "US",
    }
    created = client.post(
        f"/api/store/demo/skus/{sku['id']}/views",
        headers=headers,
        json={"event_id": event_id},
    )
    assert created.status_code == 204, created.text
    repeated = client.post(
        f"/api/store/demo/skus/{sku['id']}/views",
        headers=headers,
        json={"event_id": event_id},
    )
    assert repeated.status_code == 204, repeated.text

    with SessionLocal() as session:
        events = session.scalars(
            select(StorefrontProductViewEventRow).where(
                StorefrontProductViewEventRow.tenant_id == DEFAULT_TENANT_ID,
                StorefrontProductViewEventRow.event_id == event_id,
            )
        ).all()
        assert len(events) == 1
        # TestClient does not run through the trusted production proxy chain;
        # invalid non-IP scope values fail closed instead of trusting headers.
        assert events[0].ip_address == "0.0.0.0"
        assert events[0].country_code == "ZZ"
        after_total = int(
            session.scalar(
                select(
                    func.coalesce(
                        func.sum(StorefrontProductViewDailyRow.view_count),
                        0,
                    )
                ).where(
                    StorefrontProductViewDailyRow.tenant_id == DEFAULT_TENANT_ID
                )
            )
            or 0
        )
        assert after_total == before_total + 1

    dashboard = client.get("/api/v1/storefront-analytics", params={"days": 30})
    assert dashboard.status_code == 200, dashboard.text
    payload = dashboard.json()
    assert payload["summary"]["total_views"] >= 1
    assert payload["summary"]["unique_visitors"] >= 1
    assert payload["summary"]["viewed_products"] >= 1
    assert any(item["sku_id"] == sku["id"] for item in payload["products"])
    assert any(item["country_code"] == "ZZ" for item in payload["countries"])
    assert "ip_address" not in dashboard.text

    with SessionLocal() as session:
        with pytest.raises(ApplicationError) as denied:
            storefront_analytics_use_cases.get_storefront_analytics(
                session,
                tenant_id=DEFAULT_TENANT_ID,
                permissions=frozenset(),
                days=30,
            )
        assert denied.value.code == "PERMISSION_REQUIRED"


def test_storefront_visitor_headers_require_a_matching_trusted_client_ip(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("TRUST_CLOUDFLARE_VISITOR_HEADERS", "true")
    request = Request(
        {
            "type": "http",
            "http_version": "1.1",
            "method": "POST",
            "scheme": "https",
            "path": "/api/store/demo/skus/example/views",
            "raw_path": b"/api/store/demo/skus/example/views",
            "query_string": b"",
            "headers": [
                (b"cf-connecting-ip", b"203.0.113.17"),
                (b"cf-ipcountry", b"US"),
            ],
            "client": ("203.0.113.17", 43120),
            "server": ("4everapi.top", 443),
        }
    )
    visitor_ip = request_visitor_ip(request)
    assert visitor_ip == "203.0.113.17"
    assert request_country_code(request, visitor_ip=visitor_ip) == "US"
    assert request_country_code(request, visitor_ip="203.0.113.18") == "ZZ"


class _CatalogTranslationTestProvider:
    identity = TranslationIdentity(provider="deeplx", version="v1")

    def __init__(self) -> None:
        self.calls = 0

    def translate(
        self,
        text: str,
        *,
        source_locale: str,
        target_locale: str,
    ) -> str:
        assert source_locale == "zh-CN"
        assert target_locale == "en-US"
        self.calls += 1
        translated = text
        for source, target in (
            ("宠物缓存失效测试", "Pet cache invalidation test"),
            ("宠物无线饮水机（不锈钢款）", "Stainless-steel wireless pet fountain"),
            ("八片带门宠物围栏", "Eight-panel pet fence with door"),
            ("智能宠物喂食器 6L", "Smart 6L pet feeder"),
            ("饮水与喂食", "Water and feeding"),
            ("围栏与玩具", "Fences and toys"),
            ("智能硬件", "Smart hardware"),
            ("宠物饮水", "Pet drinking"),
            ("宠物围栏", "Pet fence"),
            ("智能喂食", "Smart feeding"),
            ("宠物用品", "Pet supplies"),
            ("智能用品", "Smart supplies"),
            ("不锈钢", "Stainless steel"),
            ("钢材", "Steel"),
        ):
            translated = translated.replace(source, target)
        return re.sub(r"[\u3400-\u9fff]+", "Translated", translated)


class _BlockingCatalogTranslationTestProvider(_CatalogTranslationTestProvider):
    def __init__(self) -> None:
        super().__init__()
        self.started = Event()
        self.release = Event()
        self.lock = Lock()

    def translate(
        self,
        text: str,
        *,
        source_locale: str,
        target_locale: str,
    ) -> str:
        assert source_locale == "zh-CN"
        assert target_locale == "en-US"
        with self.lock:
            self.calls += 1
        self.started.set()
        assert self.release.wait(timeout=2)
        return text.replace("宠物用品", "Pet supplies")


class _CheckpointCatalogTranslationTestProvider(_CatalogTranslationTestProvider):
    """Succeed once, then emulate an outage until the test restores it."""

    def __init__(self) -> None:
        super().__init__()
        self.outage_after_first = True

    def translate(
        self,
        text: str,
        *,
        source_locale: str,
        target_locale: str,
    ) -> str:
        if self.outage_after_first and self.calls >= 1:
            self.calls += 1
            raise TranslationProviderError(
                "translation provider returned HTTP 429"
            )
        return super().translate(
            text,
            source_locale=source_locale,
            target_locale=target_locale,
        )


class _SplitRecoveryCatalogTranslationTestProvider:
    identity = TranslationIdentity(provider="split-recovery-test", version="v1")

    def translate(
        self,
        text: str,
        *,
        source_locale: str,
        target_locale: str,
    ) -> str:
        if text.count("[[ATCV_") > 1:
            raise TranslationProviderError(
                "damaged batch boundary",
                recover_with_smaller_batches=True,
            )
        return text.replace("宠物用品", "Pet supplies").replace(
            "智能用品",
            "Smart supplies",
        )


def test_translation_memory_recovers_failed_batches_in_smaller_groups() -> None:
    successes, failures = translation_memory_service._translate_uncached_values(
        _SplitRecoveryCatalogTranslationTestProvider(),
        ["宠物用品", "智能用品"],
        source_locale="zh-CN",
        target_locale="en-US",
    )

    assert successes == {
        "宠物用品": "Pet supplies",
        "智能用品": "Smart supplies",
    }
    assert failures == {}


class _NonRecoverableCatalogTranslationTestProvider:
    identity = TranslationIdentity(provider="provider-failure-test", version="v1")

    def __init__(self) -> None:
        self.calls = 0

    def translate(
        self,
        text: str,
        *,
        source_locale: str,
        target_locale: str,
    ) -> str:
        self.calls += 1
        raise TranslationProviderError("translation provider returned HTTP 429")


def test_translation_memory_does_not_amplify_provider_failures() -> None:
    translator = _NonRecoverableCatalogTranslationTestProvider()

    successes, failures = translation_memory_service._translate_uncached_values(
        translator,
        ["宠物用品", "智能用品"],
        source_locale="zh-CN",
        target_locale="en-US",
    )

    assert successes == {}
    assert set(failures) == {"宠物用品", "智能用品"}
    assert translator.calls == 1


def test_translation_memory_chunks_large_database_reads_and_writes() -> None:
    provider = "large-memory-test"
    values = [f"批量翻译字段 {index}" for index in range(1_205)]
    translated = {value: f"Translated field {index}" for index, value in enumerate(values)}
    try:
        translation_memory_service._database_store_many(
            tenant_id=DEFAULT_TENANT_ID,
            source_locale="zh-CN",
            target_locale="en-US",
            provider=provider,
            provider_version="v1",
            translations=translated,
        )
        loaded = translation_memory_service._database_get_many(
            tenant_id=DEFAULT_TENANT_ID,
            source_locale="zh-CN",
            target_locale="en-US",
            provider=provider,
            provider_version="v1",
            sources_by_hash={
                translation_memory_service.translation_source_hash(value): value
                for value in values
            },
        )
        assert loaded == translated
    finally:
        with SessionLocal() as session:
            session.execute(
                delete(CatalogTextTranslationRow).where(
                    CatalogTextTranslationRow.tenant_id == DEFAULT_TENANT_ID,
                    CatalogTextTranslationRow.provider == provider,
                )
            )
            session.commit()


class _MixedLanguageCatalogTranslationTestProvider:
    identity = TranslationIdentity(provider="mixed-language-test", version="v1")

    def translate(
        self,
        text: str,
        *,
        source_locale: str,
        target_locale: str,
    ) -> str:
        assert target_locale == "pt"
        if source_locale == "zh-CN":
            replacements = {
                "MC宠物包": "MC Pet Pack",
                "宠物包": "Bolsa para animais",
                "产品简介": "Descrição do produto",
                "货号": "Número do artigo",
                "颜色": "Cor",
                "红色": "Vermelho",
            }
        elif source_locale == "en-US":
            replacements = {
                "MC Pet Pack": "Pacote de animais MC",
                "ITEM NO": "NÚMERO DO ITEM",
                "Travel goods": "Artigos de viagem",
            }
        else:  # pragma: no cover - protects the intended source routing
            raise AssertionError(f"unexpected source locale: {source_locale}")
        translated = text
        for source, target in replacements.items():
            translated = translated.replace(source, target)
        return translated


def test_public_catalog_translates_mixed_chinese_and_english_content() -> None:
    provider = _MixedLanguageCatalogTranslationTestProvider()
    translation_memory_service._reset_translation_memory_for_tests()
    with SessionLocal() as session:
        session.execute(
            delete(CatalogTextTranslationRow).where(
                CatalogTextTranslationRow.tenant_id == DEFAULT_TENANT_ID,
                CatalogTextTranslationRow.provider == "mixed-language-test",
            )
        )
        session.commit()

    try:
        translated, complete = (
            public_catalog_use_cases._translate_public_catalog_values(
                tenant_id=DEFAULT_TENANT_ID,
                translator=provider,
                values=[
                    "MC宠物包 3068",
                    "产品简介:货号 ITEM NO:3068",
                    "Travel goods",
                    "颜色",
                    "红色",
                ],
                source_locale="zh-CN",
                target_locale="pt",
                normalize_provider_output=True,
            )
        )

        assert translated["MC宠物包 3068"] == "Pacote de animais MC 3068"
        assert "ITEM NO" not in translated["产品简介:货号 ITEM NO:3068"]
        assert "Número do artigo" in translated[
            "产品简介:货号 ITEM NO:3068"
        ]
        assert translated["产品简介:货号 ITEM NO:3068"].startswith(
            "Descrição do produto:"
        )
        assert "Descriçãoção" not in translated[
            "产品简介:货号 ITEM NO:3068"
        ]
        assert translated["Travel goods"] == "Artigos de viagem"
        assert translated["颜色"] == "Cor"
        assert translated["红色"] == "Vermelho"
        assert complete == {
            "MC宠物包 3068",
            "产品简介:货号 ITEM NO:3068",
            "Travel goods",
            "颜色",
            "红色",
        }
    finally:
        translation_memory_service._reset_translation_memory_for_tests()
        with SessionLocal() as session:
            session.execute(
                delete(CatalogTextTranslationRow).where(
                    CatalogTextTranslationRow.tenant_id == DEFAULT_TENANT_ID,
                    CatalogTextTranslationRow.provider == "mixed-language-test",
                )
            )
            session.commit()


def test_mixed_language_capable_provider_skips_english_repair_pass(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    provider = _CatalogTranslationTestProvider()
    provider.translates_mixed_language_text = True
    calls: list[tuple[str, tuple[str, ...]]] = []

    def translate_once(**kwargs):
        source_locale = kwargs["source_locale"]
        values = tuple(kwargs["values"])
        calls.append((source_locale, values))
        return {
            value: (
                "Descrição do produto: Número do item:3068 "
                "Tamanho do produto:42x33x28.5"
            )
            for value in values
        }

    monkeypatch.setattr(
        public_catalog_use_cases,
        "translate_values_with_memory",
        translate_once,
    )
    source_value = (
        "产品简介:货号 ITEM NO:3068 产品尺寸SIZE:42x33x28.5"
    )

    translated, complete = (
        public_catalog_use_cases._translate_public_catalog_values(
            tenant_id=DEFAULT_TENANT_ID,
            translator=provider,
            values=[source_value],
            source_locale="zh-CN",
            target_locale="pt",
            normalize_provider_output=True,
        )
    )

    assert len(calls) == 1
    assert calls[0][0] == "zh-CN"
    localized = translated[source_value]
    assert "ITEM NO" not in localized
    assert "SIZE" not in localized
    assert source_value in complete


def test_public_product_option_labels_and_values_are_localized() -> None:
    translation = public_catalog_use_cases.PublicProductTranslation(
        name="Bolsa para animais",
        description=None,
        category=None,
        tags=(),
        display_tag=None,
        specifications={},
        option_labels={"颜色": "Cor", "尺寸": "Tamanho"},
        option_values={"红色": "Vermelho", "小号": "Pequeno"},
        complete=True,
    )

    localized = public_catalog_use_cases._localized_public_option_values(
        {
            "_sku2quotation": {
                "source": "PRODUCT_TEMPLATE",
                "variant_option_keys": ["颜色", "尺寸"],
            },
            "颜色": "红色",
            "尺寸": "小号",
            "商品型号": "MC-3068",
        },
        translation=translation,
    )

    assert localized["Cor"] == "Vermelho"
    assert localized["Tamanho"] == "Pequeno"
    assert localized["商品型号"] == "MC-3068"
    assert localized["_sku2quotation"]["variant_option_keys"] == [
        "Cor",
        "Tamanho",
    ]
    assert "颜色" not in localized
    assert "尺寸" not in localized


def test_translation_memory_coalesces_identical_inflight_text(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.delenv("REDIS_URL", raising=False)
    provider = _BlockingCatalogTranslationTestProvider()
    translation_memory_service._reset_translation_memory_for_tests()
    with SessionLocal() as session:
        session.execute(
            delete(CatalogTextTranslationRow).where(
                CatalogTextTranslationRow.tenant_id == DEFAULT_TENANT_ID
            )
        )
        session.commit()

    def translate() -> dict[str, str]:
        return translation_memory_service.translate_values_with_memory(
            tenant_id=DEFAULT_TENANT_ID,
            translator=provider,
            values=["宠物用品"],
            source_locale="zh-CN",
            target_locale="en-US",
        )

    try:
        with ThreadPoolExecutor(max_workers=2) as executor:
            first = executor.submit(translate)
            assert provider.started.wait(timeout=2)
            second = executor.submit(translate)
            sleep(0.05)
            provider.release.set()
            assert first.result(timeout=2)["宠物用品"] == "Pet supplies"
            assert second.result(timeout=2)["宠物用品"] == "Pet supplies"
        assert provider.calls == 1
    finally:
        provider.release.set()
        translation_memory_service._reset_translation_memory_for_tests()
        with SessionLocal() as session:
            session.execute(
                delete(CatalogTextTranslationRow).where(
                    CatalogTextTranslationRow.tenant_id == DEFAULT_TENANT_ID
                )
            )
            session.commit()


def test_public_catalog_reuses_on_demand_translation_memory(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    provider = _CatalogTranslationTestProvider()
    monkeypatch.setattr(
        public_catalog_use_cases,
        "catalog_translation_is_configured",
        lambda: True,
    )
    monkeypatch.setattr(
        public_catalog_use_cases,
        "configured_catalog_translator",
        lambda: provider,
    )
    with SessionLocal() as session:
        session.execute(
            delete(CatalogTextTranslationRow).where(
                CatalogTextTranslationRow.tenant_id == DEFAULT_TENANT_ID
            )
        )
        session.commit()

    try:
        store_response = client.get(
            "/api/store/demo",
            params={"locale": "en-US"},
        )
        assert store_response.status_code == 200, store_response.text
        assert "en-US" in store_response.json()["available_locales"]

        first_page = client.get(
            "/api/store/demo/skus",
            params={
                "locale": "en-US",
                "page": 1,
                "page_size": 1,
                "include_facets": "false",
            },
        )
        assert first_page.status_code == 200, first_page.text
        first_item = first_page.json()["items"][0]
        assert first_item["translation_status"] == "TRANSLATED"
        first_page_calls = provider.calls
        assert first_page_calls > 0

        repeated_first_page = client.get(
            "/api/store/demo/skus",
            params={
                "locale": "en-US",
                "page": 1,
                "page_size": 1,
                "include_facets": "false",
            },
        )
        assert repeated_first_page.status_code == 200, repeated_first_page.text
        assert provider.calls == first_page_calls

        second_page = client.get(
            "/api/store/demo/skus",
            params={
                "locale": "en-US",
                "page": 2,
                "page_size": 1,
                "include_facets": "false",
            },
        )
        assert second_page.status_code == 200, second_page.text
        assert second_page.json()["items"][0]["translation_status"] == "TRANSLATED"
        second_page_calls = provider.calls
        assert second_page_calls > first_page_calls

        chinese_page = client.get(
            "/api/store/demo/skus",
            params={"page": 1, "page_size": 1, "include_facets": "false"},
        )
        assert chinese_page.status_code == 200, chinese_page.text
        assert chinese_page.json()["items"][0]["translation_status"] == "SOURCE"
        assert provider.calls == second_page_calls

        detail = client.get(
            f"/api/store/demo/skus/{first_item['id']}",
            params={"locale": "en-US"},
        )
        assert detail.status_code == 200, detail.text
        assert detail.json()["translation_status"] == "TRANSLATED"
        assert provider.calls == second_page_calls

        with SessionLocal() as session:
            changed_sku = session.get(SkuRow, UUID(first_item["id"]))
            assert changed_sku is not None
            original_name = changed_sku.name
            original_version = changed_sku.version
            changed_sku.name = "宠物缓存失效测试"
            changed_sku.version += 1
            session.commit()
        try:
            changed_detail = client.get(
                f"/api/store/demo/skus/{first_item['id']}",
                params={"locale": "en-US"},
            )
            assert changed_detail.status_code == 200, changed_detail.text
            assert changed_detail.json()["name"].startswith("Pet")
            assert provider.calls > second_page_calls
        finally:
            with SessionLocal() as session:
                changed_sku = session.get(SkuRow, UUID(first_item["id"]))
                assert changed_sku is not None
                changed_sku.name = original_name
                changed_sku.version = original_version
                session.commit()

        with SessionLocal() as session:
            assert (
                session.scalar(
                    select(func.count(CatalogTextTranslationRow.id)).where(
                        CatalogTextTranslationRow.tenant_id == DEFAULT_TENANT_ID
                    )
                )
                or 0
            ) > 0
            assert (
                session.scalar(
                    select(func.count(CatalogSkuTranslationRow.id)).where(
                        CatalogSkuTranslationRow.tenant_id == DEFAULT_TENANT_ID
                    )
                )
                == 0
            )
    finally:
        with SessionLocal() as session:
            session.execute(
                delete(CatalogTextTranslationRow).where(
                    CatalogTextTranslationRow.tenant_id == DEFAULT_TENANT_ID
                )
            )
            session.commit()


def test_public_product_detail_reuses_one_product_translation_batch(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    provider = _CatalogTranslationTestProvider()
    monkeypatch.setattr(
        public_catalog_use_cases,
        "catalog_translation_is_configured",
        lambda: True,
    )
    monkeypatch.setattr(
        public_catalog_use_cases,
        "configured_catalog_translator",
        lambda: provider,
    )

    def fail_duplicate_sku_translation(*_args, **_kwargs):
        raise AssertionError(
            "product detail must not run a second per-SKU translation batch"
        )

    monkeypatch.setattr(
        public_catalog_use_cases,
        "_live_sku_translation_map",
        fail_duplicate_sku_translation,
    )
    with SessionLocal() as session:
        session.execute(
            delete(CatalogTextTranslationRow).where(
                CatalogTextTranslationRow.tenant_id == DEFAULT_TENANT_ID
            )
        )
        session.commit()

    try:
        chinese_listing = client.get(
            "/api/store/demo/products",
            params={"page": 1, "page_size": 1, "include_facets": "false"},
        )
        assert chinese_listing.status_code == 200, chinese_listing.text
        product_id = chinese_listing.json()["items"][0]["id"]

        detail = client.get(
            f"/api/store/demo/products/{product_id}",
            params={"locale": "en-US"},
        )
        assert detail.status_code == 200, detail.text
        payload = detail.json()
        assert payload["translation_status"] == "TRANSLATED"
        assert payload["skus"]
        assert all(
            sku["translation_status"] == "TRANSLATED"
            for sku in payload["skus"]
        )
        assert provider.calls > 0
    finally:
        with SessionLocal() as session:
            session.execute(
                delete(CatalogTextTranslationRow).where(
                    CatalogTextTranslationRow.tenant_id == DEFAULT_TENANT_ID
                )
            )
            session.commit()


def test_catalog_translation_job_reports_progress_and_caches_results(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    provider = _CatalogTranslationTestProvider()
    monkeypatch.setenv("TRANSLATION_PACKAGE_STORAGE_BACKEND", "local")
    monkeypatch.setenv("TRANSLATION_PACKAGE_LOCAL_ROOT", str(tmp_path))
    monkeypatch.delenv("TRANSLATION_PACKAGE_PUBLIC_BASE_URL", raising=False)
    monkeypatch.setattr(
        catalog_translation_use_cases,
        "catalog_translation_is_configured",
        lambda: True,
    )
    monkeypatch.setattr(
        catalog_translation_use_cases,
        "configured_catalog_translator",
        lambda: provider,
    )
    monkeypatch.setattr(
        catalog_translation_use_cases,
        "_dispatch_translation_job",
        lambda **kwargs: catalog_translation_use_cases._run_translation_job(
            **kwargs
        ),
    )

    try:
        initial = client.get("/api/v1/catalog/translations/status")
        assert initial.status_code == 200, initial.text
        initial_payload = initial.json()
        assert initial_payload["provider_configured"] is True
        assert initial_payload["package_storage_configured"] is True
        assert initial_payload["pending_skus"] == initial_payload["total_skus"] == 3

        started = client.post(
            "/api/v1/catalog/translations/jobs",
            json={
                "target_locale": "en-US",
                "mode": "INCREMENTAL",
                "confirm_full_rebuild": False,
            },
        )
        assert started.status_code == 202, started.text
        job_id = started.json()["id"]

        finished = client.get(
            f"/api/v1/catalog/translations/jobs/{job_id}"
        )
        assert finished.status_code == 200, finished.text
        finished_payload = finished.json()
        assert finished_payload["status"] == "SUCCEEDED"
        assert finished_payload["processed_skus"] == 3
        assert finished_payload["failed_skus"] == 0
        assert finished_payload["progress_percent"] == 100.0
        assert finished_payload["stage"] == "PUBLISHED"
        assert finished_payload["package_published"] is True
        assert finished_payload["package_version"] == 1

        final = client.get("/api/v1/catalog/translations/status")
        assert final.status_code == 200, final.text
        assert final.json()["pending_skus"] == 0
        assert final.json()["translated_skus"] == 3
        assert final.json()["package_outdated"] is False
        package = final.json()["package"]
        assert package["version"] == 1
        assert package["sku_count"] == 3
        assert package["product_count"] >= 1

        manifest = client.get("/api/store/demo/language-packages/en-US")
        assert manifest.status_code == 200, manifest.text
        assert manifest.json()["content_sha256"] == package["content_sha256"]
        download = client.get(manifest.json()["download_url"])
        assert download.status_code == 200, download.text
        assert download.headers["cache-control"].endswith("immutable")
        language_payload = download.json()
        assert language_payload["schema"] == "atc-catalog-language-pack"
        assert language_payload["version"] == 1
        source_products = client.get(
            "/api/store/demo/products",
            params={"page": 1, "page_size": 24},
        )
        assert source_products.status_code == 200, source_products.text
        for product in source_products.json()["items"]:
            assert product["translation_source_hash"] == (
                language_payload["products"][product["id"]]["source_hash"]
            )
        source_skus = client.get(
            "/api/store/demo/skus",
            params={"page": 1, "page_size": 24},
        )
        assert source_skus.status_code == 200, source_skus.text
        for sku in source_skus.json()["items"]:
            assert sku["translation_source_hash"] == (
                language_payload["skus"][sku["id"]]["source_hash"]
            )

    finally:
        with SessionLocal() as session:
            session.execute(
                delete(CatalogLanguagePackRow).where(
                    CatalogLanguagePackRow.tenant_id == DEFAULT_TENANT_ID
                )
            )
            session.execute(
                delete(CatalogSkuTranslationRow).where(
                    CatalogSkuTranslationRow.tenant_id == DEFAULT_TENANT_ID
                )
            )
            session.execute(
                delete(CatalogTranslationJobRow).where(
                    CatalogTranslationJobRow.tenant_id == DEFAULT_TENANT_ID
                )
            )
            session.commit()


def test_catalog_translation_job_can_pause_and_resume_at_safe_checkpoint(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    provider = _CatalogTranslationTestProvider()
    monkeypatch.setenv("TRANSLATION_PACKAGE_STORAGE_BACKEND", "local")
    monkeypatch.setenv("TRANSLATION_PACKAGE_LOCAL_ROOT", str(tmp_path))
    monkeypatch.setattr(
        catalog_translation_use_cases,
        "catalog_translation_is_configured",
        lambda: True,
    )
    monkeypatch.setattr(
        catalog_translation_use_cases,
        "configured_catalog_translator",
        lambda: provider,
    )
    dispatches: list[dict[str, object]] = []
    monkeypatch.setattr(
        catalog_translation_use_cases,
        "_dispatch_translation_job",
        lambda **kwargs: dispatches.append(kwargs),
    )

    started = client.post(
        "/api/v1/catalog/translations/jobs",
        json={
            "target_locale": "es",
            "mode": "FULL_REBUILD",
            "confirm_full_rebuild": True,
        },
    )
    assert started.status_code == 202, started.text
    job_id = UUID(started.json()["id"])
    assert len(dispatches) == 1

    try:
        with SessionLocal() as session:
            row = session.get(CatalogTranslationJobRow, job_id)
            assert row is not None
            assert len(row.remaining_sku_ids) == row.total_skus == 3
            row.status = "RUNNING"
            row.stage = "TRANSLATING"
            session.commit()

        requested = client.post(
            f"/api/v1/catalog/translations/jobs/{job_id}/pause"
        )
        assert requested.status_code == 200, requested.text
        assert requested.json()["status"] == "RUNNING"
        assert requested.json()["pause_requested"] is True

        with SessionLocal() as session:
            row = session.get(CatalogTranslationJobRow, job_id)
            assert row is not None
            assert catalog_translation_use_cases._pause_at_safe_checkpoint(
                session,
                row,
            ) is True

        paused = client.get(
            f"/api/v1/catalog/translations/jobs/{job_id}"
        )
        assert paused.status_code == 200, paused.text
        assert paused.json()["status"] == "PAUSED"
        assert paused.json()["stage"] == "PAUSED"
        assert paused.json()["pause_requested"] is False
        assert paused.json()["paused_at"] is not None

        duplicate = client.post(
            "/api/v1/catalog/translations/jobs",
            json={
                "target_locale": "es",
                "mode": "INCREMENTAL",
                "confirm_full_rebuild": False,
            },
        )
        assert duplicate.status_code == 202, duplicate.text
        assert duplicate.json()["id"] == str(job_id)
        assert duplicate.json()["status"] == "PAUSED"

        resumed = client.post(
            f"/api/v1/catalog/translations/jobs/{job_id}/resume"
        )
        assert resumed.status_code == 200, resumed.text
        assert resumed.json()["status"] == "QUEUED"
        assert resumed.json()["pause_requested"] is False
        assert len(dispatches) == 2

        duplicate_resume = client.post(
            f"/api/v1/catalog/translations/jobs/{job_id}/resume"
        )
        assert duplicate_resume.status_code == 409, duplicate_resume.text
        assert len(dispatches) == 2

        paused_again = client.post(
            f"/api/v1/catalog/translations/jobs/{job_id}/pause"
        )
        assert paused_again.status_code == 200, paused_again.text
        assert paused_again.json()["status"] == "PAUSED"
    finally:
        with SessionLocal() as session:
            session.execute(
                delete(CatalogTranslationJobRow).where(
                    CatalogTranslationJobRow.id == job_id
                )
            )
            session.commit()


def test_catalog_translation_job_resumes_failed_provider_from_last_batch(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    provider = _CheckpointCatalogTranslationTestProvider()
    monkeypatch.setenv("TRANSLATION_PACKAGE_STORAGE_BACKEND", "local")
    monkeypatch.setenv("TRANSLATION_PACKAGE_LOCAL_ROOT", str(tmp_path))
    monkeypatch.setenv("CATALOG_TRANSLATION_BATCH_SIZE", "1")
    monkeypatch.setenv("CATALOG_TRANSLATION_PROVIDER_RETRIES", "1")
    monkeypatch.setattr(catalog_translation_use_cases.time, "sleep", lambda _s: None)
    monkeypatch.setattr(
        catalog_translation_use_cases,
        "catalog_translation_is_configured",
        lambda: True,
    )
    monkeypatch.setattr(
        catalog_translation_use_cases,
        "configured_catalog_translator",
        lambda: provider,
    )
    monkeypatch.setattr(
        catalog_translation_use_cases,
        "_dispatch_translation_job",
        lambda **kwargs: catalog_translation_use_cases._run_translation_job(
            **kwargs
        ),
    )

    started = client.post(
        "/api/v1/catalog/translations/jobs",
        json={
            "target_locale": "en-US",
            "mode": "FULL_REBUILD",
            "confirm_full_rebuild": True,
        },
    )
    assert started.status_code == 202, started.text
    job_id = UUID(started.json()["id"])

    try:
        interrupted = client.get(
            f"/api/v1/catalog/translations/jobs/{job_id}"
        )
        assert interrupted.status_code == 200, interrupted.text
        interrupted_payload = interrupted.json()
        assert interrupted_payload["status"] == "FAILED"
        assert interrupted_payload["processed_skus"] == 1
        assert interrupted_payload["remaining_skus"] == 2
        assert interrupted_payload["resumable"] is True
        assert "断点" in interrupted_payload["error_message"]
        assert provider.calls == 3

        provider.outage_after_first = False
        resumed = client.post(
            f"/api/v1/catalog/translations/jobs/{job_id}/resume"
        )
        assert resumed.status_code == 200, resumed.text

        finished = client.get(
            f"/api/v1/catalog/translations/jobs/{job_id}"
        )
        assert finished.status_code == 200, finished.text
        finished_payload = finished.json()
        assert finished_payload["status"] == "SUCCEEDED"
        assert finished_payload["processed_skus"] == 3
        assert finished_payload["remaining_skus"] == 0
        assert finished_payload["resumable"] is False
        assert finished_payload["package_published"] is True
    finally:
        with SessionLocal() as session:
            session.execute(
                delete(CatalogLanguagePackRow).where(
                    CatalogLanguagePackRow.tenant_id == DEFAULT_TENANT_ID
                )
            )
            session.execute(
                delete(CatalogSkuTranslationRow).where(
                    CatalogSkuTranslationRow.tenant_id == DEFAULT_TENANT_ID
                )
            )
            session.execute(
                delete(CatalogTranslationJobRow).where(
                    CatalogTranslationJobRow.id == job_id
                )
            )
            session.commit()


def test_catalog_translation_startup_recovery_preserves_checkpoint(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    provider = _CatalogTranslationTestProvider()
    monkeypatch.setenv("TRANSLATION_PACKAGE_STORAGE_BACKEND", "local")
    monkeypatch.setenv("TRANSLATION_PACKAGE_LOCAL_ROOT", str(tmp_path))
    monkeypatch.setattr(
        catalog_translation_use_cases,
        "catalog_translation_is_configured",
        lambda: True,
    )
    monkeypatch.setattr(
        catalog_translation_use_cases,
        "configured_catalog_translator",
        lambda: provider,
    )
    monkeypatch.setattr(
        catalog_translation_use_cases,
        "_dispatch_translation_job",
        lambda **_kwargs: None,
    )

    started = client.post(
        "/api/v1/catalog/translations/jobs",
        json={
            "target_locale": "en-US",
            "mode": "FULL_REBUILD",
            "confirm_full_rebuild": True,
        },
    )
    assert started.status_code == 202, started.text
    job_id = UUID(started.json()["id"])

    try:
        with SessionLocal() as session:
            row = session.get(CatalogTranslationJobRow, job_id)
            assert row is not None
            row.status = "RUNNING"
            row.stage = "TRANSLATING"
            row.started_at = datetime.now(UTC)
            row.processed_skus = 1
            row.remaining_sku_ids = row.remaining_sku_ids[1:]
            session.commit()

        recovered = (
            catalog_translation_use_cases.recover_interrupted_translation_jobs()
        )
        assert recovered >= 1

        payload = client.get(
            f"/api/v1/catalog/translations/jobs/{job_id}"
        ).json()
        assert payload["status"] == "PAUSED"
        assert payload["stage"] == "PAUSED"
        assert payload["processed_skus"] == 1
        assert payload["remaining_skus"] == 2
        assert payload["resumable"] is True
        assert "服务重启" in payload["error_message"]
    finally:
        with SessionLocal() as session:
            session.execute(
                delete(CatalogTranslationJobRow).where(
                    CatalogTranslationJobRow.id == job_id
                )
            )
            session.commit()


def test_public_category_and_all_products_follow_managed_sort_order() -> None:
    initial_response = client.get("/api/store/demo/skus")
    assert initial_response.status_code == 200, initial_response.text
    initial_categories = initial_response.json()["categories"]
    assert len(initial_categories) > 1
    desired_categories = list(reversed(initial_categories))
    originals: dict[UUID, int] = {}

    with SessionLocal() as session:
        rows = list(
            session.scalars(
                select(ProductCategoryRow).where(
                    ProductCategoryRow.tenant_id == DEFAULT_TENANT_ID,
                    ProductCategoryRow.path.in_(initial_categories)
                )
            ).all()
        )
        assert len(rows) == len(initial_categories)
        rows_by_path = {row.path: row for row in rows}
        assert all(row.parent_id is None for row in rows)
        originals = {row.id: row.sort_order for row in rows}
        for position, path in enumerate(desired_categories):
            rows_by_path[path].sort_order = position
        session.commit()

    try:
        reordered_response = client.get(
            "/api/store/demo/skus",
            params={"page": 1, "page_size": 1},
        )
        assert reordered_response.status_code == 200, reordered_response.text
        reordered_payload = reordered_response.json()
        assert reordered_payload["categories"] == desired_categories
        assert reordered_payload["items"][0]["category"] == desired_categories[0]
    finally:
        with SessionLocal() as session:
            rows = list(
                session.scalars(
                    select(ProductCategoryRow).where(
                        ProductCategoryRow.tenant_id == DEFAULT_TENANT_ID,
                        ProductCategoryRow.id.in_(list(originals))
                    )
                ).all()
            )
            for row in rows:
                row.sort_order = originals[row.id]
            session.commit()


def test_all_products_position_is_merchant_controlled_and_publicly_projected() -> None:
    initial_layout_response = client.get("/api/v1/categories/layout")
    assert initial_layout_response.status_code == 200, initial_layout_response.text
    initial_layout = initial_layout_response.json()
    root_count = initial_layout["root_category_count"]
    target_position = min(2, root_count)

    try:
        updated_response = client.patch(
            "/api/v1/categories/layout",
            json={"all_products_position": target_position},
        )
        assert updated_response.status_code == 200, updated_response.text
        assert updated_response.json() == {
            "all_products_position": target_position,
            "root_category_count": root_count,
        }

        persisted_response = client.get("/api/v1/categories/layout")
        assert persisted_response.status_code == 200, persisted_response.text
        assert (
            persisted_response.json()["all_products_position"]
            == target_position
        )

        public_response = client.get("/api/store/demo/skus")
        assert public_response.status_code == 200, public_response.text
        public_payload = public_response.json()
        visible_roots = {
            path.replace("／", "/").split("/", 1)[0].strip()
            for path in public_payload["categories"]
        }
        with SessionLocal() as session:
            ordered_roots = list(
                session.scalars(
                    select(ProductCategoryRow)
                    .where(
                        ProductCategoryRow.tenant_id == DEFAULT_TENANT_ID,
                        ProductCategoryRow.parent_id.is_(None),
                    )
                    .order_by(
                        ProductCategoryRow.sort_order,
                        ProductCategoryRow.name,
                    )
                ).all()
            )
        expected_public_position = sum(
            1 for row in ordered_roots[:target_position] if row.name in visible_roots
        )
        assert (
            public_payload["all_products_position"]
            == expected_public_position
        )

        invalid_response = client.patch(
            "/api/v1/categories/layout",
            json={"all_products_position": root_count + 1},
        )
        assert invalid_response.status_code == 409
        assert (
            invalid_response.json()["detail"]["code"]
            == "CATEGORY_LAYOUT_POSITION_INVALID"
        )
    finally:
        restored = client.patch(
            "/api/v1/categories/layout",
            json={
                "all_products_position": min(
                    initial_layout["all_products_position"],
                    root_count,
                )
            },
        )
        assert restored.status_code == 200, restored.text


def test_public_catalog_paginates_in_database_and_can_skip_facets() -> None:
    first = client.get(
        "/api/store/demo/skus",
        params={
            "page": 1,
            "page_size": 1,
            "include_facets": "false",
        },
    )
    second = client.get(
        "/api/store/demo/skus",
        params={
            "page": 2,
            "page_size": 1,
            "include_facets": "false",
        },
    )

    assert first.status_code == 200, first.text
    assert second.status_code == 200, second.text
    assert first.json()["total"] == second.json()["total"] == 3
    assert first.json()["pages"] == second.json()["pages"] == 3
    assert first.json()["categories"] == []
    assert first.json()["tags"] == []
    assert first.json()["items"][0]["id"] != second.json()["items"][0]["id"]


class _EmptyCompileResult:
    def all(self):
        return []


class _PostgresCompileBind:
    dialect = type("_DialectName", (), {"name": "postgresql"})()


class _CompileOnlyPostgresSession:
    bind = _PostgresCompileBind()

    def __init__(self) -> None:
        self.statements = []

    def get_bind(self):
        return self.bind

    def execute(self, statement):
        statement.compile(dialect=postgresql.dialect())
        self.statements.append(statement)
        return _EmptyCompileResult()

    def scalars(self, statement):
        statement.compile(dialect=postgresql.dialect())
        self.statements.append(statement)
        return _EmptyCompileResult()


def test_public_catalog_page_query_has_database_limit_and_offset() -> None:
    session = _CompileOnlyPostgresSession()

    public_catalog_repository.list_public_catalog_page(
        session,
        tenant_id=uuid4(),
        now=datetime.now(UTC),
        query="",
        category=None,
        tags={"宠物"},
        page=400,
        page_size=24,
    )

    assert len(session.statements) == 1
    sql = str(
        session.statements[0].compile(dialect=postgresql.dialect())
    ).upper()
    assert " LIMIT " in sql
    assert " OFFSET " in sql
    assert "JSONB_ARRAY_ELEMENTS_TEXT" in sql


def test_postgres_hybrid_search_bounds_candidates_before_orm_hydration() -> None:
    session = _CompileOnlyPostgresSession()

    result = hybrid_product_search(
        session,
        tenant_id=uuid4(),
        query="无线宠物饮水机",
        limit=24,
        embedder=DeterministicFeatureHashEmbedding(),
    )

    assert result["results"] == []
    assert len(session.statements) == 2
    semantic_sql, lexical_sql = (
        str(statement.compile(dialect=postgresql.dialect())).upper()
        for statement in session.statements
    )
    assert "FROM EMBEDDINGS" in semantic_sql
    assert " LIMIT " in semantic_sql
    assert " LIMIT " in lexical_sql


def test_category_template_download_and_incremental_import_are_idempotent() -> None:
    download = client.get("/api/v1/category-template.xlsx")
    assert download.status_code == 200, download.text
    assert download.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    downloaded_workbook = load_workbook(BytesIO(download.content), read_only=True)
    try:
        assert CATEGORY_TEMPLATE_SHEET in downloaded_workbook.sheetnames
        downloaded_sheet = downloaded_workbook[CATEGORY_TEMPLATE_SHEET]
        assert tuple(downloaded_sheet.cell(1, column).value for column in (1, 2)) == (
            CATEGORY_TEMPLATE_HEADERS
        )
    finally:
        downloaded_workbook.close()

    suffix = uuid4().hex[:10].upper()
    original_category_state: dict[UUID, tuple[int, int]] = {}
    original_reorder_event_ids: set[UUID] = set()
    with SessionLocal() as session:
        original_category_state = {
            row.id: (row.sort_order, row.version)
            for row in session.scalars(
                select(ProductCategoryRow).where(
                    ProductCategoryRow.tenant_id == DEFAULT_TENANT_ID
                )
            ).all()
        }
        original_reorder_event_ids = set(
            session.scalars(
                select(ProductAuditEventRow.id).where(
                    ProductAuditEventRow.tenant_id == DEFAULT_TENANT_ID,
                    ProductAuditEventRow.action == "category.import_reordered",
                )
            ).all()
        )
    primary_one = f"导入一级甲-{suffix}"
    primary_two = f"导入一级乙-{suffix}"
    secondary_one = f"导入二级甲-{suffix}"
    secondary_two = f"导入二级乙-{suffix}"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "客户分类数据"
    sheet.append(list(CATEGORY_TEMPLATE_HEADERS))
    sheet.append([primary_one, secondary_one])
    sheet.append([primary_one, secondary_two])
    sheet.append([primary_two, None])
    sheet.append([primary_one, secondary_one])
    sheet.append([None, None])
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    created_ids: list[str] = []
    try:
        first = client.post(
            "/api/v1/categories/import",
            files={
                "file": (
                    "分类模板.xlsx",
                    content.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert first.status_code == 200, first.text
        assert first.json() == {
            "processed_rows": 3,
            "primary_created": 2,
            "secondary_created": 2,
            "primary_existing": 0,
            "secondary_existing": 0,
            "duplicate_rows_ignored": 1,
            "blank_rows_ignored": 1,
        }

        second = client.post(
            "/api/v1/categories/import",
            files={
                "file": (
                    "分类模板.xlsx",
                    content.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert second.status_code == 200, second.text
        assert second.json()["primary_created"] == 0
        assert second.json()["secondary_created"] == 0
        assert second.json()["primary_existing"] == 2
        assert second.json()["secondary_existing"] == 2

        reordered_workbook = Workbook()
        reordered_sheet = reordered_workbook.active
        reordered_sheet.title = "客户分类新顺序"
        reordered_sheet.append(list(CATEGORY_TEMPLATE_HEADERS))
        reordered_sheet.append([primary_two, None])
        reordered_sheet.append([primary_one, secondary_two])
        reordered_sheet.append([primary_one, secondary_one])
        reordered_content = BytesIO()
        reordered_workbook.save(reordered_content)
        reordered_workbook.close()
        reordered = client.post(
            "/api/v1/categories/import",
            files={
                "file": (
                    "分类新顺序.xlsx",
                    reordered_content.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert reordered.status_code == 200, reordered.text
        assert reordered.json()["primary_created"] == 0
        assert reordered.json()["secondary_created"] == 0
        assert reordered.json()["primary_existing"] == 2
        assert reordered.json()["secondary_existing"] == 2

        with SessionLocal() as session:
            imported = list(
                session.scalars(
                    select(ProductCategoryRow).where(
                        ProductCategoryRow.tenant_id == DEFAULT_TENANT_ID,
                        ProductCategoryRow.name.in_(
                            [primary_one, primary_two, secondary_one, secondary_two]
                        ),
                    )
                ).all()
            )
            created_ids = [str(row.id) for row in imported]
            root_by_name = {
                row.name: row for row in imported if row.parent_id is None
            }
            children = [row for row in imported if row.parent_id is not None]
            assert set(root_by_name) == {primary_one, primary_two}
            assert [
                row.name
                for row in sorted(
                    root_by_name.values(), key=lambda row: row.sort_order
                )
            ] == [primary_two, primary_one]
            assert {row.name for row in children} == {secondary_one, secondary_two}
            assert all(row.parent_id == root_by_name[primary_one].id for row in children)
            assert [row.name for row in sorted(children, key=lambda row: row.sort_order)] == [
                secondary_two,
                secondary_one,
            ]
    finally:
        with SessionLocal() as session:
            current_reorder_event_ids = set(
                session.scalars(
                    select(ProductAuditEventRow.id).where(
                        ProductAuditEventRow.tenant_id == DEFAULT_TENANT_ID,
                        ProductAuditEventRow.action == "category.import_reordered",
                    )
                ).all()
            )
            new_reorder_event_ids = (
                current_reorder_event_ids - original_reorder_event_ids
            )
            if new_reorder_event_ids:
                session.execute(
                    delete(ProductAuditEventRow).where(
                        ProductAuditEventRow.id.in_(new_reorder_event_ids)
                    )
                )
            if not created_ids:
                created_ids = [
                    str(value)
                    for value in session.scalars(
                        select(ProductCategoryRow.id).where(
                            ProductCategoryRow.tenant_id == DEFAULT_TENANT_ID,
                            ProductCategoryRow.name.in_(
                                [primary_one, primary_two, secondary_one, secondary_two]
                            ),
                        )
                    ).all()
                ]
            if created_ids:
                category_ids = [UUID(value) for value in created_ids]
                session.execute(
                    delete(ProductAuditEventRow).where(
                        ProductAuditEventRow.entity_type == "CATEGORY",
                        ProductAuditEventRow.entity_id.in_(created_ids),
                    )
                )
                session.execute(
                    delete(ProductCategoryRow).where(
                        ProductCategoryRow.id.in_(category_ids),
                        ProductCategoryRow.parent_id.is_not(None),
                    )
                )
                session.flush()
                session.execute(
                    delete(ProductCategoryRow).where(
                        ProductCategoryRow.id.in_(category_ids)
                    )
                )
            if original_category_state:
                original_rows = session.scalars(
                    select(ProductCategoryRow).where(
                        ProductCategoryRow.id.in_(list(original_category_state))
                    )
                ).all()
                for row in original_rows:
                    row.sort_order, row.version = original_category_state[row.id]
            session.commit()


def test_category_delete_cascades_children_without_deleting_products() -> None:
    suffix = uuid4().hex[:10].upper()
    category_ids: list[str] = []
    product_ids = [uuid4(), uuid4(), uuid4()]
    definition_id = uuid4()
    attribute_id = uuid4()
    try:
        root_response = client.post(
            "/api/v1/categories",
            json={
                "code": f"DELETE-ROOT-{suffix}",
                "name": f"待删除一级-{suffix}",
                "sort_order": 900,
            },
        )
        assert root_response.status_code == 201, root_response.text
        root = root_response.json()
        category_ids.append(root["id"])

        children = []
        for position in range(2):
            response = client.post(
                "/api/v1/categories",
                json={
                    "parent_id": root["id"],
                    "code": f"DELETE-CHILD-{position}-{suffix}",
                    "name": f"待删除二级-{position}-{suffix}",
                    "sort_order": position,
                },
            )
            assert response.status_code == 201, response.text
            children.append(response.json())
            category_ids.append(response.json()["id"])

        with SessionLocal() as session:
            session.add_all(
                [
                    ProductRow(
                        id=product_ids[0],
                        tenant_id=DEFAULT_TENANT_ID,
                        product_code=f"DELETE-CATEGORY-PRODUCT-A-{suffix}",
                        name=f"分类删除保留商品甲-{suffix}",
                        category_id=UUID(root["id"]),
                        status="ACTIVE",
                        default_unit="PCS",
                        current_version=1,
                        search_document_version=8,
                        created_by=DEFAULT_OWNER_USER_ID,
                        updated_by=DEFAULT_OWNER_USER_ID,
                    ),
                    ProductRow(
                        id=product_ids[1],
                        tenant_id=DEFAULT_TENANT_ID,
                        product_code=f"DELETE-CATEGORY-PRODUCT-B-{suffix}",
                        name=f"分类删除保留商品乙-{suffix}",
                        category_id=UUID(children[0]["id"]),
                        status="ACTIVE",
                        default_unit="PCS",
                        current_version=1,
                        search_document_version=9,
                        created_by=DEFAULT_OWNER_USER_ID,
                        updated_by=DEFAULT_OWNER_USER_ID,
                    ),
                    ProductRow(
                        id=product_ids[2],
                        tenant_id=DEFAULT_TENANT_ID,
                        product_code=f"DELETE-CATEGORY-ARCHIVED-{suffix}",
                        name=f"已归档历史商品-{suffix}",
                        category_id=UUID(children[0]["id"]),
                        status="ARCHIVED",
                        default_unit="PCS",
                        current_version=1,
                        search_document_version=0,
                        created_by=DEFAULT_OWNER_USER_ID,
                        updated_by=DEFAULT_OWNER_USER_ID,
                    ),
                ]
            )
            session.add(
                AttributeDefinitionRow(
                    id=definition_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    category_id=UUID(children[0]["id"]),
                    attribute_key=f"delete_category_{suffix.lower()}",
                    display_name="待保留的属性值",
                    data_type="TEXT",
                    is_required=False,
                    is_variant=False,
                    is_filterable=True,
                    is_matchable=True,
                    status="ACTIVE",
                    version=1,
                )
            )
            session.flush()
            session.add(
                ProductAttributeRow(
                    id=attribute_id,
                    tenant_id=DEFAULT_TENANT_ID,
                    product_id=product_ids[1],
                    attribute_definition_id=definition_id,
                    attribute_key=f"delete_category_{suffix.lower()}",
                    value_text="属性值继续保留",
                    review_status="CONFIRMED",
                )
            )
            session.commit()

        impact_response = client.get(
            f"/api/v1/categories/{root['id']}/delete-impact"
        )
        assert impact_response.status_code == 200, impact_response.text
        assert impact_response.json() == {
            "category_id": root["id"],
            "category_name": root["name"],
            "is_primary": True,
            "child_category_count": 2,
            "affected_product_count": 2,
            "attribute_definition_count": 1,
            "attribute_value_count": 1,
        }

        stale_response = client.delete(
            f"/api/v1/categories/{root['id']}",
            params={"expected_version": root["version"] + 1},
        )
        assert stale_response.status_code == 409
        assert stale_response.json()["detail"]["code"] == "CATEGORY_VERSION_CONFLICT"

        delete_response = client.delete(
            f"/api/v1/categories/{root['id']}",
            params={"expected_version": root["version"]},
        )
        assert delete_response.status_code == 200, delete_response.text
        assert delete_response.json()["deleted_category_count"] == 3
        assert delete_response.json()["unclassified_product_count"] == 2
        assert delete_response.json()["deleted_attribute_definition_count"] == 1
        assert delete_response.json()["detached_attribute_value_count"] == 1

        with SessionLocal() as session:
            assert not session.scalars(
                select(ProductCategoryRow).where(
                    ProductCategoryRow.id.in_([UUID(value) for value in category_ids])
                )
            ).all()
            products = list(
                session.scalars(
                    select(ProductRow)
                    .where(ProductRow.id.in_(product_ids))
                    .order_by(ProductRow.product_code)
                ).all()
            )
            assert len(products) == 3
            assert all(product.category_id is None for product in products)
            assert all(product.search_document_version == 0 for product in products)
            assert session.get(AttributeDefinitionRow, definition_id) is None
            attribute = session.get(ProductAttributeRow, attribute_id)
            assert attribute is not None
            assert attribute.attribute_definition_id is None
            assert attribute.value_text == "属性值继续保留"
            assert session.scalar(
                select(ProductAuditEventRow).where(
                    ProductAuditEventRow.entity_type == "CATEGORY",
                    ProductAuditEventRow.entity_id == root["id"],
                    ProductAuditEventRow.action == "category.deleted",
                )
            ) is not None
    finally:
        with SessionLocal() as session:
            session.execute(
                delete(ProductAuditEventRow).where(
                    ProductAuditEventRow.entity_type == "CATEGORY",
                    ProductAuditEventRow.entity_id.in_(category_ids),
                )
            )
            session.execute(
                delete(ProductAttributeRow).where(
                    ProductAttributeRow.id == attribute_id
                )
            )
            session.execute(
                delete(ProductRow).where(ProductRow.id.in_(product_ids))
            )
            session.execute(
                delete(AttributeDefinitionRow).where(
                    AttributeDefinitionRow.id == definition_id
                )
            )
            if category_ids:
                category_uuids = [UUID(value) for value in category_ids]
                session.execute(
                    delete(ProductCategoryRow).where(
                        ProductCategoryRow.id.in_(category_uuids),
                        ProductCategoryRow.parent_id.is_not(None),
                    )
                )
                session.flush()
                session.execute(
                    delete(ProductCategoryRow).where(
                        ProductCategoryRow.id.in_(category_uuids)
                    )
                )
            session.commit()


def test_category_api_enforces_two_levels_and_updates_human_paths() -> None:
    suffix = uuid4().hex[:10].upper()
    created_ids: list[str] = []
    try:
        root_response = client.post(
            "/api/v1/categories",
            json={
                "code": f"TEST-ROOT-{suffix}",
                "name": f"测试一级-{suffix}",
                "sort_order": 3,
                "display_color": "#287d6e",
            },
        )
        assert root_response.status_code == 201, root_response.text
        root = root_response.json()
        created_ids.append(root["id"])
        assert root["parent_id"] is None
        assert root["path"] == f"测试一级-{suffix}"
        assert root["display_color"] == "#287D6E"

        invalid_color_response = client.post(
            "/api/v1/categories",
            json={
                "code": f"TEST-INVALID-COLOR-{suffix}",
                "name": f"无效颜色-{suffix}",
                "sort_order": 4,
                "display_color": "green",
            },
        )
        assert invalid_color_response.status_code == 422

        child_response = client.post(
            "/api/v1/categories",
            json={
                "parent_id": root["id"],
                "code": f"TEST-CHILD-{suffix}",
                "name": "测试二级",
                "sort_order": 1,
            },
        )
        assert child_response.status_code == 201, child_response.text
        child = child_response.json()
        created_ids.append(child["id"])
        assert child["path"] == f"测试一级-{suffix}/测试二级"
        assert child["display_color"] is None

        third_response = client.post(
            "/api/v1/categories",
            json={
                "parent_id": child["id"],
                "code": f"TEST-THIRD-{suffix}",
                "name": "不允许的第三级",
                "sort_order": 0,
            },
        )
        assert third_response.status_code == 409
        assert third_response.json()["detail"]["code"] == "CATEGORY_DEPTH_EXCEEDED"

        renamed_response = client.patch(
            f"/api/v1/categories/{child['id']}",
            json={
                "expected_version": child["version"],
                "parent_id": root["id"],
                "name": "更新后的二级",
                "sort_order": 2,
                "status": "ACTIVE",
            },
        )
        assert renamed_response.status_code == 200, renamed_response.text
        assert (
            renamed_response.json()["path"]
            == f"测试一级-{suffix}/更新后的二级"
        )
        renamed_child = renamed_response.json()

        second_child_response = client.post(
            "/api/v1/categories",
            json={
                "parent_id": root["id"],
                "code": f"TEST-CHILD-2-{suffix}",
                "name": "测试二级乙",
                "sort_order": 3,
            },
        )
        assert second_child_response.status_code == 201, second_child_response.text
        second_child = second_child_response.json()
        created_ids.append(second_child["id"])

        third_child_response = client.post(
            "/api/v1/categories",
            json={
                "parent_id": root["id"],
                "code": f"TEST-CHILD-3-{suffix}",
                "name": "测试二级丙",
                "sort_order": 4,
            },
        )
        assert third_child_response.status_code == 201, third_child_response.text
        third_child = third_child_response.json()
        created_ids.append(third_child["id"])

        reordered_response = client.patch(
            "/api/v1/categories/reorder",
            json={
                "items": [
                    {
                        "id": third_child["id"],
                        "expected_version": third_child["version"],
                    },
                    {
                        "id": renamed_child["id"],
                        "expected_version": renamed_child["version"],
                    },
                    {
                        "id": second_child["id"],
                        "expected_version": second_child["version"],
                    },
                ]
            },
        )
        assert reordered_response.status_code == 200, reordered_response.text
        reordered = reordered_response.json()
        assert [row["name"] for row in reordered] == [
            "测试二级丙",
            "更新后的二级",
            "测试二级乙",
        ]
        assert [row["sort_order"] for row in reordered] == [0, 1, 2]

        recolored_response = client.patch(
            f"/api/v1/categories/{root['id']}",
            json={
                "expected_version": root["version"],
                "parent_id": None,
                "name": root["name"],
                "sort_order": root["sort_order"],
                "status": "ACTIVE",
                "display_color": "#a45f3e",
            },
        )
        assert recolored_response.status_code == 200, recolored_response.text
        assert recolored_response.json()["display_color"] == "#A45F3E"
    finally:
        if created_ids:
            category_uuids = [UUID(value) for value in created_ids]
            with SessionLocal() as session:
                session.execute(
                    delete(ProductAuditEventRow).where(
                        ProductAuditEventRow.entity_type == "CATEGORY",
                        ProductAuditEventRow.entity_id.in_(created_ids),
                    )
                )
                for category_id in reversed(category_uuids):
                    session.execute(
                        delete(ProductCategoryRow).where(
                            ProductCategoryRow.id == category_id
                        )
                    )
                session.commit()


def test_public_media_uses_tenant_scoped_relative_proxy_when_no_cdn_is_configured(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.delenv("PUBLIC_MEDIA_BASE_URL", raising=False)
    listing = client.get("/api/store/demo/skus", params={"q": "PF-8G01"})
    assert listing.status_code == 200
    image_url = listing.json()["items"][0]["image_url"]
    assert image_url.startswith("/api/store/demo/media/")
    media = client.get(image_url)
    assert media.status_code == 200
    assert media.headers["content-type"].startswith("image/")
    assert media.headers["x-content-type-options"] == "nosniff"
    assert client.get(f"/api/store/demo/media/{uuid4()}").status_code == 404


def test_merchant_publication_requires_active_sku_and_defaults_missing_price_to_zero() -> None:
    root_category_id = uuid4()
    category_id = uuid4()
    product_id = uuid4()
    sku_id = uuid4()
    sku_code = f"PUB-{uuid4().hex[:10].upper()}"
    now = datetime.now(UTC)
    with SessionLocal() as session:
        supplier = session.scalar(
            select(SupplierRow).where(
                SupplierRow.tenant_id == DEFAULT_TENANT_ID,
                SupplierRow.status == "ACTIVE",
            )
        )
        assert supplier is not None
        session.add(
            ProductCategoryRow(
                id=root_category_id,
                tenant_id=DEFAULT_TENANT_ID,
                code=f"PUB-CATEGORY-{uuid4().hex[:10].upper()}",
                name="公开测试分类",
                path="公开测试分类",
                display_color="#3F6F9C",
                status="ACTIVE",
            )
        )
        session.flush()
        session.add(
            ProductCategoryRow(
                id=category_id,
                tenant_id=DEFAULT_TENANT_ID,
                parent_id=root_category_id,
                code=f"PUB-SUBCATEGORY-{uuid4().hex[:10].upper()}",
                name="公开测试子分类",
                path="公开测试分类/公开测试子分类",
                status="ACTIVE",
            )
        )
        session.flush()
        session.add(
            ProductRow(
                id=product_id,
                tenant_id=DEFAULT_TENANT_ID,
                product_code=sku_code,
                name="Explicit Public Price Product",
                category_id=category_id,
                status="ACTIVE",
                default_unit="piece",
            )
        )
        session.flush()
        session.add(
            SkuRow(
                id=sku_id,
                tenant_id=DEFAULT_TENANT_ID,
                product_id=product_id,
                sku_code=sku_code,
                name="Explicit Public Price SKU",
                default_moq=Decimal("2"),
                moq_unit="piece",
                status="DRAFT",
            )
        )
        session.flush()
        source = SupplierProductRow(
            tenant_id=DEFAULT_TENANT_ID,
            supplier_id=supplier.id,
            product_id=product_id,
            sku_id=sku_id,
            supplier_sku=sku_code,
            supplier_product_name="Supplier Private Name",
            status="ACTIVE",
        )
        session.add(source)
        session.flush()
        session.add(
            SupplierPriceRow(
                tenant_id=DEFAULT_TENANT_ID,
                supplier_product_id=source.id,
                sku_id=sku_id,
                min_quantity=Decimal("2"),
                unit_price=Decimal("12.34"),
                currency="CNY",
                unit_code="piece",
                valid_from=now,
                status="CONFIRMED",
                confirmed_by_membership_id=DEFAULT_MEMBERSHIP_ID,
                confirmed_at=now,
            )
        )
        session.commit()

        with pytest.raises(ApplicationError) as denied:
            upsert_public_offer_use_case(
                session,
                tenant_id=DEFAULT_TENANT_ID,
                membership_id=DEFAULT_MEMBERSHIP_ID,
                permissions=frozenset({"product.edit"}),
                sku_id=sku_id,
                request=PublicCatalogOfferUpsertRequest(
                    unit_price=Decimal("88.00"),
                    currency="CNY",
                    tags=["公开标签"],
                    publication_status="PUBLISHED",
                ),
            )
        assert denied.value.code == "PERMISSION_REQUIRED"

    missing_price = client.put(
        f"/api/v1/skus/{sku_id}/public-offer",
        json={"currency": "CNY", "tags": ["公开标签"], "publication_status": "PUBLISHED"},
    )
    assert missing_price.status_code == 409
    assert missing_price.json()["detail"]["code"] == "PUBLIC_SKU_NOT_ACTIVE"

    invalid_tag_color = client.put(
        f"/api/v1/skus/{sku_id}/public-offer",
        json={
            "unit_price": "88.00",
            "currency": "CNY",
            "tags": ["公开标签"],
            "tag_color": "orange",
            "publication_status": "DRAFT",
        },
    )
    assert invalid_tag_color.status_code == 422

    invalid_display_tag = client.put(
        f"/api/v1/skus/{sku_id}/public-offer",
        json={
            "unit_price": "88.00",
            "currency": "CNY",
            "tags": ["公开标签"],
            "display_tag": "不存在的标签",
            "publication_status": "DRAFT",
        },
    )
    assert invalid_display_tag.status_code == 422

    inactive = client.put(
        f"/api/v1/skus/{sku_id}/public-offer",
        json={
            "unit_price": "88.00",
            "currency": "cny",
            "tags": ["公开标签", "公开标签"],
            "publication_status": "PUBLISHED",
        },
    )
    assert inactive.status_code == 409
    assert inactive.json()["detail"]["code"] == "PUBLIC_SKU_NOT_ACTIVE"

    activated = client.patch(
        f"/api/v1/skus/{sku_id}",
        json={"expected_version": 1, "status": "ACTIVE"},
    )
    assert activated.status_code == 200, activated.text

    zero_price = client.put(
        f"/api/v1/skus/{sku_id}/public-offer",
        json={
            "currency": "CNY",
            "tags": ["公开标签"],
            "publication_status": "PUBLISHED",
        },
    )
    assert zero_price.status_code == 200, zero_price.text
    assert Decimal(str(zero_price.json()["unit_price"])) == Decimal("0")
    assert zero_price.json()["publication_status"] == "PUBLISHED"

    published = client.put(
        f"/api/v1/skus/{sku_id}/public-offer",
        json={
            "unit_price": "88.00",
            "currency": "cny",
            "tags": ["公开标签", "新品"],
            "display_tag": "新品",
            "tag_color": "#b65a3a",
            "publication_status": "PUBLISHED",
        },
    )
    assert published.status_code == 200, published.text
    assert Decimal(str(published.json()["unit_price"])) == Decimal("88.00")
    assert published.json()["currency"] == "CNY"
    assert published.json()["display_tag"] == "新品"
    assert published.json()["tag_color"] == "#B65A3A"
    assert published.json()["publication_status"] == "PUBLISHED"

    offers = client.get(f"/api/v1/products/{product_id}/public-offers")
    assert offers.status_code == 200
    assert len(offers.json()) == 1
    assert offers.json()[0]["display_tag"] == "新品"
    assert offers.json()[0]["tag_color"] == "#B65A3A"
    assert "supplier_product_id" not in offers.json()[0]
    assert "unit_cost" not in offers.json()[0]

    public_listing = client.get("/api/store/demo/skus", params={"q": sku_code})
    assert public_listing.status_code == 200
    public_item = public_listing.json()["items"][0]
    assert Decimal(str(public_item["price"])) == Decimal("88.00")
    assert public_item["display_tag"] == "新品"
    assert public_item["tag_color"] == "#B65A3A"
    assert public_item["category_color"] == "#3F6F9C"
    assert Decimal(str(public_item["price"])) != Decimal("12.34")
    assert "supplier_product_id" not in public_item

    suspended = client.put(
        f"/api/v1/skus/{sku_id}/public-offer",
        json={
            "unit_price": "88.00",
            "currency": "CNY",
            "tags": ["公开标签", "新品"],
            "display_tag": "新品",
            "tag_color": "#B65A3A",
            "publication_status": "SUSPENDED",
        },
    )
    assert suspended.status_code == 200
    assert client.get("/api/store/demo/skus", params={"q": sku_code}).json()["total"] == 0

    with SessionLocal() as session:
        audit = session.scalar(
            select(ProductAuditEventRow)
            .where(
                ProductAuditEventRow.tenant_id == DEFAULT_TENANT_ID,
                ProductAuditEventRow.product_id == product_id,
                ProductAuditEventRow.action == "public_offer.published",
            )
            .order_by(ProductAuditEventRow.occurred_at.desc())
        )
        assert audit is not None
        assert audit.after["unit_price"] == "88.00"


def test_supplier_catalog_import_requires_a_tenant_scoped_supplier() -> None:
    response = client.post(
        "/api/v1/imports",
        files={"file": ("unbound.csv", b"name,sku,moq\nUnbound,UB-1,10\n", "text/csv")},
        data={"source_type": "SUPPLIER_CATALOG"},
    )
    assert response.status_code == 422
    assert response.json()["detail"]["code"] == "SUPPLIER_REQUIRED"


def test_supplier_manage_creates_unique_tenant_supplier_before_import() -> None:
    suffix = uuid4().hex[:10].upper()
    supplier_code = f"NEW-{suffix}"
    supplier_name = f"New Tenant Supplier {suffix}"
    request = SupplierCreateRequest(
        supplier_code=supplier_code.lower(),
        name=supplier_name,
        category="家居用品",
        country_code="cn",
        website="https://supplier.example.test",
    )
    with SessionLocal() as session:
        with pytest.raises(ApplicationError) as denied:
            create_supplier_use_case(
                session,
                tenant_id=DEFAULT_TENANT_ID,
                permissions=frozenset({"supplier.view"}),
                request=request,
            )
        assert denied.value.code == "PERMISSION_DENIED"

    created = client.post(
        "/api/v1/supplier-profiles", json=request.model_dump(mode="json")
    )
    assert created.status_code == 201, created.text
    payload = created.json()
    assert payload["supplier_code"] == supplier_code
    assert payload["name"] == supplier_name
    assert payload["country_code"] == "CN"
    assert payload["active_products"] == 0
    assert payload["active_skus"] == 0

    duplicate_code = client.post(
        "/api/v1/supplier-profiles",
        json={**request.model_dump(mode="json"), "name": f"Different {supplier_name}"},
    )
    assert duplicate_code.status_code == 409
    assert duplicate_code.json()["detail"]["code"] == "SUPPLIER_CODE_CONFLICT"

    duplicate_name = client.post(
        "/api/v1/supplier-profiles",
        json={**request.model_dump(mode="json"), "supplier_code": f"ALT-{suffix}"},
    )
    assert duplicate_name.status_code == 409
    assert duplicate_name.json()["detail"]["code"] == "SUPPLIER_NAME_CONFLICT"

    other_organization_id = uuid4()
    other_tenant_id = uuid4()
    with SessionLocal() as session:
        session.add(
            OrganizationRow(
                id=other_organization_id,
                code=f"SUP-{suffix}",
                name=f"Supplier Test Organization {suffix}",
            )
        )
        session.flush()
        session.add(
            TenantRow(
                id=other_tenant_id,
                organization_id=other_organization_id,
                slug=f"supplier-{suffix.casefold()}",
                name=f"Supplier Test Tenant {suffix}",
            )
        )
        session.commit()
        other = create_supplier_use_case(
            session,
            tenant_id=other_tenant_id,
            permissions=frozenset({"supplier.manage"}),
            request=request,
        )
        assert other.supplier_code == supplier_code
        assert other.id != payload["id"]

    current_directory = client.get("/api/v1/supplier-profiles")
    assert current_directory.status_code == 200
    matching = [
        row for row in current_directory.json() if row["supplier_code"] == supplier_code
    ]
    assert [row["id"] for row in matching] == [payload["id"]]


def test_public_quote_draft_snapshot_hashed_expiring_downloads_and_formula_safety() -> None:
    listing = client.get("/api/store/demo/skus", params={"q": "PF-8G01"})
    assert listing.status_code == 200
    sku_data = listing.json()["items"][0]
    original_name = sku_data["name"]
    original_price = Decimal(str(sku_data["price"]))

    rejected_without_privacy_acknowledgment = client.post(
        "/api/store/demo/quotes",
        json={
            "customer_name": "No Privacy Acknowledgment",
            "privacy_acknowledged": False,
            "items": [{"sku_id": sku_data["id"], "quantity": 1}],
        },
    )
    assert rejected_without_privacy_acknowledgment.status_code == 422

    with SessionLocal() as session:
        sku = session.get(SkuRow, UUID(sku_data["id"]))
        assert sku is not None
        original_moq = sku.default_moq
        sku.default_moq = Decimal("500")
        session.commit()
    try:
        create_response = client.post(
            "/api/store/demo/quotes",
            json={
                "customer_name": "=2+2",
                "customer_company": "+Formula Company",
                "customer_email": "buyer@example.test",
                "customer_phone": "@PHONE",
                "notes": "@SUM(A1:A2)",
                "privacy_acknowledged": True,
                "items": [{"sku_id": sku_data["id"], "quantity": 2}],
            },
        )
    finally:
        with SessionLocal() as session:
            sku = session.get(SkuRow, UUID(sku_data["id"]))
            assert sku is not None
            sku.default_moq = original_moq
            session.commit()
    assert create_response.status_code == 201, create_response.text
    assert create_response.headers["cache-control"] == "no-store"
    assert create_response.headers["pragma"] == "no-cache"
    draft = create_response.json()
    assert draft["status"] == "PENDING_CONFIRMATION"
    assert draft["quote_number"].startswith("QD-")
    assert "不构成" in draft["disclaimer"]
    assert Decimal(str(draft["total"])) == original_price * 2
    raw_token = draft["download_token"]
    assert raw_token and raw_token.startswith(f"{DEFAULT_TENANT_ID}.")
    assert "token=" not in draft["pdf_url"]
    assert "token=" not in draft["xlsx_url"]
    assert "minimum_order_quantity" not in draft["items"][0]
    quote_id = UUID(draft["id"])
    download_headers = {"X-Quote-Download-Token": raw_token}

    with SessionLocal() as session:
        stored = session.scalar(
            select(PublicQuoteDownloadTokenRow).where(
                PublicQuoteDownloadTokenRow.quote_draft_id == quote_id
            )
        )
        snapshot_item = session.scalar(
            select(PublicQuoteDraftItemRow).where(
                PublicQuoteDraftItemRow.quote_draft_id == quote_id
            )
        )
        stored_draft = session.get(PublicQuoteDraftRow, quote_id)
        assert stored is not None and snapshot_item is not None and stored_draft is not None
        assert stored.token_hash == hash_secret(raw_token)
        assert stored.token_hash != raw_token
        assert len(stored.token_hash) == 64
        assert snapshot_item.name_snapshot == original_name
        assert snapshot_item.unit_price_snapshot == original_price
        assert stored_draft.snapshot["status"] == "PENDING_CONFIRMATION"
        assert stored_draft.snapshot["privacy_notice"]["acknowledged"] is True
        assert stored_draft.snapshot["privacy_notice"]["version"] == "privacy-v1"

        sku = session.get(SkuRow, UUID(sku_data["id"]))
        offer = session.scalar(
            select(PublicCatalogOfferRow).where(
                PublicCatalogOfferRow.sku_id == UUID(sku_data["id"])
            )
        )
        assert sku is not None and offer is not None
        sku.name = "MUTATED AFTER SUBMISSION"
        offer.unit_price = Decimal("9999.00")
        session.commit()
    try:
        pdf_response = client.get(draft["pdf_url"], headers=download_headers)
        assert pdf_response.status_code == 200, pdf_response.text
        assert pdf_response.content.startswith(b"%PDF")
        assert "x-quote-status" not in pdf_response.headers
        assert pdf_response.headers["cache-control"] == "private, no-store"
        assert pdf_response.headers["pragma"] == "no-cache"
        assert f'{draft["quote_number"]}.pdf' in pdf_response.headers["content-disposition"]
        assert "PENDING-CONFIRMATION" not in pdf_response.headers["content-disposition"]

        xlsx_response = client.get(draft["xlsx_url"], headers=download_headers)
        assert xlsx_response.status_code == 200, xlsx_response.text
        workbook = load_workbook(BytesIO(xlsx_response.content), data_only=False)
        sheet = workbook.active
        assert sheet["A1"].value == "报价单 / QUOTATION"
        assert sheet["B3"].value == "'=2+2"
        assert all(
            "待人工确认" not in str(cell.value or "")
            for row in sheet.iter_rows()
            for cell in row
        )
        header_row = next(
            row_index
            for row_index in range(1, sheet.max_row + 1)
            if sheet.cell(row_index, 1).value == "序号"
        )
        assert [sheet.cell(header_row, column).value for column in range(1, 15)] == [
            "序号",
            "图片",
            "SKU",
            "商品名称",
            "数量",
            "单位",
            "装箱数量",
            "装箱尺寸",
            "毛重(kg)",
            "立方(m³)",
            "单价",
            "总价",
            "总立方(m³)",
            "总毛重(kg)",
        ]
        assert sheet.cell(header_row + 1, 4).value == original_name
        assert Decimal(str(sheet.cell(header_row + 1, 11).value)) == original_price
        assert all(cell.data_type != "f" for row in sheet.iter_rows() for cell in row)

        merchant_list = client.get("/api/v1/public-quote-drafts")
        assert merchant_list.status_code == 200
        assert merchant_list.headers["cache-control"] == "no-store"
        assert merchant_list.headers["pragma"] == "no-cache"
        assert str(quote_id) in {item["id"] for item in merchant_list.json()}
        merchant_detail = client.get(f"/api/v1/public-quote-drafts/{quote_id}")
        assert merchant_detail.status_code == 200
        assert merchant_detail.headers["cache-control"] == "no-store"
        assert merchant_detail.headers["pragma"] == "no-cache"
        assert merchant_detail.json()["download_token"] is None
        assert merchant_detail.json()["pdf_url"] is None
        assert merchant_detail.json()["items"][0]["name_snapshot"] == original_name
        merchant_pdf = client.get(
            f"/api/v1/public-quote-drafts/{quote_id}/pdf"
        )
        assert merchant_pdf.status_code == 200
        assert merchant_pdf.content.startswith(b"%PDF")
        assert merchant_pdf.headers["cache-control"] == "private, no-store"
        assert merchant_pdf.headers["pragma"] == "no-cache"
        merchant_xlsx = client.get(
            f"/api/v1/public-quote-drafts/{quote_id}/xlsx"
        )
        assert merchant_xlsx.status_code == 200
        assert merchant_xlsx.content.startswith(b"PK")
        assert merchant_xlsx.headers["cache-control"] == "private, no-store"
        assert merchant_xlsx.headers["pragma"] == "no-cache"
    finally:
        with SessionLocal() as session:
            sku = session.get(SkuRow, UUID(sku_data["id"]))
            offer = session.scalar(
                select(PublicCatalogOfferRow).where(
                    PublicCatalogOfferRow.sku_id == UUID(sku_data["id"])
                )
            )
            assert sku is not None and offer is not None
            sku.name = original_name
            offer.unit_price = original_price
            session.commit()

    forged_tenant_token = f"{uuid4()}.{raw_token.split('.', 1)[1]}"
    assert client.get(
        f"/api/quotes/{quote_id}/pdf",
        headers={"X-Quote-Download-Token": forged_tenant_token},
    ).status_code == 404
    query_only = client.get(
        f"/api/quotes/{quote_id}/pdf", params={"token": raw_token}
    )
    assert query_only.status_code == 422

    with SessionLocal() as session:
        stored = session.scalar(
            select(PublicQuoteDownloadTokenRow).where(
                PublicQuoteDownloadTokenRow.quote_draft_id == quote_id
            )
        )
        assert stored is not None
        stored.expires_at = datetime.now(UTC) - timedelta(seconds=1)
        session.commit()
    expired = client.get(
        f"/api/quotes/{quote_id}/pdf",
        headers={"X-Quote-Download-Token": raw_token},
    )
    assert expired.status_code == 410
    assert expired.json()["detail"]["code"] == "DOWNLOAD_EXPIRED"


def test_custom_quote_excel_template_upload_mapping_and_rendering() -> None:
    system_template = client.get(
        "/api/v1/quote-excel-templates/system-default.xlsx"
    )
    assert system_template.status_code == 200, system_template.text
    assert system_template.headers["cache-control"] == "no-store"
    default_workbook = load_workbook(BytesIO(system_template.content))
    default_sheet = default_workbook["报价单"]
    default_header_row = next(
        row_index
        for row_index in range(1, default_sheet.max_row + 1)
        if default_sheet.cell(row_index, 1).value == "序号"
    )
    assert default_sheet.cell(default_header_row, 2).value == "图片"
    assert default_sheet.cell(default_header_row, 14).value == "总毛重(kg)"
    default_workbook.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商家报价单"
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "报价单 {{quote_number}}"
    sheet["A1"].font = Font(size=18, bold=True, color="D4AF37")
    sheet["A1"].fill = PatternFill("solid", fgColor="2D1B69")
    sheet["A2"] = "客户"
    sheet["B2"] = "{{customer_name}}"
    sheet["D2"] = "{{quote_date}}"
    headers = [
        "SKU代码",
        "品名",
        "规格",
        "数量",
        "单位",
        "单价",
        "金额",
        "客户自定义列",
    ]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(4, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D1B69")
    sample = [
        "SAMPLE-001",
        "示例商品",
        "红色 / M",
        1,
        "件",
        1.5,
        1.5,
        "商家示例值",
    ]
    for column, value in enumerate(sample, 1):
        cell = sheet.cell(5, column, value)
        cell.fill = PatternFill("solid", fgColor="F4EFFA")
    sheet["A6"] = "合计"
    sheet["G6"] = "=SUM(G5:G5)"
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    upload = client.post(
        "/api/v1/quote-excel-templates",
        files={
            "file": (
                "商家自定义报价单.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload.status_code == 201, upload.text
    template = upload.json()
    template_id = template["id"]
    assert template["sheet_name"] == "商家报价单"
    assert template["header_row"] == 4
    assert template["data_start_row"] == 5
    assert {column["header"] for column in template["columns"]} == set(headers)

    mappings = {
        "A": "sku_code",
        "B": "product_name",
        "C": "specification",
        "D": "quantity",
        "E": "unit_code",
        "F": "unit_price",
        "G": "line_total",
    }
    configured = client.put(
        f"/api/v1/quote-excel-templates/{template_id}",
        json={
            "name": "外贸标准报价单",
            "column_mappings": mappings,
            "is_default": True,
        },
    )
    assert configured.status_code == 200, configured.text
    assert configured.json()["is_default"] is True

    try:
        listing = client.get("/api/store/demo/skus", params={"limit": 2})
        assert listing.status_code == 200, listing.text
        skus = listing.json()["items"][:2]
        assert len(skus) == 2
        created = client.post(
            "/api/store/demo/quotes",
            json={
                "customer_name": "=Formula Customer",
                "privacy_acknowledged": True,
                "items": [
                    {"sku_id": skus[0]["id"], "quantity": 2},
                    {"sku_id": skus[1]["id"], "quantity": 3},
                ],
            },
        )
        assert created.status_code == 201, created.text
        draft = created.json()
        downloaded = client.get(
            draft["xlsx_url"],
            headers={"X-Quote-Download-Token": draft["download_token"]},
        )
        assert downloaded.status_code == 200, downloaded.text
        rendered = load_workbook(BytesIO(downloaded.content), data_only=False)
        rendered_sheet = rendered["商家报价单"]
        assert rendered_sheet["A1"].value == f"报价单 {draft['quote_number']}"
        assert rendered_sheet["B2"].value == "'=Formula Customer"
        assert rendered_sheet["A5"].value == skus[0]["sku_code"]
        assert rendered_sheet["A6"].value == skus[1]["sku_code"]
        assert rendered_sheet["B5"].value == skus[0]["name"]
        assert rendered_sheet["D5"].value == 2
        assert rendered_sheet["D6"].value == 3
        assert rendered_sheet["G7"].value == "=SUM(G5:G6)"
        assert rendered_sheet["H5"].value is None
        assert rendered_sheet["H6"].value is None
        assert rendered_sheet["A5"].fill.fgColor.rgb.endswith("F4EFFA")
        assert rendered_sheet["A6"].fill.fgColor.rgb.endswith("F4EFFA")
        assert all(
            "待人工确认" not in str(cell.value or "")
            for row in rendered_sheet.iter_rows()
            for cell in row
        )
        rendered.close()
    finally:
        deleted = client.delete(f"/api/v1/quote-excel-templates/{template_id}")
        assert deleted.status_code == 204, deleted.text


def test_public_quote_drafts_are_tenant_scoped_for_public_and_authenticated_reads() -> None:
    organization_id = uuid4()
    tenant_b = uuid4()
    product_b = uuid4()
    sku_b = uuid4()
    draft_b = uuid4()
    now = datetime.now(UTC)
    with SessionLocal() as session:
        session.add(
            OrganizationRow(
                id=organization_id,
                code=f"PUBLIC-{organization_id.hex[:8]}",
                name="Public Draft Tenant B Organization",
            )
        )
        session.flush()
        session.add(
            TenantRow(
                id=tenant_b,
                organization_id=organization_id,
                slug=f"public-{tenant_b.hex[:8]}",
                name="Public Draft Tenant B",
            )
        )
        session.flush()
        session.add(
            ProductRow(
                id=product_b,
                tenant_id=tenant_b,
                product_code=f"PB-{product_b.hex[:8]}",
                name="Tenant B Private Product",
                status="ACTIVE",
            )
        )
        session.flush()
        session.add(
            SkuRow(
                id=sku_b,
                tenant_id=tenant_b,
                product_id=product_b,
                sku_code=f"SB-{sku_b.hex[:8]}",
                name="Tenant B Private SKU",
                default_moq=Decimal("1"),
                moq_unit="piece",
                status="ACTIVE",
            )
        )
        session.add(
            PublicQuoteDraftRow(
                id=draft_b,
                tenant_id=tenant_b,
                request_number=f"QD-B-{draft_b.hex[:8]}",
                status="PENDING_CONFIRMATION",
                customer_name="Tenant B Customer",
                currency="CNY",
                subtotal_amount=Decimal("10.00"),
                estimated_total=Decimal("10.00"),
                expires_at=now + timedelta(days=7),
                snapshot={"status": "PENDING_CONFIRMATION", "items": []},
                content_hash="b" * 64,
                disclaimer_version="public-draft-v1",
            )
        )
        session.commit()

    cross_tenant_cart = client.post(
        "/api/store/demo/quotes",
        json={
            "customer_name": "Cross Tenant Attempt",
            "privacy_acknowledged": True,
            "items": [{"sku_id": str(sku_b), "quantity": 1}],
        },
    )
    assert cross_tenant_cart.status_code == 422
    assert cross_tenant_cart.json()["detail"]["code"] == "PUBLIC_SKU_NOT_FOUND"

    merchant_list = client.get("/api/v1/public-quote-drafts")
    assert merchant_list.status_code == 200
    assert str(draft_b) not in {item["id"] for item in merchant_list.json()}
    merchant_detail = client.get(f"/api/v1/public-quote-drafts/{draft_b}")
    assert merchant_detail.status_code == 404
    merchant_pdf = client.get(f"/api/v1/public-quote-drafts/{draft_b}/pdf")
    assert merchant_pdf.status_code == 404


def test_public_catalog_migration_is_reversible_on_sqlite(tmp_path: Path) -> None:
    database_path = tmp_path / "public-catalog-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)
    public_tables = {
        "tenant_public_profiles",
        "public_catalog_offers",
        "public_quote_drafts",
        "public_quote_draft_items",
        "public_quote_download_tokens",
        "quote_excel_templates",
        "storefront_announcements",
        "catalog_delete_jobs",
        "storefront_chat_conversations",
        "storefront_chat_messages",
    }
    customer_account_tables = {
        "local_account_credentials",
        "customer_account_access_events",
    }
    analytics_tables = {
        "storefront_product_view_events",
        "storefront_product_view_daily",
    }

    command.upgrade(config, "20260718_0019")
    before_engine = create_engine(migration_url)
    assert public_tables.isdisjoint(inspect(before_engine).get_table_names())
    assert customer_account_tables.isdisjoint(inspect(before_engine).get_table_names())
    assert analytics_tables.isdisjoint(inspect(before_engine).get_table_names())
    before_engine.dispose()

    command.upgrade(config, "head")
    upgraded_engine = create_engine(migration_url)
    assert public_tables.issubset(inspect(upgraded_engine).get_table_names())
    assert customer_account_tables.issubset(inspect(upgraded_engine).get_table_names())
    assert analytics_tables.issubset(inspect(upgraded_engine).get_table_names())
    assert "submitted_by_membership_id" in {
        column["name"]
        for column in inspect(upgraded_engine).get_columns("public_quote_drafts")
    }
    assert "latest_import_job_id" in {
        column["name"]
        for column in inspect(upgraded_engine).get_columns("skus")
    }
    assert {"specification_snapshot", "option_values_snapshot"}.issubset({
        column["name"]
        for column in inspect(upgraded_engine).get_columns("public_quote_draft_items")
    })
    announcement_columns = {
        column["name"]: column
        for column in inspect(upgraded_engine).get_columns("storefront_announcements")
    }
    assert "related_sku_ids" in announcement_columns
    assert "ticker_speed_px_per_second" in announcement_columns
    assert "repeat_interval_hours" not in announcement_columns
    assert announcement_columns["title"]["nullable"] is True
    profile_columns = {
        column["name"]: column
        for column in inspect(upgraded_engine).get_columns(
            "tenant_public_profiles"
        )
    }
    assert "storefront_locales" in profile_columns
    assert profile_columns["storefront_locales"]["nullable"] is False
    assert "hot_products_enabled" in profile_columns
    assert profile_columns["hot_products_enabled"]["nullable"] is False
    assert "support_widget_config" in profile_columns
    assert profile_columns["support_widget_config"]["nullable"] is False
    product_columns = {
        column["name"]: column
        for column in inspect(upgraded_engine).get_columns("products")
    }
    assert "storefront_pinned_at" in product_columns
    assert product_columns["storefront_pinned_at"]["nullable"] is True
    delete_job_columns = {
        column["name"]
        for column in inspect(upgraded_engine).get_columns("catalog_delete_jobs")
    }
    assert {
        "status",
        "stage",
        "progress",
        "total_products",
        "total_skus",
        "deleted_product_count",
        "deleted_sku_count",
        "error_message",
    }.issubset(delete_job_columns)
    translation_job_columns = {
        column["name"]
        for column in inspect(upgraded_engine).get_columns(
            "catalog_translation_jobs"
        )
    }
    assert {
        "stage",
        "package_version",
        "package_published",
        "package_byte_size",
        "source_cutoff_at",
        "remaining_sku_ids",
        "pause_requested_at",
        "paused_at",
    }.issubset(translation_job_columns)
    translation_settings_columns = {
        column["name"]
        for column in inspect(upgraded_engine).get_columns(
            "translation_provider_settings"
        )
    }
    assert "requests_per_minute" in translation_settings_columns
    with upgraded_engine.connect() as connection:
        assert connection.exec_driver_sql(
            "SELECT version_num FROM alembic_version"
        ).scalar() == "20260809_0059"
    upgraded_engine.dispose()
    command.check(config)

    command.downgrade(config, "20260718_0019")
    downgraded_engine = create_engine(migration_url)
    assert public_tables.isdisjoint(inspect(downgraded_engine).get_table_names())
    assert customer_account_tables.isdisjoint(inspect(downgraded_engine).get_table_names())
    assert analytics_tables.isdisjoint(inspect(downgraded_engine).get_table_names())
    downgraded_engine.dispose()
    command.upgrade(config, "head")
    command.check(config)


def test_catalog_tag_color_migration_is_reversible_on_sqlite(tmp_path: Path) -> None:
    database_path = tmp_path / "catalog-tag-color-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)

    command.upgrade(config, "20260726_0030")
    before_engine = create_engine(migration_url)
    assert "tag_color" not in {
        column["name"]
        for column in inspect(before_engine).get_columns("public_catalog_offers")
    }
    before_engine.dispose()

    command.upgrade(config, "head")
    upgraded_engine = create_engine(migration_url)
    assert "tag_color" in {
        column["name"]
        for column in inspect(upgraded_engine).get_columns("public_catalog_offers")
    }
    upgraded_engine.dispose()

    command.downgrade(config, "20260726_0030")
    downgraded_engine = create_engine(migration_url)
    assert "tag_color" not in {
        column["name"]
        for column in inspect(downgraded_engine).get_columns("public_catalog_offers")
    }
    downgraded_engine.dispose()


def test_catalog_display_tag_migration_is_reversible_on_sqlite(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "catalog-display-tag-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)

    command.upgrade(config, "20260726_0034")
    before_engine = create_engine(migration_url)
    assert "display_tag" not in {
        column["name"]
        for column in inspect(before_engine).get_columns("public_catalog_offers")
    }
    before_engine.dispose()

    command.upgrade(config, "head")
    upgraded_engine = create_engine(migration_url)
    assert "display_tag" in {
        column["name"]
        for column in inspect(upgraded_engine).get_columns("public_catalog_offers")
    }
    upgraded_engine.dispose()

    command.downgrade(config, "20260726_0034")
    downgraded_engine = create_engine(migration_url)
    assert "display_tag" not in {
        column["name"]
        for column in inspect(downgraded_engine).get_columns("public_catalog_offers")
    }
    downgraded_engine.dispose()


def test_category_display_color_migration_is_reversible_on_sqlite(tmp_path: Path) -> None:
    database_path = tmp_path / "category-display-color-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)

    command.upgrade(config, "20260726_0031")
    before_engine = create_engine(migration_url)
    assert "display_color" not in {
        column["name"]
        for column in inspect(before_engine).get_columns("product_categories")
    }
    before_engine.dispose()

    command.upgrade(config, "head")
    upgraded_engine = create_engine(migration_url)
    assert "display_color" in {
        column["name"]
        for column in inspect(upgraded_engine).get_columns("product_categories")
    }
    upgraded_engine.dispose()

    command.downgrade(config, "20260726_0031")
    downgraded_engine = create_engine(migration_url)
    assert "display_color" not in {
        column["name"]
        for column in inspect(downgraded_engine).get_columns("product_categories")
    }
    downgraded_engine.dispose()


def test_storefront_category_layout_migration_is_reversible_on_sqlite(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "storefront-category-layout-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)

    command.upgrade(config, "20260728_0038")
    before_engine = create_engine(migration_url)
    assert "all_products_position" not in {
        column["name"]
        for column in inspect(before_engine).get_columns("tenant_public_profiles")
    }
    before_engine.dispose()

    command.upgrade(config, "head")
    upgraded_engine = create_engine(migration_url)
    columns = {
        column["name"]: column
        for column in inspect(upgraded_engine).get_columns("tenant_public_profiles")
    }
    assert "all_products_position" in columns
    with upgraded_engine.connect() as connection:
        assert connection.exec_driver_sql(
            "SELECT COUNT(*) FROM tenant_public_profiles "
            "WHERE all_products_position != 0"
        ).scalar() == 0
    upgraded_engine.dispose()

    command.downgrade(config, "20260728_0038")
    downgraded_engine = create_engine(migration_url)
    assert "all_products_position" not in {
        column["name"]
        for column in inspect(downgraded_engine).get_columns(
            "tenant_public_profiles"
        )
    }
    downgraded_engine.dispose()


def test_catalog_translation_migration_is_reversible_on_sqlite(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "catalog-translation-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)

    command.upgrade(config, "20260729_0039")
    before_engine = create_engine(migration_url)
    assert {
        "catalog_sku_translations",
        "catalog_text_translations",
        "catalog_translation_jobs",
    }.isdisjoint(inspect(before_engine).get_table_names())
    before_engine.dispose()

    command.upgrade(config, "head")
    upgraded_engine = create_engine(migration_url)
    tables = set(inspect(upgraded_engine).get_table_names())
    assert {
        "catalog_sku_translations",
        "catalog_text_translations",
        "catalog_translation_jobs",
    }.issubset(tables)
    translation_columns = {
        column["name"]
        for column in inspect(upgraded_engine).get_columns(
            "catalog_sku_translations"
        )
    }
    assert {
        "source_hash",
        "source_category",
        "name",
        "description",
        "category",
        "tags",
        "display_tag",
    }.issubset(translation_columns)
    memory_columns = {
        column["name"]
        for column in inspect(upgraded_engine).get_columns(
            "catalog_text_translations"
        )
    }
    assert {
        "source_locale",
        "target_locale",
        "source_hash",
        "source_text",
        "translated_text",
        "provider",
        "provider_version",
        "last_accessed_at",
    }.issubset(memory_columns)
    upgraded_engine.dispose()

    command.downgrade(config, "20260729_0039")
    downgraded_engine = create_engine(migration_url)
    assert {
        "catalog_sku_translations",
        "catalog_text_translations",
        "catalog_translation_jobs",
    }.isdisjoint(inspect(downgraded_engine).get_table_names())
    downgraded_engine.dispose()


def test_product_tags_migration_is_reversible_and_uuid_compatible_on_sqlite(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "product-tags-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)

    command.upgrade(config, "20260726_0032")
    before_engine = create_engine(migration_url)
    assert "product_tags" not in inspect(before_engine).get_table_names()
    before_engine.dispose()

    command.upgrade(config, "head")
    upgraded_engine = create_engine(migration_url)
    assert "product_tags" in inspect(upgraded_engine).get_table_names()
    columns = {
        column["name"]: str(column["type"]).upper()
        for column in inspect(upgraded_engine).get_columns("product_tags")
    }
    assert columns["id"] == "CHAR(32)"
    assert columns["tenant_id"] == "CHAR(32)"

    organization_id = uuid4().hex
    tenant_id = uuid4().hex
    tag_id = uuid4().hex
    with upgraded_engine.begin() as connection:
        connection.exec_driver_sql("PRAGMA foreign_keys=ON")
        connection.exec_driver_sql(
            """
            INSERT INTO organizations
                (id, code, name, status, created_at, updated_at)
            VALUES (?, ?, ?, 'active', CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """,
            (organization_id, f"TAG-{organization_id[:8]}", "Tag migration organization"),
        )
        connection.exec_driver_sql(
            """
            INSERT INTO tenants
                (id, organization_id, slug, name, default_locale,
                 default_currency, timezone, status, created_at, updated_at)
            VALUES (?, ?, ?, ?, 'zh-CN', 'CNY', 'Asia/Shanghai',
                    'active', CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """,
            (
                tenant_id,
                organization_id,
                f"tag-{tenant_id[:8]}",
                "Tag migration tenant",
            ),
        )
        connection.exec_driver_sql(
            """
            INSERT INTO product_tags
                (id, tenant_id, name, normalized_name, usage_count,
                 created_at, updated_at)
            VALUES (?, ?, '防水', '防水', 0, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """,
            (tag_id, tenant_id),
        )
        assert connection.exec_driver_sql(
            "SELECT COUNT(*) FROM product_tags WHERE tenant_id = ?",
            (tenant_id,),
        ).scalar_one() == 1
    upgraded_engine.dispose()

    command.downgrade(config, "20260726_0032")
    downgraded_engine = create_engine(migration_url)
    assert "product_tags" not in inspect(downgraded_engine).get_table_names()
    downgraded_engine.dispose()


def test_embedding_management_migration_is_reversible_on_sqlite(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "embedding-management-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)
    tables = {"embedding_provider_settings", "knowledge_index_jobs"}

    command.upgrade(config, "20260726_0033")
    before_engine = create_engine(migration_url)
    assert tables.isdisjoint(inspect(before_engine).get_table_names())
    before_engine.dispose()

    command.upgrade(config, "head")
    upgraded_engine = create_engine(migration_url)
    assert tables.issubset(inspect(upgraded_engine).get_table_names())
    job_columns = {
        column["name"]
        for column in inspect(upgraded_engine).get_columns("knowledge_index_jobs")
    }
    assert {
        "total_products",
        "processed_products",
        "failed_products",
        "current_product_name",
        "error_message",
    }.issubset(job_columns)
    upgraded_engine.dispose()

    command.downgrade(config, "20260726_0033")
    downgraded_engine = create_engine(migration_url)
    assert tables.isdisjoint(inspect(downgraded_engine).get_table_names())
    downgraded_engine.dispose()


def test_product_tags_api_supports_crud() -> None:
    tag_name = f"Tag-{uuid4().hex[:10]}"
    created = client.post(
        "/api/tags",
        json={
            "name": tag_name,
            "description": "Temporary tag description",
            "category": "特性",
        },
    )
    assert created.status_code == 201
    created_tag = created.json()
    assert created_tag["name"] == tag_name
    assert created_tag["description"] == "Temporary tag description"
    assert created_tag["category"] == "特性"

    listed = client.get("/api/tags", params={"category": "特性"})
    assert listed.status_code == 200
    assert created_tag["id"] in {row["id"] for row in listed.json()["tags"]}

    updated = client.patch(
        f"/api/tags/{created_tag['id']}",
        json={"description": None, "category": None},
    )
    assert updated.status_code == 200
    assert updated.json()["description"] is None
    assert updated.json()["category"] is None

    deleted = client.delete(f"/api/tags/{created_tag['id']}")
    assert deleted.status_code == 204
    assert created_tag["id"] not in {
        row["id"] for row in client.get("/api/tags").json()["tags"]
    }


def test_merchant_path_migration_uses_name_and_keeps_previous_slug(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "merchant-path-migration.db"
    migration_url = f"sqlite:///{database_path.as_posix()}"
    config = Config(str(API_ROOT / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", migration_url)
    command.upgrade(config, "20260724_0026")

    migration_engine = create_engine(migration_url)
    metadata = MetaData()
    metadata.reflect(migration_engine)
    organization_id = uuid4()
    tenant_id = uuid4()
    organization_key = organization_id.hex
    tenant_key = tenant_id.hex
    now = datetime.now(UTC)
    with migration_engine.begin() as connection:
        connection.execute(
            metadata.tables["organizations"].insert(),
            {
                "id": organization_key,
                "code": "MIGRATION",
                "name": "Migration Organization",
                "status": "active",
                "created_at": now,
                "updated_at": now,
                "deleted_at": None,
            },
        )
        connection.execute(
            metadata.tables["tenants"].insert(),
            {
                "id": tenant_key,
                "organization_id": organization_key,
                "slug": "qingwan",
                "name": "澄湾选品",
                "default_locale": "zh-CN",
                "default_currency": "CNY",
                "timezone": "Asia/Shanghai",
                "status": "active",
                "created_at": now,
                "updated_at": now,
                "deleted_at": None,
            },
        )
        connection.execute(
            metadata.tables["tenant_public_profiles"].insert(),
            {
                "tenant_id": tenant_key,
                "slug": "qingwan",
                "description": None,
                "logo_url": None,
                "contact_email": None,
                "contact_phone": None,
                "publication_status": "PUBLISHED",
                "created_at": now,
                "updated_at": now,
                "deleted_at": None,
            },
        )
    migration_engine.dispose()

    command.upgrade(config, "20260724_0027")
    migrated_engine = create_engine(migration_url)
    migrated_metadata = MetaData()
    migrated_metadata.reflect(migrated_engine)
    with migrated_engine.connect() as connection:
        tenant = connection.execute(
            select(migrated_metadata.tables["tenants"]).where(
                migrated_metadata.tables["tenants"].c.id == tenant_key
            )
        ).mappings().one()
        profile = connection.execute(
            select(migrated_metadata.tables["tenant_public_profiles"]).where(
                migrated_metadata.tables["tenant_public_profiles"].c.tenant_id
                == tenant_key
            )
        ).mappings().one()
        assert tenant["slug"] == "澄湾选品"
        assert profile["slug"] == "澄湾选品"
        assert profile["legacy_slugs"] == ["qingwan"]
    migrated_engine.dispose()


def _add_tenant_member_with_role(
    *,
    role_code: str,
    display_name: str,
) -> tuple[UUID, UUID]:
    user_id = uuid4()
    membership_id = uuid4()
    with SessionLocal() as session:
        role = session.scalar(
            select(RoleRow).where(
                RoleRow.tenant_id == DEFAULT_TENANT_ID,
                RoleRow.code == role_code,
            )
        )
        assert role is not None
        session.add(
            UserRow(
                id=user_id,
                email_normalized=f"{user_id.hex}@access-control.test",
                display_name=display_name,
                identity_provider="local-bootstrap",
                identity_subject=str(user_id),
                status="active",
                is_platform_admin=False,
            )
        )
        session.add(
            MembershipRow(
                id=membership_id,
                tenant_id=DEFAULT_TENANT_ID,
                user_id=user_id,
                status="active",
            )
        )
        session.flush()
        session.add(
            MembershipRoleRow(
                tenant_id=DEFAULT_TENANT_ID,
                membership_id=membership_id,
                role_id=role.id,
                assigned_by_user_id=DEFAULT_OWNER_USER_ID,
            )
        )
        session.commit()
    return user_id, membership_id


def _local_access_token(test_client: TestClient, user_id: UUID) -> str:
    response = test_client.post(
        "/api/v1/auth/login",
        json={
            "provider": "local_fake",
            "authorization_code": f"fake:{user_id}",
            "code_verifier": "R" * 43,
            "redirect_uri": "http://127.0.0.1:5173/login/callback",
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    if not data["requires_tenant_selection"]:
        return data["access_token"]
    headers = {"Authorization": f"Bearer {data['access_token']}"}
    memberships = test_client.get(
        "/api/v1/auth/memberships", headers=headers
    )
    assert memberships.status_code == 200, memberships.text
    membership = next(
        row
        for row in memberships.json()
        if row["tenant_id"] == str(DEFAULT_TENANT_ID)
    )
    switched = test_client.post(
        "/api/v1/auth/tenant-context",
        headers={**headers, "X-CSRF-Token": data["csrf_token"]},
        json={"membership_id": membership["id"]},
    )
    assert switched.status_code == 200, switched.text
    return switched.json()["data"]["access_token"]


def test_customer_subaccount_is_restricted_and_orders_remain_owner_read_only(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """An owner creates a child; the child sees only its portal and own order trail."""

    suffix = uuid4().hex[:10]
    created = client.post(
        "/api/v1/customer-accounts",
        json={
            "display_name": f"Downstream Customer {suffix}",
            "login_identifier": f"customer-{suffix}",
            "password": f"Customer{suffix}9",
            "email": f"customer-{suffix}@subaccount.test",
        },
    )
    assert created.status_code == 201, created.text
    account = created.json()
    assert account["status"] == "active"
    assert account["login_count_30d"] == 0

    listing = client.get("/api/store/demo/skus", params={"page_size": 1})
    assert listing.status_code == 200, listing.text
    sku_id = listing.json()["items"][0]["id"]

    with monkeypatch.context() as auth_environment:
        auth_environment.setenv("AUTH_TEST_BYPASS", "false")
        with TestClient(app) as child_client:
            login = child_client.post(
                "/api/v1/auth/login",
                json={
                    "grant_type": "password",
                    "identifier": f"customer-{suffix}",
                    "password": f"Customer{suffix}9",
                },
            )
            assert login.status_code == 200, login.text
            token = login.json()["data"]["access_token"]
            assert login.json()["data"]["context"]["account_scope"] == "CUSTOMER_SUBACCOUNT"
            headers = {"Authorization": f"Bearer {token}"}

            portal = child_client.get("/api/v1/customer-portal/overview", headers=headers)
            assert portal.status_code == 200, portal.text
            assert portal.json()["display_name"] == f"Downstream Customer {suffix}"
            assert child_client.get("/api/v1/customer-accounts", headers=headers).status_code == 403

            submitted = child_client.post(
                "/api/store/demo/quotes",
                headers=headers,
                json={
                    "customer_name": f"Downstream Customer {suffix}",
                    "privacy_acknowledged": True,
                    "items": [{"sku_id": sku_id, "quantity": 1}],
                },
            )
            assert submitted.status_code == 201, submitted.text
            quote_id = submitted.json()["id"]
            own_orders = child_client.get("/api/v1/customer-portal/orders", headers=headers)
            assert own_orders.status_code == 200, own_orders.text
            assert [row["id"] for row in own_orders.json()] == [quote_id]

    owner_dashboard = client.get("/api/v1/customer-accounts")
    assert owner_dashboard.status_code == 200, owner_dashboard.text
    owner_account = next(
        row for row in owner_dashboard.json()["accounts"] if row["id"] == account["id"]
    )
    assert owner_account["login_count_30d"] >= 1
    assert owner_account["order_count"] == 1
    owner_orders = client.get(
        "/api/v1/customer-accounts/orders", params={"page": 1, "page_size": 100}
    )
    assert owner_orders.status_code == 200, owner_orders.text
    assert owner_orders.json()["total"] >= 1
    assert quote_id in {row["id"] for row in owner_orders.json()["items"]}

    suspended = client.patch(
        f"/api/v1/customer-accounts/{account['id']}/status",
        json={"status": "suspended"},
    )
    assert suspended.status_code == 200, suspended.text
    assert suspended.json()["status"] == "suspended"


def test_product_template_import_requires_edit_and_publish_permissions(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    sales_user_id, _ = _add_tenant_member_with_role(
        role_code="SALES",
        display_name="Template Import Sales Guard",
    )
    purchasing_user_id, _ = _add_tenant_member_with_role(
        role_code="PURCHASING",
        display_name="Template Import Purchasing Guard",
    )
    template_bytes = client.get("/api/v1/product-template.xlsx").content
    with SessionLocal() as session:
        before_count = session.scalar(select(func.count()).select_from(ImportJobRow))

    monkeypatch.setenv("AUTH_TEST_BYPASS", "false")
    with TestClient(app) as scoped_client:
        cases = (
            (sales_user_id, "product.edit"),
            (purchasing_user_id, "catalog.publish"),
        )
        for user_id, missing_permission in cases:
            token = _local_access_token(scoped_client, user_id)
            response = scoped_client.post(
                "/api/v1/imports",
                headers={"Authorization": f"Bearer {token}"},
                files={
                    "file": (
                        "商品模版.xlsx",
                        template_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={"source_type": "PRODUCT_TEMPLATE"},
            )
            assert response.status_code == 403, response.text
            assert response.json()["detail"] == {
                "code": "PERMISSION_DENIED",
                "message": f"Permission is required: {missing_permission}",
            }

    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ImportJobRow)) == before_count


def test_tenant_access_control_manages_custom_roles_and_isolates_members() -> None:
    _user_id, membership_id = _add_tenant_member_with_role(
        role_code="VIEWER",
        display_name="Access Control Target",
    )
    role_code = f"AUDITOR_{uuid4().hex[:8].upper()}"
    created = client.post(
        "/api/v1/access-control/roles",
        json={
            "code": role_code,
            "name": "报价审阅",
            "description": "只读报价与图册",
            "permission_codes": ["quotation.view"],
        },
    )
    assert created.status_code == 201, created.text
    custom_role = created.json()
    assert custom_role["is_system"] is False
    assert custom_role["permission_codes"] == ["quotation.view"]

    updated = client.patch(
        f"/api/v1/access-control/roles/{custom_role['id']}",
        json={"permission_codes": ["quotation.view", "catalog.view"]},
    )
    assert updated.status_code == 200, updated.text
    assert updated.json()["permission_codes"] == ["catalog.view", "quotation.view"]

    before = next(
        row
        for row in client.get("/api/v1/access-control/members").json()
        if row["id"] == str(membership_id)
    )
    assigned = client.put(
        f"/api/v1/access-control/members/{membership_id}/roles",
        json={"role_ids": [custom_role["id"]]},
    )
    assert assigned.status_code == 200, assigned.text
    assert [role["code"] for role in assigned.json()["roles"]] == [role_code]
    assert assigned.json()["permission_version"] == before["permission_version"] + 1

    owner_role = next(
        role
        for role in client.get("/api/v1/access-control/roles").json()
        if role["code"] == "OWNER"
    )
    immutable = client.patch(
        f"/api/v1/access-control/roles/{owner_role['id']}",
        json={"name": "Mutable Owner"},
    )
    assert immutable.status_code == 409
    assert immutable.json()["detail"]["code"] == "SYSTEM_ROLE_IMMUTABLE"

    organization_id = uuid4()
    other_tenant_id = uuid4()
    other_user_id = uuid4()
    other_membership_id = uuid4()
    with SessionLocal() as session:
        session.add(
            OrganizationRow(
                id=organization_id,
                code=f"ACL-{organization_id.hex[:8]}",
                name="Cross Tenant ACL",
            )
        )
        session.add(
            UserRow(
                id=other_user_id,
                email_normalized=f"{other_user_id.hex}@cross-acl.test",
                display_name="Cross Tenant Member",
                identity_provider="local-bootstrap",
                identity_subject=str(other_user_id),
                status="active",
            )
        )
        session.flush()
        session.add(
            TenantRow(
                id=other_tenant_id,
                organization_id=organization_id,
                slug=f"acl-{other_tenant_id.hex[:10]}",
                name="Cross Tenant ACL",
            )
        )
        session.flush()
        session.add(
            MembershipRow(
                id=other_membership_id,
                tenant_id=other_tenant_id,
                user_id=other_user_id,
                status="active",
            )
        )
        session.commit()

    cross_tenant = client.put(
        f"/api/v1/access-control/members/{other_membership_id}/roles",
        json={"role_ids": [custom_role["id"]]},
    )
    assert cross_tenant.status_code == 404
    assert cross_tenant.json()["detail"]["code"] == "MEMBERSHIP_NOT_FOUND"


def test_viewer_cannot_write_or_manage_access(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    viewer_user_id, _membership_id = _add_tenant_member_with_role(
        role_code="VIEWER",
        display_name="Read Only Viewer",
    )
    monkeypatch.setenv("AUTH_TEST_BYPASS", "false")
    with TestClient(app) as viewer_client:
        token = _local_access_token(viewer_client, viewer_user_id)
        headers = {"Authorization": f"Bearer {token}"}
        assert (
            viewer_client.get(
                "/api/v1/product-center/skus", headers=headers
            ).status_code
            == 200
        )
        denied_supplier = viewer_client.post(
            "/api/v1/supplier-profiles",
            headers=headers,
            json={
                "supplier_code": f"NO-{uuid4().hex[:8]}",
                "name": "Viewer Cannot Create",
                "category": "test",
                "country_code": "CN",
            },
        )
        denied_role = viewer_client.post(
            "/api/v1/access-control/roles",
            headers=headers,
            json={
                "code": f"NO_{uuid4().hex[:8].upper()}",
                "name": "No escalation",
                "permission_codes": ["product.view"],
            },
        )
    assert denied_supplier.status_code == 403
    assert denied_role.status_code == 403
    assert denied_role.json()["detail"]["code"] == "PERMISSION_DENIED"


def test_access_control_blocks_privilege_escalation_and_owner_assignment(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    manager_user_id = uuid4()
    manager_membership_id = uuid4()
    manager_role_id = uuid4()
    target_user_id, target_membership_id = _add_tenant_member_with_role(
        role_code="VIEWER",
        display_name="Escalation Target",
    )
    del target_user_id
    with SessionLocal() as session:
        manager_role = RoleRow(
            id=manager_role_id,
            tenant_id=DEFAULT_TENANT_ID,
            code=f"ROLE_MANAGER_{uuid4().hex[:6].upper()}",
            name="Delegated Role Manager",
            is_system=False,
            status="active",
        )
        session.add(
            UserRow(
                id=manager_user_id,
                email_normalized=f"{manager_user_id.hex}@delegated-manager.test",
                display_name="Delegated Manager",
                identity_provider="local-bootstrap",
                identity_subject=str(manager_user_id),
                status="active",
            )
        )
        session.add(
            MembershipRow(
                id=manager_membership_id,
                tenant_id=DEFAULT_TENANT_ID,
                user_id=manager_user_id,
                status="active",
            )
        )
        session.add(manager_role)
        session.flush()
        for code in ("system.user_manage", "system.role_manage"):
            permission = session.scalar(
                select(PermissionRow).where(PermissionRow.code == code)
            )
            assert permission is not None
            session.add(
                RolePermissionRow(
                    tenant_id=DEFAULT_TENANT_ID,
                    role_id=manager_role.id,
                    permission_id=permission.id,
                )
            )
        session.add(
            MembershipRoleRow(
                tenant_id=DEFAULT_TENANT_ID,
                membership_id=manager_membership_id,
                role_id=manager_role.id,
                assigned_by_user_id=DEFAULT_OWNER_USER_ID,
            )
        )
        session.commit()

    monkeypatch.setenv("AUTH_TEST_BYPASS", "false")
    with TestClient(app) as manager_client:
        manager_token = _local_access_token(manager_client, manager_user_id)
        escalation = manager_client.post(
            "/api/v1/access-control/roles",
            headers={"Authorization": f"Bearer {manager_token}"},
            json={
                "code": f"COST_{uuid4().hex[:8].upper()}",
                "name": "Forbidden Cost Role",
                "permission_codes": ["product.cost.write"],
            },
        )
        governance_lockout = manager_client.patch(
            f"/api/v1/access-control/roles/{manager_role_id}",
            headers={"Authorization": f"Bearer {manager_token}"},
            json={"permission_codes": ["system.role_manage"]},
        )
    assert escalation.status_code == 403
    assert escalation.json()["detail"]["code"] == "PRIVILEGE_ESCALATION_FORBIDDEN"
    assert governance_lockout.status_code == 409
    assert governance_lockout.json()["detail"]["code"] == "SELF_LOCKOUT_FORBIDDEN"

    admin_user_id, _admin_membership_id = _add_tenant_member_with_role(
        role_code="ADMIN",
        display_name="Tenant Admin But Not Owner",
    )
    with TestClient(app) as admin_client:
        admin_token = _local_access_token(admin_client, admin_user_id)
        owner_role = next(
            role
            for role in admin_client.get(
                "/api/v1/access-control/roles",
                headers={"Authorization": f"Bearer {admin_token}"},
            ).json()
            if role["code"] == "OWNER"
        )
        owner_escalation = admin_client.put(
            f"/api/v1/access-control/members/{target_membership_id}/roles",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={"role_ids": [owner_role["id"]]},
        )
    assert owner_escalation.status_code == 403
    assert owner_escalation.json()["detail"]["code"] == "OWNER_ASSIGNMENT_FORBIDDEN"


def test_access_control_prevents_self_lockout_and_last_owner_removal(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _second_owner_user_id, second_owner_membership_id = _add_tenant_member_with_role(
        role_code="OWNER",
        display_name="Second Tenant Owner",
    )
    roles = {
        role["code"]: role
        for role in client.get("/api/v1/access-control/roles").json()
    }
    self_lockout = client.put(
        f"/api/v1/access-control/members/{DEFAULT_MEMBERSHIP_ID}/roles",
        json={"role_ids": [roles["VIEWER"]["id"]]},
    )
    assert self_lockout.status_code == 409
    assert self_lockout.json()["detail"]["code"] == "SELF_LOCKOUT_FORBIDDEN"

    admin_user_id, _admin_membership_id = _add_tenant_member_with_role(
        role_code="ADMIN",
        display_name="Owner Removal Guard Admin",
    )
    with monkeypatch.context() as auth_environment:
        auth_environment.setenv("AUTH_TEST_BYPASS", "false")
        with TestClient(app) as admin_client:
            admin_token = _local_access_token(admin_client, admin_user_id)
            forbidden_owner_removal = admin_client.put(
                f"/api/v1/access-control/members/{second_owner_membership_id}/roles",
                headers={"Authorization": f"Bearer {admin_token}"},
                json={"role_ids": [roles["ADMIN"]["id"]]},
            )
    assert forbidden_owner_removal.status_code == 403
    assert (
        forbidden_owner_removal.json()["detail"]["code"]
        == "OWNER_ASSIGNMENT_FORBIDDEN"
    )

    downgrade_second = client.put(
        f"/api/v1/access-control/members/{second_owner_membership_id}/roles",
        json={"role_ids": [roles["VIEWER"]["id"]]},
    )
    assert downgrade_second.status_code == 200, downgrade_second.text

    remove_last_owner = client.put(
        f"/api/v1/access-control/members/{DEFAULT_MEMBERSHIP_ID}/roles",
        json={"role_ids": [roles["ADMIN"]["id"]]},
    )
    assert remove_last_owner.status_code == 409
    assert remove_last_owner.json()["detail"]["code"] == "LAST_OWNER_REQUIRED"


def test_member_role_change_invalidates_existing_access_token(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    viewer_user_id, viewer_membership_id = _add_tenant_member_with_role(
        role_code="VIEWER",
        display_name="Permission Version Viewer",
    )
    monkeypatch.setenv("AUTH_TEST_BYPASS", "false")
    with TestClient(app) as viewer_client, TestClient(app) as owner_client:
        viewer_token = _local_access_token(viewer_client, viewer_user_id)
        owner_token = _local_access_token(owner_client, DEFAULT_OWNER_USER_ID)
        owner_headers = {"Authorization": f"Bearer {owner_token}"}
        role_response = owner_client.get(
            "/api/v1/access-control/roles", headers=owner_headers
        )
        assert role_response.status_code == 200, role_response.text
        sales_role = next(
            role
            for role in role_response.json()
            if role["code"] == "SALES"
        )
        changed = owner_client.put(
            f"/api/v1/access-control/members/{viewer_membership_id}/roles",
            headers=owner_headers,
            json={"role_ids": [sales_role["id"]]},
        )
        assert changed.status_code == 200, changed.text
        stale = viewer_client.get(
            "/api/v1/me",
            headers={"Authorization": f"Bearer {viewer_token}"},
        )
    assert stale.status_code == 401
    assert stale.json()["detail"]["code"] == "AUTH_PERMISSION_STALE"


def test_inventory_purchase_sales_and_transfer_flow() -> None:
    warehouses_response = client.get("/api/v1/inventory/warehouses")
    assert warehouses_response.status_code == 200, warehouses_response.text
    default_warehouse = warehouses_response.json()[0]
    assert default_warehouse["is_default"] is True

    stock_response = client.get(
        "/api/v1/inventory/stocks",
        params={"warehouse_id": default_warehouse["id"], "q": "AQ-320S"},
    )
    assert stock_response.status_code == 200, stock_response.text
    initial_stock = stock_response.json()["items"][0]
    sku_id = initial_stock["sku_id"]
    initial_on_hand = Decimal(str(initial_stock["on_hand_quantity"]))
    initial_average_cost = Decimal(str(initial_stock["average_cost"]))

    adjustment_key = f"adjust-{uuid4()}"
    adjusted = client.post(
        "/api/v1/inventory/adjustments",
        json={
            "warehouse_id": default_warehouse["id"],
            "reason": "期初库存盘点",
            "idempotency_key": adjustment_key,
            "items": [
                {
                    "sku_id": sku_id,
                    "quantity_delta": 10,
                    "unit_cost": 5,
                }
            ],
        },
    )
    assert adjusted.status_code == 201, adjusted.text
    duplicate_adjustment = client.post(
        "/api/v1/inventory/adjustments",
        json={
            "warehouse_id": default_warehouse["id"],
            "reason": "重复请求不得再次入账",
            "idempotency_key": adjustment_key,
            "items": [
                {
                    "sku_id": sku_id,
                    "quantity_delta": 10,
                    "unit_cost": 5,
                }
            ],
        },
    )
    assert duplicate_adjustment.status_code == 201, duplicate_adjustment.text
    assert duplicate_adjustment.json()["id"] == adjusted.json()["id"]

    sales = client.post(
        "/api/v1/sales-orders",
        json={
            "customer_name": "库存流程测试客户",
            "warehouse_id": default_warehouse["id"],
            "currency": "CNY",
            "items": [{"sku_id": sku_id, "quantity": 4, "unit_price": 8}],
        },
    )
    assert sales.status_code == 201, sales.text
    sales_order = sales.json()
    assert sales_order["status"] == "DRAFT"

    confirmed_sale = client.post(
        f"/api/v1/sales-orders/{sales_order['id']}/confirm",
        json={"expected_version": sales_order["version"]},
    )
    assert confirmed_sale.status_code == 200, confirmed_sale.text
    confirmed_sales_order = confirmed_sale.json()
    assert confirmed_sales_order["status"] == "CONFIRMED"

    reserved_stock = client.get(
        "/api/v1/inventory/stocks",
        params={"warehouse_id": default_warehouse["id"], "q": "AQ-320S"},
    ).json()["items"][0]
    assert Decimal(str(reserved_stock["on_hand_quantity"])) == initial_on_hand + 10
    assert Decimal(str(reserved_stock["reserved_quantity"])) == 4

    sales_item = confirmed_sales_order["items"][0]
    shipment_key = f"shipment-{uuid4()}"
    shipped = client.post(
        f"/api/v1/sales-orders/{sales_order['id']}/ship",
        json={
            "expected_version": confirmed_sales_order["version"],
            "idempotency_key": shipment_key,
            "items": [{"order_item_id": sales_item["id"], "quantity": 2}],
        },
    )
    assert shipped.status_code == 200, shipped.text
    shipped_order = shipped.json()
    assert shipped_order["status"] == "PARTIALLY_SHIPPED"
    assert Decimal(str(shipped_order["items"][0]["shipped_quantity"])) == 2
    assert Decimal(str(shipped_order["items"][0]["reserved_quantity"])) == 2

    duplicate_shipment = client.post(
        f"/api/v1/sales-orders/{sales_order['id']}/ship",
        json={
            "expected_version": confirmed_sales_order["version"],
            "idempotency_key": shipment_key,
            "items": [{"order_item_id": sales_item["id"], "quantity": 2}],
        },
    )
    assert duplicate_shipment.status_code == 200, duplicate_shipment.text
    assert Decimal(
        str(duplicate_shipment.json()["items"][0]["shipped_quantity"])
    ) == 2

    cancelled_sale = client.post(
        f"/api/v1/sales-orders/{sales_order['id']}/cancel",
        json={
            "expected_version": shipped_order["version"],
            "reason": "客户取消未发数量",
        },
    )
    assert cancelled_sale.status_code == 200, cancelled_sale.text
    assert cancelled_sale.json()["status"] == "CANCELLED"

    after_sale_stock = client.get(
        "/api/v1/inventory/stocks",
        params={"warehouse_id": default_warehouse["id"], "q": "AQ-320S"},
    ).json()["items"][0]
    on_hand_after_sale = initial_on_hand + 8
    assert Decimal(str(after_sale_stock["on_hand_quantity"])) == on_hand_after_sale
    assert Decimal(str(after_sale_stock["reserved_quantity"])) == 0

    purchase = client.post(
        "/api/v1/purchases",
        json={
            "supplier_name": "库存流程测试供应商",
            "warehouse_id": default_warehouse["id"],
            "currency": "CNY",
            "items": [{"sku_id": sku_id, "quantity": 5, "unit_cost": 6}],
        },
    )
    assert purchase.status_code == 201, purchase.text
    purchase_order = purchase.json()
    confirmed_purchase = client.post(
        f"/api/v1/purchases/{purchase_order['id']}/confirm",
        json={"expected_version": purchase_order["version"]},
    )
    assert confirmed_purchase.status_code == 200, confirmed_purchase.text
    purchase_order = confirmed_purchase.json()
    purchase_item = purchase_order["items"][0]

    first_receipt_key = f"receipt-{uuid4()}"
    first_receipt = client.post(
        f"/api/v1/purchases/{purchase_order['id']}/receive",
        json={
            "expected_version": purchase_order["version"],
            "idempotency_key": first_receipt_key,
            "items": [{"order_item_id": purchase_item["id"], "quantity": 2}],
        },
    )
    assert first_receipt.status_code == 200, first_receipt.text
    partially_received = first_receipt.json()
    assert partially_received["status"] == "PARTIALLY_RECEIVED"

    duplicate_receipt = client.post(
        f"/api/v1/purchases/{purchase_order['id']}/receive",
        json={
            "expected_version": purchase_order["version"],
            "idempotency_key": first_receipt_key,
            "items": [{"order_item_id": purchase_item["id"], "quantity": 2}],
        },
    )
    assert duplicate_receipt.status_code == 200, duplicate_receipt.text
    assert Decimal(
        str(duplicate_receipt.json()["items"][0]["received_quantity"])
    ) == 2

    final_receipt = client.post(
        f"/api/v1/purchases/{purchase_order['id']}/receive",
        json={
            "expected_version": partially_received["version"],
            "idempotency_key": f"receipt-{uuid4()}",
            "items": [{"order_item_id": purchase_item["id"], "quantity": 3}],
        },
    )
    assert final_receipt.status_code == 200, final_receipt.text
    assert final_receipt.json()["status"] == "RECEIVED"

    after_purchase_stock = client.get(
        "/api/v1/inventory/stocks",
        params={"warehouse_id": default_warehouse["id"], "q": "AQ-320S"},
    ).json()["items"][0]
    expected_on_hand = on_hand_after_sale + 5
    assert Decimal(str(after_purchase_stock["on_hand_quantity"])) == expected_on_hand
    adjusted_cost = (
        initial_on_hand * initial_average_cost + Decimal("10") * Decimal("5")
    ) / (initial_on_hand + 10)
    expected_average_cost = (
        on_hand_after_sale * adjusted_cost + Decimal("5") * Decimal("6")
    ) / expected_on_hand
    assert Decimal(str(after_purchase_stock["average_cost"])) == expected_average_cost.quantize(
        Decimal("0.000001")
    )

    second_warehouse = client.post(
        "/api/v1/inventory/warehouses",
        json={
            "code": f"T{uuid4().hex[:7]}",
            "name": "库存流程测试分仓",
            "currency": "CNY",
        },
    )
    assert second_warehouse.status_code == 201, second_warehouse.text
    destination = second_warehouse.json()

    transfer = client.post(
        "/api/v1/inventory/transfers",
        json={
            "from_warehouse_id": default_warehouse["id"],
            "to_warehouse_id": destination["id"],
            "reason": "补充分仓库存",
            "idempotency_key": f"transfer-{uuid4()}",
            "items": [{"sku_id": sku_id, "quantity": 3}],
        },
    )
    assert transfer.status_code == 201, transfer.text
    assert transfer.json()["document_type"] == "STOCK_TRANSFER"

    destination_stock = client.get(
        "/api/v1/inventory/stocks",
        params={"warehouse_id": destination["id"], "q": "AQ-320S"},
    ).json()["items"][0]
    assert Decimal(str(destination_stock["on_hand_quantity"])) == 3

    excessive_adjustment = client.post(
        "/api/v1/inventory/adjustments",
        json={
            "warehouse_id": destination["id"],
            "reason": "测试库存不可为负",
            "items": [{"sku_id": sku_id, "quantity_delta": -4}],
        },
    )
    assert excessive_adjustment.status_code == 409
    assert (
        excessive_adjustment.json()["detail"]["code"]
        == "INSUFFICIENT_AVAILABLE_STOCK"
    )

    movements = client.get(
        "/api/v1/inventory/movements",
        params={"q": "AQ-320S", "page_size": 100},
    )
    assert movements.status_code == 200, movements.text
    movement_types = {row["movement_type"] for row in movements.json()["items"]}
    assert {
        "MANUAL_ADJUSTMENT",
        "SALES_RESERVATION",
        "SALES_SHIPMENT",
        "SALES_RELEASE",
        "PURCHASE_RECEIPT",
        "TRANSFER_OUT",
        "TRANSFER_IN",
    }.issubset(movement_types)
