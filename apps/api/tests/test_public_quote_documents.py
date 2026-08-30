from __future__ import annotations

from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook
from PIL import Image as PillowImage
from pydantic import ValidationError
from pypdf import PdfReader

from app.public_catalog_schemas import (
    PublicQuoteDocument,
    PublicQuoteDraftItemResponse,
    PublicQuoteDraftResponse,
    PublicQuoteDraftSettingsUpdate,
)
from app.quote_template_schemas import (
    QuoteExcelColumn,
    QuoteExcelTemplateRenderSpec,
    QuoteExcelTemplateUpdateRequest,
)
from app.services.public_quote_documents import (
    DEFAULT_QUOTE_HEADERS,
    render_default_quote_template_xlsx,
    render_public_quote_draft_pdf,
    render_public_quote_draft_xlsx,
)
from app.use_cases import public_catalog as public_catalog_use_cases


def _image_bytes() -> bytes:
    output = BytesIO()
    PillowImage.new("RGB", (120, 80), color=(45, 27, 105)).save(
        output,
        format="PNG",
    )
    return output.getvalue()


def test_quote_reuses_version_compatible_sku_translation_after_offer_metadata_change() -> None:
    """Changing a tag must not make the translated quote name fall back to Chinese."""

    tenant_id = uuid4()
    sku_id = uuid4()
    product_id = uuid4()
    row = (
        SimpleNamespace(tags=["新品"], display_tag="新品"),
        SimpleNamespace(
            id=sku_id,
            tenant_id=tenant_id,
            product_id=product_id,
            sku_code="SKU-001",
            name="大型犬牵引绳",
            option_values={"规格名称": "白色"},
            version=2,
        ),
        SimpleNamespace(
            id=product_id,
            tenant_id=tenant_id,
            name="大型犬牵引绳",
            description="适用大型犬",
            current_version=3,
        ),
        SimpleNamespace(path="宠物用品/牵引绳", name="牵引绳", code="LEASH"),
    )
    stored = SimpleNamespace(
        source_hash="0" * 64,
        source_category="宠物用品/牵引绳",
        name="Large Dog Leash",
        description="For large dogs",
        category="Pet Supplies/Leashes",
        tags=["Featured"],
        display_tag="Featured",
        product_version=3,
        sku_version=2,
    )

    translated = public_catalog_use_cases._stored_quote_sku_translation(stored, row)

    assert translated is not None
    assert translated.name == "Large Dog Leash"
    assert translated.description == "For large dogs"
    assert translated.category == "Pet Supplies/Leashes"
    # Tags belong to the changed offer metadata and must not be copied from the
    # stale translation row.
    assert translated.tags == ("新品",)
    assert translated.complete is False

    row[1].version = 3
    assert public_catalog_use_cases._stored_quote_sku_translation(stored, row) is None


def _document(
    *,
    template: QuoteExcelTemplateRenderSpec | None = None,
) -> PublicQuoteDocument:
    now = datetime(2026, 8, 1, 12, 0, tzinfo=UTC)
    item = PublicQuoteDraftItemResponse(
        id=uuid4(),
        sku_id=uuid4(),
        position=1,
        quantity=Decimal("40"),
        sku_code_snapshot="SKU-IMAGE-001",
        name_snapshot="带图片的测试商品",
        description_snapshot="用于验证默认和自定义报价模板。",
        specification_snapshot="紫色 / 大号",
        option_values_snapshot={
            "一箱个数": "20",
            "装箱尺寸": "50 × 40 × 30 cm",
            "毛重": "12.5 kg",
            "立方": "0.06 m³",
        },
        category_snapshot="测试分类",
        tags_snapshot=["图片", "物流"],
        image_url_snapshot="memory://product.png",
        unit_code_snapshot="件",
        currency_snapshot="USD",
        unit_price_snapshot=Decimal("2.50"),
        line_total=Decimal("100.00"),
        product_version=2,
        sku_version=3,
    )
    quote = PublicQuoteDraftResponse(
        id=uuid4(),
        tenant_id=uuid4(),
        quote_number="QD-20260801-0001",
        status="PENDING_CONFIRMATION",
        customer_name="Example Buyer",
        customer_company="Example Company",
        customer_email="buyer@example.test",
        customer_phone="+1 555 0100",
        notes="测试备注",
        currency="USD",
        subtotal=Decimal("100.00"),
        total=Decimal("100.00"),
        total_amount=Decimal("100.00"),
        valid_until=now + timedelta(days=7),
        created_at=now,
        updated_at=now,
        content_hash="a" * 64,
        items=[item],
    )
    return PublicQuoteDocument(
        tenant_name="示例商家",
        contact_email="sales@example.test",
        contact_phone="+86 10000",
        quote=quote,
        excel_template=template,
    )


def test_default_quote_xlsx_embeds_image_and_calculates_logistics_totals() -> None:
    content = render_public_quote_draft_xlsx(
        _document(),
        image_loader=lambda _url: _image_bytes(),
    )

    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["报价单"]
    header_row = next(
        row
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, 1).value == "序号"
    )
    assert tuple(cell.value for cell in sheet[header_row]) == DEFAULT_QUOTE_HEADERS
    item_row = header_row + 1
    total_row = item_row + 1
    assert sheet.cell(item_row, 3).value == "SKU-IMAGE-001"
    assert sheet.cell(item_row, 7).value == 20
    assert sheet.cell(item_row, 8).value == "50 × 40 × 30 cm"
    assert sheet.cell(item_row, 9).value == 12.5
    assert sheet.cell(item_row, 10).value == 0.06
    assert sheet.cell(item_row, 12).value == 100
    assert sheet.cell(item_row, 13).value == 0.12
    assert sheet.cell(item_row, 14).value == 25
    assert sheet.cell(total_row, 12).value == 100
    assert sheet.cell(total_row, 13).value == 0.12
    assert sheet.cell(total_row, 14).value == 25
    assert len(sheet._images) == 1
    assert all(
        cell.data_type != "f"
        for row in sheet.iter_rows()
        for cell in row
    )
    workbook.close()


def test_quote_pdf_embeds_item_thumbnail() -> None:
    content = render_public_quote_draft_pdf(
        _document(),
        image_loader=lambda _url: _image_bytes(),
    )

    reader = PdfReader(BytesIO(content))
    assert len(reader.pages) == 1
    page = reader.pages[0]
    xobjects = page.get("/Resources", {}).get("/XObject", {})
    assert any(
        obj.get_object().get("/Subtype") == "/Image"
        for obj in xobjects.values()
    )


def test_quote_pdf_settings_limit_visible_columns_to_five() -> None:
    accepted = PublicQuoteDraftSettingsUpdate(
        visible_columns=[
            "product_image",
            "product_name",
            "quantity",
            "unit_price",
            "line_total",
        ],
    )
    assert len(accepted.visible_columns or []) == 5

    with pytest.raises(ValidationError):
        PublicQuoteDraftSettingsUpdate(
            visible_columns=[
                "serial_number",
                "sku_code",
                "product_name",
                "quantity",
                "unit_price",
                "line_total",
            ],
        )


def test_quote_excel_template_only_maps_product_region_fields() -> None:
    accepted = QuoteExcelTemplateUpdateRequest(
        name="商品区域",
        column_mappings={"A": "product_image", "B": "product_name"},
    )
    assert accepted.column_mappings["B"] == "product_name"

    with pytest.raises(ValidationError):
        QuoteExcelTemplateUpdateRequest(
            name="错误的整单模板",
            column_mappings={"A": "quote_number"},
        )


def test_custom_quote_xlsx_can_embed_image_and_leave_unmapped_column_blank(
    tmp_path,
) -> None:
    template_path = tmp_path / "custom-quote.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "客户模板"
    headers = ["图片", "装箱数量", "客户自定义列", "总立方", "总毛重"]
    sheet.append(headers)
    sheet.append([None, "示例", "这段示例数据应被清空", None, None])
    workbook.save(template_path)
    workbook.close()

    columns = [
        QuoteExcelColumn(
            key=chr(64 + index),
            index=index,
            header=header,
            samples=[],
        )
        for index, header in enumerate(headers, start=1)
    ]
    spec = QuoteExcelTemplateRenderSpec(
        object_key="quotes/custom-quote.xlsx",
        sheet_name="客户模板",
        header_row=1,
        data_start_row=2,
        data_end_row=2,
        columns=columns,
        column_mappings={
            "A": "product_image",
            "B": "packing_quantity",
            "D": "total_volume",
            "E": "total_gross_weight",
        },
    )
    content = render_public_quote_draft_xlsx(
        _document(template=spec),
        template_path=template_path,
        image_loader=lambda _url: _image_bytes(),
    )

    rendered = load_workbook(BytesIO(content), data_only=False)
    rendered_sheet = rendered["报价单"]
    assert rendered_sheet["A1"].value == "报价单"
    assert rendered_sheet["B2"].value == "示例商家"
    assert rendered_sheet["B8"].value == 20
    assert rendered_sheet["C8"].value is None
    assert rendered_sheet["D8"].value == 0.12
    assert rendered_sheet["E8"].value == 25
    assert len(rendered_sheet._images) == 1
    rendered.close()


def test_downloadable_system_template_exposes_all_default_columns() -> None:
    workbook = load_workbook(
        BytesIO(render_default_quote_template_xlsx()),
        data_only=False,
    )
    sheet = workbook["商品明细模板"]
    header_row = next(
        row
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, 1).value == "序号"
    )
    assert tuple(cell.value for cell in sheet[header_row]) == DEFAULT_QUOTE_HEADERS
    assert "保留为空" in str(sheet.cell(header_row + 3, 2).value)
    assert not any(
        cell.value == "商家"
        for row in sheet.iter_rows()
        for cell in row
    )
    assert all(
        cell.data_type != "f"
        for row in sheet.iter_rows()
        for cell in row
    )
    workbook.close()
