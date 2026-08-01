from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.services.category_template_import import (
    CATEGORY_TEMPLATE_HEADERS,
    CATEGORY_TEMPLATE_PATH,
    CATEGORY_TEMPLATE_SHEET,
    CategoryTemplateValidationError,
    parse_category_template,
)


def _workbook_bytes(rows: list[list[object]], *, sheet_name: str = CATEGORY_TEMPLATE_SHEET) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(list(CATEGORY_TEMPLATE_HEADERS))
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_category_template_groups_repeated_primary_categories_in_file_order() -> None:
    parsed = parse_category_template(
        _workbook_bytes(
            [
                ["宠物用品", "宠物服饰"],
                ["宠物用品", "宠物饮水"],
                ["家居用品", None],
                ["宠物用品", "宠物服饰"],
                [None, None],
            ]
        )
    )

    assert parsed.processed_rows == 3
    assert parsed.duplicate_rows_ignored == 1
    assert parsed.blank_rows_ignored == 1
    assert [(group.primary_name, group.secondary_names) for group in parsed.groups] == [
        ("宠物用品", ("宠物服饰", "宠物饮水")),
        ("家居用品", ()),
    ]


def test_category_template_requires_primary_category_when_secondary_is_present() -> None:
    with pytest.raises(CategoryTemplateValidationError) as caught:
        parse_category_template(_workbook_bytes([[None, "宠物服饰"]]))

    assert caught.value.issues[0].row_number == 2
    assert caught.value.issues[0].code == "PRIMARY_CATEGORY_REQUIRED"


@pytest.mark.parametrize(
    ("rows", "expected_code"),
    [
        ([["宠物用品/宠物服饰", None]], "CATEGORY_NAME_CONTAINS_SEPARATOR"),
        ([["宠物用品", "=CONCAT(\"宠物\",\"服饰\")"]], "CATEGORY_TEMPLATE_FORMULA_NOT_ALLOWED"),
    ],
)
def test_category_template_rejects_hierarchy_separators_and_formulas(
    rows: list[list[object]],
    expected_code: str,
) -> None:
    with pytest.raises(CategoryTemplateValidationError) as caught:
        parse_category_template(_workbook_bytes(rows))

    assert caught.value.issues[0].code == expected_code


def test_category_template_requires_the_fixed_sheet_name() -> None:
    with pytest.raises(CategoryTemplateValidationError, match="缺少工作表"):
        parse_category_template(_workbook_bytes([["宠物用品", None]], sheet_name="Sheet1"))


def test_downloadable_category_template_has_the_expected_contract() -> None:
    assert CATEGORY_TEMPLATE_PATH.is_file()
    workbook = load_workbook(CATEGORY_TEMPLATE_PATH, read_only=False, data_only=False)
    try:
        assert workbook.sheetnames == [CATEGORY_TEMPLATE_SHEET, "填写说明"]
        sheet = workbook[CATEGORY_TEMPLATE_SHEET]
        assert tuple(sheet.cell(1, column).value for column in (1, 2)) == CATEGORY_TEMPLATE_HEADERS
        assert all(
            sheet.cell(row, column).value is None
            for row in range(2, sheet.max_row + 1)
            for column in (1, 2)
        )
    finally:
        workbook.close()
